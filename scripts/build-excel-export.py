#!/usr/bin/env python3
"""
build-excel-export.py — UNIVERSAL ForgeOS run -> multi-tab Excel review surface.

Turns any ``out/<run>/`` directory into ``out/<run>/dossier.xlsx``: a workbook that
REPLACES the HTML dossier as the review surface AND surfaces computational errors
via LIVE Excel formulas (edit a yellow input -> the whole chain recomputes; a
"Checks" tab goes RED where the engine's numbers do not reconcile).

USAGE
    .venv/bin/python scripts/build-excel-export.py out/ras-v26-verify
        argv[1] = run directory (defaults to out/ras-v26-verify if omitted)
        argv[2] = optional output path (defaults to <run>/dossier.xlsx)

DESIGN PRINCIPLES
  * UNIVERSAL: no ``if run == 'ras'`` logic. Everything is data-driven off
    state.json's documented data model. A class-specific value (e.g. the 1336
    per-tank inlet flow) is *discovered* by scanning the contract, never hardcoded.
  * READ-ONLY on engine code: this script imports nothing from the engine and
    modifies no engine file. It only reads out/<run>/*.json and *.png/*.svg.
  * LIVE where possible: STRUCTURED worked-calcs (those carrying inputs[]) become
    real Excel formulas referencing labelled yellow input cells, chained within a
    tool by exact value-equality. LEGACY calcs (no inputs[]) render as static text.

DATA MODEL (established by the feasibility pass — see module docstring map below)
  state.toolsUsedPage.tools[]                : list of 15+ tools, each
      .tool_id / .tool_name                  : identity
      .worked[]                              : the calcs. Two shapes:
        STRUCTURED {label, formula, substitution, inputs:[{symbol,value,unit}],
                    result:{value,unit}, assumptions}   (~67 calcs / 11 tools)
        LEGACY     {label, formula, substitution, result, result_unit}  (no inputs)
  state.orchestratorContract.quantities      : DICT name -> {value,unit,family,basis,
                                               scope,source,source_detail}
  state.orchestratorContract.closures        : engine's own balance closures
  state.requirementsBom                      : LIST of {tag,requirement,status,part,
                                               qty,unit_gbp,line_gbp,basis}
  state.costBasis                            : {lines:[...], rollup:{...}, methodology}
  out/<run>/quality-scorecard.json           : {floor,mean,allPass,sections:[{name,
                                               score,defects}], iteration}
  out/<run>/parts-ledger.json                : {grand_total_gbp, coverage_by_drawing,
                                               connection_coverage, ...}

FORMULA -> EXCEL TRANSLATION (verified with the `formulas` lib in the POC)
    x -> *   |   ^ stays ^   |   log10()/pi survive   |   parentheses survive
    GOTCHA: any cell whose *text* starts with '=' is parsed as a formula. All
    textual / display columns are written WITHOUT a leading '='.
"""

from __future__ import annotations

import json
import os
import re
import subprocess
import sys
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

# The SHARED deterministic-check library — the SAME pure-arithmetic checks the
# standalone CLI (scripts/deterministic-checks.py) runs, so the workbook's Checks
# tab and the instant CLI can never diverge. Imported by absolute path so the
# exporter works regardless of the caller's cwd.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import deterministic_checks_lib as dcl  # noqa: E402

# The DETERMINISTIC per-tab self-audit + ship gate (scripts/lib/dossier_audit.py). It is the
# SHIP GATE: the dossier is not "validated" unless its scorecard is clean. Imported by absolute
# path (scripts/lib is the sibling dir) so the exporter works regardless of the caller's cwd.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib"))
from dossier_audit import audit_dossier, tab_scores, tab_scorecard_summary  # noqa: E402
from dossier_repair import repair_dossier  # noqa: E402

# ----------------------------------------------------------------------------
# STYLES — light theme only (Tristan: light mode, never dark)
# ----------------------------------------------------------------------------
FILL_HEADER = PatternFill("solid", fgColor="1F3A5F")      # deep navy header band
FILL_SUB = PatternFill("solid", fgColor="DCE6F1")          # pale blue sub-header
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")        # YELLOW = editable input
FILL_RESULT = PatternFill("solid", fgColor="E2EFDA")       # pale green = live result
FILL_CONST = PatternFill("solid", fgColor="FCE4D6")        # peach = shared constant
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")         # green pass
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")         # red fail
FILL_ADVISORY = PatternFill("solid", fgColor="FFF2CC")     # amber = advisory (non-gating LLM)
FILL_LEGACY = PatternFill("solid", fgColor="F2F2F2")       # grey = static legacy calc
FILL_TITLE = PatternFill("solid", fgColor="2E5A88")

FONT_TITLE = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUB = Font(name="Calibri", size=11, bold=True, color="1F3A5F")
FONT_PASS = Font(bold=True, color="006100")
FONT_FAIL = Font(bold=True, color="9C0006")
FONT_ADVISORY = Font(bold=True, color="9C6500")            # amber text = advisory
FONT_NOTE = Font(italic=True, size=9, color="666666")
FONT_MONO = Font(name="Menlo", size=10)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
LEFT_TOP = Alignment(horizontal="left", vertical="top")


# ============================================================================
# Small generic helpers
# ============================================================================
def load_json(path: str) -> Optional[Any]:
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r") as fh:
            return json.load(fh)
    except Exception as exc:  # noqa: BLE001 — never let one bad file kill the export
        print(f"  ! could not parse {path}: {exc}")
        return None


def git_short_sha() -> str:
    """READ-only git: short SHA for provenance. Never mutate the repo."""
    try:
        out = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True, text=True, timeout=10,
        )
        return out.stdout.strip() or "unknown"
    except Exception:  # noqa: BLE001
        return "unknown"


def num(v: Any) -> Optional[float]:
    """Coerce a value (possibly a display string like '13,360' or '£8.15 M') to float."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("£", "").replace("$", "")
        m = re.search(r"-?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s)
        if m:
            try:
                return float(m.group(0))
            except ValueError:
                return None
    return None


def qval(quantities: Dict[str, Any], key: str) -> Optional[float]:
    """Numeric value of a contract quantity (handles {value:..} or bare scalar)."""
    if key not in quantities:
        return None
    v = quantities[key]
    return num(v.get("value") if isinstance(v, dict) else v)


def qunit(quantities: Dict[str, Any], key: str) -> str:
    v = quantities.get(key)
    return (v.get("unit", "") if isinstance(v, dict) else "") or ""


def set_widths(ws: Worksheet, widths: Dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# Per-tab deterministic scorecard (Tristan 2026-06-26): set in build() from dossier_audit.tab_scores,
# keyed by Excel sheet name. title_row stamps each tab's score banner from this — every tab shows its
# deterministic quality vs the ≥8 floor, on the tab.
_TAB_SCORES: dict = {}
_RUN_DIR: str = ""           # set in build() so the banner can read parts-ledger coverage
_COV_CACHE: dict = {}


def _ledger_coverage(run_dir: str) -> dict:
    if run_dir in _COV_CACHE:
        return _COV_CACHE[run_dir]
    cov = {}
    try:
        d = load_json(os.path.join(run_dir, "parts-ledger.json")) or {}
        cov = d.get("coverage_by_drawing") or {}
    except Exception:  # noqa: BLE001
        cov = {}
    _COV_CACHE[run_dir] = cov
    return cov


def _aux_tab_score(title: str, run_dir: str):
    """A deterministic quality score for a DRAWING / RENDER / META sheet that the per-tab scorecard
    (the 16 data tabs) doesn't cover — so EVERY sheet shows a quality number (Tristan 2026-06-27).
    Drawings score from parts-ledger drawing coverage (an EMPTY P&ID → 0); renders use coverage as a
    proxy + flag that object-level visual quality is an open check (no fake clean PASS); ⚠ Checks uses
    the invariant pass-rate; ⚠ Audit the ship verdict. Pure navigation tabs (Contents) return None."""
    t = (title or "").lower()
    cov = _ledger_coverage(run_dir)

    def _cov(key: str, label: str, advisory: str = ""):
        c = cov.get(key)
        if not isinstance(c, dict):
            return None
        pct = c.get("pct")
        if not isinstance(pct, (int, float)):
            return None
        sc = max(0, min(10, round(pct / 10)))
        st = "PASS" if sc >= 8 else "FAIL"
        iss = []
        if advisory:
            iss.append(advisory)
        iss.append(f"{label} part coverage {c.get('present')}/{c.get('expected')} ({pct:.0f}%)"
                   + ("" if sc >= 8 else " — under-covered: the drawing must render its parts so their tags match the BoM"))
        return {"score": sc, "target": 8, "status": st, "issues": iss,
                "fix": "drawing generator must emit every principal part's tag so parts_ledger coverage ≥ 80%"}

    if "p&id" in t or t == "pid":
        return _cov("pid", "P&ID")
    if "block flow" in t or "bfd" in t:
        return _cov("block-flow-diagram", "Block-flow")
    if "general arrangement" in t or t.startswith("ga "):
        return _cov("general-arrangement", "GA")
    if "single-line" in t or "single line" in t:
        return _cov("single-line-diagram", "Single-line")
    if "isometric" in t:
        return _cov("isometric-index", "Isometric")
    if "render" in t or "interior layout" in t or "building exterior" in t:
        return _cov("blender", "Render",
                    advisory="ADVISORY: object-level visual quality (sizing / scatter / GA-vs-render consistency) is an OPEN check — coverage proxy only")
    if "checks" in t:  # ⚠ Checks — deterministic-invariant pass rate
        try:
            import deterministic_checks_lib as _dcl
            _chk = _dcl.run_all_checks(run_dir, None)
            _n = len(_chk)
            _nf = len([c for c in _chk if str(getattr(c, "status", "")).upper() == "FAIL"])
            if _n:
                # ANY failing invariant means this tab is NOT clean — never round a handful of
                # fails away into a green 10 (the fake-8). 0 fails → 10; else a hard FAIL.
                _sc = 10 if _nf == 0 else max(0, 8 - 2 * _nf)
                return {"score": _sc, "target": 8, "status": "PASS" if _sc >= 8 else "FAIL",
                        "issues": [f"{_nf} of {_n} deterministic invariants FAIL — this tab cannot be a clean 10 over failures"] if _nf else [f"{_n}/{_n} invariants pass"],
                        "fix": "fix each failing invariant at source"}
        except Exception:  # noqa: BLE001
            pass
        return None
    if "audit" in t:  # ⚠ Audit — the dossier ship verdict = does every scored tab reach ≥8?
        _scored = [v.get("score") for v in _TAB_SCORES.values() if isinstance(v.get("score"), (int, float))]
        if _scored:
            _worst = min(_scored)
            _ok = _worst >= 8
            return {"score": 8 if _ok else max(2, _worst), "target": 8, "status": "PASS" if _ok else "FAIL",
                    "issues": [] if _ok else [f"the dossier does NOT ship — the weakest tab scores {_worst}/10; every tab must reach ≥8"],
                    "fix": "resolve the failing tabs (see the per-tab punch-list / ⚠ Audit findings)"}
        return None
    if "hvac" in t:
        # a drawing we can't deterministically score yet → ADVISORY, never a fake PASS.
        return {"score": None, "target": 8, "status": "UNSCORED",
                "issues": ["no deterministic quality check covers the HVAC drawing yet (open coverage gap)"],
                "fix": "add a deterministic check for the HVAC drawing (duct sizing / placement / class-appropriateness)"}
    return None  # Contents / ⭐ Scorecard navigation tabs


def _tab_quality_banner(title: str):
    """The quality-banner spec ({text, fill, height}) for a sheet, or None if the tab isn't scored."""
    v = _TAB_SCORES.get(title)
    if not isinstance(v, dict):
        v = _aux_tab_score(title, _RUN_DIR)   # drawing / render / meta sheets → on-the-fly score
    if not isinstance(v, dict):
        return None
    tgt = v.get("target", 8)
    status = v.get("status")
    if status == "UNSCORED":
        return {
            "text": f"⬤ TAB QUALITY — UNSCORED: no deterministic check covers this tab yet, so it cannot be certified ≥{tgt}. {v.get('fix', '')}",
            "fill": FILL_ADVISORY, "height": 30,
        }
    sc = v.get("score")
    issues = v.get("issues") or []
    tail = (f"  →  {issues[0]}" if (status == "FAIL" and issues) else "")
    return {
        "text": f"⬤ TAB QUALITY: {sc}/10   ·   target ≥{tgt}   ·   {status}{tail}",
        "fill": FILL_PASS if status == "PASS" else FILL_FAIL,
        "height": 30 if tail else 18,
    }


def title_row(ws: Worksheet, text: str, span: int, subtitle: str = "") -> int:
    """Write a full-width title band (+ a deterministic per-tab quality banner); return next free row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26
    nxt = 2
    # Quality banner — every scored tab shows its deterministic score vs the ≥8 floor, ON the tab.
    qb = _tab_quality_banner(ws.title)
    if qb is not None:
        ws.merge_cells(start_row=nxt, start_column=1, end_row=nxt, end_column=span)
        bc = ws.cell(nxt, 1, qb["text"])
        bc.fill = qb["fill"]
        bc.font = FONT_SUB
        bc.alignment = LEFT_TOP
        ws.row_dimensions[nxt].height = qb["height"]
        nxt += 1
    if subtitle:
        ws.merge_cells(start_row=nxt, start_column=1, end_row=nxt, end_column=span)
        s = ws.cell(nxt, 1, subtitle)
        s.font = FONT_NOTE
        s.alignment = LEFT_TOP
        # Grow the merged subtitle row so a 2-3 line purpose note never clips
        # (Excel does NOT auto-grow a MERGED wrapped cell). Estimate lines from
        # the text length vs the merged width (~5.5 chars per column-width unit).
        approx_cols_wide = max(1, sum(_col_width(ws, ci) for ci in range(1, span + 1)))
        chars_per_line = max(40, approx_cols_wide * 1.05)
        lines = max(1, int(_math.ceil(len(subtitle) / chars_per_line)))
        ws.row_dimensions[nxt].height = 15 + 13 * min(lines, 4)
        nxt += 1
    return nxt + 1  # leave a blank spacer row


def _col_width(ws: Worksheet, col_idx: int) -> float:
    """Best-effort current width of a column (default 8.43 if unset)."""
    try:
        from openpyxl.utils import get_column_letter
        cd = ws.column_dimensions.get(get_column_letter(col_idx))
        if cd is not None and cd.width:
            return float(cd.width)
    except Exception:  # noqa: BLE001
        pass
    return 8.43


def header(ws: Worksheet, row: int, cols: List[str]) -> None:
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row, i, name)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)


def sub_banner(ws: Worksheet, row: int, text: str, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    # A deliberate "=..." live-formula heading (e.g. a Spec-sheet banner referencing the
    # master tag) is written raw; everything else is defanged for display.
    val = text if (isinstance(text, str) and text.startswith("=")) else clean_cell(text)
    c = ws.cell(row, 1, val)
    c.font = FONT_SUB
    c.fill = FILL_SUB
    c.alignment = LEFT_TOP


def unverified_banner(ws: Worksheet, row: int, span: int, text: str) -> int:
    """A red, wrapped warning band spanning `span` columns at `row` — used to flag an
    economics tab as presenting an UNVERIFIED model (no real sale price derivable for
    this class). Returns the next free row. Two rows tall for the wrapped message."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=span)
    c = ws.cell(row, 1, clean_cell(text))
    c.fill = FILL_FAIL
    c.font = FONT_FAIL
    c.alignment = WRAP_TOP
    return row + 2


# Number formats (#37) — kill General/scientific. Excel display masks only;
# they never change the stored value, so they are safe on live-formula cells.
FMT_GBP = "£#,##0"          # money, no decimals
FMT_GBP2 = "£#,##0.00"      # money with pence (small unit prices)
FMT_INT = "#,##0"           # counts / integers with thousands separators
FMT_DEC1 = "#,##0.0"        # one decimal (velocities, %drop, ratings)
FMT_DEC2 = "#,##0.00"       # two decimals (ratios, m/s, areas)
FMT_NUM = "#,##0.##"        # thousands sep, drops trailing zeros (62 -> "62", 2.58 -> "2.58")

# CONTENTS hyperlink target / back-link constant.
CONTENTS_SHEET = "Contents"
FONT_LINK = Font(name="Calibri", size=11, color="1F3A5F", underline="single", bold=True)

# One-line descriptions for the Contents index (#26). Image/module tabs fall
# back to _default_desc(); these cover the data tabs by their exact sheet name.
_TAB_DESCRIPTIONS: Dict[str, str] = {
    "Overview": "Quality scorecard, headline metrics & run provenance.",
    "Brief": "The original client brief and the engine's enhanced, structured interpretation that drives the design.",
    "⚠ Checks": "Live arithmetic invariants (== the CLI verifier; RED = numbers don't reconcile), then Brief-compliance (target vs achieved) and Tool-provenance (USED/STALE/ORPHANED) sections.",
    "Part names": "The master parts list — every part named once; every other tab references these cells.",
    "Connection trace": "Which part connects to what — live input/output cell-references with completeness & integrity counts.",
    "Quantities": "Every sized contract quantity with family, basis & source.",
    "Calculations": "Worked calcs grouped by tool — live Excel formulas where structured.",
    "Bill of Materials (Ledger)": "THE BILL — what to buy (tag · item · qty · unit £ · live Σ line £), with two collapsible column-groups: 'Cost basis' (how each £ was derived) and 'Engineering spec' (why each principal is this size).",
    "Cost waterfall": "BoM → assembly → factory COGS → install → installed ASP (live running totals).",
    "Inputs & Assumptions": "Editable yellow drivers (price/feed/energy/labour/capex) — the economics model inputs.",
    "Financial model": "The whole commercial model on one sheet: base-case Economics (revenue / opex / EBITDA / NPV / IRR + charts), the scale sweep + Low/Central/High price sensitivity, and the Investment-analysis sweet-spot finder — all live off the Inputs tab.",
    "Panel schedule": "Electrical panel / load schedule as a real sortable table.",
    "Process schedules": "Process line list, valve list & instrument index — three sortable sections cross-referenced to the P&ID.",
    "Line & velocity": "Every sized run with velocity / volt-drop & within-spec flagging.",
    "Glossary": "Plain-English meaning of every abbreviation (DN, ISA tags, FC/FO, status codes, units).",
    "Risk & Regulatory": "Live hazard & risk register (physics critic, gate flags, cost, equipment) + the compliance-gate verdict and statutory duties, on one sheet.",
    "Assembly sequence": "Order-of-works: civils → tankage → mechanical → pipework → electrical → I&C → commissioning.",
}

# A whitelist of CONTROL-CHARS Excel rejects inside a worksheet cell string —
# md tables occasionally carry stray control bytes; strip them so .save never throws.
_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Formula-injection defang (CWE-1236): a worksheet cell whose visible text starts with
# one of these is interpreted as a FORMULA by Excel/LibreOffice when the file is opened.
# Display text from the brief / LLM / md-tables is externally-sourced, so a stray
# "=cmd|..." / "+.." / "-.." / "@.." could execute on the customer's machine. We prefix
# such display strings with a zero-width space — they render identically but are inert
# text. Deliberate live formulas are written directly via cell.value = "=..." and NEVER
# pass through clean_cell (verified: no formula write routes through it).
_FORMULA_TRIGGERS = ("=", "+", "-", "@")


def clean_cell(v: Any) -> Any:
    """Make any value safe + tidy for a worksheet DISPLAY cell. Strings get control
    chars stripped, are trimmed, and any leading formula-trigger is defanged with a
    zero-width space (CWE-1236). Non-strings pass through untouched (numbers stay
    numbers). NEVER call this on a deliberate "=..." live-formula string."""
    if isinstance(v, str):
        s = _CTRL.sub("", v).strip().replace("`", "")   # strip markdown code-ticks (Tristan: `201-PR` → 201-PR)
        if s and s[0] in _FORMULA_TRIGGERS:
            s = "​" + s
        return s
    return v


def back_link(ws: Worksheet, span: int) -> None:
    """Write a '↑ Contents' internal hyperlink at the top-right of a data tab —
    column ``span+1`` (immediately right of the merged title band, so it never
    collides with the title merge). UNIVERSAL: called on every non-image tab."""
    col = span + 1
    c = ws.cell(1, col, "↑ Contents")
    # quote the sheet name so spaces/symbols in CONTENTS_SHEET are always valid
    c.hyperlink = f"#'{CONTENTS_SHEET}'!A1"
    c.font = FONT_LINK
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = 13


def apply_col_formats(ws: Worksheet, first_row: int, fmt_by_col: Dict[int, str],
                      last_row: Optional[int] = None) -> None:
    """Apply a number format to numeric cells in the given columns from
    ``first_row`` down. Only touches cells whose stored value is a number OR a
    live formula string (formulas resolve to numbers, so a money/decimal mask is
    correct + kills the General default). Text cells are left alone. ROBUST:
    a missing/empty column is silently skipped."""
    lr = last_row if last_row is not None else ws.max_row
    for col, fmt in fmt_by_col.items():
        letter = get_column_letter(col)
        for row in range(first_row, lr + 1):
            cell = ws[f"{letter}{row}"]
            v = cell.value
            if v is None:
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.number_format = fmt
            elif isinstance(v, str) and v.startswith("="):
                cell.number_format = fmt


# ============================================================================
# FORMULA TRANSLATION:  engine symbolic ASCII  ->  live Excel formula
# ============================================================================
# Functions the engine emits that Excel understands as-is.
_FUNC_OK = {"log10", "log", "ln", "exp", "sqrt", "abs", "sin", "cos", "tan", "max", "min"}


def formula_to_excel(rhs: str, symbol_cell: Dict[str, str]) -> Optional[str]:
    """
    Translate the right-hand side of a worked-calc formula into a live Excel
    formula string (WITHOUT the leading '='), substituting each input symbol
    with its yellow cell reference.

    rhs            : e.g. "rho x g x (Q_m3h / 3600) x H_total"
    symbol_cell    : {symbol_name -> cell coordinate}, e.g. {"rho":"$B$5", "Q_m3h":"C12"}

    Returns None if any symbol in the expression has no cell mapping (we then
    fall back to static text — never emit a half-bound, wrong formula).
    """
    if not rhs:
        return None
    expr = rhs
    # 1) operator translation
    expr = expr.replace("×", "*").replace("·", "*")
    expr = re.sub(r"(?<=[\w\)\s])x(?=[\s\(])", "*", expr)  # ' x ' multiplication
    expr = expr.replace(" x ", " * ")
    # ln -> LN, exp -> EXP etc. are fine; Excel uses ^ for power already.
    expr = expr.replace("π", "PI()")
    # standalone 'pi' token -> PI()
    expr = re.sub(r"\bpi\b", "PI()", expr)

    # 2) collect identifier-like tokens, decide which are symbols vs functions
    #    identifier = letters/digits/underscore, must start with a letter or underscore.
    tokens = set(re.findall(r"[A-Za-z_][A-Za-z0-9_]*", expr))
    # which tokens are immediately followed by '(' -> they are function calls
    func_tokens = set(re.findall(r"([A-Za-z_][A-Za-z0-9_]*)\s*\(", expr))
    # A function token Excel does NOT have (e.g. ceil_to_standard — a tool-internal
    # helper that snaps a value to the next IEC standard rating) cannot become a live
    # Excel formula. Emit None so the caller renders the STATIC computed value, never a
    # broken "=ceil_to_standard(...)" that shows #NAME?. (PI/LOG10/SQRT/... stay live.)
    _excel_funcs = _FUNC_OK | {"pi", "ceiling", "floor", "round", "power", "sum"}
    if any(fn.lower() not in _excel_funcs for fn in func_tokens):
        return None

    replacements: List[Tuple[str, str]] = []
    for tok in tokens:
        low = tok.lower()
        if tok in func_tokens or low in _FUNC_OK:
            continue  # it's a function call, leave it
        if low in ("pi", "e"):
            continue
        if tok not in symbol_cell:
            # An unbound symbol -> cannot make a faithful live formula.
            return None
        replacements.append((tok, symbol_cell[tok]))

    # 3) substitute longest-symbol-first so 'H_total' isn't clobbered by 'H'
    for tok, cell in sorted(replacements, key=lambda kv: -len(kv[0])):
        expr = re.sub(rf"(?<![A-Za-z0-9_]){re.escape(tok)}(?![A-Za-z0-9_])", cell, expr)

    # 4) uppercase the known math functions for Excel
    for fn in _FUNC_OK:
        expr = re.sub(rf"\b{fn}\b", fn.upper(), expr, flags=re.IGNORECASE)
    expr = expr.replace("LOG10", "LOG10").replace("LN(", "LN(")
    return expr


def rhs_of(formula: str) -> str:
    """Return the right-hand side of 'lhs = rhs' (or the whole thing if no '=')."""
    if formula and "=" in formula:
        return formula.split("=", 1)[1].strip()
    return (formula or "").strip()


def lhs_symbol(formula: str) -> str:
    """Return the OUTPUT symbol a calc produces — the left-hand side of 'lhs = rhs',
    normalised to the bare identifier. Used to key within-tool chaining by IDENTITY
    (bug #19: chaining by numeric value grabbed the wrong producing cell when two
    calcs produced the same number, e.g. q_wall and q_roof both = 4920)."""
    if not formula or "=" not in formula:
        return ""
    lhs = formula.split("=", 1)[0].strip()
    m = re.match(r"[A-Za-z_][A-Za-z0-9_]*", lhs)
    return m.group(0) if m else ""


# ============================================================================
# COST BREAKDOWN BY CATEGORY — deterministic "where does my money go?" (Tristan 2026-06-24)
# ============================================================================
# A founder opening a 30-tab BoM is overwhelmed; a founder who sees "62% of capex is vessels +
# heat exchangers, 11% pumps" gets value in five seconds. Deterministic: reads parts-ledger.json
# (the engine's own classified equipment[] + connections[]), groups by a founder-friendly category,
# sums line_gbp, returns sorted [(category, gbp, pct, n_lines)]. UNIVERSAL — name-based classifier,
# no per-class table. Instruments/Valves are tested BEFORE Vessels so "reactor pH probe" classifies
# as an instrument, not a vessel (the same shared-keyword trap fixed in bom-cost-grounding.ts).
_COST_CAT_RULES = [
    (r'reboil|condenser|\bcooler\b|exchanger|\bhx\b|chiller|economiser|heat[- ]?recovery|\bheater\b', 'Heat exchangers & thermal'),
    (r'dryer|kiln|calciner|hot[- ]?air', 'Drying & thermal'),
    (r'centrifuge|filter press|cyclone|decanter|clarifier|belt filter|\bscreen\b|membrane|\bfilter\b', 'Filtration & separation'),
    (r'bagging|packaging|palletis|conveyor|\bfeeder\b|hopper|\bsilo\b|\bsack\b|material handling', 'Material handling & packaging'),
    (r'\bprobe\b|\bsensor\b|\bmeter\b|\bgauge\b|transmitter|analy[sz]er|detector|thermowell|sight\s?glass', 'Instruments'),
    (r'\bvalve\b|solenoid|\bactuator\b|\bdamper\b', 'Valves'),
    (r'reactor|crystallis|\bvessel\b|\btank\b|\bcolumn\b|\btower\b|absorber|stripper|\bdrum\b|reservoir|receiver|launder|\bsump\b', 'Vessels, reactors & columns'),
    (r'\bpump\b|blower|compressor|\bfan\b|agitator|mixer|skimmer|aerat', 'Pumps, blowers & compressors'),
    (r'\bmcc\b|motor control|switchgear|transformer|\bpanel\b|busbar|breaker|\brelay\b|\bups\b|electrical|\bcable\b|generator|genset', 'Electrical & power'),
    (r'\bplc\b|scada|control|\bi/o\b|\bhmi\b|gateway', 'Control system'),
    (r'\bskid\b|\bframe\b|structure|foundation|plinth|containment|\bbund\b|platform|walkway', 'Structure & containment'),
]
_COST_TYPE_FALLBACK = {
    'vessel': 'Vessels, reactors & columns', 'rotating': 'Pumps, blowers & compressors',
    'exchanger': 'Heat exchangers & thermal', 'separator': 'Filtration & separation',
    'instrument': 'Instruments', 'valve': 'Valves', 'electrical': 'Electrical & power',
    'control': 'Control system', 'structural': 'Structure & containment',
}


def _cost_category(name: str, tag: str, etype: str) -> str:
    blob = f"{name} {tag}".lower()
    for rx, cat in _COST_CAT_RULES:
        if re.search(rx, blob):
            return cat
    return _COST_TYPE_FALLBACK.get((etype or '').lower(), 'Balance of plant')


# Universal noun-keyed equipment-category map (2026-06-25 fix). Keyed on NOUNS that appear in
# the requirement/part text of ANY product class — no per-class hardcoding. Rules are matched in
# order; first hit wins, so the most specific principal nouns (cells, racks, power conversion)
# precede the generic connection/enclosure buckets. The catch-all is "Other equipment".
_EQUIP_CAT_RULES = [
    (r'\bcell\b|\bcells\b|\bmodule\b|\bmodules\b|\bbattery pack\b|\bprismatic\b|\bpouch\b', 'Battery cells & modules'),
    (r'\brack\b|\bframe\b|\bbusbar\b|\bbus[- ]?bar\b|\bstructure\b|\bchassis\b|\bskid\b|\bplinth\b|\bsupport\b', 'Racks & structure'),
    (r'\binverter\b|\bpcs\b|\btransformer\b|\bswitchgear\b|\bbreaker\b|\bcontactor\b|\brectifier\b|\bconverter\b|\bmcc\b|\bgenset\b|\bgenerator\b', 'Power conversion & electrical'),
    (r'\bpump\b|\bchiller\b|\bcooling\b|\bcoolant\b|\bfan\b|\bhvac\b|\bradiator\b|\bcompressor\b|\bcondenser\b|\bevaporator\b|cold[- ]?plate', 'Thermal management'),
    (r'\bbms\b|\bsensor\b|\bcontroller\b|\bmonitor\b|\bplc\b|\bscada\b|\bhmi\b|\bthermistor\b|\bgauge\b|\bmeter\b|\btransmitter\b|\bprobe\b|\binstrument\b', 'Controls & instrumentation'),
    (r'\bcable\b|\bcabling\b|\bwire\b|\bwiring\b|\bharness\b|\bloom\b|\bconduit\b|\blug\b|\bterminal\b', 'Cabling & power runs'),
    (r'\bpipe\b|\bpipework\b|\bvalve\b|\bfitting\b|\bmanifold\b|\bhose\b|\bflange\b', 'Pipework'),
    (r'\bcontainer\b|\benclosure\b|\bdoor\b|\bcabinet\b|\bhousing\b|\bcanopy\b|\bpanel\b', 'Enclosure'),
]


def _equip_category(name: str) -> str:
    """UNIVERSAL noun-keyed category for a BoM line (no per-class tables). First rule wins;
    everything unmatched falls into a single honest catch-all."""
    blob = str(name or "").lower()
    for rx, cat in _EQUIP_CAT_RULES:
        if re.search(rx, blob):
            return cat
    return 'Other equipment'


def cost_breakdown_by_category(rows: List[dict]) -> List[tuple]:
    """Returns [(category, gbp, pct_of_total, n_lines)] sorted by cost desc, [] if no rows.
    (2026-06-25 fix) Categorise the WHOLE assembled bill of materials by EQUIPMENT category so the
    breakdown SUMS TO THE BoM GRAND TOTAL — was reading parts-ledger.json's `equipment` list, which
    is empty for most runs, leaving only the connection (cabling/pipework) lines and a ~£11.6k total
    against a ~£797k bill. Groups the principal `requirementsBom` rows (excluding SUB-COMPONENT
    children whose cost rolls into a parent) by a universal noun map; percentages are of the BoM
    grand total."""
    agg: Dict[str, List[float]] = {}   # cat -> [gbp, n]
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        if str(r.get("status", "") or "").strip().upper() == "SUB-COMPONENT":
            continue
        v = r.get("line_gbp") or 0
        if not isinstance(v, (int, float)) or v <= 0:
            continue
        cat = _equip_category(
            f"{r.get('requirement', '')} {r.get('part', '')}")
        a = agg.setdefault(cat, [0.0, 0]); a[0] += float(v); a[1] += 1
    total = sum(a[0] for a in agg.values()) or 1.0
    out = [(cat, gn[0], 100.0 * gn[0] / total, int(gn[1])) for cat, gn in agg.items()]
    out.sort(key=lambda x: x[1], reverse=True)
    return out


# ============================================================================
# TAB 0 — EXECUTIVE SUMMARY (the "wow" cover — gridlines off, image-led)
# ============================================================================
# Tristan 2026-06-24: a 30-tab spreadsheet says "audit"; the first thing a reader sees must say
# "wow". This is a pitch-deck-style cover (NOT a grid): hero render + headline cards (what · output ·
# cost · status) + the capex breakdown + "what's inside" + the concierge ladder (your next steps).
# The detailed review surfaces (Overview, Checks, Ledger) follow. UNIVERSAL — no per-class content.
# proper display names for known class slugs (an acronym like "bess" reads badly as "Bess").
_CLASS_DISPLAY = {
    "bess": "Battery Energy Storage System",
    "bess-utility-scale": "Battery Energy Storage System",
    "energy_storage": "Battery Energy Storage System",
    "haps": "High-Altitude Pseudo-Satellite",
    "ras": "Recirculating Aquaculture System",
    "aquaculture_ras": "Recirculating Aquaculture System",
    "auv": "Autonomous Underwater Vehicle",
    "co2_mineralisation": "CO₂ Mineralisation Plant",
}


def _humanize_class(c: Any) -> str:
    raw = str(c or "").strip().lower()
    if raw in _CLASS_DISPLAY:
        return _CLASS_DISPLAY[raw]
    s = str(c or "this system").replace("_", " ").replace("-", " ").strip()
    s = re.sub(r"\bco2\b", "CO₂", s, flags=re.I)
    # uppercase common engineering acronyms so they don't read as Title-case words
    s = re.sub(r"\b(bess|haps|ras|auv|pcs|hvac|saf|uv|dc|ac|lfp|nmc)\b",
               lambda m: m.group(1).upper(), s, flags=re.I)
    return (s[:1].upper() + s[1:]) if s else "This system"


def _headline_build_cost(state: dict) -> tuple:
    """(label, gbp) for the cover's cost card — the most defensible 'what it costs', labelled honestly."""
    cs = state.get("costStack") or {}
    for k, label in (("installed_asp_gbp", "All-in installed cost"),
                     ("oem_transfer_price_gbp", "Build cost (ex-works)")):
        v = cs.get(k)
        if isinstance(v, (int, float)) and v > 0:
            return label, float(v)
    cr = (state.get("cost_reality") or {}).get("bom_total_gbp")
    if isinstance(cr, (int, float)) and cr > 0:
        return "Materials cost (bill of materials)", float(cr)
    return None, None


def _exec_validation_verdict(state: dict) -> tuple:
    """(short_status, long_sentence) from the DETERMINISTIC validation signals — the benchmark net
    (independent top-down market check) + the cost-sanity gate. Universal; no LLM. Returns
    ('Engineering-validated' | 'N checks flagged' | '', sentence)."""
    bd = state.get("benchmarkDivergence") or {}
    worst = bd.get("worst")
    nflag = sum(1 for f in (bd.get("findings") or []) if f.get("verdict") and f.get("verdict") != "ok")
    if worst == "ok":
        return ("Engineering-validated",
                "An independent top-down market benchmark agrees with the engine on every checked "
                "dimension (cost, output, sizing).")
    if worst in ("warn", "radical"):
        return (f"{nflag} check{'s' if nflag != 1 else ''} flagged",
                f"An independent market benchmark flags {nflag} dimension"
                f"{'s' if nflag != 1 else ''} where the design diverges from the brief/market "
                f"expectation — see the ⚠ Checks tab for the routed detail.")
    return ("", "")


def _exec_synopsis(state: dict) -> str:
    """A deterministic 1-paragraph synopsis assembled ENTIRELY from state — no LLM prose. Every
    clause is a state value (class, headline output, build cost, benchmark verdict), so the prose
    can never drift from the numbers in the tabs. Universal across product classes."""
    km = state.get("keyMetrics") or {}
    pc = ((state.get("orchestratorContract") or {}).get("product_class")
          or (state.get("parsedBrief") or {}).get("product_class"))
    proj = _humanize_class(pc)
    parts = []
    ho = km.get("headline_output") or {}
    if ho.get("value") not in (None, ""):
        out = f"{ho.get('value')} {ho.get('unit', '')}".strip()
        lbl = str(ho.get("label", "")).strip()
        parts.append(f"This dossier specifies a {proj} — {out}"
                     + (f" ({lbl.lower()})" if lbl else "") + ".")
    else:
        parts.append(f"This dossier specifies a {proj}.")
    clabel, cgbp = _headline_build_cost(state)
    if cgbp:
        parts.append(f"The engine values the build at £{round(cgbp):,} ({clabel.lower()}).")
    # The deterministic ship gate has the last word: when it FAILS, the synopsis must NOT claim the
    # dossier is validated (2026-06-25). It carries the honest DRAFT clause instead of the benchmark
    # "agrees on every dimension" sentence.
    _aud = state.get("_dossierAudit") or {}
    if _aud and _aud.get("ship_ok") is False:
        _open = int(_aud.get("total") or 0)
        parts.append(f"This is a DRAFT: the deterministic self-audit flags {_open} open "
                     f"issue{'s' if _open != 1 else ''} ({_aud.get('high', 0)} high-severity) — "
                     f"see the ⚠ Audit tab. It is not yet engineering-validated.")
    else:
        _, vsent = _exec_validation_verdict(state)
        if vsent:
            parts.append(vsent)
    parts.append("Every figure in this workbook — the bill of materials, the costs, and the "
                 "specifications below — is derived deterministically from the engineering "
                 "contract, computed rather than estimated by hand.")
    return " ".join(parts)


def tab_executive_summary(wb: Workbook, state: dict, run_dir: str, sha: str) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 26, "B": 20, "C": 18, "D": 18, "E": 16, "F": 16, "G": 16})
    km = state.get("keyMetrics") or {}
    pc = ((state.get("orchestratorContract") or {}).get("product_class")
          or (state.get("parsedBrief") or {}).get("product_class"))
    proj = _humanize_class(pc)

    nxt = title_row(
        ws, f"ForgeOS Engineering Dossier — {proj}", 7,
        "The engineering reality of your hardware idea — a buildable design, a real bill of "
        "materials, the true cost, and who can make it. The detail is in the tabs; this page is "
        "the summary.",
    )

    # ---- hero render (right side) ----
    try:
        from openpyxl.drawing.image import Image as XLImage
        hero = next((p for p in (os.path.join(run_dir, "00-hero.png"),
                                 os.path.join(run_dir, "blender-cover.png"))
                     if os.path.exists(p)), None)
        if hero:
            ds = downscale_png(hero, run_dir, max_px=900)
            im = XLImage(ds)
            if im.width and im.width > 470:
                r = 470 / float(im.width)
                im.width = int(im.width * r)
                im.height = int(im.height * r)
            ws.add_image(im, "E" + str(nxt))
    except Exception:  # never let the cover image break the build
        pass

    # ---- headline cards (left column) ----
    FONT_CARD_L = Font(name="Calibri", size=9, bold=True, color="888888")
    FONT_CARD_V = Font(name="Calibri", size=16, bold=True, color="1F3A5F")
    row = nxt

    def card(label: str, value: str, sub: str = "") -> None:
        nonlocal row
        cl = ws.cell(row, 1, label.upper())
        cl.font = FONT_CARD_L
        row += 1
        cv = ws.cell(row, 1, value)
        cv.font = FONT_CARD_V
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        if sub:
            cs = ws.cell(row, 1, sub)
            cs.font = FONT_NOTE
            cs.alignment = WRAP_TOP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            row += 1
        row += 1

    card("What it is", proj)
    ho = km.get("headline_output") or {}
    if ho.get("value") not in (None, ""):
        card("Output", f"{ho.get('value')} {ho.get('unit', '')}".strip(), str(ho.get("label", "")))
    clabel, cgbp = _headline_build_cost(state)
    if cgbp:
        card(clabel, f"£{round(cgbp):,}")
    sc_obj = state.get("qualityScorecard") or load_json(os.path.join(run_dir, "quality-scorecard.json")) or {}
    floor = sc_obj.get("floor")
    vstatus, _vsent = _exec_validation_verdict(state)
    # ---- the deterministic SHIP GATE overrides every other status (2026-06-25). When the per-tab
    # self-audit (scripts/lib/dossier_audit.py) says ship_ok=False, the dossier is a DRAFT — it may
    # NOT claim "Engineering-validated" no matter what the quality floor or benchmark say. ----
    _aud = state.get("_dossierAudit") or {}
    _open = int(_aud.get("total") or 0)
    if _aud and _aud.get("ship_ok") is False:
        card("Status", f"DRAFT — {_open} open issue{'s' if _open != 1 else ''}",
             "Deterministic self-audit ship gate FAILED — see the ⚠ Audit tab.")
    elif isinstance(floor, (int, float)):
        card("Status",
             "Engineering-validated" if floor >= 8 else f"Quality floor {floor}/10",
             "Scored against deterministic engineering gates." if floor >= 8 else "")
    elif vstatus:
        card("Validation", vstatus, "Independent top-down market benchmark vs the engine.")

    # clear the hero image before full-width sections
    row = max(row, nxt + 16) + 1

    # ---- deterministic synopsis (every clause is a state value — prose cannot drift) ----
    syn = _exec_synopsis(state)
    if syn:
        sc = ws.cell(row, 1, syn)
        sc.font = Font(name="Calibri", size=11, color="333333")
        sc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=7)
        ws.row_dimensions[row].height = 56
        row += 4

    # ---- key specifications & brief compliance (deterministic, brief-driven spec sheet) ----
    cmp_start = row
    sub_banner(ws, row, "Key specifications & brief compliance — every target deterministically verified", 7)
    nr = _render_brief_compliance_section(ws, state, row + 1)
    if nr:
        row = nr + 1
    else:
        # no brief metrics → drop the empty banner we just wrote
        for _c in range(1, 8):
            ws.cell(cmp_start, _c).value = None
        row = cmp_start

    # ---- where the money goes (top categories) ----
    cb = cost_breakdown_by_category(state.get("requirementsBom") or [])
    if cb:
        sub_banner(ws, row, "Where the money goes — capex by category", 7)
        row += 1
        _bar_font = Font(name="Menlo", size=10, color="2E5A88")
        for cat, gbp, pct, _n in cb[:7]:
            ws.cell(row, 1, cat).font = FONT_SUB
            cg = ws.cell(row, 2, round(gbp))
            cg.number_format = "#,##0"
            cp = ws.cell(row, 4, round(pct, 1))
            cp.number_format = '0.0"%"'
            cbar = ws.cell(row, 5, "█" * int(round(pct / 2.5)) if pct >= 1.25 else "▏")
            cbar.font = _bar_font
            row += 1
        row += 1

    # ---- what's inside ----
    sub_banner(ws, row, "What's inside this workbook", 7)
    row += 1
    for tab in ("Bill of Materials (Ledger)", "Cost waterfall", "Financial model",
                "Calculations", "Risk & Regulatory", "Connection trace"):
        desc = _TAB_DESCRIPTIONS.get(tab)
        if desc:
            ws.cell(row, 1, "•  " + tab).font = FONT_SUB
            d = ws.cell(row, 2, desc)
            d.font = FONT_NOTE
            d.alignment = WRAP_TOP
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            row += 1
    row += 1

    # ---- your next steps (the concierge ladder) ----
    sub_banner(ws, row, "Your next steps — from design to a funded factory", 7)
    row += 1
    ladder = [
        ("✓  You have the engineering dossier", "A buildable design, a real bill of materials, and the true cost — this document.", FONT_PASS),
        ("→  Talk to the experts this design needs", "We connect you to vetted specialists for the open questions the design raises.", FONT_SUB),
        ("→  Get real supplier quotes (RFQ)", "We take this bill of materials to suppliers and bring back real quotes.", FONT_SUB),
        ("→  Raise the money on these numbers", "We help you turn the validated design and costs into a fundraise.", FONT_SUB),
    ]
    for head, sub, font in ladder:
        h = ws.cell(row, 1, head)
        h.font = font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        s = ws.cell(row, 1, "      " + sub)
        s.font = FONT_NOTE
        s.alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
    row += 1
    contact = ws.cell(row, 1, "Fractional Forge — Tristan Fischer, Founder.  Reply to take the next step.")
    contact.font = FONT_SUB


# ============================================================================
# TAB 1 — OVERVIEW (quality scorecard + headline metrics + provenance)
# ============================================================================
def tab_overview(wb: Workbook, state: dict, run_dir: str, sha: str) -> None:
    ws = wb.create_sheet("Overview")
    set_widths(ws, {"A": 30, "B": 16, "C": 14, "D": 60})
    run_name = os.path.basename(os.path.normpath(run_dir))
    row = title_row(
        ws, f"ForgeOS Review Workbook — {run_name}", 4,
        "Live Excel review surface. Yellow cells are editable inputs; green cells are "
        "live formulas that recompute downstream. See the '⚠ Checks' tab for invariants "
        "that go RED where the engine's numbers do not reconcile.",
    )

    # ---- hero render thumbnail (top-right) — the cover answers "what does it look like?" without
    # a click; the full-resolution hero is its own tab. Floats over the empty cols right of D. ----
    try:
        from openpyxl.drawing.image import Image as XLImage
        _hero = next((p for p in (os.path.join(run_dir, "00-hero.png"),
                                  os.path.join(run_dir, "blender-cover.png"))
                      if os.path.exists(p)), None)
        if _hero:
            _hds = downscale_png(_hero, run_dir, max_px=900)
            _himg = XLImage(_hds)
            if _himg.width and _himg.width > 560:
                _hr = 560 / float(_himg.width)
                _himg.width = int(_himg.width * _hr)
                _himg.height = int(_himg.height * _hr)
            ws.add_image(_himg, "F2")
    except Exception:  # never let the cover image break the Overview
        pass

    # ---- provenance block ----
    sub_banner(ws, row, "Run provenance", 4)
    row += 1
    prov = [
        ("Run directory", run_name),
        ("Git short SHA", sha),
        ("State saved at", state.get("savedAt", "—")),
        ("Project id", state.get("projectId", "—")),
        ("Product class",
         (state.get("orchestratorContract") or {}).get("product_class", "—")),
    ]
    for label, val in prov:
        ws.cell(row, 1, label).font = FONT_SUB
        ws.cell(row, 2, str(val))
        row += 1
    row += 1

    # ---- quality scorecard ----
    sc = state.get("qualityScorecard") or load_json(
        os.path.join(run_dir, "quality-scorecard.json")
    )
    sub_banner(ws, row, "Quality scorecard", 4)
    row += 1
    if sc:
        floor = sc.get("floor")
        mean = sc.get("mean")
        all_pass = sc.get("allPass")
        # allPass may be absent -> derive from floor >= 8 (the ≥8-everywhere rule)
        if all_pass is None and floor is not None:
            all_pass = floor >= 8
        for label, val in [
            ("Floor (min section)", floor),
            ("Mean", mean),
            ("All sections ≥ 8 (allPass)", all_pass),
            ("Iteration", sc.get("iteration")),
        ]:
            ws.cell(row, 1, label).font = FONT_SUB
            c = ws.cell(row, 2, val if val is not None else "—")
            if label.startswith("Floor") and isinstance(floor, (int, float)):
                c.fill = FILL_PASS if floor >= 8 else FILL_FAIL
                c.font = FONT_PASS if floor >= 8 else FONT_FAIL
            if label.startswith("All sections") and isinstance(all_pass, bool):
                c.value = "PASS" if all_pass else "FAIL"
                c.fill = FILL_PASS if all_pass else FILL_FAIL
                c.font = FONT_PASS if all_pass else FONT_FAIL
            row += 1
        row += 1

        # per-section table. A section flagged `advisory` is the LLM SEMANTIC self-audit — it is
        # NON-GATING (the gating floor/allPass already exclude it) and noisy, so a sub-8 advisory
        # section is shown AMBER "ADVISORY", NOT red FAIL — red is reserved for a real DETERMINISTIC
        # gate breach (Tristan 2026-06-23: advisory LLM scores were painting the dossier red while
        # the authoritative gates all pass). The 'advisory' flag comes straight from the scorecard.
        header(ws, row, ["Section", "Score", "≥8?", "Defects"])
        row += 1
        note = ws.cell(row, 1, "Sections marked (advisory) are the LLM self-audit — non-gating; "
                               "amber = advisory concern, red = a deterministic gate breach.")
        note.font = FONT_NOTE
        note.alignment = WRAP_TOP
        row += 1
        for sec in sc.get("sections", []):
            name = sec.get("name", "")
            score = sec.get("score")
            advisory = bool(sec.get("advisory"))
            defects = sec.get("defects") or []
            ws.cell(row, 1, name + ("  (advisory)" if advisory else "")).border = BORDER
            cs = ws.cell(row, 2, score)
            cs.border = BORDER
            ok = isinstance(score, (int, float)) and score >= 8
            if ok:
                verdict, vfill, vfont = "PASS", FILL_PASS, FONT_PASS
            elif advisory:
                verdict, vfill, vfont = "ADVISORY", FILL_ADVISORY, FONT_ADVISORY
            else:
                verdict, vfill, vfont = "FAIL", FILL_FAIL, FONT_FAIL
            cp = ws.cell(row, 3, verdict)
            cp.fill = vfill
            cp.font = vfont
            cp.border = BORDER
            cd = ws.cell(row, 4, "; ".join(str(d) for d in defects))
            cd.alignment = WRAP_TOP
            cd.border = BORDER
            row += 1
        row += 1
    else:
        ws.cell(row, 1, "No quality-scorecard found.").font = FONT_NOTE
        row += 2

    # ---- deterministic computational-checks summary (Tristan 2026-06-23: a one-glance pass/fail
    # count ON the Overview; the per-invariant detail stays on the ⚠ Checks tab). Computed LIVE via
    # the SAME dcl.run_all_checks the ⚠ Checks tab renders, so the count can never drift from it. ----
    try:
        _all_checks = dcl.run_all_checks(run_dir, state)
        _fails = [c for c in _all_checks
                  if str(getattr(c, "status", "")).upper() == "FAIL"]
        _n, _nf = len(_all_checks), len(_fails)
        if _n:
            sub_banner(ws, row, "Computational checks", 4)
            row += 1
            ws.cell(row, 1, "Deterministic invariants").font = FONT_SUB
            _cc = ws.cell(row, 2, f"{_n - _nf} / {_n} pass"
                          + (f"  ·  {_nf} FAIL" if _nf else "  ·  0 fail"))
            _cc.fill = FILL_PASS if _nf == 0 else FILL_FAIL
            _cc.font = FONT_PASS if _nf == 0 else FONT_FAIL
            _nt = ws.cell(row, 3, "full detail on the ⚠ Checks tab")
            _nt.font = FONT_NOTE
            row += 1
            for _c in _fails[:12]:
                ws.cell(row, 1, "✗ " + str(getattr(_c, "name", ""))).font = FONT_FAIL
                row += 1
            row += 1
    except Exception:  # never let the summary break the Overview
        pass

    # ---- headline metrics ----
    km = state.get("keyMetrics") or {}
    sub_banner(ws, row, "Headline metrics", 4)
    row += 1
    header(ws, row, ["Metric", "Value", "Unit", "Notes / source"])
    row += 1

    def metric_row(m: dict) -> None:
        nonlocal row
        ws.cell(row, 1, m.get("label", "")).border = BORDER
        ws.cell(row, 2, m.get("value", "")).border = BORDER
        ws.cell(row, 3, m.get("unit", "")).border = BORDER
        n = ws.cell(row, 4, m.get("notes", m.get("source", "")))
        n.alignment = WRAP_TOP
        n.font = FONT_NOTE
        n.border = BORDER
        row += 1

    if isinstance(km, dict):
        if km.get("headline_output"):
            metric_row(km["headline_output"])
        if km.get("headline_constraint"):
            metric_row(km["headline_constraint"])
        for m in (km.get("supporting_metrics") or [])[:12]:
            metric_row(m)
    row += 1

    # ---- where the money goes (deterministic capex breakdown by category) — the five-second
    # "where does my budget go?" answer a founder needs before they read 30 tabs. ----
    _cb = cost_breakdown_by_category(state.get("requirementsBom") or [])
    if _cb:
        sub_banner(ws, row, "Where the money goes — capex by category", 4)
        row += 1
        _note = ws.cell(row, 1, "Deterministic roll-up of every bill-of-materials line, grouped by "
                                "equipment category. Sums to the bill-of-materials grand total.")
        _note.font = FONT_NOTE
        _note.alignment = WRAP_TOP
        row += 1
        header(ws, row, ["Category", "Cost (£)", "% of capex", "Share"])
        row += 1
        _bar_font = Font(name="Menlo", size=10, color="2E5A88")
        for cat, gbp, pct, _n in _cb:
            ws.cell(row, 1, cat).border = BORDER
            cg = ws.cell(row, 2, round(gbp))
            cg.number_format = "#,##0"
            cg.border = BORDER
            cp = ws.cell(row, 3, round(pct, 1))
            cp.number_format = '0.0"%"'
            cp.border = BORDER
            cb = ws.cell(row, 4, "█" * int(round(pct / 2.5)) if pct >= 1.25 else "▏")
            cb.font = _bar_font
            cb.border = BORDER
            row += 1
        ws.cell(row, 1, "Total (bill of materials)").font = FONT_SUB
        _ct = ws.cell(row, 2, round(sum(r[1] for r in _cb)))
        _ct.number_format = "#,##0"
        _ct.font = FONT_SUB
        row += 2

    back_link(ws, 4)


# ============================================================================
# TAB — BRIEF (original client brief vs the engine-enhanced structured brief)
# ============================================================================
def _brief_paragraphs(md: str) -> List[str]:
    """Split a markdown/prose brief into readable paragraphs, one per row. Blank-line
    separated blocks become paragraphs; a leading '# heading' / list bullet is kept
    verbatim (lightly de-marked). Universal — no class assumptions."""
    out: List[str] = []
    block: List[str] = []
    for raw in str(md or "").splitlines():
        ln = raw.rstrip()
        if not ln.strip():
            if block:
                out.append(" ".join(block).strip())
                block = []
            continue
        s = ln.strip()
        # a markdown heading or list bullet is its OWN paragraph (don't fold into prose)
        if re.match(r"^#{1,6}\s", s) or re.match(r"^[-*]\s", s):
            if block:
                out.append(" ".join(block).strip())
                block = []
            s = re.sub(r"^#{1,6}\s+", "", s)        # drop the ### markers, keep the text
            out.append(s)
        else:
            block.append(s)
    if block:
        out.append(" ".join(block).strip())
    return [p for p in out if p]


def tab_brief(wb: Workbook, run_dir: str) -> None:
    """The ORIGINAL client brief (left) vs the engine's ENHANCED, structured brief
    (right). Original = 0-original-brief.md (prose, one paragraph per row, wrapped);
    enhanced = 1-brief-expanded.json (labelled sections; lists one item per row).
    Falls back gracefully when a file is missing. Universal."""
    ws = wb.create_sheet("Brief")
    set_widths(ws, {"A": 80, "B": 4, "C": 46})
    title_row(
        ws, "Brief — original vs engine-enhanced", 3,
        "The verbatim client brief (left) and the engine's structured, enhanced "
        "interpretation (right) that drives the whole design.",
    )

    # ---- read both sources (gracefully) ----
    orig = ""
    op = os.path.join(run_dir, "0-original-brief.md")
    if os.path.exists(op):
        try:
            orig = open(op, "r").read()
        except Exception:  # noqa: BLE001
            orig = ""
    enh = load_json(os.path.join(run_dir, "1-brief-expanded.json")) or {}

    # ---- column headers ----
    r0 = 4
    oc = ws.cell(r0, 1, "Original brief (verbatim)")
    oc.font = FONT_SUB
    oc.fill = FILL_SUB
    oc.alignment = LEFT_TOP
    ec = ws.cell(r0, 3, "Engine-enhanced brief (structured)")
    ec.font = FONT_SUB
    ec.fill = FILL_SUB
    ec.alignment = LEFT_TOP
    r0 += 1

    # ---- LEFT: original prose, one paragraph per row, wrapped ----
    left_r = r0
    paras = _brief_paragraphs(orig)
    if not paras:
        c = ws.cell(left_r, 1, "— 0-original-brief.md not found —")
        c.font = FONT_NOTE
        c.alignment = WRAP_TOP
        left_r += 1
    else:
        for p in paras:
            c = ws.cell(left_r, 1, clean_cell(p))
            c.alignment = WRAP_TOP
            # grow the row so a long paragraph never clips (col A ≈ 80 wide)
            ws.row_dimensions[left_r].height = 14.5 * min(8, max(1, -(-len(p) // 95)))
            left_r += 1

    # ---- RIGHT: enhanced brief, each key a labelled section ----
    right_r = r0
    SCALAR_KEYS = [
        ("product_summary", "Product summary"),
        ("primary_product", "Primary product"),
        ("construction_materials", "Construction materials"),
    ]
    LIST_KEYS = [
        ("derived_requirements", "Derived requirements"),
        ("operating_conditions", "Operating conditions"),
    ]

    def _label(text: str) -> None:
        nonlocal right_r
        lc = ws.cell(right_r, 3, clean_cell(text))
        lc.font = FONT_SUB
        right_r += 1

    def _value(text: str) -> None:
        nonlocal right_r
        vc = ws.cell(right_r, 3, clean_cell(text))
        vc.alignment = WRAP_TOP
        vc.font = FONT_NOTE
        s = str(text or "")
        ws.row_dimensions[right_r].height = 14.5 * min(6, max(1, -(-len(s) // 55)))
        right_r += 1

    if not enh:
        c = ws.cell(right_r, 3, "— 1-brief-expanded.json not found —")
        c.font = FONT_NOTE
        c.alignment = WRAP_TOP
        right_r += 1
    else:
        for key, label in SCALAR_KEYS:
            v = enh.get(key)
            if v:
                _label(label)
                _value(str(v))
                right_r += 1  # spacer
        for key, label in LIST_KEYS:
            items = enh.get(key)
            if isinstance(items, list) and items:
                _label(label)
                for it in items:
                    if isinstance(it, dict):
                        lbl = (it.get("label") or it.get("key") or "").strip()
                        val = it.get("value")
                        unit = it.get("unit") or ""
                        line = f"{lbl}: {val} {unit}".strip() if lbl else f"{val} {unit}".strip()
                        prov = it.get("provenance") or it.get("confidence")
                        if prov:
                            line += f"   ({prov})"
                    else:
                        line = str(it)
                    _value("• " + line)
                right_r += 1  # spacer
        notes = enh.get("notes")
        if notes:
            _label("Notes")
            _value(str(notes))

    ws.freeze_panes = "A4"
    back_link(ws, 3)


# ============================================================================
# TAB 2 — "⚠ Checks"  (THE ERROR-SURFACING TAB — the point of the exercise)
# ============================================================================
def _render_lib_checks(ws: Worksheet, state: dict, run_dir: str, r: int,
                       data_r: int, data_col_a: str,
                       fail_labels: List[str]) -> Tuple[int, int, int]:
    """Render every deterministic_checks_lib.Check as a LIVE Excel row, grouped by
    family. Returns (next_row, next_data_row, fail_count).

    Layout per check (one visible row per lib Check; STATUS recomputes live):
      * hidden data cells: K=actual-base (or per-unit), L=expected, M=count
      * visible: B=ACTUAL formula, C=EXPECTED ref, D=Δ, E=Tol/band, F=STATUS, G=detail
      * the STATUS formula is chosen to MIRROR the lib's own decision for that check
        so the live recompute can never disagree with the CLI:
          ge    -> =IF(B>=C-E,"PASS","FAIL")        (rating must clear duty)
          le    -> =IF(B<=C+E,"PASS","FAIL")        (within a ceiling)
          tally -> =IF(B<=C,"PASS","FAIL")          (out-of-spec count must be 0)
          band  -> =IF(OR(B/C>E,B/C<1/E),"FAIL","PASS")  (COST ×N price band; the lib
                   tags these relation="eq" but decides by a RATIO band with tol==0 —
                   here E holds the band factor so the column stays meaningful)
          eq    -> =IF(ABS(D)<E,"PASS","FAIL")      (equality within tolerance)
    A COST check with relation "eq" and tol 0 IS a ratio-band check (the lib's COST1
    `unit price within xN of <ref>`); every other eq check carries a non-zero tol.
    """
    checks = dcl.run_all_checks(run_dir, state)
    fail_count = 0
    if not checks:
        return r, data_r, fail_count

    fam_titles = {
        "CONSISTENCY": "E1 · CONSISTENCY — per-unit×count, Σsub==line, rating==quantity",
        "ADEQUACY": "E2 · ADEQUACY — rating ≥ duty (motor, breaker, cable, vessel, chiller)",
        "BALANCE": "E3 · BALANCE — mass/energy/flow closures & continuity",
        "COST": "E4 · COST — per-line price band & Σ lines == cover total",
        "CONNECTIVITY": "E5 · CONNECTIVITY/SPEC — out-of-spec tally & velocity limit",
        "BRIEF": "E6 · BRIEF — design realises each stated brief target metric (±5%)",
        "PROVENANCE": "E7 · PROVENANCE — every engineering-tool output is actually USED by the "
                      "design (no stale/orphaned tool computations)",
    }
    for fam in ("CONSISTENCY", "ADEQUACY", "BALANCE", "COST", "CONNECTIVITY", "BRIEF", "PROVENANCE"):
        fam_checks = [c for c in checks if c.category == fam]
        if not fam_checks:
            continue
        sub_banner(ws, r, fam_titles[fam], 7)
        r += 1
        for c in fam_checks:
            if c.actual is None or c.expected is None:
                # No numeric target → no LIVE formula possible. But a check the ENGINE
                # itself flagged FAIL (e.g. a closure with status=warn/fail such as
                # capex_within_ceiling: £8.15 M design vs the £5.0 M brief ceiling) must
                # still surface — otherwise the tab silently disagrees with the CLI and
                # hides a real breach. Render it as a STATIC red row (no recompute).
                if c.status == dcl.FAIL:
                    ws.cell(r, 1, clean_cell(c.name)).border = BORDER
                    ws.cell(r, 1).alignment = WRAP_TOP
                    ws.cell(r, 2, num(c.actual) if c.actual is not None else "—").border = BORDER
                    ws.cell(r, 3, "(engine verdict)").border = BORDER
                    ws.cell(r, 4, "—").border = BORDER
                    ws.cell(r, 5, "—").border = BORDER
                    cs = ws.cell(r, 6, "FAIL")
                    cs.border = BORDER
                    cs.font = Font(bold=True)
                    cdet = ws.cell(r, 7, clean_cell(c.detail))
                    cdet.alignment = WRAP_TOP
                    cdet.font = FONT_NOTE
                    cdet.border = BORDER
                    fail_count += 1
                    fail_labels.append(f"[{fam}] {c.name}")
                    for col in range(1, 8):
                        ws.cell(r, col).fill = FILL_FAIL
                    r += 1
                continue  # N/A, or static-fail already rendered — nothing live
            data_r += 1
            # --- hidden editable data cells ---
            ws.cell(data_r, 10, c.name)
            if c.a_factors is not None:
                per_unit, count = c.a_factors
                ws.cell(data_r, 11, per_unit).fill = FILL_INPUT      # K = per-unit
                ws.cell(data_r, 13, count).fill = FILL_INPUT         # M = count
                actual_formula = f"=${data_col_a}${data_r}*$M${data_r}"
            else:
                ws.cell(data_r, 11, c.actual).fill = FILL_INPUT      # K = actual
                actual_formula = f"=${data_col_a}${data_r}"
            ws.cell(data_r, 12, c.expected).fill = FILL_INPUT        # L = expected

            # A COST check tagged relation="eq" with tol==0 is decided by the lib as a
            # RATIO BAND (unit price within xN of its reference), NOT an equality —
            # render it as a live band test so the recompute matches the CLI exactly.
            is_band = (c.category == "COST" and c.relation == "eq" and not c.tol)
            tol_cell_val = dcl.COST_BAND_FACTOR if is_band else c.tol

            # --- visible live row ---
            ws.cell(r, 1, c.name).border = BORDER
            ws.cell(r, 1).alignment = WRAP_TOP
            ws.cell(r, 2, actual_formula).border = BORDER
            ws.cell(r, 3, f"=$L${data_r}").border = BORDER
            ws.cell(r, 4, f"=B{r}-C{r}").border = BORDER
            ws.cell(r, 5, tol_cell_val).border = BORDER
            # STATUS formula chosen to MIRROR the lib's own decision (see docstring)
            if is_band:                     # COST ratio band: flag >xN or <1/N of ref
                status_f = (f'=IF(OR(C{r}=0,AND(B{r}/C{r}<=E{r},'
                            f'B{r}/C{r}>=1/E{r})),"PASS","FAIL")')
            elif c.relation == "ge":        # actual must be >= expected
                status_f = f'=IF(B{r}>=C{r}-E{r},"PASS","FAIL")'
            elif c.relation == "le":        # actual must be <= expected
                status_f = f'=IF(B{r}<=C{r}+E{r},"PASS","FAIL")'
            elif c.relation == "tally":     # actual count must be 0
                status_f = f'=IF(B{r}<=C{r},"PASS","FAIL")'
            else:                            # equality within tolerance
                status_f = f'=IF(ABS(D{r})<E{r},"PASS","FAIL")'
            cs = ws.cell(r, 6, status_f)
            cs.border = BORDER
            cs.font = Font(bold=True)
            cdet = ws.cell(r, 7, c.detail)
            cdet.alignment = WRAP_TOP
            cdet.font = FONT_NOTE
            cdet.border = BORDER

            # pre-evaluate now so we can colour + count FAILs (and so the CLI and
            # the tab agree before Excel recomputes on open)
            if c.status == dcl.FAIL:
                fail_count += 1
                fail_labels.append(f"[{fam}] {c.name}")
                for col in range(1, 8):
                    cc = ws.cell(r, col)
                    cc.fill = FILL_FAIL
            r += 1
        r += 1  # spacer between families
    return r, data_r, fail_count


def tab_checks(wb: Workbook, state: dict, run_dir: str) -> int:
    """
    Live-formula invariants — ONE SOURCE OF TRUTH with the CLI verifier.

    This tab renders EXACTLY the checks ``deterministic_checks_lib.run_all_checks``
    produces — the SAME pure-arithmetic suite the standalone CLI
    (``scripts/deterministic-checks.py``) prints — one visible row per applicable
    check, in the SAME five families (CONSISTENCY / ADEQUACY / BALANCE / COST /
    CONNECTIVITY), with the SAME PASS/FAIL verdict. There is NO exporter-side check
    machinery any more: previously this tab ALSO carried four hand-rolled sections
    (A per-unit×count, B Σ-lines==cover, C emitter-render-vs-contract, D closures)
    that DUPLICATED the lib (B, D) or emitted FALSE POSITIVES the lib does not (the
    stale "BoM per-unit flow 1336 m³/h × train count 8" — 1,336 is the correct
    per-TANK inlet flow ÷10 tanks, not a per-pump flow ÷8 trains — and the
    "Pump motor power: worked-calc render vs contract" full-loop-vs-per-pump basis
    comparison). Both have been removed so the tab can never disagree with the CLI.

    Each row:
        Check | ACTUAL (live =formula over the data cells) | EXPECTED
              | Δ (=ACTUAL-EXPECTED) | Tol | STATUS (=IF(...)) | Detail
    The numbers live in a hidden, editable DATA block (cols J+) so the visible
    ACTUAL/EXPECTED are genuine cell references: edit a yellow input and STATUS
    recomputes live. Conditional formatting colours STATUS on open.

    Returns the number of FAIL rows (for the final report) — identical to the
    CLI's fail count on the same run.
    """
    ws = wb.create_sheet("⚠ Checks")
    set_widths(ws, {"A": 46, "B": 16, "C": 16, "D": 14, "E": 10, "F": 12, "G": 52})
    title_row(
        ws, "⚠ Computational checks — live invariants (== the CLI verifier)", 7,
        "EXACTLY the checks scripts/deterministic-checks.py reports, rendered as "
        "LIVE Excel formulas over the data cells (cols J+). Edit a data cell and "
        "STATUS recomputes. RED = the engine's numbers do not reconcile.",
    )
    hdr = 4
    header(ws, hdr, ["Check", "ACTUAL", "EXPECTED", "Δ (actual-exp)",
                     "Tol", "STATUS", "Detail / why"])
    r = hdr + 1

    # Hidden data columns: J (label), K (a-value), L (b-value/expected), M (count)
    DATA_COL_A = "K"
    data_r = hdr  # parallel pointer into the data block
    ws.cell(hdr, 10, "DATA (live inputs — editable)").font = FONT_SUB

    fail_count = 0
    fail_labels: List[str] = []

    # ---- THE SHARED deterministic check suite — the ONLY source of checks --------
    # Every check from deterministic_checks_lib.run_all_checks — the SAME arithmetic
    # the standalone CLI runs — rendered as live Excel rows, grouped by family.
    # ACTUAL/EXPECTED/Δ/STATUS are formulas over editable yellow data cells;
    # per-unit×count checks make the product itself live (=K*M). This is the WHOLE
    # tab: identical set + identical PASS/FAIL to the CLI, no exporter-side extras.
    r, data_r, fc = _render_lib_checks(
        ws, state, run_dir, r, data_r, DATA_COL_A, fail_labels)
    fail_count += fc

    # conditional formatting so STATUS cells colour live on open / recompute. GUARD: only when
    # there is >=1 check row — an archetype that yields ZERO checks would make an invalid F5:F4
    # range and crash the whole Excel build (Tristan 2026-06-23 universality audit: SAF/satellite).
    if r - 1 >= hdr + 1:
        from openpyxl.formatting.rule import CellIsRule
        status_range = f"F{hdr+1}:F{r-1}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL, font=FONT_FAIL),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"PASS"'], fill=FILL_PASS, font=FONT_PASS),
        )
    ws.cell(2, 1)  # keep title intact

    # ---- FOLDED SECTIONS (Tristan 2026-06-24 consolidation): Brief compliance + Tool
    # provenance now live UNDER the checks as labelled sections rather than two standalone
    # tabs — all three are "did the engine's numbers hold up?" audit content. Each section
    # renders its own header + table; rows go below the checks block (cols A–H), clear of the
    # hidden J–M data block which is parallel to the checks rows above. ----
    sect_r = r + 2
    big1 = ws.cell(sect_r, 1, "Brief compliance — every target vs achieved vs PASS/FAIL")
    big1.font = FONT_TITLE; big1.fill = FILL_TITLE
    big1.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=sect_r, start_column=1, end_row=sect_r, end_column=8)
    ws.row_dimensions[sect_r].height = 24
    sect_r += 2
    bc_next = _render_brief_compliance_section(ws, state, sect_r)
    if bc_next is None:
        ws.cell(sect_r, 1, "— the brief carries no target_performance metrics to verify —").font = FONT_NOTE
        sect_r += 2
    else:
        # live conditional formatting on the section's STATUS column (G)
        from openpyxl.formatting.rule import CellIsRule as _CIR
        _rng = f"G{sect_r + 1}:G{bc_next - 1}"
        if bc_next - 1 >= sect_r + 1:
            ws.conditional_formatting.add(_rng, _CIR(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL, font=FONT_FAIL))
            ws.conditional_formatting.add(_rng, _CIR(operator="equal", formula=['"PASS"'], fill=FILL_PASS, font=FONT_PASS))
        sect_r = bc_next + 1

    big2 = ws.cell(sect_r, 1, "Tool provenance — input → output traceability (USED / STALE / ORPHANED)")
    big2.font = FONT_TITLE; big2.fill = FILL_TITLE
    big2.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=sect_r, start_column=1, end_row=sect_r, end_column=8)
    ws.row_dimensions[sect_r].height = 24
    sect_r += 2
    ti_next = _render_tool_io_section(ws, state, run_dir, sect_r)
    if ti_next is None:
        ws.cell(sect_r, 1, "— no orchestrator tool log for this run —").font = FONT_NOTE

    # ↑ Contents back-link at col span+1 (=H), clear of the hidden J–M data block.
    back_link(ws, 7)
    ws.freeze_panes = "A5"

    # stash the fail summary on the object for the caller's report
    ws._forge_fail_count = fail_count       # type: ignore[attr-defined]
    ws._forge_fail_labels = fail_labels     # type: ignore[attr-defined]
    return fail_count


# ============================================================================
# TAB — "⚠ Audit"  (THE DETERMINISTIC SHIP GATE — per-tab self-audit findings)
# ============================================================================
def tab_audit(wb: Workbook, report) -> None:
    """Render the deterministic per-tab self-audit (scripts/lib/dossier_audit.py) as a worksheet.
    A scorecard header line + every Finding grouped by its target tab. HIGH = red, MED = amber,
    LOW = grey — the SHIP GATE made visible: a FAIL verdict means the dossier is NOT validated."""
    ws = wb.create_sheet("⚠ Audit")
    set_widths(ws, {"A": 10, "B": 30, "C": 60, "D": 22, "E": 22})
    sc = report.scorecard()
    title_row(
        ws, "⚠ Dossier self-audit — the deterministic ship gate", 5,
        "Every per-tab check (scripts/lib/dossier_audit.py) — pass/fail FLAGS, no LLM. A FAIL "
        "verdict (any HIGH finding) means the dossier is NOT validated. HIGH = red, MED = amber, "
        "LOW = grey.",
    )
    r = 4
    # ---- scorecard header line ----
    verdict = sc.get("verdict", "?")
    sline = (f"VERDICT: {verdict}  ·  {sc.get('high', 0)} HIGH  ·  {sc.get('med', 0)} MED  ·  "
             f"{sc.get('low', 0)} LOW  ·  {sc.get('total', 0)} total"
             f"  ·  ship_ok={sc.get('ship_ok')}")
    hc = ws.cell(r, 1, sline)
    hc.font = FONT_TITLE
    hc.fill = FILL_FAIL if verdict == "FAIL" else (FILL_ADVISORY if verdict == "REVIEW" else FILL_PASS)
    hc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 22
    r += 2

    by_tab = report.by_tab()
    if not by_tab:
        ws.cell(r, 1, "No findings — every deterministic check passed.").font = FONT_PASS
        back_link(ws, 5)
        ws.freeze_panes = "A5"
        return

    _SEV_STYLE = {
        "HIGH": (FILL_FAIL, FONT_FAIL),
        "MED": (FILL_ADVISORY, FONT_ADVISORY),
        "LOW": (FILL_LEGACY, FONT_NOTE),
    }
    for tab_name, findings in by_tab.items():
        # tab band
        tb = ws.cell(r, 1, tab_name)
        tb.font = FONT_TITLE
        tb.fill = FILL_TITLE
        tb.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 20
        r += 1
        header(ws, r, ["Severity", "Check", "Message", "Actual", "Expected"])
        r += 1
        for f in findings:
            sev = str(getattr(f, "severity", "") or "").upper()
            fill, font = _SEV_STYLE.get(sev, (FILL_LEGACY, FONT_NOTE))
            sc1 = ws.cell(r, 1, sev or "—")
            sc1.fill = fill
            sc1.font = font
            sc1.border = BORDER
            ws.cell(r, 2, clean_cell(getattr(f, "check", "") or "")).border = BORDER
            mc = ws.cell(r, 3, clean_cell(getattr(f, "message", "") or ""))
            mc.alignment = WRAP_TOP
            mc.border = BORDER
            ac = ws.cell(r, 4, clean_cell(getattr(f, "actual", "") or ""))
            ac.alignment = WRAP_TOP
            ac.border = BORDER
            ec = ws.cell(r, 5, clean_cell(getattr(f, "expected", "") or ""))
            ec.alignment = WRAP_TOP
            ec.border = BORDER
            r += 1
        r += 1

    back_link(ws, 5)
    ws.freeze_panes = "A5"


# ============================================================================
# TAB 3 — QUANTITIES (orchestratorContract.quantities)
# ============================================================================
def tab_quantities(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("Quantities")
    set_widths(ws, {"A": 34, "B": 16, "C": 12, "D": 16, "E": 12, "F": 12, "G": 62})
    title_row(ws, "Contract quantities", 7,
              "Every sized contract quantity with its family, basis and source.")
    header(ws, 4, ["Name", "Value", "Unit", "Family", "Basis", "Source", "Source detail"])
    quantities = (state.get("orchestratorContract") or {}).get("quantities") or {}
    r = 5
    for name in quantities:
        v = quantities[name]
        if isinstance(v, dict):
            value, unit = v.get("value"), v.get("unit", "")
            family, basis = v.get("family", ""), v.get("basis", "")
            source, detail = v.get("source", ""), v.get("source_detail", "")
        else:
            value, unit, family, basis, source, detail = v, "", "", "", "", ""
        ws.cell(r, 1, clean_cell(name)).border = BORDER
        ws.cell(r, 2, value).border = BORDER
        ws.cell(r, 3, clean_cell(unit)).border = BORDER
        ws.cell(r, 4, clean_cell(family)).border = BORDER
        ws.cell(r, 5, clean_cell(basis)).border = BORDER
        ws.cell(r, 6, clean_cell(source)).border = BORDER
        cd = ws.cell(r, 7, clean_cell(detail))
        cd.alignment = WRAP_TOP
        cd.font = FONT_NOTE
        cd.border = BORDER
        r += 1
    # number formats (#37) on the Value column — FMT_NUM drops trailing .00 so an
    # integer-valued metric reads "62" not "62.00", while 2.58 still shows "2.58".
    apply_col_formats(ws, 5, {2: FMT_NUM}, r - 1)
    ws.auto_filter.ref = f"A4:G{r - 1}"
    ws.freeze_panes = "A5"
    back_link(ws, 7)


# ============================================================================
# TAB 4 — CALCULATIONS (worked-calcs grouped by tool; live where structured)
# ============================================================================
# ============================================================================
# UNIVERSAL HELPERS — chart styling + worked-calc verification (class-agnostic;
# NO per-archetype logic — these run for every dossier the engine produces)
# ============================================================================
import math as _math
import ast as _ast
import operator as _op

_CHART_PALETTE = ["2F5496", "C0392B", "548235", "7F4FC9", "1F9BD6",
                  "E67E22", "16A085", "8E44AD"]


def style_chart(ch, *, legend="auto", data_labels=False, gridlines=False):
    """Universal chart hygiene. Excel/openpyxl default to varying colour PER
    DATA POINT on a single-series chart — which renders a line as rainbow
    segments AND dumps every category into the legend (the 12-entry "£2.0M…"
    legends Tristan flagged). Force ONE solid colour per series, drop the heavy
    black gridlines, keep axis values, and show a legend only for genuine
    multi-series charts. Applies to every class."""
    try:
        ch.varyColors = False
    except Exception:  # noqa: BLE001
        pass
    series = list(getattr(ch, "series", []) or [])
    for i, s in enumerate(series):
        col = _CHART_PALETTE[i % len(_CHART_PALETTE)]
        try:
            s.graphicalProperties.line.solidFill = col
            s.graphicalProperties.line.width = 28000  # EMU ≈ 2.2 pt
        except Exception:  # noqa: BLE001
            pass
        try:
            s.graphicalProperties.solidFill = col  # bar/area fill (line series ignore)
        except Exception:  # noqa: BLE001
            pass
        try:
            # Excel renders a NEGATIVE bar HOLLOW/white by default (invertIfNegative).
            # A loss-making EBITDA bar must stay the SAME solid colour as the positives
            # (just pointing down) — Tristan flagged the white "EBITDA Low" bar.
            s.invertIfNegative = False
        except Exception:  # noqa: BLE001
            pass
        try:
            s.smooth = False
        except Exception:  # noqa: BLE001
            pass
        try:
            if getattr(s, "marker", None) is not None:
                s.marker.symbol = "none"
        except Exception:  # noqa: BLE001
            pass
    for ax_name in ("x_axis", "y_axis"):
        ax = getattr(ch, ax_name, None)
        if ax is None:
            continue
        try:
            ax.delete = False
            # openpyxl writes an axis TITLE that Excel renders OVERLAPPING the tick labels
            # (no layout reserve) — drop it; the chart title + the formatted tick labels carry
            # the meaning. (Tristan: "Output (t/yr)" / "Capex £" overwriting the tick labels.)
            ax.title = None
            if not gridlines:
                ax.majorGridlines = None
        except Exception:  # noqa: BLE001
            pass
    want = (legend is True) or (legend == "auto" and len(series) > 1)
    try:
        if want:
            from openpyxl.chart.legend import Legend
            if ch.legend is None:
                ch.legend = Legend()
            ch.legend.position = "r"
            ch.legend.overlay = False
        else:
            ch.legend = None
    except Exception:  # noqa: BLE001
        pass
    if data_labels:
        try:
            from openpyxl.chart.label import DataLabelList
            dl = DataLabelList()
            dl.showVal = True            # the value ONLY
            dl.showSerName = False       # NEVER "Series1" (Tristan: ugly + overwrites the title)
            dl.showCatName = False       # the category axis already names each bar
            dl.showLegendKey = False
            dl.showPercent = False
            dl.showBubbleSize = False
            ch.dataLabels = dl
        except Exception:  # noqa: BLE001
            pass
    return ch


# --- worked-calc verification: turn "static" legacy calcs into LIVE, checked ---
_ARITH_OPS = {_ast.Add: _op.add, _ast.Sub: _op.sub, _ast.Mult: _op.mul,
              _ast.Div: _op.truediv, _ast.Pow: _op.pow, _ast.Mod: _op.mod,
              _ast.USub: _op.neg, _ast.UAdd: _op.pos}
_ARITH_FUNCS = {"sqrt": _math.sqrt, "abs": abs, "exp": _math.exp,
                "log": _math.log, "ln": _math.log, "log10": _math.log10,
                "sin": _math.sin, "cos": _math.cos, "tan": _math.tan,
                "min": min, "max": max, "pow": pow}
_EXCEL_FUNCS = {"sqrt": "SQRT", "abs": "ABS", "exp": "EXP", "log": "LN",
                "ln": "LN", "log10": "LOG10", "sin": "SIN", "cos": "COS",
                "tan": "TAN", "min": "MIN", "max": "MAX", "pow": "POWER"}


def _eval_arith(node):
    if isinstance(node, _ast.Constant):
        if isinstance(node.value, (int, float)):
            return node.value
        raise ValueError("non-numeric constant")
    if isinstance(node, _ast.BinOp):
        return _ARITH_OPS[type(node.op)](_eval_arith(node.left), _eval_arith(node.right))
    if isinstance(node, _ast.UnaryOp):
        return _ARITH_OPS[type(node.op)](_eval_arith(node.operand))
    if isinstance(node, _ast.Call):
        fn = getattr(node.func, "id", "")
        if fn.lower() in _ARITH_FUNCS and not node.keywords:
            return _ARITH_FUNCS[fn.lower()](*[_eval_arith(a) for a in node.args])
        raise ValueError("unsupported call")
    if isinstance(node, _ast.Name) and node.id.lower() == "pi":
        return _math.pi
    raise ValueError("unsupported node")


def _clean_substitution(sub):
    """Normalise a worked-calc substitution to a bare arithmetic expression.
    Handles BOTH engine formats universally: pure '(a*b)/c' AND the structured
    'lhs = a x b x c = result UNIT' form (keep the middle expression; × / spaced-x
    → *). No class-specific parsing."""
    s = re.sub(r"^=\s*", "", str(sub or "").strip())
    if s.count("=") >= 2:                       # 'lhs = EXPR = result UNIT'
        s = s.split("=")[1]
    elif "=" in s:                              # 'lhs = expr' OR 'expr = result'
        a, b = s.split("=", 1)
        s = b if re.search(r"[-+*/x^()]", b) else a
    s = s.replace("×", "*").replace(",", "")
    s = re.sub(r"(?<=[\d\)\s])x(?=[\s\d(])", "*", s)   # spaced 'x' = multiply, never max()/exp()
    return s.strip()


def safe_eval_substitution(sub):
    """Evaluate a worked-calc expression (numbers + - * / ^, sqrt/abs/exp/log/trig,
    pi). Returns float, or None when it isn't pure arithmetic. Universal."""
    s = _clean_substitution(sub).replace("^", "**")
    if not s:
        return None
    try:
        return float(_eval_arith(_ast.parse(s, mode="eval").body))
    except Exception:  # noqa: BLE001
        return None


def substitution_to_excel(sub):
    """Convert a worked-calc substitution to a LIVE Excel formula ('=…'), or None
    when it isn't pure arithmetic. Lets a 'static' calc recompute itself in the
    sheet so the value is verifiable + editable. Universal."""
    if safe_eval_substitution(sub) is None:
        return None
    s = _clean_substitution(sub)
    s = re.sub(r"\bpi\b", "PI()", s, flags=re.I)
    for fn, xl in _EXCEL_FUNCS.items():
        s = re.sub(r"\b" + fn + r"\s*\(", xl + "(", s, flags=re.I)
    return "=" + s


def _qty_tokens(s):
    STOP = {"the", "a", "of", "per", "and", "to", "for", "at", "in", "kg", "day",
            "hr", "h", "m", "m2", "m3", "mm", "cm", "kw", "kwh", "gbp", "pct",
            "ratio", "yr", "year", "rate", "total", "design", "value", "load"}
    return {t for t in re.split(r"[^a-z0-9]+", str(s).lower())
            if t and t not in STOP and not t.isdigit()}


def index_design_quantities(state):
    """Flat (key, tokens, value, unit) list from the engineering contract — the
    authoritative design quantities a worked-calc result can be cross-checked
    against. Universal (reads orchestratorContract.quantities for any class)."""
    out = []
    q = ((state.get("orchestratorContract") or {}).get("quantities")) or {}
    if isinstance(q, dict):
        items = list(q.items())
    elif isinstance(q, list):
        items = [(e.get("key"), e) for e in q if isinstance(e, dict)]
    else:
        items = []
    for key, e in items:
        if not isinstance(e, dict):
            continue
        v = e.get("value")
        if isinstance(v, (int, float)):
            out.append((key, _qty_tokens(key), float(v), e.get("unit") or ""))
    return out


def match_design_quantity(label, res_val, qindex):
    """SAFE positive-only cross-check: returns (key, value) ONLY when EVERY label
    token appears in the quantity key AND the values agree within 10% — an
    unambiguous confirmation. Never claims a divergence (a similarly-named but
    DIFFERENT quantity, e.g. MBBR 'Tank Volume' 205 m³ vs rearing
    total_tank_volume 851 m³, must not raise a false alarm). Universal."""
    if not isinstance(res_val, (int, float)):
        return None
    lt = _qty_tokens(label)
    if not lt:
        return None
    for (key, ktoks, val, _unit) in qindex:
        if lt.issubset(ktoks) and val and abs(res_val - val) / abs(val) <= 0.10:
            return (key, val)
    return None


def _unit_from_symbol(sym: str) -> str:
    """Best-effort unit from a snake_case symbol's trailing unit token (solids_load_kg_day →
    kg/day, throughput_m3_h → m³/h). Empty when none recognised. Universal."""
    table = [
        ("kg_day", "kg/day"), ("kg_d", "kg/day"), ("mg_l", "mg/L"), ("g_kg", "g/kg"),
        ("m3_h", "m³/h"), ("m3h", "m³/h"), ("l_day", "L/day"), ("l_s", "L/s"),
        ("kwh", "kWh"), ("kw", "kW"), ("kpa", "kPa"), ("mpa", "MPa"), ("bar", "bar"),
        ("mm", "mm"), ("m2", "m²"), ("m3", "m³"), ("deg_c", "°C"), ("ppt", "ppt"),
        ("pct", "%"), ("hz", "Hz"), ("rpm", "rpm"),
    ]
    s = str(sym or "").lower()
    for suf, u in table:
        if s.endswith("_" + suf) or s == suf:
            return u
    return ""


def _infer_inputs_from_formula(formula, substitution):
    """Recover a legacy worked-calc's named inputs by aligning its FORMULA symbol tokens with
    its SUBSTITUTION number tokens (same expression structure) — so a calc that ships only
    formula+substitution text renders in the SAME yellow-input + live-result style as a
    structured calc (not a divergent 'recomputed live' text block). Returns
    [{symbol, value, unit}] or None when the two don't align 1:1. Universal — no class logic."""
    f = re.sub(r"^\s*[A-Za-z_]\w*\s*=\s*", "", str(formula or "").strip())   # drop a leading 'lhs ='
    s = re.sub(r"^\s*[A-Za-z_]\w*\s*=\s*", "", str(substitution or "").strip())
    s = re.split(r"=", s)[0].strip()                                          # drop a trailing '= result unit'
    if not f or not s:
        return None
    tokre = r"[A-Za-z_]\w*|\d+\.?\d*(?:[eE][-+]?\d+)?|[-+*/^()]|×|x(?=[\s\d(])"
    ft, st = re.findall(tokre, f), re.findall(tokre, s)
    if not ft or len(ft) != len(st):
        return None
    SYM, NUM = re.compile(r"^[A-Za-z_]\w*$"), re.compile(r"^-?\d+\.?\d*(?:[eE][-+]?\d+)?$")
    inputs: Dict[str, Dict] = {}
    for a, b in zip(ft, st):
        a_sym, b_num = bool(SYM.match(a)), bool(NUM.match(b))
        if a_sym and b_num:
            if a not in inputs:
                inputs[a] = {"symbol": a, "value": float(b), "unit": _unit_from_symbol(a)}
        elif a_sym and not b_num:
            if a.lower() != str(b).lower():     # a function name etc. must match verbatim
                return None
        elif not a_sym and a != b:              # operator/constant must be identical in both
            try:
                if float(a) != float(b):
                    return None
            except Exception:
                return None
    return list(inputs.values()) or None


def tab_calculations(wb: Workbook, state: dict, run_dir: str) -> Tuple[int, int]:
    """
    Returns (live_calc_count, static_calc_count).

    Layout per STRUCTURED calc:
      * each input symbol -> a labelled YELLOW cell (editable)
      * the formula -> a LIVE Excel formula referencing those cells (green)
      * within a tool, if an input's value equals a prior result's value, the
        input cell REFERENCES the producing result cell (chained, not duplicated)
      * the live result sits next to the engine's stored result + a Δ
    LEGACY calcs (no inputs[]) render as static text (formula + substitution +
    result), clearly marked "static — no input map".
    Shared constants (rho, g, pi, ...) live in a constants block at the top and
    are referenced by absolute address.
    """
    ws = wb.create_sheet("Calculations")
    set_widths(ws, {"A": 30, "B": 18, "C": 12, "D": 46, "E": 16, "F": 12, "G": 10, "H": 64})
    title_row(
        ws, "Worked calculations — every value recomputed live + checked", 8,
        "Yellow = editable input. Green col B = LIVE formula. 'Engine value' = the "
        "value the engine stored; Δ should be ~0 (inputs are display-rounded to 4 s.f.). "
        "Legacy calcs (no input map) are RECOMPUTED LIVE from their substitution and "
        "cross-checked against the engine value (col H verdict).",
    )
    r = 4

    # ---- shared constants block ----
    CONSTANTS = {
        "rho": (1025.0, "kg/m³", "seawater density"),
        "rho_air": (1.2, "kg/m³", "air density"),
        "g": (9.8066, "m/s²", "gravity"),
        "pi": (3.141592653589793, "-", "pi (use PI() in formulas)"),
        "mu": (0.001, "Pa·s", "dynamic viscosity (water)"),
    }
    sub_banner(ws, r, "Shared constants (referenced by formulas below)", 8)
    r += 1
    header(ws, r, ["Constant", "Value", "Unit", "Note"])
    r += 1
    const_cell: Dict[str, str] = {}
    for sym, (val, unit, note) in CONSTANTS.items():
        ws.cell(r, 1, sym).font = FONT_MONO
        cc = ws.cell(r, 2, val)
        cc.fill = FILL_CONST
        const_cell[sym] = f"$B${r}"  # absolute ref so it survives anywhere
        ws.cell(r, 3, unit)
        ws.cell(r, 4, note).font = FONT_NOTE
        r += 1
    r += 1

    # ---- Tool data-flow section (INPUT → TOOL → OUTPUT destination) ----------
    # Tristan 2026-06-25: "the spreadsheet should have all of the tools with a clear
    # input of where the information came from to start that tool, and then a clear
    # exit as to where the information from that tool goes to." That flow lived ONLY
    # on the ⚠ Checks tab — the founder looks for calcs HERE. Reuse the SAME
    # _render_tool_io_section (Tool | Output field | Value | Unit | Input—from |
    # → Consumed by | Status) so each tool's inputs and output destinations are
    # visible ABOVE the worked-calc transcript. Kept on ⚠ Checks too; this is an
    # ADDITIONAL surface, not a move.
    df_banner = ws.cell(
        r, 1,
        "Tool data-flow — where each tool's inputs came from and where its outputs go")
    df_banner.font = FONT_TITLE
    df_banner.fill = FILL_TITLE
    df_banner.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(
        r, 1,
        "Each row: a tool OUTPUT, the INPUT it was computed from (— from), and where "
        "that value GOES (→ Consumed by). Status: USED/TRACED = the output flows on; "
        "ORPHANED = it appears nowhere downstream; STALE = the downstream value disagrees.",
    ).font = FONT_NOTE
    r += 2
    io_next = _render_tool_io_section(ws, state, run_dir, r)
    if io_next is None:
        ws.cell(r, 1, "— no orchestrator tool log for this run —").font = FONT_NOTE
        r += 2
    else:
        # live conditional formatting on the Status column (G) for this section
        from openpyxl.formatting.rule import CellIsRule as _CIR
        _rng = f"G{r + 1}:G{io_next - 1}"
        if io_next - 1 >= r + 1:
            ws.conditional_formatting.add(
                _rng, _CIR(operator="equal", formula=['"ORPHANED"'], fill=FILL_FAIL, font=FONT_FAIL))
            ws.conditional_formatting.add(
                _rng, _CIR(operator="equal", formula=['"STALE"'], fill=FILL_FAIL, font=FONT_FAIL))
            ws.conditional_formatting.add(
                _rng, _CIR(operator="equal", formula=['"USED"'], fill=FILL_PASS, font=FONT_PASS))
            ws.conditional_formatting.add(
                _rng, _CIR(operator="equal", formula=['"TRACED"'], fill=FILL_PASS, font=FONT_PASS))
        r = io_next + 1

    # ---- worked-calc transcript (formula + substitution + result, per tool) --
    wc_banner = ws.cell(
        r, 1, "Worked calculations — every value recomputed live + cross-checked")
    wc_banner.font = FONT_TITLE
    wc_banner.fill = FILL_TITLE
    wc_banner.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 24
    r += 2

    tools = (state.get("toolsUsedPage") or {}).get("tools", [])
    qindex = index_design_quantities(state)   # authoritative design quantities to cross-check against
    live_count = 0
    static_count = 0

    for tool in tools:
        worked = tool.get("worked") or []
        if not worked:
            continue
        tname = tool.get("tool_name") or tool.get("tool_id") or "tool"
        tid = tool.get("tool_id", "")
        sub_banner(ws, r, f"{tname}   ·   {tid}", 8)
        r += 1
        header(ws, r, ["Calc / input", "Value (live)", "Unit",
                       "Formula (text)", "Engine value", "Δ", "Unit", "Assumptions"])
        r += 1

        # produced[output_symbol] -> (producing result cell, its value), for
        # within-tool chaining. BUG #19 FIX: key by the producing calc's OUTPUT
        # SYMBOL (formula LHS), not by numeric value — two calcs producing the same
        # number (q_wall and q_roof both 4920) must not collapse to one cell, and a
        # later input must chain to the result with the MATCHING symbol+label, not
        # whichever happened to register last at that value.
        produced: Dict[str, Tuple[str, float]] = {}

        for w in worked:
            label = w.get("label", "")
            formula = w.get("formula", "")
            inputs = w.get("inputs") or []
            if not inputs:
                # recover named inputs from formula+substitution so a LEGACY calc renders in the
                # SAME yellow-input + live-result style as a structured one (not a divergent
                # 'recomputed live' text block). Falls through to the legacy text only if it can't.
                inferred = _infer_inputs_from_formula(formula, w.get("substitution", ""))
                if inferred:
                    inputs = inferred
            res = w.get("result")
            res_val = res.get("value") if isinstance(res, dict) else res
            res_unit = (res.get("unit") if isinstance(res, dict) else w.get("result_unit", "")) or ""

            structured = bool(inputs)

            # ---- calc title row ----
            tc = ws.cell(r, 1, label)
            tc.font = FONT_SUB
            ws.cell(r, 4, formula).font = FONT_MONO
            sub_text = str(w.get("substitution", "") or "")
            xl_formula = substitution_to_excel(sub_text) if not structured else None
            ev_val = safe_eval_substitution(sub_text) if not structured else None
            if not structured:
                marker = "  [build-verified]" if ev_val is not None else "  [static — no auto-check]"
                ws.cell(r, 1, label + marker).font = Font(bold=True, italic=True, color="7F7F7F")
            calc_title_row = r
            r += 1

            if not structured:
                # LEGACY calc: show formula + substitution as transparent text, then
                # RECOMPUTE the substitution as a LIVE Excel formula and check it
                # against the engine's stored result. This turns an unchecked
                # printed number ("these values have not been checked") into a
                # self-verifying, editable cell. Universal — works for any tool's
                # worked-calc with a numeric substitution.
                ws.cell(r, 1, "formula").font = FONT_NOTE
                ws.cell(r, 4, str(formula)).font = FONT_MONO
                ws.cell(r, 4).fill = FILL_LEGACY
                r += 1
                ws.cell(r, 1, "substitution").font = FONT_NOTE
                sc = ws.cell(r, 4, sub_text)
                sc.font = FONT_MONO
                sc.fill = FILL_LEGACY
                r += 1
                ws.cell(r, 1, "result").font = FONT_NOTE
                if isinstance(res_val, (int, float)):
                    _bval = ev_val if isinstance(ev_val, (int, float)) else res_val
                    lc = ws.cell(r, 2, _bval)             # STATIC computed result — never a live worked-calc formula (Excel-safe)
                    lc.fill = FILL_RESULT
                    lc.number_format = "#,##0.0000"
                    ws.cell(r, 5, res_val)                # engine value
                    ws.cell(r, 6, f"=B{r}-E{r}").number_format = "#,##0.0000"
                    ws.cell(r, 7, res_unit)
                    drift = (abs(ev_val - res_val) / abs(res_val)
                             if (ev_val is not None and res_val) else 0.0)
                    if ev_val is None:
                        verdict, vfont = "— result shown (substitution not auto-evaluable)", FONT_NOTE
                        static_count += 1
                    elif drift <= 0.02:
                        verdict, vfont = "✓ maths checks out (build-verified)", FONT_PASS
                        live_count += 1
                    else:
                        verdict = (f"⚠ engine result ≠ its own substitution "
                                   f"({ev_val:.4g} vs {res_val:.4g})")
                        vfont = FONT_FAIL
                        static_count += 1
                    dq = match_design_quantity(label, res_val, qindex)
                    if dq:
                        verdict += f"   ·   confirmed = design {dq[0]}"
                    vc = ws.cell(r, 8, verdict)
                    vc.font = vfont
                    vc.alignment = WRAP_TOP
                else:
                    # non-arithmetic substitution (rare) — keep honest static text
                    static_count += 1
                    ws.cell(r, 5, res_val)
                    ws.cell(r, 7, res_unit)
                    ws.cell(r, 8, "— not auto-checkable (non-numeric substitution)").font = FONT_NOTE
                r += 1
                r += 1  # spacer
                continue

            # ---- STRUCTURED: lay each input as a labelled yellow cell ----
            symbol_cell: Dict[str, str] = {}
            for inp in inputs:
                sym = inp.get("symbol", "")
                ival = num(inp.get("value"))
                iunit = inp.get("unit", "") or ""
                ws.cell(r, 1, f"  {sym}").font = FONT_MONO
                cell = ws.cell(r, 2)
                # chain by IDENTITY (bug #19): reference a prior result ONLY when
                # this input's SYMBOL matches that result's producing output symbol
                # AND the values agree. Symbol-keying means q_wall and q_roof (both
                # 4920) each chain to their OWN producing cell, never the other's.
                ref = None
                if sym and sym in produced:
                    pcell, pval = produced[sym]
                    if ival is None or abs(pval - ival) <= max(1e-9, abs(pval) * 1e-6):
                        ref = pcell
                # also reference a shared constant if the symbol is one
                if ref is None and sym in const_cell and ival is not None:
                    cval = CONSTANTS[sym][0]
                    if abs(cval - ival) <= max(1e-9, abs(cval) * 1e-4):
                        ref = const_cell[sym]
                if ref is not None:
                    cell.value = f"={ref}"
                    cell.fill = FILL_RESULT  # chained / constant -> not a free input
                    ws.cell(r, 8, f"↳ chained from {ref}").font = FONT_NOTE
                else:
                    cell.value = ival
                    cell.fill = FILL_INPUT   # free editable input
                symbol_cell[sym] = cell.coordinate
                ws.cell(r, 3, iunit)
                r += 1

            # ---- result row (STATIC computed value — NEVER a live worked-calc formula) ----
            # A live worked-calc formula (formula_to_excel / substitution_to_excel) can be Excel-INVALID
            # for some classes — unbound physics symbols (a bare "E" joint-efficiency), sci-notation, or
            # patterns Excel's file loader rejects though openpyxl + LibreOffice both accept them. Result:
            # "Removed Records: Formula from sheetN.xml" and a workbook that won't open (the recurring SAF
            # break). The Calculations tab is a TRANSCRIPT (formula + substitution + result), so the result
            # is emitted as the STATIC computed value and the formula-vs-result reconciliation is done HERE
            # in Python and shown as a static verdict — which is exactly the "do the Excel numbers reconcile
            # with the physics" check. Universal across every archetype. (Economics/Scenarios stay live.)
            rhs = rhs_of(formula)
            sub_s = str(w.get("substitution", "") or "")
            ev_s = safe_eval_substitution(sub_s)
            _bval = ev_s if isinstance(ev_s, (int, float)) else (res_val if isinstance(res_val, (int, float)) else None)
            ws.cell(r, 1, "  = result").font = FONT_SUB
            live_cell = ws.cell(r, 2)
            produced_ok = False
            if _bval is not None:
                live_cell.value = _bval
                live_cell.fill = FILL_RESULT
                produced_ok = True
                if isinstance(ev_s, (int, float)) and isinstance(res_val, (int, float)) and res_val:
                    drift_s = abs(ev_s - res_val) / abs(res_val)
                    if drift_s <= 0.02:
                        live_count += 1
                    else:
                        ws.cell(r, 8, f"⚠ engine result ≠ substitution "
                                f"({ev_s:.4g} vs {res_val:.4g})").font = FONT_FAIL
                        static_count += 1
                else:
                    live_count += 1
            else:
                static_count += 1
            ws.cell(r, 3, res_unit)
            ws.cell(r, 4, rhs).font = FONT_MONO
            # engine's stored value + delta
            ws.cell(r, 5, res_val)
            if isinstance(res_val, (int, float)):
                # Δ = live - engine ; if live is a formula, reference it
                ws.cell(r, 6, f"=B{r}-E{r}")
            ws.cell(r, 7, res_unit)
            for a in (w.get("assumptions") or [])[:1]:
                ac = ws.cell(r, 8, str(a))
                ac.font = FONT_NOTE
                ac.alignment = WRAP_TOP

            # register this result under its OUTPUT SYMBOL so later inputs chain by
            # symbol+label identity (bug #19), not by a numeric-value collision.
            out_sym = lhs_symbol(formula)
            if out_sym and isinstance(res_val, (int, float)):
                produced[out_sym] = (f"$B${r}", float(res_val))
            r += 2  # spacer between calcs

        r += 1  # spacer between tools

    # number formats (#37): value (B) + engine-value (E) + Δ (F) columns. These
    # carry per-calc results in mixed units, so a thousands-separated 2-dp mask
    # (not £, not General/scientific) is the safe universal display.
    apply_col_formats(ws, 5, {2: FMT_DEC2, 5: FMT_DEC2, 6: FMT_DEC2}, r)
    ws.freeze_panes = "A4"
    back_link(ws, 8)
    return live_count, static_count


# ── connection-trace phantom-reference resolver (Tristan 2026-06-25) ─────────
# A pure transform over the trace adjacency that guarantees the connection trace can
# only reference parts that EXIST. The ledger authors endpoint labels as derived
# strings (title-cased / abbreviated / detail-stripped / "[N]"-indexed); this maps each
# back to the real bill-of-materials part name it was derived from, and DROPS any label
# that resolves to no real part (a pure routing-graph artefact). UNIVERSAL — keyed on
# token-subset matching against state.requirementsBom, never a per-class rename table.

# Explicit process-boundary endpoints (a connection's far side IS the plant battery limit,
# legitimately not a part) — these are KEPT as honest boundary labels, not dropped as
# phantoms. Universal boundary nouns, independent of class.
_TRACE_BOUNDARY_RE = re.compile(
    r"\b(battery limit|boundary|b/?l|to grid|from grid|grid tie|"
    r"atmosphere|ambient|disposal|effluent|drain|vent to|flare|"
    r"utility|off-?site|external supply|feedstock supply|product export)\b",
    re.I)


def _trace_tokens(s) -> List[str]:
    """Lowercase alphanumeric tokens of a part / endpoint name, detail tail (· …) and the
    ↳ child marker stripped, '[N]' index dropped — the join key for endpoint resolution."""
    import unicodedata as _ud
    t = _ud.normalize("NFKC", str(s or "")).replace("↳", " ")
    for sep in ("·", "•"):
        if sep in t:
            t = t.split(sep)[0]
    return [x for x in re.findall(r"[a-z0-9]+", t.lower())]


def _build_real_part_index(state: dict) -> List[Tuple[str, set]]:
    """(display_name, token_set) for every REAL principal part in the bill of materials —
    skips connection/pipework lines (the '→' / 'connection' signal) and ↳ sub-components,
    mirroring tab_parts_master's principal filter. The set the trace must resolve against."""
    out: List[Tuple[str, set]] = []
    seen: set = set()
    for b in (state.get("requirementsBom") or []):
        if not isinstance(b, dict):
            continue
        req = str(b.get("requirement") or "")
        if "→" in req or re.search(r"\bconnection\b", req, re.I):
            continue
        if req.strip().startswith("↳"):
            continue
        nm = b.get("name_human") or req
        ts = set(_trace_tokens(nm))
        if not ts:
            continue
        disp = _clean_name(nm) or str(nm).strip()
        key = (disp.lower(), frozenset(ts))
        if key in seen:
            continue
        seen.add(key)
        out.append((disp, ts))
    return out


def _resolve_endpoint_name(ep: str, real_index: List[Tuple[str, set]]) -> Optional[str]:
    """Resolve a derived endpoint label to the REAL part name it came from, or None when it
    matches no real part. Rule (universal): the endpoint's tokens must be a SUBSET of a real
    part's tokens, where a short endpoint token may match a real token by PREFIX so the
    engine's abbreviations resolve ('bm'⊂'bms', 'pc'⊂'pcs', 'step up transformer'⊂'step-up
    transformer'). Among candidates the SHORTEST real name wins (the tightest match). A pure
    routing artefact ('Heat Rejection', '(Busway)', 'Rack Block', 'Rack[0]') subsets nothing
    → None (the caller drops it)."""
    ets = set(_trace_tokens(ep))
    if not ets or not any(len(t) >= 2 for t in ets):
        return None
    best: Optional[str] = None
    best_len = 10 ** 9
    for disp, rt in real_index:
        ok = True
        for et in ets:
            if et in rt:
                continue
            if len(et) >= 2 and any(rtok.startswith(et) for rtok in rt):
                continue
            ok = False
            break
        if ok and len(rt) < best_len:
            best, best_len = disp, len(rt)
    return best


def _resolve_trace_endpoints(adj: dict, state: dict) -> dict:
    """Rewrite the trace adjacency so EVERY node + every input/output endpoint references a
    real bill-of-materials part (or an explicit process boundary): derived labels are mapped
    to the real part name, phantoms are dropped along with the edges that referenced them.
    Nodes whose own name is a phantom are removed entirely; resolved nodes are merged when two
    derived labels collapse onto the same real part. Pure transform — returns a fresh dict."""
    real_index = _build_real_part_index(state)

    def _canon(name: str) -> Optional[str]:
        s = str(name or "").strip()
        if not s:
            return None
        r = _resolve_endpoint_name(s, real_index)
        if r:
            return r
        if _TRACE_BOUNDARY_RE.search(s):
            return s  # honest boundary, kept as-is
        return None  # phantom — drop

    out: Dict[str, Dict[str, list]] = {}
    for node, a in (adj or {}).items():
        cn = _canon(node)
        if not cn:
            continue  # node itself is a phantom — drop it and its edges
        slot = out.setdefault(cn, {"inputs": [], "outputs": []})
        a = a or {}
        for i in (a.get("inputs") or []):
            src = _canon(i.get("from"))
            if src:
                slot["inputs"].append({"from": src, "service": i.get("service", "")})
        for o in (a.get("outputs") or []):
            dst = _canon(o.get("to"))
            if dst:
                slot["outputs"].append({"to": dst, "service": o.get("service", "")})
    # de-dup merged inputs/outputs (two derived labels collapsing onto one real part can
    # double an identical edge)
    for slot in out.values():
        seen_i, di = set(), []
        for i in slot["inputs"]:
            k = (i["from"], i.get("service", ""))
            if k not in seen_i:
                seen_i.add(k)
                di.append(i)
        slot["inputs"] = di
        seen_o, do = set(), []
        for o in slot["outputs"]:
            k = (o["to"], o.get("service", ""))
            if k not in seen_o:
                seen_o.add(k)
                do.append(o)
        slot["outputs"] = do
    return out


# ============================================================================
# TAB 5 — BoM (requirementsBom + coverage_by_drawing + Σ check)
# ============================================================================
def tab_connection_trace(wb: Workbook, state: dict, run_dir: str) -> bool:
    """Connection Trace — the ledger's per-part adjacency (Tristan 2026-06-20: "if line 1
    connects to line 3, line 3 should say its input is from line 1; in Excel we should
    trace whole systems this way"). One row per part: what feeds INTO it (by name) and
    what it feeds OUT to (by name), bidirectionally consistent. Ctrl-F any part name to
    follow the chain. Reads connection-ledger.json (authored by connection_ledger.py).
    Returns False (tab skipped) when no ledger is present."""
    led = load_json(os.path.join(run_dir, "connection-ledger.json"))
    if not isinstance(led, dict) or not led.get("adjacency"):
        # UNIVERSAL fallback (Tristan 2026-06-24): connection-ledger.json is Blender-authored and absent
        # in PDF-off / Excel-only mode, so this "decks" trace tab was RAS-only. parts_ledger.py writes a
        # flat `connections` list + a `connectivity` completeness block for EVERY archetype (bpy-free) —
        # fold it into the same adjacency shape so the connector deck is universal.
        _pl = load_json(os.path.join(run_dir, "parts-ledger.json"))
        _conns = _pl.get("connections") if isinstance(_pl, dict) else None
        if not _conns:
            return False
        _adj: Dict[str, Dict[str, list]] = {}
        for _c in _conns:
            _fp = str(_c.get("from_part") or "").strip()
            _tp = str(_c.get("to_part") or "").strip()
            _svc = str(_c.get("service") or _c.get("mech") or _c.get("line_number") or "")
            if _fp:
                _adj.setdefault(_fp, {"inputs": [], "outputs": []})
                if _tp:
                    _adj[_fp]["outputs"].append({"to": _tp, "service": _svc})
            if _tp:
                _adj.setdefault(_tp, {"inputs": [], "outputs": []})
                if _fp:
                    _adj[_tp]["inputs"].append({"from": _fp, "service": _svc})
        _cc = (_pl.get("connectivity") or {}) if isinstance(_pl, dict) else {}
        led = {"adjacency": _adj, "count": len(_conns),
               "completeness": {
                   "n_concerns": int(_cc.get("n_concerns") or len(_cc.get("concerns") or [])),
                   "concerns": [{"part": (x.get("name") or x.get("tag") or "?"),
                                 "missing": [x.get("issue")] if x.get("issue") else (x.get("missing") or [])}
                                for x in (_cc.get("concerns") or [])]},
               "referential_integrity": {}}
    adj = led["adjacency"]
    # ── PHANTOM-REFERENCE FIX (Tristan 2026-06-25, dossier_audit phantom_reference HIGH):
    # the ledger / parts-ledger author endpoint labels as abbreviated / title-cased /
    # detail-stripped / indexed derived strings ("Bm Ctrl", "Pc Inverter 1 Mw …",
    # "Heat Rejection", "(Busway)", "Rack[0]"). Some derive from a REAL bill-of-materials
    # part (just re-cased / abbreviated); others are pure routing-graph artefacts that match
    # NO real part. The connection trace must only ever reference parts that EXIST, so we
    # resolve every endpoint against the real part-name set: a derived label is rewritten to
    # the real part name it came from, and a label that resolves to nothing real (and is not
    # an explicit process boundary) is DROPPED — never shown as a phantom. UNIVERSAL: keyed
    # purely on token-subset matching against state.requirementsBom, no per-class rename table.
    adj = _resolve_trace_endpoints(adj, state)
    comp = led.get("completeness") or {}
    ri = led.get("referential_integrity") or {}
    incomplete = {c.get("part"): c.get("missing", []) for c in (comp.get("concerns") or [])}

    # tag lookup: part name -> BoM tag (so the trace carries the same tag as the
    # BoM / drawings). Exact name first, then a loose contains-match.
    tag_by_name: Dict[str, str] = {}
    for b in (state.get("requirementsBom") or []):
        if not isinstance(b, dict):
            continue
        tg = clean_cell(b.get("tag", ""))
        if not tg:
            continue
        for key in (b.get("name_human"), b.get("requirement")):
            if key:
                tag_by_name.setdefault(str(key).strip().lower(), tg)

    def _tag_for(nm: str) -> str:
        n = str(nm).strip().lower()
        if n in tag_by_name:
            return tag_by_name[n]
        for k, t in tag_by_name.items():
            if k and (k in n or n in k):
                return t
        return ""

    ws = wb.create_sheet("Connection trace")
    set_widths(ws, {"A": 38, "B": 12, "C": 10, "D": 46, "E": 46, "F": 22})
    title_row(ws, "Connection trace — which part connects to what", 6,
              "The ledger authors every connection. INPUTS / OUTPUTS are LIVE cell-references "
              "to the connected part's own row (col A): select a cell and use Formulas ▸ Trace "
              "Precedents / Trace Dependents to draw the arrows, or click through to navigate. "
              "Each part also carries its Tag. Bidirectionally consistent.")
    # health banner
    sub_banner(ws, 4,
               f"{led.get('count', len(adj))} authored connections   ·   "
               f"completeness: {comp.get('n_concerns', 0)} part(s) missing a required tie   ·   "
               f"referential integrity: {ri.get('n_violations', 0)} broken reference(s)", 6)
    header(ws, 5, ["Part", "Tag", "Status", "Inputs ← (from)", "Outputs → (to)", "Services"])

    # PASS 1 — assign every part a row so inputs/outputs can reference it by cell.
    names_sorted = sorted(adj.keys(), key=lambda s: str(s).lower())
    first_row = 6
    name_to_row = {nm: first_row + i for i, nm in enumerate(names_sorted)}

    def _ref_formula(conn_names) -> Optional[str]:
        """A formula referencing each connected part's A-cell (so Excel can trace it)
        — unknown / boundary endpoints fall back to a quoted literal."""
        parts = []
        for nm in dict.fromkeys(conn_names):
            row = name_to_row.get(nm)
            if row:
                parts.append(f"A{row}")
            else:
                parts.append('"' + str(nm).replace('"', "'") + '"')
        if not parts:
            return None
        return "=" + '&", "&'.join(parts)

    # PASS 2 — write the rows.
    r = first_row
    for name in names_sorted:
        a = adj[name] or {}
        ins = a.get("inputs") or []
        outs = a.get("outputs") or []
        in_list = [str(i.get("from")) for i in ins]
        out_list = [str(o.get("to")) for o in outs]
        svcs = ", ".join(sorted({(i.get("service") or "") for i in ins} |
                                {(o.get("service") or "") for o in outs} - {""})) or "—"
        miss = incomplete.get(name)
        status = "OK" if not miss else "missing: " + ", ".join(miss)
        # Part name + tag REFERENCE the master "Part names" tab where the part is a
        # registered principal (one identity across the workbook); literal fallback for
        # boundary / aggregate endpoints the master doesn't carry.
        _pn = name_ref(name)
        ws.cell(r, 1, _pn if _pn else clean_cell(name)).border = BORDER
        _pt = name_ref(name, "tag")
        ws.cell(r, 2, _pt if _pt else clean_cell(_tag_for(name))).border = BORDER
        sc = ws.cell(r, 3, clean_cell(status))
        sc.border = BORDER
        if miss:
            sc.font = Font(color="C00000", bold=True)
        in_f = _ref_formula(in_list)
        ci = ws.cell(r, 4, in_f if in_f else "—")
        ci.alignment = WRAP_TOP
        ci.border = BORDER
        out_f = _ref_formula(out_list)
        co = ws.cell(r, 5, out_f if out_f else "—")
        co.alignment = WRAP_TOP
        co.border = BORDER
        ws.cell(r, 6, clean_cell(svcs)).border = BORDER
        r += 1
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:F{max(6, r - 1)}"
    # ── Cabinets section (Tristan 2026-06-24 consolidation): the deterministic proof
    # that every small electrical / control device is HOUSED in a cabinet and that the
    # cabinet + its contents show their IN/OUT connectors, all connected. Reads the SAME
    # parts-ledger connectivity verdict as the trace above, so the two cannot disagree. --
    r += 1  # spacer
    big = ws.cell(r, 1, "Cabinets — housed devices + connector proof")
    big.font = FONT_TITLE
    big.fill = FILL_TITLE
    big.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 24
    r += 2
    nxt = _render_cabinet_section(ws, run_dir, r)
    if nxt is None:
        ws.cell(r, 1, "— no electrical / control cabinets itemised for this design —").font = FONT_NOTE
    back_link(ws, 6)
    return True


# ── CANONICAL PART REGISTRY (the BoM tab IS the single source) ────────────────
# tab_bom populates this {tag: {field: "'BoM'!$X$n"}} map as it renders each line,
# so every OTHER part-listing tab (Spec sheets, Panel schedule, Process line/valve/
# instrument lists) can REFERENCE the part's name / tag / unit £ / qty BY CELL rather
# than repeating the string — one canonical identity, fully traceable, and a price
# edit on the BoM propagates everywhere (Tristan 2026-06-21: "referenced parts rather
# than repeated again and again"). Tags that do not resolve fall back to the literal,
# so a tab never breaks. Cleared at the start of each build so a re-run is never stale.
_BOM_REGISTRY: Dict[str, Dict[str, str]] = {}


def _norm_tag(tag) -> str:
    """Normalise a tag for registry lookup — trim + lowercase so 'P-101' / 'p-101 '
    resolve identically across the BoM and the consumer tabs."""
    return str(tag or "").strip().lower()


def bom_ref(tag, field: str = "name") -> Optional[str]:
    """The cross-sheet cell-reference FORMULA (e.g. "='BoM'!$B$7") for a part's field
    on the canonical BoM registry, or None if the tag is not a BoM line. field ∈
    {tag, name, qty, part, unit, line}. Callers write the returned string straight into
    a cell (it carries the leading '='); on None they keep their own literal."""
    rec = _BOM_REGISTRY.get(_norm_tag(tag))
    if not rec or field not in rec:
        return None
    return "=" + rec[field]


# ── MASTER PART-NAME REGISTRY (the "Part names" tab IS the single source) ──────
# Tristan 2026-06-21, stated four times: "a master series of NAMES, and all other
# names referenced in the spreadsheet on different tabs will be referencing the
# names, not using the names from scratch again … when the name is first introduced,
# every single other reference to that name needs to reference that name."
#
# The BoM registry above is keyed by TAG — but the engine carries TWO independent tag
# namespaces (requirementsBom mints catalogue-class tags B-201/FCV-201/SYS; the parts
# manifest mints shape-class tags K-103/I-114), so a TAG can't be the cross-tab join.
# The PART NAME is the one identity every tab and both namespaces share ("Aeration
# Blower" is "Aeration Blower" in the BoM, the drawings, the ledger and the spec sheet).
# So the master is keyed by NORMALISED NAME: the "Part names" tab introduces each
# principal's name ONCE; every other tab references that cell. Edit a name there and it
# propagates across the whole workbook; Formulas ▸ Trace Dependents follows it.
# Principals only — sub-component nouns ("Casing", "Impeller", "Top Head") repeat across
# parents and are NOT global identities; they stay exploded under their parent in the BoM.
_NAME_REG: Dict[str, Dict[str, str]] = {}
# Parallel index by NORMALISED TAG, so a tab that knows only the part's tag (the
# Line & velocity From/To endpoints, a schedule's served-equipment column) can also
# reference the master row. Populated alongside _NAME_REG in tab_parts_master.
_TAG_REG: Dict[str, Dict[str, str]] = {}


def _norm_name(s) -> str:
    """Normalise a human part name to its cross-tab identity key: NFKC-fold, strip the
    ↳ child marker, cut the ' · detail' requirement tail, lowercase, collapse runs of
    whitespace, drop edge punctuation. 'Recirc Pump · 110 kW (78 kW shaft)' and
    'recirc  pump' both → 'recirc pump'."""
    import unicodedata
    t = unicodedata.normalize("NFKC", str(s or "")).replace("↳", " ")
    for sep in ("·", "•", "—", " - ", " | "):
        if sep in t:
            t = t.split(sep)[0]
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t.strip(" .,:;-")


def _clean_name(s) -> str:
    """The display name = the requirement head before the ' · ' detail tail, ↳ stripped.
    Keeps original casing/unicode (e.g. 'Dissolved-O₂ Control Valve')."""
    t = str(s or "").replace("↳", "").strip()
    for sep in ("·", "•"):
        if sep in t:
            t = t.split(sep)[0]
    return t.strip(" .,:;-")


def _detail_tail(s) -> str:
    """Everything AFTER the name head — the ' · 110 kW …' descriptor — or '' if none."""
    t = str(s or "").replace("↳", "").strip()
    for sep in ("·", "•"):
        if sep in t:
            return sep + " " + sep.join(t.split(sep)[1:]).strip()
    return ""


def name_ref(name, field: str = "name") -> Optional[str]:
    """The cross-sheet cell-reference FORMULA (e.g. "='Part names'!$B$7") for a part's
    master NAME, or None when the name is not a registered principal (caller keeps its
    literal). field ∈ {name, tag}. Carries the leading '=' — write it straight to a cell."""
    rec = _NAME_REG.get(_norm_name(name))
    if not rec or field not in rec:
        return None
    return "=" + rec[field]


def tag_ref(tag, field: str = "tag") -> Optional[str]:
    """The cross-sheet cell-reference FORMULA for a part identified by its TAG on the
    master "Part names" tab, or None when the tag is not a registered principal (caller
    keeps its literal). field ∈ {tag, name}. Carries the leading '='."""
    rec = _TAG_REG.get(_norm_tag(tag))
    if not rec or field not in rec:
        return None
    return "=" + rec[field]


def tab_parts_master(wb: Workbook, state: dict, run_dir: str) -> None:
    """THE master list of part NAMES (Tristan 2026-06-21). One row per distinct principal
    part: its canonical Tag + Name typed ONCE here. Populates _NAME_REG so every other
    tab references these cells instead of repeating the string. Built FIRST so all
    consumers can resolve. Principals only (a tag with no '.' suffix whose requirement is
    not a ↳ sub-component)."""
    _NAME_REG.clear()
    _TAG_REG.clear()
    ws = wb.create_sheet("Part names")
    set_widths(ws, {"A": 14, "B": 46, "C": 8, "D": 64})
    title_row(ws, "Part names — the master list", 4,
              "THE single source of every part's name. Each name is typed ONCE here; every "
              "other tab (BoM, Cost, Spec sheets, Connection trace, schedules) REFERENCES these "
              "cells, never repeats the string — edit a name here and it updates across the whole "
              "workbook. Select a Name cell and use Formulas ▸ Trace Dependents to see everywhere "
              "it is used. One row per principal part; sub-components are listed under their parent "
              "on the BoM.")
    header(ws, 4, ["Tag", "Name", "Qty", "Requirement (as stated)"])
    bom = state.get("requirementsBom") or []
    # collect principals in document order, dedup by normalised name (first wins)
    seen: set = set()
    principals: List[dict] = []
    for row in bom:
        if not isinstance(row, dict):
            continue
        tag = str(row.get("tag", "") or "").strip()
        req = str(row.get("requirement", "") or "")
        if "." in tag or req.strip().startswith("↳"):
            continue                         # sub-component / instance — not a master name
        # Connection / pipework lines ("air connection: A → B", tag C01…) are priced in the
        # BoM but are not PARTS — keep them off the name master (the arrow / "connection" noun
        # is the universal signal, independent of class).
        if "→" in req or re.search(r"\bconnection\b", req, re.I):
            continue
        nm = _clean_name(req)
        key = _norm_name(req)
        if not nm or not key or key in seen:
            continue
        seen.add(key)
        principals.append({"tag": tag, "name": nm, "qty": row.get("qty"), "req": req})
    # stable, readable order: by name (the thing the reader scans)
    principals.sort(key=lambda p: p["name"].lower())
    r = 5
    for p in principals:
        ws.cell(r, 1, clean_cell(p["tag"]) or "—").border = BORDER
        nc = ws.cell(r, 2, clean_cell(p["name"]))
        nc.font = FONT_SUB
        nc.border = BORDER
        ws.cell(r, 3, p["qty"]).border = BORDER
        rq = ws.cell(r, 4, clean_cell(p["req"]))
        rq.alignment = WRAP_TOP
        rq.font = FONT_NOTE
        rq.border = BORDER
        _rec = {
            "tag":  f"'Part names'!$A${r}",
            "name": f"'Part names'!$B${r}",
            "row":  str(r),
        }
        _NAME_REG[_norm_name(p["req"])] = _rec
        if p["tag"] and p["tag"] != "—":
            _TAG_REG.setdefault(_norm_tag(p["tag"]), _rec)
        r += 1
    apply_col_formats(ws, 5, {3: FMT_INT}, r - 1)
    ws.auto_filter.ref = f"A4:D{r - 1}"
    ws.freeze_panes = "A5"
    back_link(ws, 4)


_LEDGER_SHEET = "Bill of Materials (Ledger)"


def _fmt_cost_items(items) -> str:
    """Render a cost-basis inputs/factors list (each a {name,value,unit} dict) as clean
    'name: value unit; …' text instead of a raw Python dict dump in the Ledger."""
    if not isinstance(items, list):
        return str(items or "")
    out = []
    for it in items:
        if isinstance(it, dict):
            nm = it.get("name") or it.get("label") or ""
            val = it.get("value", it.get("amount", it.get("rate", "")))
            unit = it.get("unit", "")
            s = str(nm)
            if val not in (None, ""):
                s += f": {val}"
            if unit:
                s += f" {unit}"
            out.append(s.strip(": ").strip())
        elif it not in (None, ""):
            out.append(str(it))
    return "; ".join(x for x in out if x)


def _build_costbasis_by_name(state: dict) -> Dict[str, dict]:
    """Index costBasis.lines by NORMALISED NAME (cost lines carry a `label` + `word_id`
    but NO tag — name is the only join key shared with the BoM's `requirement`)."""
    out: Dict[str, dict] = {}
    cb = state.get("costBasis") or {}
    for ln in (cb.get("lines") or []):
        if not isinstance(ln, dict):
            continue
        key = _norm_name(ln.get("label", ""))
        if key and key not in out:
            out[key] = ln
    return out


def _build_mpn_by_word(state: dict) -> Dict[str, str]:
    """tag/word(lower) -> 'Manufacturer MPN' from partVerifications (Spec-sheet MPN col)."""
    out: Dict[str, str] = {}
    for pv in (state.get("partVerifications") or []):
        if not isinstance(pv, dict):
            continue
        mfr = (pv.get("manufacturer") or "").strip()
        mpn = (pv.get("part_number") or "").strip()
        if not mpn:
            continue
        label = f"{mfr} {mpn}".strip()
        for k in (pv.get("word_name"), pv.get("word_id")):
            if k:
                out.setdefault(str(k).strip().lower(), label)
    return out


def tab_bom(wb: Workbook, state: dict, run_dir: str) -> None:
    """STAGE 4 consolidation (2026-06-24) — the headline merge: BoM + Cost + Spec sheets
    on ONE sheet "Bill of Materials (Ledger)" with two collapsible Excel column-GROUPS.
    ALWAYS VISIBLE: Tag · Item · Qty · Unit £ · Line £ (the buy-list). COLLAPSIBLE GROUP
    "Cost basis" (cols F–J, from costBasis.lines, joined by NAME — cost lines carry no
    tag): method · key inputs · factors · estimate class · confidence. COLLAPSIBLE GROUP
    "Engineering spec" (cols K–N, principals only): duty/rating · material · sizing calc ·
    MPN / datasheet. Commodity lines leave the spec columns blank (correct). Both detail
    groups are collapsed-by-default. Keeps the LIVE Σ + the cost-reconciliation note."""
    _BOM_REGISTRY.clear()
    ws = wb.create_sheet(_LEDGER_SHEET)
    set_widths(ws, {"A": 12, "B": 46, "C": 8, "D": 12, "E": 12,           # always-visible
                    "F": 18, "G": 30, "H": 22, "I": 10, "J": 12,          # Cost-basis group
                    "K": 30, "L": 18, "M": 50, "N": 30})                  # Engineering-spec group
    title_row(ws, "Bill of Materials (Ledger)", 14,
              "THE BILL + its provenance + its engineering, on one sheet. ALWAYS shown: tag · "
              "item · qty · unit £ · LIVE line £ (the buy-list). Two COLLAPSIBLE column-groups "
              "(click the [+] above cols F and K to expand): 'Cost basis' = HOW each £ was "
              "derived (method / inputs / factors / estimate class / confidence); 'Engineering "
              "spec' = WHY each principal is this size (duty / material / sizing calc / MPN). "
              "Σ line £ is live at the foot; commodity lines leave the spec columns blank.")
    header(ws, 4, ["Tag", "Item", "Qty", "Unit £", "Line £",
                   "Cost method", "Key inputs", "Factors", "Est class", "Confidence",
                   "Duty / rating", "Material", "Sizing calc (basis)", "MPN / datasheet"])
    bom = state.get("requirementsBom") or []
    cost_by_name = _build_costbasis_by_name(state)
    mpn_by_word = _build_mpn_by_word(state)
    r = 5
    first_line_row = r
    for row in bom:
        # Tag column REFERENCES the master "Part names" tag where this is a principal (one
        # identity; the master is the single source of the tag too). Sub-components (P-101.1)
        # and connection lines keep their literal tag (not master principals).
        _btref = tag_ref(row.get("tag", ""))
        ws.cell(r, 1, _btref if _btref else clean_cell(row.get("tag", ""))).border = BORDER
        # Item column — the NAME is referenced from the master "Part names" tab
        # (introduced once), with this row's own detail tail (" · 110 kW …") appended, so
        # editing the master name updates the ledger too. Sub-components (↳) and unmatched
        # names keep their literal text (not global identities).
        _req_raw = clean_cell(row.get("requirement", ""))
        _nref = name_ref(row.get("requirement", "")) if not str(
            row.get("requirement", "") or "").strip().startswith("↳") else None
        if _nref:
            _tail = _detail_tail(row.get("requirement", ""))
            _val = (_nref + ' & "  ' + _tail.replace('"', '""') + '"') if _tail else _nref
            rq = ws.cell(r, 2, _val)
        else:
            rq = ws.cell(r, 2, _req_raw)
        rq.alignment = WRAP_TOP
        rq.border = BORDER
        ws.cell(r, 3, row.get("qty")).border = BORDER
        # A sub-component row (requirement marked "↳") whose cost is rolled into its
        # parent (line_gbp == 0) must NOT show a standalone unit price: the parametric
        # intermediate is often larger than the parent itself (e.g. a £137k "110 kW
        # drive motor" sub-line under a £67.9k pump) and reads to a customer as a wrong
        # number. Show "incl. in parent" instead. Totals are unaffected — the LIVE
        # Σ(line_gbp) ignores the text cell exactly as it ignored the 0.
        _line_raw = row.get("line_gbp")
        _line_num = _line_raw if isinstance(_line_raw, (int, float)) else 0
        _is_subcomp = str(row.get("requirement", "") or "").strip().startswith("↳")
        if _is_subcomp and not _line_num:
            ws.cell(r, 4, "incl. in parent").border = BORDER
            ws.cell(r, 5, "—").border = BORDER
        else:
            ws.cell(r, 4, num(row.get("unit_gbp"))).border = BORDER
            ws.cell(r, 5, num(row.get("line_gbp"))).border = BORDER

        # ── COST-BASIS group (cols F–J) — joined by name; blank when no cost line ──
        _cl = cost_by_name.get(_norm_name(row.get("requirement", "")))
        _cbasis = (_cl.get("basis") or {}) if _cl else {}
        if _cbasis:
            ws.cell(r, 6, clean_cell(_cbasis.get("method", ""))).border = BORDER
            ic = ws.cell(r, 7, clean_cell(_fmt_cost_items(_cbasis.get("inputs"))))
            ic.alignment = WRAP_TOP; ic.font = FONT_NOTE; ic.border = BORDER
            fc = ws.cell(r, 8, clean_cell(_fmt_cost_items(_cbasis.get("factors"))))
            fc.alignment = WRAP_TOP; fc.font = FONT_NOTE; fc.border = BORDER
            ws.cell(r, 9, clean_cell(_cbasis.get("estimate_class", ""))).border = BORDER
            ws.cell(r, 10, clean_cell(_cbasis.get("confidence", ""))).border = BORDER

        # ── ENGINEERING-SPEC group (cols K–N) — PRINCIPALS ONLY (a top-level line that
        #    carries a real line cost); commodity / sub-component lines leave it blank ──
        _is_principal = (not _is_subcomp) and bool(_line_num)
        if _is_principal:
            # duty / rating = the requirement text (carries the sizing, '132 kW motor …')
            dc = ws.cell(r, 11, _req_raw); dc.alignment = WRAP_TOP; dc.font = FONT_NOTE; dc.border = BORDER
            ws.cell(r, 12, clean_cell(row.get("material", ""))).border = BORDER
            sc = ws.cell(r, 13, clean_cell(row.get("basis", "")))
            sc.alignment = WRAP_TOP; sc.font = FONT_NOTE; sc.border = BORDER
            # MPN / datasheet: partVerifications by tag, then by the requirement head noun;
            # fall back to the BoM `part` text.
            _mpn = mpn_by_word.get(str(row.get("tag", "")).strip().lower()) or ""
            if not _mpn and _req_raw:
                _head = re.split(r"[·\-(]", _req_raw)[0].strip().lower()
                _mpn = mpn_by_word.get(_head, "")
            pc = ws.cell(r, 14, clean_cell(_mpn) or clean_cell(row.get("part", "")) or "—")
            pc.alignment = WRAP_TOP; pc.border = BORDER

        # Register this part's canonical cells so other tabs reference (not repeat) it.
        # Sheet name = the merged ledger (was 'BoM') so every bom_ref() still resolves.
        _rtag = str(row.get("tag", "") or "").strip()
        if _rtag:
            _BOM_REGISTRY.setdefault(_norm_tag(_rtag), {
                "tag":  f"'{_LEDGER_SHEET}'!$A${r}",
                "name": f"'{_LEDGER_SHEET}'!$B${r}",
                "qty":  f"'{_LEDGER_SHEET}'!$C${r}",
                "part": f"'{_LEDGER_SHEET}'!$N${r}",
                "unit": f"'{_LEDGER_SHEET}'!$D${r}",
                "line": f"'{_LEDGER_SHEET}'!$E${r}",
            })
        r += 1
    last_line_row = r - 1

    # LIVE Σ of line £ (col E)
    ws.cell(r, 1, "Σ TOTAL").font = FONT_SUB
    tot = ws.cell(r, 5, f"=SUM(E{first_line_row}:E{last_line_row})")
    tot.font = Font(bold=True)
    tot.fill = FILL_RESULT
    sum_row = r
    r += 2

    # BUILDING & CIVILS — listed in the BoM/equipment list (Tristan 2026-06-22: "additional items
    # must be in the BoM and equipment list") but DELIBERATELY NOT summed into the raw-materials Σ
    # above: it is INSTALLED civils that bypasses the OEM manufacturing stack. The all-in project
    # capex (equipment installed + building) is on the Cost waterfall.
    _cs_b = state.get("costStack") or {}
    _bld = num(_cs_b.get("building_civils_gbp"))
    if _bld:
        sub_banner(ws, r, "Building & Civils (installed — separate from the raw-materials Σ above)", 14)
        r += 1
        ws.cell(r, 1, "BLDG-001").border = BORDER
        ws.cell(r, 2, "Insulated steel-frame industrial building — clad walls, insulated roof, "
                      "floor slab, drainage, roller + personnel doors").border = BORDER
        ws.cell(r, 3, 1).border = BORDER
        ws.cell(r, 4, _bld).border = BORDER
        ws.cell(r, 5, _bld).border = BORDER
        ws.cell(r, 6, "derived").border = BORDER
        _bn = ws.cell(r, 11, f"installed civils: {int(_cs_b.get('building_footprint_m2') or 0):,} m² "
                            f"footprint × UK-2026 rates; NOT in the raw Σ (bypasses the OEM stack). "
                            f"All-in project capex on the Cost waterfall.")
        _bn.font = FONT_NOTE
        _bn.border = BORDER
        r += 2

    # coverage_by_drawing from parts-ledger.json
    pl = load_json(os.path.join(run_dir, "parts-ledger.json")) or {}
    cov = pl.get("coverage_by_drawing") or {}
    if cov:
        sub_banner(ws, r, "Coverage by drawing (parts-ledger.json)", 14)
        r += 1
        header(ws, r, ["Drawing", "Expected", "Present", "% present", "", "", "",
                       "", "", "", "", "", "", ""])
        r += 1
        for dname, c in cov.items():
            ws.cell(r, 1, dname).border = BORDER
            ws.cell(r, 2, c.get("expected")).border = BORDER
            ws.cell(r, 3, c.get("present")).border = BORDER
            pc = ws.cell(r, 4, c.get("pct"))
            pc.border = BORDER
            if isinstance(c.get("pct"), (int, float)):
                pc.fill = FILL_PASS if c["pct"] >= 90 else (FILL_CONST if c["pct"] >= 70 else FILL_FAIL)
            r += 1
        r += 1

    # grand-total reconciliation note
    grand = num(pl.get("grand_total_gbp"))
    if grand is not None:
        ws.cell(r, 1, "parts-ledger grand_total_gbp").font = FONT_SUB
        gc = ws.cell(r, 5, grand)
        gc.number_format = FMT_GBP
        ws.cell(r, 6, f"Compare against the LIVE Σ line £ at row {sum_row}. "
                      f"See the ⚠ Checks tab for the reconciliation row.").font = FONT_NOTE
    # number formats (#37): Qty (#,##0) + Unit/Line £ (£#,##0) + the cost-basis Est class
    apply_col_formats(ws, first_line_row, {3: FMT_INT}, last_line_row)
    apply_col_formats(ws, first_line_row, {4: FMT_GBP, 5: FMT_GBP}, sum_row)
    ws.auto_filter.ref = f"A4:N{last_line_row}"
    ws.freeze_panes = "A5"

    # ── COLLAPSIBLE column-groups (collapsed by default). The two detail blocks fold
    #    behind a [+] so the default view is the clean buy-list (A–E); a click expands
    #    'Cost basis' (F–J) or 'Engineering spec' (K–N). DO NOT delete the data — it is
    #    present, just hidden. Set the outline level + hidden flag on EVERY column in the
    #    range (openpyxl's range-group() only stamps the endpoints, leaving the middle
    #    columns ungrouped + visible), then collapse the group at its right edge. ──
    for _grp in ("FGHIJ", "KLMN"):
        for _cl in _grp:
            cd = ws.column_dimensions[_cl]
            cd.outline_level = 1
            cd.hidden = True
        # the collapse control sits on the column just RIGHT of the group
        right = get_column_letter(column_index_from_string(_grp[-1]) + 1)
        ws.column_dimensions[right].collapsed = True
    ws.sheet_properties.outlinePr.summaryRight = True

    back_link(ws, 14)


# ============================================================================
# TAB — BRIEF-COMPLIANCE MATRIX (#45)
# ============================================================================
# Every parsedBrief.constraints.target_performance.metrics[] target vs the
# ACHIEVED contract quantity vs a LIVE =IF PASS/FAIL. Names differ between the
# brief metric (e.g. production_capacity_tpy) and the contract quantity
# (annual_production_t_yr) — so matching is by VALUE + UNIT-FAMILY, never name.
# ----------------------------------------------------------------------------

# Canonical unit-family map so 204 tpy (brief) reconciles with 204 t/yr
# (contract) and 60 kg/m3 reconciles with 60 kg/m³. Each family maps a raw
# unit string (lower-cased, whitespace-stripped) to a (family, to_canonical)
# pair; the value is converted to the family's canonical unit before compare.
def _unit_family(unit: str) -> Tuple[str, float]:
    """Return (family_key, multiplier_to_canonical) for a unit string.
    Unknown units fall back to family='?'+the raw token so two identical raw
    units still match each other."""
    u = (unit or "").strip().lower().replace(" ", "")
    # m3 ≡ m³ ≡ m^3 (the brief parser is non-deterministic about exponent notation — a caret
    # form "m^3" otherwise reads as an unknown family and breaks compliance matching, Codema v7).
    u = u.replace("³", "3").replace("²", "2").replace("^", "").replace("·", ".").replace("μ", "u").replace("°", "")
    table = {
        # throughput / production per year -> canonical tonnes/yr
        "tpy": ("t_per_yr", 1.0), "t/yr": ("t_per_yr", 1.0),
        "t/y": ("t_per_yr", 1.0), "tonnes/yr": ("t_per_yr", 1.0),
        "tonne/yr": ("t_per_yr", 1.0), "te/yr": ("t_per_yr", 1.0),
        "kg/yr": ("t_per_yr", 0.001),
        # volume -> canonical m3
        "m3": ("volume_m3", 1.0), "litre": ("volume_m3", 0.001),
        "l": ("volume_m3", 0.001), "litres": ("volume_m3", 0.001),
        # volumetric flow -> canonical m3/h (a brief metric in m3/hr must match a
        # contract quantity in m3/h / m³/h — the slash/spelling must not split them)
        "m3/h": ("flow_m3h", 1.0), "m3/hr": ("flow_m3h", 1.0), "m3perhr": ("flow_m3h", 1.0),
        "m3/hour": ("flow_m3h", 1.0), "m3ph": ("flow_m3h", 1.0),
        "l/h": ("flow_m3h", 0.001), "lph": ("flow_m3h", 0.001),
        "l/s": ("flow_m3h", 3.6), "lps": ("flow_m3h", 3.6),
        # count / quantity -> canonical count (a "count"-unit metric must match a
        # contract quantity whose name is a …_count even when it carries no unit token)
        "count": ("count", 1.0), "nr": ("count", 1.0), "no": ("count", 1.0),
        "qty": ("count", 1.0), "ea": ("count", 1.0), "off": ("count", 1.0),
        "pcs": ("count", 1.0), "pieces": ("count", 1.0), "#": ("count", 1.0),
        # density -> canonical kg/m3
        "kg/m3": ("density", 1.0), "g/l": ("density", 1.0),
        # time / cycle -> canonical days
        "days": ("time_days", 1.0), "day": ("time_days", 1.0),
        "d": ("time_days", 1.0), "hr": ("time_days", 1 / 24.0),
        "hours": ("time_days", 1 / 24.0), "h": ("time_days", 1 / 24.0),
        # mass -> canonical kg
        "kg": ("mass_kg", 1.0), "g": ("mass_kg", 0.001), "t": ("mass_kg", 1000.0),
        "tonne": ("mass_kg", 1000.0), "tonnes": ("mass_kg", 1000.0),
        # energy -> canonical kWh (2026-06-25 fix: a brief metric in MWh must match a
        # contract quantity in kWh — the family string MUST be identical for both, or the
        # compliance matcher's `a_fam == b_fam` test fails and the row shows UNVERIFIED).
        "wh": ("energy_kwh", 0.001), "kwh": ("energy_kwh", 1.0),
        "mwh": ("energy_kwh", 1000.0), "gwh": ("energy_kwh", 1_000_000.0),
        # power -> canonical kW (w/kw/mw/gw all map to ONE family)
        "w": ("power_kw", 0.001), "kw": ("power_kw", 1.0),
        "mw": ("power_kw", 1000.0), "gw": ("power_kw", 1_000_000.0),
        # voltage -> canonical V
        "v": ("voltage_v", 1.0), "kv": ("voltage_v", 1000.0), "mv": ("voltage_v", 0.001),
        # current -> canonical A
        "a": ("current_a", 1.0), "ka": ("current_a", 1000.0), "ma": ("current_a", 0.001),
        # temperature -> canonical °C (° already stripped above; K not offset-converted, rare)
        "c": ("temp_c", 1.0), "degc": ("temp_c", 1.0), "celsius": ("temp_c", 1.0),
        # dimensionless / ratio
        "ratio": ("ratio", 1.0), "": ("ratio", 1.0), "-": ("ratio", 1.0),
    }
    return table.get(u, ("?" + u, 1.0))


# brief-ECHO quantity suffixes — these carry the REQUESTED target value, not the ACHIEVED design
# value, so a compliance match against them is a guaranteed false PASS. Excluded from matching.
_ECHO_SUFFIXES = ("_requested", "_request", "_target", "_demand", "_brief", "_spec")
_QTY_UNIT_SUFFIX = re.compile(
    r"_(kwh|mwh|gwh|wh|kw|mw|gw|w|kva|mva|kv|mv|v|ka|ma|a|percent|pct|cycles?|kg|t|m2|m3|c)$")
_QTY_STOP_TOKENS = {"the", "of", "per", "system", "total", "design", "rated", "nominal"}


def _norm_qty_name(s: str) -> str:
    return _QTY_UNIT_SUFFIX.sub("", str(s or "").lower())


def _match_quantity(metric: dict, quantities: Dict[str, Any]) -> Optional[Tuple[str, float, str]]:
    """Find the ACHIEVED contract quantity that fulfils a brief metric, by NAME + UNIT FAMILY.
    (2026-06-25 fix) Match by NAME — NOT by which value is closest to the target. Closeness-to-
    target is Goodhart: it grabs a brief-ECHO quantity (usable_capacity_kwh_REQUESTED=5000) over
    the real achieved value (nameplate_capacity_kwh=2912), manufacturing a false PASS in the
    compliance table. Strategy mirrors the benchmark net's engineValueForMetric: (1) exact/
    unit-normalised NAME match to the achieved quantity (echoes excluded); (2) same unit-family,
    best name-token overlap, de-prioritising echoes + peak/max. Returns (name, achieved, unit)."""
    b_val = num(metric.get("value"))
    if b_val is None:
        return None
    b_fam, _ = _unit_family(metric.get("unit", ""))
    b_key = (metric.get("key_metric") or metric.get("metric") or metric.get("name") or "").lower().strip()
    b_norm = _norm_qty_name(b_key)
    b_tokens = set(t for t in re.findall(r"[a-z]+", b_norm) if t not in _QTY_STOP_TOKENS)

    def _fam_ok(a_fam: str, qname: str) -> bool:
        """Same unit family, OR a COUNT metric vs a count-named unitless quantity
        (a …_count / …_qty quantity carries no unit token so its family reads as
        'ratio'/unknown — match it by the count noun in its name, gated by the
        token-overlap test below)."""
        if a_fam == b_fam:
            return True
        if b_fam == "count":
            return (a_fam in ("count", "ratio") or a_fam.startswith("?")) and \
                   bool(re.search(r"(count|qty|number|_nr|valves?|containers?|units?)$", qname.lower()))
        return False

    # (1) exact / unit-normalised NAME match — the achieved quantity of the SAME name (no echoes)
    for qname, qv in quantities.items():
        if not isinstance(qv, dict) or any(e in qname.lower() for e in _ECHO_SUFFIXES):
            continue
        a_val = num(qv.get("value"))
        if a_val is None:
            continue
        a_fam, _ = _unit_family(qv.get("unit", ""))
        if not _fam_ok(a_fam, qname):
            continue
        if qname.lower() == b_key or _norm_qty_name(qname) == b_norm:
            return qname, a_val, qv.get("unit", "")

    # (2) same-family, best NAME-TOKEN overlap (NOT target-closeness). Requirement-ECHO
    # quantities (…_demand / …_target) are EXCLUDED — not merely penalised — so a brief
    # metric matches the DELIVERED quantity (irrigation_pump_flow=12), never the
    # requirement it restates (irrigation_demand=90); matching the demand would
    # manufacture a false PASS and hide an undersized design. Mirrors the audit oracle.
    best = None  # (-overlap, penalty, name, value, unit)
    # Require at least HALF the brief metric's identity tokens to be covered (mirrors the
    # audit oracle's token-subset threshold). A single SHARED generic token is NOT enough:
    # gac_softener_throughput must NOT match cloth_filter_throughput on "throughput" alone
    # (a wrong-subsystem false PASS — Codema v5 showed 80 m³/h "PASS" for a 14.5 target).
    need = max(1, (len(b_tokens) + 1) // 2)
    for qname, qv in quantities.items():
        if not isinstance(qv, dict) or any(e in qname.lower() for e in _ECHO_SUFFIXES):
            continue
        a_val = num(qv.get("value"))
        if a_val is None:
            continue
        a_fam, _ = _unit_family(qv.get("unit", ""))
        if not _fam_ok(a_fam, qname):
            continue
        ql = qname.lower()
        overlap = len(b_tokens & set(re.findall(r"[a-z]+", _norm_qty_name(ql))))
        if overlap < need:
            continue
        penalty = (1 if re.search(r"peak|max|surge|inrush", ql) else 0)
        cand = (-overlap, penalty, qname, a_val, qv.get("unit", ""))
        if best is None or cand < best:
            best = cand
    if best is None:
        return None
    return best[2], best[3], best[4]


def _render_brief_compliance_section(ws: Worksheet, state: dict, start_row: int) -> Optional[int]:
    """Render the Brief-compliance matrix onto `ws` starting at `start_row` (header row).
    Returns the next free row, or None when the brief carries no target metrics. (Was
    tab_brief_compliance; refactored into a section folded under "⚠ Checks", 2026-06-24.)"""
    pb = state.get("parsedBrief") or {}
    con = pb.get("constraints") or {}
    tp = con.get("target_performance") or {}
    metrics = tp.get("metrics") or []
    # if the brief carries only the headline (no metrics[]), synthesise one row
    if not metrics and tp.get("value") is not None:
        metrics = [tp]
    if not metrics:
        return None
    quantities = (state.get("orchestratorContract") or {}).get("quantities") or {}

    header(ws, start_row, ["Brief metric", "Target", "Unit", "Matched contract quantity",
                           "Achieved", "Direction", "STATUS", "Note"])
    r = start_row + 1
    first = r
    for m in metrics:
        key = (m.get("key_metric") or m.get("metric") or m.get("name") or "").strip()
        tgt = num(m.get("value"))
        unit = m.get("unit", "") or ""
        category = (m.get("category") or "").lower()
        matched = _match_quantity(m, quantities)

        ws.cell(r, 1, clean_cell(key) or "(unnamed metric)").border = BORDER
        tc = ws.cell(r, 2, tgt)
        tc.fill = FILL_INPUT
        tc.border = BORDER
        ws.cell(r, 3, clean_cell(unit)).border = BORDER

        if matched is None:
            ws.cell(r, 4, "— no matching quantity —").border = BORDER
            ac = ws.cell(r, 5, "—")
            ac.border = BORDER
            ws.cell(r, 6, "—").border = BORDER
            sc = ws.cell(r, 7, "UNVERIFIED")
            sc.fill = FILL_CONST
            sc.font = Font(bold=True, color="7F5B00")
            sc.border = BORDER
            nt = ws.cell(r, 8, "No contract quantity in the same unit family within "
                              "±50% of target — cannot auto-verify.")
            nt.alignment = WRAP_TOP
            nt.font = FONT_NOTE
            nt.border = BORDER
            r += 1
            continue

        qname, ach, qunit_s = matched
        ws.cell(r, 4, clean_cell(qname)).border = BORDER
        ac = ws.cell(r, 5, ach)
        ac.fill = FILL_RESULT
        ac.border = BORDER

        # Direction of the PASS test (2026-06-25 fix): HIGHER-is-better is the default (scale,
        # power, capacity, ENERGY efficiency %, cycle LIFE). LOWER-is-better ONLY for genuine
        # minimise targets — a feed-conversion ratio, a time/duration to complete, a cost-per-unit,
        # or a cycle TIME (NOT cycle life). The old `category=='efficiency'` wrongly made round-trip
        # efficiency lower-better (98.5% vs 88% → false FAIL), and `'cycle' in key` wrongly caught
        # cycle_life (more cycles is better → false direction).
        kl = key.lower()
        lower_better = (
            "fcr" in kl
            or "feed_conversion" in kl
            or "conversion_ratio" in kl
            or "_days" in kl or "duration" in kl or "lead_time" in kl
            or "lcoe" in kl or "cost_per" in kl
            or ("cycle" in kl and bool(re.search(r"\btime\b|hour|minute|second|_s\b", kl)))
        )
        # tolerance band: ±2% of target (display rounding + sizing granularity)
        tol = abs(tgt) * 0.02 if tgt else 0.0
        if lower_better:
            ws.cell(r, 6, "≤ target (lower better)").border = BORDER
            status_f = f'=IF(E{r}<=B{r}+{tol:.6g},"PASS","FAIL")'
        else:
            ws.cell(r, 6, "≥ target (meet/exceed)").border = BORDER
            status_f = f'=IF(E{r}>=B{r}-{tol:.6g},"PASS","FAIL")'
        sc = ws.cell(r, 7, status_f)
        sc.font = Font(bold=True)
        sc.border = BORDER

        # pre-evaluate so the cell colours on open (before Excel recomputes)
        if tgt is not None and ach is not None:
            passed = (ach <= tgt + tol) if lower_better else (ach >= tgt - tol)
            sc_fill = FILL_PASS if passed else FILL_FAIL
            for col in range(1, 9):
                if not isinstance(ws.cell(r, col).fill, PatternFill) or \
                        ws.cell(r, col).fill.fgColor.rgb in (None, "00000000"):
                    pass
            sc.fill = sc_fill
            sc.font = FONT_PASS if passed else FONT_FAIL
        nt = ws.cell(r, 8, f"family={_unit_family(unit)[0]}")
        nt.alignment = WRAP_TOP
        nt.font = FONT_NOTE
        nt.border = BORDER
        r += 1

    # live conditional formatting on STATUS
    from openpyxl.formatting.rule import CellIsRule
    rng = f"G{first}:G{r-1}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"FAIL"'], fill=FILL_FAIL, font=FONT_FAIL))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"PASS"'], fill=FILL_PASS, font=FONT_PASS))

    # FMT_NUM on the Target / Achieved metric columns: integer-valued metrics (62 t/yr)
    # read "62" not "62.00", while genuine decimals (FCR 1.37) still show their places.
    apply_col_formats(ws, first, {2: FMT_NUM, 5: FMT_NUM}, r - 1)
    return r


# ============================================================================
# TAB — COST WATERFALL (#44)
# ============================================================================
# BoM -> +assembly -> factory COGS -> +install -> installed ASP, from
# state.costStack, as LIVE running formulas (each running total references the
# prior running total + the step delta, so editing any step recomputes the ASP).
# ----------------------------------------------------------------------------
def tab_cost_waterfall(wb: Workbook, state: dict) -> bool:
    cs = state.get("costStack")
    if not cs or not isinstance(cs, dict):
        return False

    # Ordered build-up: (label, base-key for the additive STEP value, is_running_anchor).
    # We render successive steps; each running total = previous running + this step.
    # The anchors (raw BoM, factory COGS, OEM transfer, channel list, installed ASP)
    # are taken from costStack directly and shown beside the running total so any
    # divergence is visible.
    raw = num(cs.get("raw_materials_bom_gbp"))
    if raw is None:
        return False

    ws = wb.create_sheet("Cost waterfall")
    set_widths(ws, {"A": 38, "B": 18, "C": 18, "D": 56})
    title_row(
        ws, "Cost waterfall — BoM → installed ASP", 4,
        "Running build-up from state.costStack. Yellow = editable step £; the "
        "Running total column is LIVE (each = previous running + this step), so "
        "editing any step recomputes the installed price. 'costStack anchor' shows "
        "the engine's stored figure at that milestone for cross-check.",
    )
    header(ws, 4, ["Step", "Step £ (editable)", "Running total £ (live)",
                   "costStack anchor / note"])
    r = 5

    # Build the additive ladder. Each entry: (label, step_amount, anchor_value_or_None, note)
    steps: List[Tuple[str, Optional[float], Optional[float], str]] = []
    steps.append(("Raw materials (BoM)", raw, raw,
                  "raw_materials_bom_gbp — start of the build-up"))
    asm = num(cs.get("assembly_labour_gbp")) or 0.0
    steps.append(("+ Assembly / erection labour", asm, None, "assembly_labour_gbp"))
    ovh = num(cs.get("factory_overhead_gbp")) or 0.0
    if ovh:
        steps.append(("+ Factory overhead", ovh, None, "factory_overhead_gbp"))
    steps.append(("= Factory COGS", None,
                  num(cs.get("factory_cogs_gbp")),
                  "factory_cogs_gbp (anchor — compare running total)"))
    margin = num(cs.get("manufacturer_margin_gbp")) or 0.0
    if margin:
        steps.append(("+ Manufacturer margin", margin, None, "manufacturer_margin_gbp"))
    steps.append(("= OEM transfer price", None,
                  num(cs.get("oem_transfer_price_gbp")),
                  "oem_transfer_price_gbp (anchor)"))
    chan = num(cs.get("channel_markup_gbp")) or 0.0
    if chan:
        steps.append(("+ Channel markup", chan, None, "channel_markup_gbp"))
        steps.append(("= Channel list price", None,
                      num(cs.get("channel_list_price_gbp")),
                      "channel_list_price_gbp (anchor)"))
    install = num(cs.get("installation_cost_gbp")) or 0.0
    steps.append(("+ Installation / field erection", install, None,
                  "installation_cost_gbp"))
    steps.append(("= Installed ASP (process equipment)", None,
                  num(cs.get("installed_asp_gbp")),
                  "installed_asp_gbp (anchor — process plant installed price)"))
    # Building & Civils added at INSTALLED level (not through the OEM manufacturing stack) →
    # the ALL-IN project capex (Tristan 2026-06-22: building in the BoM, capex ~£7M all-in).
    bld = num(cs.get("building_civils_gbp")) or 0.0
    if bld:
        steps.append(("+ Building & Civils (installed)", bld, None,
                      "building_civils_gbp — insulated industrial building, floor slab, drainage, "
                      "doors (civils — separate from the equipment OEM stack)"))
        steps.append(("= ALL-IN PROJECT CAPEX", None, num(cs.get("all_in_capex_gbp")),
                      "all_in_capex_gbp (anchor — equipment + building, total project capex)"))

    first = r
    running_row: Optional[int] = None  # row holding the last LIVE running total
    for label, step_amt, anchor, note in steps:
        # clean_cell defangs the leading "=" sentinel on the anchor rows ("= Factory
        # COGS" etc.) so Excel doesn't read them as formulas (the #NAME? bug Tristan
        # caught). The startswith() logic below still tests the raw Python `label`.
        ws.cell(r, 1, clean_cell(label)).border = BORDER
        ws.cell(r, 1).font = FONT_SUB if label.startswith(("=", "Raw")) else Font()
        if step_amt is not None and not label.startswith("="):
            sc = ws.cell(r, 2, step_amt)
            sc.fill = FILL_INPUT
            sc.border = BORDER
            if running_row is None:          # first anchor = the running seed
                rt = ws.cell(r, 3, f"=B{r}")
            else:
                rt = ws.cell(r, 3, f"=C{running_row}+B{r}")
            rt.fill = FILL_RESULT
            rt.border = BORDER
            running_row = r
        else:
            # milestone anchor row: running total carries forward unchanged; show
            # the engine's stored anchor in the running column for direct compare
            ws.cell(r, 2, "").border = BORDER
            if running_row is not None:
                rt = ws.cell(r, 3, f"=C{running_row}")
            else:
                rt = ws.cell(r, 3, anchor)
            rt.fill = FILL_RESULT
            rt.border = BORDER
            running_row = r
            ws.cell(r, 1).font = FONT_SUB
        nt = ws.cell(r, 4, (note + (f"  ·  engine anchor £{anchor:,.0f}"
                                    if anchor is not None else "")))
        nt.alignment = WRAP_TOP
        nt.font = FONT_NOTE
        nt.border = BORDER
        r += 1

    apply_col_formats(ws, first, {2: FMT_GBP, 3: FMT_GBP}, r - 1)
    ws.freeze_panes = "A5"
    back_link(ws, 4)
    return True


# ============================================================================
# TABS — ECONOMICS MODEL  (Inputs & Assumptions / Economics / Scenarios + charts)
# ============================================================================
# A LIVE, defensible revenue/opex/EBITDA/NPV model. Universal shape:
#     revenue = output_qty × unit_price ;  opex = Σ drivers ;  EBITDA = rev − opex
# RAS is populated with grounded, cited market defaults (£/kg fish + feed/energy/
# labour); a non-RAS class still gets an Economics tab off its own output metric.
#
# Every derived number is a LIVE Excel formula referencing the yellow input cells
# on the "Inputs & Assumptions" tab (stable $-anchored addresses, also exposed as
# WORKBOOK-LEVEL DEFINED NAMES). Edit an input -> the whole chain + every chart
# recomputes. clean_cell() wraps every string; numbers are written directly.
# ----------------------------------------------------------------------------
INPUTS_SHEET = "Inputs & Assumptions"

# Defined-name keys -> the driver each maps to. The actual cell address is filled
# in when the Inputs tab is built, then read by Economics/Scenarios so the two
# tabs stay wired even if the Inputs layout shifts.
_ECON_INPUT_ADDR: Dict[str, str] = {}

# The economics-model input DEFAULTS (mirrored exactly from tab_inputs_assumptions
# so the Python-side sweet-spot pre-compute reproduces what the LIVE formulas show
# at the as-built inputs). These drive ONLY the colouring + the recommended-row pick
# + the few values too awkward to do as a live INDEX/MATCH; the workbook cells
# themselves are all live formulas, so editing an input still re-drives everything.
_ECON_DEFAULTS = {
    "sale_price_ras": 22.0, "feed_price_ras": 2.1, "fcr_ras": 1.37,
    "sale_price_generic": 1.0, "feed_price_generic": 0.0, "fcr_generic": 0.0,
    "energy_price": 0.15, "load_factor": 0.65, "hours": 8760.0,
    "labour": 300000.0, "maint_pct": 3.0, "other_opex": 120000.0,
    "discount_rate": 10.0, "hurdle_rate": 15.0, "project_life": 20.0,
}

# Number of rows in the scale sweep, and its span as a multiple of the as-built
# output. Log-spaced 0.2x .. 5x is a genuine "scale up and down from the current
# plant" curve (Tristan: "best size plant and to scale up and down from £5M capex").
SWEEP_ROWS = 16
SWEEP_LO_MULT = 0.2
SWEEP_HI_MULT = 5.0
# Capex-vs-output scaling exponent n: capex ∝ output^n.
#   n = 1.0  → LINEAR. A MODULAR plant that scales by REPLICATING identical units
#              (more tanks / pumps / treatment skids) — capex per unit of output is
#              constant. This is the DEFAULT: ForgeOS process plants are modular, and
#              the empirical 52 t→72 t RAS data scales ~linearly (exponent ≈1, not 0.6).
#              Modular/linear is the IDEAL (Tristan 2026-06-21): de-risked, testable at
#              small scale, predictable scale-up.
#   n ≈ 0.6  → the Williams/Chilton "six-tenths" law, valid ONLY for MONOLITHIC
#              equipment that scales by getting BIGGER (a single larger vessel/column).
# Exposed as a TUNABLE Inputs cell ('scale_exp') so the investor model shows the capex
# + economics live under either regime — set it to 0.6 to model a monolithic build.
SCALE_EXPONENT = 1.0


def _sweep_outputs(base_q: float) -> List[float]:
    """The log-spaced output points for the scale sweep, base_q×0.2 .. base_q×5.0
    (SWEEP_ROWS points). Log spacing puts equal resolution per doubling — the
    natural axis for a cost-capacity curve."""
    import math
    base = base_q if base_q and base_q > 0 else 1.0
    lo = math.log(base * SWEEP_LO_MULT)
    hi = math.log(base * SWEEP_HI_MULT)
    n = SWEEP_ROWS
    return [math.exp(lo + (hi - lo) * i / (n - 1)) for i in range(n)]


def _annuity_irr(ebitda: float, capex: float, n: float) -> Optional[float]:
    """IRR of a level annuity: capex out at t0, EBITDA in for n years. Same value
    Excel RATE(n, EBITDA, -capex) returns. Bisection (robust, no numpy). Returns
    None for a non-positive-EBITDA row (no real positive IRR)."""
    if ebitda is None or capex is None or ebitda <= 0 or capex <= 0 or n <= 0:
        return None
    lo, hi = 1e-7, 5.0  # 0 .. 500% — wide enough for any sane plant
    # PV(i) = EBITDA*(1-(1+i)^-n)/i is monotone decreasing in i; solve PV==capex
    for _ in range(200):
        mid = (lo + hi) / 2.0
        pv = ebitda * (1 - (1 + mid) ** -n) / mid
        if pv > capex:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2.0


def _econ_at(out: float, base_q: float, base_capex: float,
             is_ras: bool) -> Dict[str, Optional[float]]:
    """Pure-Python mirror of one sweep row's economics at output ``out`` — the SAME
    arithmetic the live Excel formulas compute, used to colour cells + pick the
    recommended row. Holds labour fixed (small plants don't shed fixed staff), scales
    energy pro-rata with output, capex via the six-tenths law."""
    d = _ECON_DEFAULTS
    # On the non-RAS path, mirror the SIGNAL-DERIVED drivers the live cells use (not the
    # RAS hardcodes), so the Python colouring/recommended-row pick matches the workbook.
    g = None if is_ras else _ECON_GENERIC
    sale = d["sale_price_ras"] if is_ras else (g["sale_price"] if g else d["sale_price_generic"])
    feed_p = d["feed_price_ras"] if is_ras else d["feed_price_generic"]
    fcr = d["fcr_ras"] if is_ras else d["fcr_generic"]
    sale_mult = 1000.0 if is_ras else 1.0
    hours = d["hours"] if is_ras else (g["hours"] if g else d["hours"])
    load_factor = d["load_factor"] if is_ras else (g["load_factor"] if g else d["load_factor"])
    energy_price = d["energy_price"] if is_ras else (g["energy_price"] if g else d["energy_price"])
    labour = d["labour"] if is_ras else (g["labour"] if g else d["labour"])
    other_opex = d["other_opex"] if is_ras else (g["other_opex"] if g else d["other_opex"])
    ratio = (out / base_q) if base_q else 1.0
    capex = base_capex * (ratio ** SCALE_EXPONENT)
    revenue = out * sale_mult * sale
    feed = out * sale_mult * fcr * feed_p
    # connected load scaled pro-rata with output (mirrors the sweep's energy formula)
    energy = _ECON_LOAD_KW * ratio * hours * load_factor * energy_price
    maint = capex * d["maint_pct"] / 100.0
    opex = feed + energy + labour + maint + other_opex
    ebitda = revenue - opex
    r = d["discount_rate"] / 100.0
    n = d["project_life"]
    npv = -capex + ebitda * (1 - (1 + r) ** -n) / r
    payback = (capex / ebitda) if ebitda > 0 else None
    irr = _annuity_irr(ebitda, capex, n)
    margin = (ebitda / revenue) if revenue else None
    return {"out": out, "capex": capex, "revenue": revenue, "opex": opex,
            "ebitda": ebitda, "payback": payback, "npv": npv, "irr": irr,
            "margin": margin}


# the connected load (kW) at the as-built plant — set when the Inputs tab resolves
# it (mirrors tab_inputs_assumptions' load_kw), so the Python sweep mirror scales
# energy correctly. Defaults to a safe placeholder until the Inputs tab runs.
_ECON_LOAD_KW = 500.0


def _sweet_spot(state: dict) -> Optional[Dict[str, Any]]:
    """Compute the sweet-spot findings in Python (the SAME maths the workbook shows
    live), for the recommended-deployment callout + threshold colouring. Returns a
    dict of scale findings, or None if no usable output metric. The workbook STILL
    renders these live via INDEX/MATCH over the sweep; this mirror just lets us write
    a human callout + colour the headline cells before Excel recomputes on open."""
    out_qty, _unit, price_unit, _noun = _econ_output_metric(state)
    if not out_qty or out_qty <= 0:
        return None
    is_ras = price_unit == "£/kg"
    cs = state.get("costStack") or {}
    # prefer the ALL-IN project capex (equipment + building) when present, so the financial
    # analysis matches the all-in headline (Tristan 2026-06-22); else the equipment installed.
    base_capex = (num(cs.get("all_in_capex_gbp")) or num(cs.get("installed_asp_gbp"))
                  or num(cs.get("factory_cogs_gbp"))
                  or num(cs.get("raw_materials_bom_gbp")) or 1_000_000.0)
    rows = [_econ_at(o, out_qty, base_capex, is_ras) for o in _sweep_outputs(out_qty)]
    d = _ECON_DEFAULTS

    def first(pred):
        for row in rows:
            if pred(row):
                return row
        return None

    break_even = first(lambda x: x["ebitda"] is not None and x["ebitda"] >= 0)
    viability = first(lambda x: x["irr"] is not None and x["irr"] >= d["discount_rate"] / 100.0)
    investable = first(lambda x: x["irr"] is not None and x["irr"] >= d["hurdle_rate"] / 100.0)
    npv_max = max(rows, key=lambda x: (x["npv"] if x["npv"] is not None else -1e18))
    # is NPV monotone increasing to the top of the range? (RAS: usually yes)
    npv_monotonic = npv_max["out"] >= rows[-1]["out"] - 1e-6

    # the £5M-affordable output and its economics
    five_out = out_qty * (5_000_000.0 / base_capex) ** (1.0 / SCALE_EXPONENT)
    five = _econ_at(five_out, out_qty, base_capex, is_ras)

    # recommended deployment = the investable scale if it exists, else NPV-max
    recommend = investable or npv_max
    return {
        "is_ras": is_ras, "base_q": out_qty, "base_capex": base_capex,
        "rows": rows, "break_even": break_even, "viability": viability,
        "investable": investable, "npv_max": npv_max, "npv_monotonic": npv_monotonic,
        "five": five, "recommend": recommend,
    }


def _econ_output_metric(state: dict) -> Tuple[float, str, str, str]:
    """Resolve the headline OUTPUT the economics model sells, universally.

    Returns (qty, unit, price_unit_label, output_noun). RAS -> (204, 't/yr',
    '£/kg', 'fish'). For a non-RAS class we read the primary brief metric /
    largest production-family contract quantity so the tab still renders a
    generic £/output-unit model. Always returns a positive qty (fallback 1.0)."""
    q = (state.get("orchestratorContract") or {}).get("quantities") or {}
    pclass = ((state.get("orchestratorContract") or {}).get("product_class")
              or "").lower()

    # 1) RAS / aquaculture: annual tonnage of fish, sold £/kg
    if "aquaculture" in pclass or "ras" in pclass or "annual_production_t_yr" in q:
        t = qval(q, "annual_production_t_yr")
        if t:
            return float(t), "t/yr", "£/kg", "fish"

    # 2) generic: take the brief's primary target_performance metric value+unit
    pb = state.get("parsedBrief") or {}
    tp = (pb.get("constraints") or {}).get("target_performance") or {}
    metrics = tp.get("metrics") or ([tp] if tp.get("value") is not None else [])
    if metrics:
        m0 = metrics[0]
        v = num(m0.get("value"))
        if v:
            unit = (m0.get("unit") or "unit").strip()
            return float(v), unit, "£/unit", "output"

    # 3) last resort: the largest production/flow-family contract quantity
    best = None
    for name, qv in q.items():
        if not isinstance(qv, dict):
            continue
        fam = (qv.get("family") or "").lower()
        if fam in ("flow_rate", "production", "throughput", "mass"):
            v = num(qv.get("value"))
            if v and (best is None or v > best[0]):
                best = (float(v), qv.get("unit", "unit"), name)
    if best:
        return best[0], best[1], "£/unit", "output"

    return 1.0, "unit", "£/unit", "output"


def _output_is_per_year(out_unit: str) -> bool:
    """Does the headline output metric carry a TIME dimension (per-year / per-day /
    per-hour throughput or production)? If so, an annual £/yr revenue–EBITDA–IRR
    model is the right frame (RAS t/yr, SAF t/yr, CO2 t/day). If NOT — a per-UNIT
    manufactured product whose spec is W / kg / m (a satellite's bus power, a HAPS
    payload mass) — '£/yr revenue / EBITDA / IRR' is the WRONG frame and the
    per-year economic tabs should be flagged as not-applicable. Signal-based: looks
    only at the unit string's time dimension, never at the archetype name."""
    u = (out_unit or "").strip().lower()
    # explicit per-time tokens anywhere in the unit (t/yr, kg/day, m³/h, /hr, p.a.)
    for tok in ("/yr", "/year", "/y", "/day", "/d", "/h", "/hr", "/hour", "/s",
                "/min", "/week", "/wk", "/month", "/mo", "per year", "per day",
                "per hour", "p.a", "pa"):
        if tok in u:
            return True
    return False


def _econ_generic_drivers(state: dict, capex: float) -> Dict[str, Any]:
    """Derive the NON-RAS economics drivers from physical signals already in state,
    NOT from RAS-shaped hardcodes. Universal: keyed on signals (capex fraction,
    operating-hours field), never on archetype name. Returns the driver values + an
    honest, RAS-free basis string for each, plus `sale_price_verified` (False when no
    real per-unit price is derivable → the economics tabs must be flagged UNVERIFIED).

    The same dict feeds BOTH the live Inputs cells (tab_inputs_assumptions) and the
    Python sweep mirror (_ECON_DEFAULTS / _econ_at) so the two stay byte-consistent."""
    q = (state.get("orchestratorContract") or {}).get("quantities") or {}
    cap = capex if (capex and capex > 0) else 1_000_000.0

    # ── operating hours: from an engine duty signal, else a labelled assumption.
    hours = qval(q, "operating_hours_yr") or qval(q, "annual_operating_hours") \
        or qval(q, "operating_hours_per_year")
    if hours and hours > 0:
        hours_basis = "from engine · operating_hours_yr"
    else:
        hours = 8000.0
        hours_basis = ("assumed — typical process-plant annual operating hours "
                       "(supply a class duty-cycle to refine)")

    # ── load factor: from a duty / availability / capacity-factor signal, else a
    # generic assumption (NOT 'continuous RAS duty').
    lf = qval(q, "load_factor") or qval(q, "capacity_factor") \
        or qval(q, "availability_factor") or qval(q, "duty_cycle")
    if lf and 0 < lf <= 1.0:
        load_factor = float(lf)
        lf_basis = "from engine · duty-cycle / load-factor signal"
    else:
        load_factor = 0.65
        lf_basis = ("assumed average/peak electrical load factor — supply a class "
                    "duty-cycle to refine")

    # ── labour: from a headcount/FTE signal if present, else a SANE fixed default.
    # (Tristan 2026-06-25: the old 4%-of-capex proxy gave ~£71k/yr for an unmanned battery
    # container — absurd. Labour is NOT a fraction of installed capex; for an unmanned /
    # skid / containerised plant with no operating headcount it is a small fixed
    # maintenance-contract figure — remote monitoring + a few annual service visits — not
    # a number that scales with how expensive the kit was.) UNIVERSAL: keyed only on the
    # presence of a headcount signal, never on the class. When no headcount is supplied we
    # default an unmanned plant to a small ASSUMED fixed annual figure, clearly flagged
    # 'assumed — supply a real figure'; supply a headcount/FTE to model a manned operation.
    fte = qval(q, "operating_headcount") or qval(q, "fte_count") \
        or qval(q, "operator_count") or qval(q, "staff_count")
    if fte and fte > 0:
        labour = round(float(fte) * 65_000.0, 0)  # £65k fully-loaded per FTE
        labour_basis = (f"from engine · {int(round(fte))} FTE × £65,000 fully-loaded "
                        "(edit the rate or headcount to refine)")
    else:
        # Unmanned / skid plant: a small fixed maintenance-contract default, NOT a % of capex.
        labour = 25_000.0
        labour_basis = ("assumed — supply a real figure. No operating headcount signal, so "
                        "this is defaulted as an UNMANNED / skid plant: ~£25,000/yr for "
                        "remote monitoring + periodic maintenance visits (NOT a % of capex). "
                        "Enter your real annual labour cost, or supply an operating headcount.")

    # ── other opex: from a consumables signal if present, else opex-fraction-of-capex.
    consum = qval(q, "consumables_cost_gbp_yr") or qval(q, "reagent_cost_gbp_yr") \
        or qval(q, "annual_consumables_gbp")
    if consum and consum > 0:
        other_opex = round(float(consum), 0)
        other_basis = "from engine · consumables / reagent annual cost"
    else:
        other_opex = round(cap * 0.02, 0)  # 2% of capex/yr — insurance/overhead proxy
        other_basis = ("assumed — 2% of installed capex/yr (consumables, insurance, "
                       "overhead; supply a class opex model to refine)")

    # ── energy price: a generic UK-industrial assumption, NO RAS micro-grid claim.
    energy_price = 0.15
    energy_basis = ("assumed — UK industrial electricity ~£0.15/kWh "
                    "(edit to your tariff)")

    # ── sale price: only verified if a real per-unit price signal exists in state.
    sale = qval(q, "sale_price_gbp_per_unit") or qval(q, "product_price_gbp") \
        or qval(q, "unit_sale_price_gbp")
    if sale and sale > 0:
        sale_price = float(sale)
        sale_verified = True
        sale_basis = "from engine · product unit sale price"
    else:
        sale_price = 0.0  # do NOT ship a degenerate £1/unit stub
        sale_verified = False
        sale_basis = ("UNVERIFIED — no per-unit market price is derivable for this "
                      "class. Enter a real sale price to make the economics tabs "
                      "meaningful; until then revenue / EBITDA / IRR are not valid.")

    return {
        "hours": hours, "hours_basis": hours_basis,
        "load_factor": load_factor, "load_factor_basis": lf_basis,
        "labour": labour, "labour_basis": labour_basis,
        "other_opex": other_opex, "other_opex_basis": other_basis,
        "energy_price": energy_price, "energy_basis": energy_basis,
        "sale_price": sale_price, "sale_price_basis": sale_basis,
        "sale_price_verified": sale_verified,
    }


# Resolved non-RAS drivers for the current run, published by tab_inputs_assumptions
# so the Python sweep mirror (_econ_at) reproduces what the live cells show. None on
# the RAS path (the RAS _ECON_DEFAULTS values are used unchanged). Also carries the
# sale-price-verified flag the economics tabs read to decide the UNVERIFIED banner.
_ECON_GENERIC: Optional[Dict[str, Any]] = None


def _econ_sale_unverified() -> bool:
    """True when the current (non-RAS) run has no real per-unit sale price, so the
    per-year economic tabs are presenting an UNVERIFIED model. False on the RAS path
    (the £/kg fish price is grounded) and whenever a real price was derived."""
    return bool(_ECON_GENERIC and not _ECON_GENERIC.get("sale_price_verified", False))


def tab_inputs_assumptions(wb: Workbook, state: dict) -> bool:
    """TAB 1 — editable yellow input cells, one per row, driving the whole model.
    Engine-derived values (tonnage, capex, connected load) are pulled from state
    and labelled 'from engine'; market defaults are grounded + cited in the Basis
    column. Every value cell is registered as a workbook DEFINED NAME and its
    address cached in _ECON_INPUT_ADDR for the Economics/Scenarios tabs."""
    q = (state.get("orchestratorContract") or {}).get("quantities") or {}
    cs = state.get("costStack") or {}

    # ---- engine-grounded values (fall back + label as assumption if absent) ----
    out_qty, out_unit, price_unit, out_noun = _econ_output_metric(state)

    capex = num(cs.get("all_in_capex_gbp")) or num(cs.get("installed_asp_gbp"))
    capex_basis = ("from engine · costStack.all_in_capex_gbp (equipment installed + building & civils)"
                   if cs.get("all_in_capex_gbp")
                   else "from engine · costStack.installed_asp_gbp (BoM + assembly + install)")
    if capex is None:
        capex = num(cs.get("factory_cogs_gbp")) or num(cs.get("raw_materials_bom_gbp"))
        capex_basis = "from engine · costStack (installed ASP absent — COGS proxy)"
    if capex is None:
        capex = 1_000_000.0
        capex_basis = "ASSUMPTION · no costStack — placeholder £1.0M, replace"

    # connected electrical load: engine field, else derive from transformer kVA,
    # else a labelled default.
    load_kw = qval(q, "connected_electrical_load_kw")
    load_basis = "from engine · connected_electrical_load_kw"
    if load_kw is None:
        load_kw = qval(q, "total_supply_demand_kw")
        load_basis = "from engine · total_supply_demand_kw (connected load absent)"
    if load_kw is None:
        kva = qval(q, "main_transformer_kva")
        if kva is not None:
            load_kw = round(kva * 0.9, 1)
            load_basis = "ASSUMPTION · main_transformer_kva × 0.9 power-factor (no load field)"
    if load_kw is None:
        load_kw = 500.0
        load_basis = "ASSUMPTION · no electrical field in state — placeholder 500 kW"

    # publish the resolved connected load to the Python sweet-spot mirror so its
    # energy term scales correctly (the workbook is live regardless; this only
    # affects pre-render colouring + the recommended-deployment callout text).
    global _ECON_LOAD_KW
    _ECON_LOAD_KW = float(load_kw)

    # FCR (feed conversion ratio) — only meaningful for a feed-driven biological
    # class; from engine where present.
    fcr = qval(q, "feed_conversion_ratio")
    fcr_basis = "from engine · feed_conversion_ratio"
    if fcr is None:
        fcr = 1.37
        fcr_basis = "ASSUMPTION · from brief (engine field absent)"

    is_ras = price_unit == "£/kg"

    # On the NON-RAS path derive the economics drivers from physical signals (capex
    # fraction / operating-hours field), with honest RAS-free basis strings; publish
    # them so the Python sweep mirror (_econ_at) reproduces the live cells. On the
    # RAS path leave _ECON_GENERIC None so the grounded £/kg defaults stand unchanged.
    global _ECON_GENERIC
    gen = None if is_ras else _econ_generic_drivers(state, capex)
    _ECON_GENERIC = gen

    ws = wb.create_sheet(INPUTS_SHEET)
    set_widths(ws, {"A": 34, "B": 14, "C": 12, "D": 74})
    title_row(
        ws, "Inputs & Assumptions — the economics model drivers", 4,
        "EVERY yellow cell is editable. Engine-derived values are labelled "
        "'from engine'; market defaults are grounded + cited in Basis. The "
        "Economics & Scenarios tabs reference these cells with LIVE formulas — "
        "edit one driver and the whole model + every chart recomputes.",
    )
    header(ws, 4, ["Driver", "Value", "Unit", "Basis / source"])
    r = 5

    # rows: (defined_name, label, value, unit, basis, number_format)
    # The fish-price / feed rows are RAS-specific; for a non-RAS class they are
    # still emitted (generic £/unit sale price; feed driver zeroed) so the model
    # shape is universal and the formulas never reference a missing cell.
    rows: List[Tuple[str, str, float, str, str, str]] = []
    rows.append(("out_qty", f"Output volume ({out_noun})", round(out_qty, 4),
                 out_unit, "from engine · primary output metric", FMT_DEC2))
    if is_ras:
        rows.append(("sale_price", "Fish sale price", 22.0, "£/kg",
                     "sashimi-grade yellowtail kingfish (Hamachi), UK food-service "
                     "wholesale; range £18–28/kg", FMT_GBP2))
        rows.append(("feed_price", "Feed price", 2.1, "£/kg",
                     "high-protein marine carnivore aquafeed; range £1.8–2.4/kg",
                     FMT_GBP2))
        rows.append(("fcr", "Feed conversion ratio (FCR)", round(fcr, 3), "kg/kg",
                     fcr_basis, FMT_DEC2))
    else:
        # NON-RAS: a real per-unit price only when the engine supplies one; otherwise
        # the price is UNVERIFIED (do NOT ship a degenerate £1/unit stub that fakes an
        # investor model — the economics tabs are flagged UNVERIFIED instead).
        rows.append(("sale_price", "Sale price (per output unit)",
                     round(gen["sale_price"], 4), f"£/{out_unit}",
                     gen["sale_price_basis"], FMT_GBP2))
        rows.append(("feed_price", "Feedstock price", 0.0, "£/unit",
                     "assumed — feedstock cost driver (0 if not feed-driven)",
                     FMT_GBP2))
        rows.append(("fcr", "Feedstock conversion ratio", 0.0, "ratio",
                     "assumed — feed-to-output ratio (0 disables the feed term)",
                     FMT_DEC2))
    # Energy price / load factor / hours / labour / other opex: RAS keeps its grounded
    # values + basis EXACTLY; the non-RAS path uses signal-derived values with honest,
    # RAS-free basis strings (no LOX / juveniles / micro-grid / 'continuous RAS duty').
    if is_ras:
        rows.append(("energy_price", "Energy price", 0.15, "£/kWh",
                     "UK industrial net of the planned on-site renewable micro-grid; "
                     "grid alone ~£0.22/kWh", FMT_GBP2))
        rows.append(("load_kw", "Connected electrical load", round(load_kw, 1), "kW",
                     load_basis, FMT_DEC1))
        rows.append(("load_factor", "Electrical load factor", 0.65, "avg/peak",
                     "continuous RAS duty, average/peak", FMT_DEC2))
        rows.append(("hours", "Operating hours", 8760.0, "h/yr", "continuous",
                     FMT_INT))
        rows.append(("labour", "Labour", 300000.0, "£/yr",
                     "≈8 FTE loaded for a small RAS; scale with tonnage", FMT_GBP))
        rows.append(("maint_pct", "Maintenance", 3.0, "% capex/yr",
                     "process-plant norm", FMT_DEC1))
        rows.append(("other_opex", "Other opex", 120000.0, "£/yr",
                     "LOX, chemicals, juveniles, insurance, overhead", FMT_GBP))
    else:
        rows.append(("energy_price", "Energy price", round(gen["energy_price"], 4),
                     "£/kWh", gen["energy_basis"], FMT_GBP2))
        rows.append(("load_kw", "Connected electrical load", round(load_kw, 1), "kW",
                     load_basis, FMT_DEC1))
        rows.append(("load_factor", "Electrical load factor",
                     round(gen["load_factor"], 3), "avg/peak",
                     gen["load_factor_basis"], FMT_DEC2))
        rows.append(("hours", "Operating hours", round(gen["hours"], 0), "h/yr",
                     gen["hours_basis"], FMT_INT))
        rows.append(("labour", "Labour", gen["labour"], "£/yr",
                     gen["labour_basis"], FMT_GBP))
        rows.append(("maint_pct", "Maintenance", 3.0, "% capex/yr",
                     "process-plant norm", FMT_DEC1))
        rows.append(("other_opex", "Other opex", gen["other_opex"], "£/yr",
                     gen["other_opex_basis"], FMT_GBP))
    rows.append(("capex", "Installed capex", round(capex, 0), "£", capex_basis,
                 FMT_GBP))
    rows.append(("scale_exp", "Scaling exponent n  (capex ∝ output^n)",
                 SCALE_EXPONENT, "n",
                 "1.0 = LINEAR — a MODULAR plant scaled by replicating identical units "
                 "(more tanks/pumps/skids); capex per tonne is constant, de-risked + "
                 "testable at small scale (the ForgeOS default; the empirical RAS data "
                 "scales ~linearly). 0.6 = the six-tenths law for MONOLITHIC equipment "
                 "that scales by getting bigger. Edit to model either regime — every "
                 "capex/economics cell on the sweep, solver and £5M anchor is live off "
                 "this.", FMT_DEC2))
    rows.append(("discount_rate", "Discount rate", 10.0, "%",
                 "real WACC for a small infrastructure project", FMT_DEC1))
    rows.append(("hurdle_rate", "Investor hurdle rate (IRR)", 15.0, "%",
                 "the minimum IRR an investor demands before deploying capital — "
                 "the 'investable' bar on the Investment Analysis tab; edit to your "
                 "fund's threshold", FMT_DEC1))
    rows.append(("project_life", "Project life", 20.0, "yr",
                 "asset economic life for the NPV horizon", FMT_DEC1))

    for name, label, value, unit, basis, fmt in rows:
        ws.cell(r, 1, clean_cell(label)).font = FONT_SUB
        vc = ws.cell(r, 2, value)
        vc.fill = FILL_INPUT
        vc.border = BORDER
        vc.number_format = fmt
        addr = f"${get_column_letter(2)}${r}"     # e.g. $B$5 — stable absolute ref
        _ECON_INPUT_ADDR[name] = f"'{INPUTS_SHEET}'!{addr}"
        # workbook-level defined name so a human can also use it in any formula
        try:
            from openpyxl.workbook.defined_name import DefinedName
            dn = f"in_{name}"
            if dn not in wb.defined_names:
                wb.defined_names[dn] = DefinedName(
                    dn, attr_text=f"'{INPUTS_SHEET}'!{addr}")
        except Exception:  # noqa: BLE001 — defined names are a nicety, never fatal
            pass
        ws.cell(r, 3, clean_cell(unit)).border = BORDER
        bs = ws.cell(r, 4, clean_cell(basis))
        bs.alignment = WRAP_TOP
        bs.font = FONT_NOTE
        bs.border = BORDER
        r += 1

    # (Removed the 'Engine bootstrap economics' stub — Tristan 2026-06-21: it only ever showed
    # '—' placeholders; the transparent live model on the Economics tab is the authoritative
    # figure, so an empty engine-stub block is noise. Don't ship a stub.)

    ws.freeze_panes = "A5"
    back_link(ws, 4)
    return True


def _ref(name: str) -> str:
    """The cross-sheet reference to a registered input cell (e.g.
    "'Inputs & Assumptions'!$B$6"). Raises if the input wasn't built — callers
    only run after tab_inputs_assumptions succeeded."""
    return _ECON_INPUT_ADDR[name]


def _render_economics_section(ws: Worksheet, state: dict, start_row: int) -> Optional[int]:
    """Render the Economics block onto `ws` from `start_row`. Returns the next free row,
    or None when the Inputs tab didn't build. (Was tab_economics; refactored into a
    section of the merged "Financial model" sheet, 2026-06-24.) Every cell is a LIVE
    formula over the 'Inputs & Assumptions' cells: revenue, the opex stack, EBITDA +
    margin, simple payback, a discounted-cashflow NPV column and a live IRR, plus an opex
    pie + a revenue/opex/EBITDA bar."""
    if not _ECON_INPUT_ADDR:
        return None  # Inputs tab didn't build -> nothing to reference

    out_qty, out_unit, price_unit, out_noun = _econ_output_metric(state)
    is_ras = price_unit == "£/kg"
    # output is sold per-kg for RAS (tonnes×1000), else per the metric's own unit
    sale_mult = "*1000" if is_ras else ""

    R = _ref  # local alias
    r = start_row
    # Flag the tab UNVERIFIED when no real per-unit sale price exists (the revenue/
    # EBITDA/IRR below are then NOT a valid investor model); and reframe when the
    # output is a per-UNIT product rather than annual throughput.
    if _econ_sale_unverified():
        r = unverified_banner(
            ws, r, 4,
            "⚠ UNVERIFIED ECONOMICS — no per-unit market sale price could be "
            "derived for this product class. Revenue, EBITDA, payback and IRR below "
            "are NOT a valid investor model until you enter a real sale price on the "
            "'Inputs & Assumptions' tab.")
    if not _output_is_per_year(out_unit):
        r = unverified_banner(
            ws, r, 4,
            "⚠ PER-UNIT PRODUCT — this output (" + clean_cell(out_unit) + ") is a "
            "manufactured-product spec, not annual throughput. A £/yr revenue / "
            "EBITDA / IRR frame does not apply; read this tab as per-unit cost, not "
            "an annual P&L.")

    # ---- revenue + opex stack ----------------------------------------------
    sub_banner(ws, r, "Annual profit & loss", 4)
    r += 1
    header(ws, r, ["Line", "£ / yr (live)", "Unit", "Formula"])
    r += 1

    # remember key rows so later cells reference them
    rows_addr: Dict[str, int] = {}

    def line(key: str, label: str, formula: str, note: str,
             fmt: str = FMT_GBP, fill=FILL_RESULT) -> None:
        nonlocal r
        ws.cell(r, 1, clean_cell(label)).font = FONT_SUB
        c = ws.cell(r, 2, formula)        # LIVE formula (string starting with '=')
        c.fill = fill
        c.border = BORDER
        c.number_format = fmt
        ws.cell(r, 3, clean_cell("£/yr"))
        nt = ws.cell(r, 4, clean_cell(note))
        nt.alignment = WRAP_TOP
        nt.font = FONT_NOTE
        rows_addr[key] = r
        r += 1

    # revenue = output × (×1000 for tonnes) × sale price
    line("revenue", "Annual revenue",
         f"={R('out_qty')}{sale_mult}*{R('sale_price')}",
         f"output × {'1000 ×' if is_ras else ''} sale price ({price_unit})")
    # feed cost = output × FCR × feed price (RAS: ×1000 kg)
    line("feed", "Feed / feedstock cost",
         f"={R('out_qty')}{sale_mult}*{R('fcr')}*{R('feed_price')}",
         "output × FCR × feed price")
    # energy = connected load × hours × load factor × energy price
    line("energy", "Energy cost",
         f"={R('load_kw')}*{R('hours')}*{R('load_factor')}*{R('energy_price')}",
         "connected load (kW) × hours × load factor × £/kWh")
    line("labour", "Labour", f"={R('labour')}", "from inputs")
    line("maint", "Maintenance",
         f"={R('capex')}*{R('maint_pct')}/100", "capex × maint% / 100")
    line("other", "Other opex", f"={R('other_opex')}", "from inputs")
    # total opex = Σ of the four+ driver rows
    line("opex", "Total opex",
         f"=B{rows_addr['feed']}+B{rows_addr['energy']}+B{rows_addr['labour']}"
         f"+B{rows_addr['maint']}+B{rows_addr['other']}",
         "Σ feed + energy + labour + maintenance + other",
         fill=FILL_CONST)
    # EBITDA = revenue − opex
    line("ebitda", "EBITDA",
         f"=B{rows_addr['revenue']}-B{rows_addr['opex']}",
         "revenue − total opex", fill=FILL_CONST)
    # Red-flag a loss-making base case (negative EBITDA) — the honest headline when
    # the plant is sub-scale (the £5M / 45 t/yr point runs at or below break-even).
    from openpyxl.formatting.rule import CellIsRule as _CIR
    _eb_cell = f"B{rows_addr['ebitda']}"
    ws.conditional_formatting.add(
        f"{_eb_cell}:{_eb_cell}",
        _CIR(operator="lessThan", formula=["0"], fill=FILL_FAIL, font=FONT_FAIL),
    )
    r += 1

    # ---- headline ratios ----------------------------------------------------
    sub_banner(ws, r, "Headline metrics (live)", 4)
    r += 1
    margin_row = r
    ws.cell(r, 1, "EBITDA margin").font = FONT_SUB
    mc = ws.cell(r, 2, f"=B{rows_addr['ebitda']}/B{rows_addr['revenue']}")
    mc.fill = FILL_RESULT
    mc.border = BORDER
    mc.number_format = "0.0%"
    ws.cell(r, 4, clean_cell("EBITDA ÷ revenue")).font = FONT_NOTE
    r += 1

    payback_row = r
    ws.cell(r, 1, "Simple payback").font = FONT_SUB
    # guard divide-by-zero / negative EBITDA -> blank (no payback)
    pc = ws.cell(r, 2,
                 f"=IF(B{rows_addr['ebitda']}>0,{R('capex')}/B{rows_addr['ebitda']},"
                 f'"n/a (EBITDA≤0)")')
    pc.fill = FILL_RESULT
    pc.border = BORDER
    pc.number_format = FMT_DEC1
    ws.cell(r, 3, clean_cell("yr"))
    ws.cell(r, 4, clean_cell("capex ÷ EBITDA (only if EBITDA > 0)")).font = FONT_NOTE
    r += 2

    # ---- NPV / IRR via a discounted cashflow column -------------------------
    sub_banner(ws, r, "Discounted cashflow — NPV & IRR (live)", 4)
    r += 1
    header(ws, r, ["Year", "Cashflow £ (live)", "Discounted £ (live)",
                   "Note"])
    r += 1
    cf_first = r
    life_int = 20
    pl = qval((state.get("orchestratorContract") or {}).get("quantities") or {},
              "project_life_years")
    # life is an editable input; for the static column count we use a fixed 20-row
    # horizon (the model's editable life caps the discount via an IF guard below).
    for y in range(0, life_int + 1):
        ws.cell(r, 1, y).border = BORDER
        if y == 0:
            cf = f"=-{R('capex')}"                       # capex outflow at year 0
            note = "capex outflow"
        else:
            # EBITDA inflow each year, but only while year ≤ editable project life
            cf = f"=IF({y}<={R('project_life')},B{rows_addr['ebitda']},0)"
            note = "EBITDA inflow (within project life)"
        cc = ws.cell(r, 2, cf)
        cc.border = BORDER
        cc.number_format = FMT_GBP
        # discounted = cashflow / (1+rate)^year
        dc = ws.cell(r, 3,
                     f"=B{r}/((1+{R('discount_rate')}/100)^A{r})")
        dc.border = BORDER
        dc.number_format = FMT_GBP
        ws.cell(r, 4, clean_cell(note)).font = FONT_NOTE
        r += 1
    cf_last = r - 1

    # NPV = Σ discounted column ; IRR = live IRR over the (undiscounted) cashflow
    ws.cell(r, 1, "NPV").font = FONT_SUB
    npv_c = ws.cell(r, 3, f"=SUM(C{cf_first}:C{cf_last})")
    npv_c.fill = FILL_RESULT
    npv_c.border = BORDER
    npv_c.number_format = FMT_GBP
    ws.cell(r, 4, clean_cell("Σ discounted cashflow (capex at yr 0 + discounted "
                             "EBITDA)")).font = FONT_NOTE
    npv_row = r
    r += 1
    ws.cell(r, 1, "IRR").font = FONT_SUB
    irr_c = ws.cell(r, 3, f"=IRR(B{cf_first}:B{cf_last})")
    irr_c.fill = FILL_RESULT
    irr_c.border = BORDER
    irr_c.number_format = "0.0%"
    ws.cell(r, 4, clean_cell("live =IRR over the year-0..N cashflow row")).font = FONT_NOTE
    r += 2

    # ---- opex breakdown sub-table (feeds the pie chart) ---------------------
    sub_banner(ws, r, "Opex breakdown (feeds the pie chart)", 4)
    r += 1
    header(ws, r, ["Driver", "£ / yr (live)", "% of opex (live)", ""])
    r += 1
    brk_first = r
    for key, lbl in [("feed", "Feed / feedstock"), ("energy", "Energy"),
                     ("labour", "Labour"), ("maint", "Maintenance"),
                     ("other", "Other")]:
        ws.cell(r, 1, clean_cell(lbl)).border = BORDER
        c = ws.cell(r, 2, f"=B{rows_addr[key]}")
        c.border = BORDER
        c.number_format = FMT_GBP
        pcc = ws.cell(r, 3, f"=B{r}/B{rows_addr['opex']}")
        pcc.border = BORDER
        pcc.number_format = "0.0%"
        r += 1
    brk_last = r - 1

    # ---- charts: opex pie + revenue/opex/EBITDA bar -------------------------
    from openpyxl.chart import PieChart, BarChart, Reference
    # Pie: opex breakdown
    pie = PieChart()
    pie.title = "Opex breakdown"
    pie.height, pie.width = 8, 17   # wider so the category legend has room
    pdata = Reference(ws, min_col=2, min_row=brk_first, max_row=brk_last)
    plabs = Reference(ws, min_col=1, min_row=brk_first, max_row=brk_last)
    pie.add_data(pdata, titles_from_data=False)
    pie.set_categories(plabs)
    from openpyxl.chart.label import DataLabelList
    # Clean pie: the CATEGORY names go in a right-side legend (no collision), the
    # slices carry the PERCENT only. (Tristan: "lots of writing overwriting each
    # other" — putting category names ON a small pie's slices is the collision.)
    dl = DataLabelList()
    dl.showPercent = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showVal = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    pie.dataLabels = dl
    from openpyxl.chart.legend import Legend
    pie.legend = Legend()
    pie.legend.position = "r"
    pie.legend.overlay = False
    ws.add_chart(pie, f"F{cf_first}")

    # Bar: revenue vs total opex vs EBITDA — build a tiny 3-row helper block so
    # the chart has clean contiguous categories+values.
    bar_first = r + 1
    ws.cell(bar_first, 1, clean_cell("Revenue")).font = FONT_NOTE
    ws.cell(bar_first, 2, f"=B{rows_addr['revenue']}").number_format = FMT_GBP
    ws.cell(bar_first + 1, 1, clean_cell("Total opex")).font = FONT_NOTE
    ws.cell(bar_first + 1, 2, f"=B{rows_addr['opex']}").number_format = FMT_GBP
    ws.cell(bar_first + 2, 1, clean_cell("EBITDA")).font = FONT_NOTE
    ws.cell(bar_first + 2, 2, f"=B{rows_addr['ebitda']}").number_format = FMT_GBP
    bar = BarChart()
    bar.title = "Revenue vs Opex vs EBITDA"
    bar.type = "col"
    bar.height, bar.width = 8, 13
    bdata = Reference(ws, min_col=2, min_row=bar_first, max_row=bar_first + 2)
    blabs = Reference(ws, min_col=1, min_row=bar_first, max_row=bar_first + 2)
    bar.add_data(bdata, titles_from_data=False)
    bar.set_categories(blabs)
    style_chart(bar, legend=None, data_labels=True)  # solid bars + £ labels (EBITDA stays visible even when near break-even)
    ws.add_chart(bar, f"F{cf_first + 16}")

    # leave room below the two stacked charts (each ~16 rows tall, anchored in col F)
    return max(r, cf_first + 34)


def _render_scenarios_section(ws: Worksheet, state: dict, start_row: int) -> Optional[int]:
    """Render the Scenarios block onto `ws` from `start_row`. Returns the next free row,
    or None when the Inputs tab didn't build. (Was tab_scenarios; refactored into a
    section of the merged "Financial model" sheet, 2026-06-24.) A live scenario explorer:
    a FINE log-spaced scale sweep carrying capex / revenue / opex / EBITDA / payback / NPV
    (annuity DCF) / IRR (RATE), plus a Low/Central/High price-driver block — all LIVE
    formulas off the Inputs cells. The sweep's cell ranges are stashed on THIS worksheet
    object (`ws._forge_sweep`, keyed off `ws.title`) so the Investment-analysis section
    below drives its sweet-spot INDEX/MATCH + curves off them. Charts anchor relative to
    `start_row` so they never collide with the Economics section above."""
    if not _ECON_INPUT_ADDR:
        return None

    out_qty, out_unit, price_unit, out_noun = _econ_output_metric(state)
    is_ras = price_unit == "£/kg"
    sale_mult = "*1000" if is_ras else ""
    R = _ref

    r = start_row
    if _econ_sale_unverified():
        r = unverified_banner(
            ws, r, 8,
            "⚠ UNVERIFIED ECONOMICS — no per-unit market sale price could be "
            "derived for this product class. The whole revenue/EBITDA/IRR sweep "
            "below is NOT a valid investor model until you enter a real sale price "
            "on the 'Inputs & Assumptions' tab.")
    if not _output_is_per_year(out_unit):
        r = unverified_banner(
            ws, r, 8,
            "⚠ PER-UNIT PRODUCT — this output (" + clean_cell(out_unit) + ") is a "
            "manufactured-product spec, not annual throughput; an annual £/yr "
            "scale-economics sweep does not apply. Read as a per-unit cost curve.")

    # ---- FINE scale sweep ---------------------------------------------------
    base_q = round(out_qty, 6) or 204.0
    sweep = [round(o, 4) for o in _sweep_outputs(base_q)]

    sub_banner(ws, r, f"Output sweep ({out_noun}, {out_unit}) — capex via the tunable "
                      f"scaling exponent (default linear/modular); payback / NPV / IRR live per row", 8)
    r += 1
    header(ws, r, [f"Output ({out_unit})", "Capex £ (live)", "Revenue £ (live)",
                   "Total opex £ (live)", "EBITDA £ (live)", "Payback yr",
                   "NPV £ (live)", "IRR (live)"])
    r += 1
    sweep_first = r
    qref = R("out_qty")
    for qv in sweep:
        # A-relative: this row's output qty drives every other cell on the row.
        ws.cell(r, 1, qv).border = BORDER
        ws.cell(r, 1).number_format = FMT_DEC1
        ws.cell(r, 1).fill = FILL_INPUT     # the sweep point is editable too
        # capex = capex_ref × (q / q_ref)^n  (n = the tunable scale_exp cell; 1.0 = linear)
        cc = ws.cell(r, 2, f"={R('capex')}*(A{r}/{qref})^{R('scale_exp')}")
        cc.border = BORDER
        cc.number_format = FMT_GBP
        # revenue = q × (×1000 for tonnes) × sale price
        rc = ws.cell(r, 3, f"=A{r}{sale_mult}*{R('sale_price')}")
        rc.border = BORDER
        rc.number_format = FMT_GBP
        # opex = feed(q) + energy(load scaled q/q_ref) + labour(FIXED) + maint(capex(q)) + other
        feed = f"A{r}{sale_mult}*{R('fcr')}*{R('feed_price')}"
        energy = (f"{R('load_kw')}*(A{r}/{qref})*{R('hours')}*"
                  f"{R('load_factor')}*{R('energy_price')}")
        maint = f"B{r}*{R('maint_pct')}/100"
        oc = ws.cell(r, 4, f"={feed}+{energy}+{R('labour')}+{maint}+{R('other_opex')}")
        oc.border = BORDER
        oc.number_format = FMT_GBP
        # EBITDA = revenue − opex
        ec = ws.cell(r, 5, f"=C{r}-D{r}")
        ec.border = BORDER
        ec.number_format = FMT_GBP
        # payback = capex / EBITDA (n/a if EBITDA ≤ 0 — small plant is sub-break-even)
        pc = ws.cell(r, 6, f'=IF(E{r}>0,B{r}/E{r},"n/a")')
        pc.border = BORDER
        pc.number_format = FMT_DEC1
        # NPV = live annuity DCF: -capex + EBITDA × (1-(1+r)^-n)/r ; one compact cell
        rr = f"({R('discount_rate')}/100)"
        nn = R("project_life")
        npv = (f"=-B{r}+E{r}*(1-(1+{rr})^-{nn})/{rr}")
        nc = ws.cell(r, 7, npv)
        nc.border = BORDER
        nc.number_format = FMT_GBP
        # IRR = live RATE of the level annuity (capex out t0, EBITDA in for n yrs).
        # RATE() is Newton-solved and DIVERGES from its default 10% guess on steep
        # annuities (a £33M/£10.8M-a-yr row -> IRR ≈ 33% returned -190% with the
        # default seed). We pass a DATA-DERIVED guess = EBITDA/capex (the cash-on-cash
        # return — always a hair above the true annuity IRR, an ideal Newton seed) so
        # it converges across the whole 0.2x-5x sweep. Guard ≤0 EBITDA -> "n/a";
        # IFERROR + MAX(-0.99,…) keep a pathological row from ever rendering absurd.
        irr = (f'=IF(E{r}>0,IFERROR(MAX(-0.99,RATE({nn},E{r},-B{r},0,0,E{r}/B{r})),'
               f'-1),"n/a")')
        ic = ws.cell(r, 8, irr)
        ic.border = BORDER
        ic.number_format = "0.0%"
        r += 1
    sweep_last = r - 1
    # Red-flag every loss-making sweep row (EBITDA < 0) so the sub-scale plants are
    # visually obvious and the sweet-spot crossing reads at a glance.
    from openpyxl.formatting.rule import CellIsRule as _CIR
    ws.conditional_formatting.add(
        f"E{sweep_first}:E{sweep_last}",
        _CIR(operator="lessThan", formula=["0"], fill=FILL_FAIL, font=FONT_FAIL),
    )
    r += 1

    # a numeric payback column the line-chart can plot (text "n/a" breaks a chart
    # series), capped so a near-break-even point doesn't blow the axis.
    sub_banner(ws, r, "Chart helper — payback capped at 40 yr (text 'n/a' breaks a "
                      "chart series)", 8)
    r += 1
    header(ws, r, [f"Output ({out_unit})", "Payback yr (chart)", "", "", "", "",
                   "", ""])
    r += 1
    pay_first = r
    for i, qv in enumerate(sweep):
        src = sweep_first + i
        ws.cell(r, 1, f"=A{src}").number_format = FMT_DEC1
        ws.cell(r, 1).border = BORDER
        # capped numeric payback: ≤0 EBITDA or >40yr -> 40 (off-the-chart marker)
        ws.cell(r, 2,
                f"=IF(E{src}>0,MIN(40,B{src}/E{src}),40)").number_format = FMT_DEC1
        ws.cell(r, 2).border = BORDER
        r += 1
    pay_last = r - 1
    r += 1

    # Stash the sweep geometry so tab_investment_analysis can build live INDEX/MATCH
    # sweet-spot cells + the curves directly off these columns (cols A..H).
    ws._forge_sweep = {                                   # type: ignore[attr-defined]
        "sheet": ws.title, "first": sweep_first, "last": sweep_last,
        "col_out": "A", "col_capex": "B", "col_rev": "C", "col_opex": "D",
        "col_ebitda": "E", "col_payback": "F", "col_npv": "G", "col_irr": "H",
    }

    # ---- Low / Central / High price-driver block ---------------------------
    price_label = "fish price" if is_ras else "sale price"
    sub_banner(ws, r, f"Low / Central / High — {price_label} ±25%, energy ±25%, "
                      "capex ±15% (at the base output)", 8)
    r += 1
    header(ws, r, ["Driver / metric", "Low", "Central", "High", "", "", "", ""])
    r += 1
    # rows: sale price, energy price, capex (the swung inputs), then EBITDA+payback
    # recomputed for each column. Columns: B=Low, C=Central, D=High.
    lo, ce, hi = "B", "C", "D"
    # sale price row
    sp_row = r
    ws.cell(r, 1, clean_cell("Sale price (±25%)")).font = FONT_SUB
    ws.cell(r, 2, f"={R('sale_price')}*0.75").number_format = FMT_GBP2
    ws.cell(r, 3, f"={R('sale_price')}").number_format = FMT_GBP2
    ws.cell(r, 4, f"={R('sale_price')}*1.25").number_format = FMT_GBP2
    for col in (2, 3, 4):
        ws.cell(r, col).border = BORDER
    r += 1
    # energy price row
    ep_row = r
    ws.cell(r, 1, clean_cell("Energy price (±25%)")).font = FONT_SUB
    ws.cell(r, 2, f"={R('energy_price')}*0.75").number_format = FMT_GBP2
    ws.cell(r, 3, f"={R('energy_price')}").number_format = FMT_GBP2
    ws.cell(r, 4, f"={R('energy_price')}*1.25").number_format = FMT_GBP2
    for col in (2, 3, 4):
        ws.cell(r, col).border = BORDER
    r += 1
    # capex row
    cx_row = r
    ws.cell(r, 1, clean_cell("Capex (±15%)")).font = FONT_SUB
    ws.cell(r, 2, f"={R('capex')}*0.85").number_format = FMT_GBP
    ws.cell(r, 3, f"={R('capex')}").number_format = FMT_GBP
    ws.cell(r, 4, f"={R('capex')}*1.15").number_format = FMT_GBP
    for col in (2, 3, 4):
        ws.cell(r, col).border = BORDER
    r += 1
    # NOTE on the sensitivity direction: Low = worst case (low price, HIGH energy,
    # HIGH capex); High = best case. We build EBITDA accordingly per column.
    # revenue per column uses that column's sale price; energy uses the OPPOSITE
    # extreme so 'Low' is genuinely the pessimistic corner.
    rev_row = r
    ws.cell(r, 1, clean_cell("Revenue £ (live)")).font = FONT_SUB
    for col, spcol in ((2, lo), (3, ce), (4, hi)):
        ws.cell(r, col,
                f"={R('out_qty')}{sale_mult}*{spcol}{sp_row}").number_format = FMT_GBP
        ws.cell(r, col).border = BORDER
    r += 1
    # opex per column: energy at the column's energy price, capex-driven maint at
    # the column's capex, feed/labour/other fixed.
    op_row = r
    ws.cell(r, 1, clean_cell("Total opex £ (live)")).font = FONT_SUB
    feed_t = f"{R('out_qty')}{sale_mult}*{R('fcr')}*{R('feed_price')}"
    for col, ecol, ccol in ((2, lo, lo), (3, ce, ce), (4, hi, hi)):
        energy_t = (f"{R('load_kw')}*{R('hours')}*{R('load_factor')}*{ecol}{ep_row}")
        maint_t = f"{ccol}{cx_row}*{R('maint_pct')}/100"
        ws.cell(r, col,
                f"={feed_t}+{energy_t}+{R('labour')}+{maint_t}+{R('other_opex')}"
                ).number_format = FMT_GBP
        ws.cell(r, col).border = BORDER
    r += 1
    # EBITDA per column. WORST corner (Low) = low sale price + HIGH energy + HIGH
    # capex; so for the 'Low' EBITDA use Low revenue but High-energy/High-capex
    # opex. We assemble each corner explicitly for an honest sensitivity.
    eb_row = r
    ws.cell(r, 1, clean_cell("EBITDA £ (live)")).font = FONT_SUB
    # Low corner: rev=Low, opex=High-energy+High-capex
    low_energy = f"{R('load_kw')}*{R('hours')}*{R('load_factor')}*{hi}{ep_row}"
    low_maint = f"{hi}{cx_row}*{R('maint_pct')}/100"
    ws.cell(r, 2,
            f"=B{rev_row}-({feed_t}+{low_energy}+{R('labour')}+{low_maint}+"
            f"{R('other_opex')})").number_format = FMT_GBP
    # Central
    ws.cell(r, 3, f"=C{rev_row}-C{op_row}").number_format = FMT_GBP
    # High corner: rev=High, opex=Low-energy+Low-capex
    hi_energy = f"{R('load_kw')}*{R('hours')}*{R('load_factor')}*{lo}{ep_row}"
    hi_maint = f"{lo}{cx_row}*{R('maint_pct')}/100"
    ws.cell(r, 4,
            f"=D{rev_row}-({feed_t}+{hi_energy}+{R('labour')}+{hi_maint}+"
            f"{R('other_opex')})").number_format = FMT_GBP
    for col in (2, 3, 4):
        ws.cell(r, col).border = BORDER
    r += 1
    # payback per column
    pb_row = r
    ws.cell(r, 1, clean_cell("Payback yr (live)")).font = FONT_SUB
    ws.cell(r, 2, f'=IF(B{eb_row}>0,D{cx_row}/B{eb_row},"n/a")'
            ).number_format = FMT_DEC1      # Low corner used High capex (D col)
    ws.cell(r, 3, f'=IF(C{eb_row}>0,C{cx_row}/C{eb_row},"n/a")'
            ).number_format = FMT_DEC1
    ws.cell(r, 4, f'=IF(D{eb_row}>0,B{cx_row}/D{eb_row},"n/a")'
            ).number_format = FMT_DEC1      # High corner used Low capex (B col)
    for col in (2, 3, 4):
        ws.cell(r, col).border = BORDER
    r += 1

    # ---- Additional what-if scenarios: single-driver sensitivity (tornado),
    #      breakeven price + combined best/worst corners. Every cell is a LIVE
    #      formula off the Inputs tab (Tristan: "a whole series of additional
    #      scenario set-ups... more scenarios done"). Universal — uses the same
    #      driver refs as the sweep, no class-specific logic. -------------------
    q = R('out_qty')

    def _ebitda(sale="", fcr="", feed="", energy="", capex="", labour=""):
        """A LIVE EBITDA formula with an optional ×factor on ONE driver (the rest
        held at the Inputs base): revenue − (feed + energy + labour + capex-maint +
        other)."""
        rev = f"{q}{sale_mult}*{R('sale_price')}{sale}"
        feed_t = f"{q}{sale_mult}*{R('fcr')}{fcr}*{R('feed_price')}{feed}"
        energy_t = f"{R('load_kw')}*{R('hours')}*{R('load_factor')}*{R('energy_price')}{energy}"
        maint_t = f"{R('capex')}{capex}*{R('maint_pct')}/100"
        opex = (f"({feed_t}+{energy_t}+{R('labour')}{labour}+{maint_t}"
                f"+{R('other_opex')})")
        return f"={rev}-{opex}"

    sub_banner(ws, r, "What-if sensitivity — one driver at a time (±20%), with the "
                      "EBITDA swing (tornado magnitude); held at the base output", 8)
    r += 1
    header(ws, r, ["Driver (±20%)", "EBITDA @ −20%", "EBITDA @ base",
                   "EBITDA @ +20%", "Swing £ (|+20 − −20|)", "", "", ""])
    r += 1
    DRIVERS = [
        ("Sale price", "sale"), ("Feed price", "feed"), ("Energy price", "energy"),
        ("Feed-conversion ratio (FCR)", "fcr"), ("Capex (→ maintenance)", "capex"),
        ("Labour", "labour"),
    ]
    for label, key in DRIVERS:
        ws.cell(r, 1, clean_cell(label)).font = FONT_SUB
        ws.cell(r, 2, _ebitda(**{key: "*0.8"})).number_format = FMT_GBP
        ws.cell(r, 3, _ebitda()).number_format = FMT_GBP
        ws.cell(r, 4, _ebitda(**{key: "*1.2"})).number_format = FMT_GBP
        ws.cell(r, 5, f"=ABS(D{r}-B{r})").number_format = FMT_GBP
        for col in range(1, 6):
            ws.cell(r, col).border = BORDER
        r += 1
    r += 1

    sub_banner(ws, r, "Breakeven & combined corners (live)", 8)
    r += 1
    base_opex = (f"({q}{sale_mult}*{R('fcr')}*{R('feed_price')}"
                 f"+{R('load_kw')}*{R('hours')}*{R('load_factor')}*{R('energy_price')}"
                 f"+{R('labour')}+{R('capex')}*{R('maint_pct')}/100+{R('other_opex')})")
    ws.cell(r, 1, clean_cell(f"Breakeven {price_label} ({price_unit})")).font = FONT_SUB
    ws.cell(r, 2, f"={base_opex}/({q}{sale_mult})").number_format = FMT_GBP2
    ws.cell(r, 3, clean_cell("EBITDA = 0 at this price (vs the Inputs price)")).font = FONT_NOTE
    r += 1
    ws.cell(r, 1, clean_cell("Combined corners → EBITDA £")).font = FONT_SUB
    for col, lbl in ((2, "Worst"), (3, "Base"), (4, "Best")):
        ws.cell(r, col, clean_cell(lbl)).font = FONT_NOTE
    r += 1
    worst = _ebitda(sale="*0.8", feed="*1.2", energy="*1.2", capex="*1.15", fcr="*1.1")
    best = _ebitda(sale="*1.2", feed="*0.8", energy="*0.8", capex="*0.85", fcr="*0.9")
    ws.cell(r, 1, clean_cell("EBITDA (all drivers swung together)")).font = FONT_SUB
    ws.cell(r, 2, worst).number_format = FMT_GBP
    ws.cell(r, 3, _ebitda()).number_format = FMT_GBP
    ws.cell(r, 4, best).number_format = FMT_GBP
    for col in range(1, 5):
        ws.cell(r, col).border = BORDER
    r += 2

    # an EBITDA-only contiguous block for the L/C/H bar chart (categories must be
    # the labels Low/Central/High in a column for a clean chart).
    lch_first = r + 1
    ws.cell(lch_first, 1, clean_cell("Low")).font = FONT_NOTE
    ws.cell(lch_first, 2, f"=B{eb_row}").number_format = FMT_GBP
    ws.cell(lch_first + 1, 1, clean_cell("Central")).font = FONT_NOTE
    ws.cell(lch_first + 1, 2, f"=C{eb_row}").number_format = FMT_GBP
    ws.cell(lch_first + 2, 1, clean_cell("High")).font = FONT_NOTE
    ws.cell(lch_first + 2, 2, f"=D{eb_row}").number_format = FMT_GBP

    # ---- charts (anchored right of the 8-col sweep, in cols J+) -------------
    from openpyxl.chart import LineChart, BarChart, Reference
    # 1) capex vs output (line)
    c1 = LineChart()
    c1.title = "Capex vs output"
    c1.height, c1.width = 8, 14
    c1.y_axis.title = "Capex £"
    c1.x_axis.title = f"Output ({out_unit})"
    d1 = Reference(ws, min_col=2, min_row=sweep_first, max_row=sweep_last)
    cats = Reference(ws, min_col=1, min_row=sweep_first, max_row=sweep_last)
    c1.add_data(d1, titles_from_data=False)
    c1.set_categories(cats)
    style_chart(c1, legend=None)  # single solid line, no rainbow, no heavy gridlines
    ws.add_chart(c1, f"J{start_row}")

    # 2) payback vs output (line) — uses the capped numeric column
    c2 = LineChart()
    c2.title = "Payback vs output"
    c2.height, c2.width = 8, 14
    c2.y_axis.title = "Payback (yr, capped 40)"
    c2.x_axis.title = f"Output ({out_unit})"
    d2 = Reference(ws, min_col=2, min_row=pay_first, max_row=pay_last)
    cats2 = Reference(ws, min_col=1, min_row=pay_first, max_row=pay_last)
    c2.add_data(d2, titles_from_data=False)
    c2.set_categories(cats2)
    style_chart(c2, legend=None)
    ws.add_chart(c2, f"J{start_row + 17}")

    # 3) Low/Central/High EBITDA (bar)
    c3 = BarChart()
    c3.title = "Scenario EBITDA — Low / Central / High"
    c3.type = "col"
    c3.height, c3.width = 8, 14
    d3 = Reference(ws, min_col=2, min_row=lch_first, max_row=lch_first + 2)
    cats3 = Reference(ws, min_col=1, min_row=lch_first, max_row=lch_first + 2)
    c3.add_data(d3, titles_from_data=False)
    c3.set_categories(cats3)
    style_chart(c3, legend=None, data_labels=True)  # Low/Central/High £ labels on each bar
    ws.add_chart(c3, f"J{start_row + 34}")

    # the three line/bar charts (cols J+) span ~start_row .. start_row+50; return below
    # the LATER of the data rows and the chart stack so the next section never overlaps.
    return max(r, start_row + 52)


# ============================================================================
# TAB — INVESTMENT ANALYSIS  (the SWEET-SPOT FINDER — the headline)
# ============================================================================
# Drives off the FINE scale sweep on the Scenarios tab (stashed cell ranges) to
# answer "what plant size should we build, and where does it become investable?".
# Break-even / viability / investable / NPV-max / £5M-anchor are LIVE INDEX/MATCH
# over the sweep columns (they re-resolve when any Input changes). Four curves
# (IRR-vs-capex with threshold lines, NPV-vs-capex, payback-vs-capex, EBITDA-margin-
# vs-scale) let an investor SEE the crossings. A recommended-deployment callout
# (Python-mirrored prose + live value cells) sits at the top.
# ----------------------------------------------------------------------------
def _render_investment_section(ws: Worksheet, state: dict, start_row: int) -> Optional[int]:
    """Render the Investment-analysis (sweet-spot finder) block onto `ws` from `start_row`.
    Returns the next free row, or None when the Scenarios sweep didn't build. (Was
    tab_investment_analysis; refactored into a section of the merged "Financial model"
    sheet, 2026-06-24.) Reads the sweep geometry stashed by the Scenarios section on the
    SAME worksheet (`ws._forge_sweep`); because that stash keys off `ws.title`, the
    INDEX/MATCH + curve references resolve to whatever this sheet is called — the merge
    needs no formula edits."""
    if not _ECON_INPUT_ADDR:
        return None
    sw = getattr(ws, "_forge_sweep", None)
    if not sw:
        return None  # Scenarios sweep didn't build -> nothing to analyse

    out_qty, out_unit, price_unit, out_noun = _econ_output_metric(state)
    is_ras = price_unit == "£/kg"
    R = _ref
    ss = _sweet_spot(state)  # Python mirror for the prose callout + colouring

    # sweep cell ranges (on THIS sheet) — quoted refs to ws.title (same-sheet refs are
    # valid Excel; they resolve correctly whatever the merged sheet is named).
    sh = f"'{sw['sheet']}'"
    f, l = sw["first"], sw["last"]
    OUT = f"{sh}!${sw['col_out']}${f}:${sw['col_out']}${l}"
    CAP = f"{sh}!${sw['col_capex']}${f}:${sw['col_capex']}${l}"
    EBI = f"{sh}!${sw['col_ebitda']}${f}:${sw['col_ebitda']}${l}"
    NPV = f"{sh}!${sw['col_npv']}${f}:${sw['col_npv']}${l}"
    IRRr = f"{sh}!${sw['col_irr']}${f}:${sw['col_irr']}${l}"
    PAY = f"{sh}!${sw['col_payback']}${f}:${sw['col_payback']}${l}"

    r = start_row
    if _econ_sale_unverified():
        r = unverified_banner(
            ws, r, 6,
            "⚠ UNVERIFIED ECONOMICS — no per-unit market sale price could be "
            "derived for this product class, so the sweet-spot / investability "
            "verdict below is NOT a valid investor model. Enter a real sale price "
            "on the 'Inputs & Assumptions' tab first.")
    if not _output_is_per_year(out_unit):
        r = unverified_banner(
            ws, r, 6,
            "⚠ PER-UNIT PRODUCT — this output (" + clean_cell(out_unit) + ") is a "
            "manufactured-product spec, not annual throughput. 'IRR / payback / "
            "investable scale' is the wrong frame for a per-unit product; treat the "
            "figures below as indicative only.")

    # ── ★ CAPEX → MAX OUTPUT SOLVER (Tristan 2026-06-20: "how do you get as much fish
    #    production out for a specific amount of capex?"). The six-tenths cost-capacity
    #    law INVERTED: for a capex budget B, the largest internally-consistent plant
    #    produces out_realised × (B / capex_realised)^(1/0.6). Type a budget in the yellow
    #    cell → read the max tonnage + its LIVE economics. Revenue is exact (qty×price);
    #    operating cost is read off the Scenarios sweep (near-linear in output) so EBITDA /
    #    payback / IRR stay consistent with the rest of the model. Universal.
    CAPX = _ECON_INPUT_ADDR.get("capex")
    OUTQ = _ECON_INPUT_ADDR.get("out_qty")
    SALE = _ECON_INPUT_ADDR.get("sale_price")
    if CAPX and OUTQ and SALE:
        OPX = f"{sh}!${sw['col_opex']}${f}:${sw['col_opex']}${l}"
        OUTr = f"{sh}!${sw['col_out']}${f}:${sw['col_out']}${l}"
        sub_banner(ws, r, "★ CAPEX → MAX OUTPUT — type a capex budget in the yellow cell; "
                          "read the biggest plant it buys + its economics (via the tunable scaling exponent)", 6)
        r += 1
        b_row = r
        ws.cell(r, 1, "Capex budget (EDIT me) £").font = FONT_SUB
        bc = ws.cell(r, 2, f"={CAPX}")
        bc.fill = FILL_INPUT; bc.number_format = FMT_GBP; bc.border = BORDER
        r += 1
        o_row = r
        ws.cell(r, 1, f"→ Max production ({out_unit})").font = FONT_SUB
        oc = ws.cell(r, 2, f"={OUTQ}*(B{b_row}/{CAPX})^(1/{R('scale_exp')})")
        oc.fill = FILL_RESULT; oc.font = Font(bold=True, size=12)
        oc.number_format = FMT_DEC1; oc.border = BORDER
        r += 1
        rev_row = r
        ws.cell(r, 1, "→ Revenue / yr £").border = BORDER
        rc = ws.cell(r, 2, f"=B{o_row}*1000*{SALE}")
        rc.number_format = FMT_GBP; rc.border = BORDER
        r += 1
        opx_row = r
        ws.cell(r, 1, "→ Operating cost / yr £ (interp. from sweep)").border = BORDER
        oxc = ws.cell(r, 2, f"=FORECAST(B{o_row},{OPX},{OUTr})")
        oxc.number_format = FMT_GBP; oxc.border = BORDER
        r += 1
        ebt_row = r
        ws.cell(r, 1, "→ EBITDA / yr £").font = FONT_SUB
        ebc = ws.cell(r, 2, f"=B{rev_row}-B{opx_row}")
        ebc.fill = FILL_RESULT; ebc.number_format = FMT_GBP; ebc.border = BORDER
        r += 1
        ws.cell(r, 1, "→ Simple payback (yr)").border = BORDER
        pbc = ws.cell(r, 2, f'=IF(B{ebt_row}>0,B{b_row}/B{ebt_row},"n/a (EBITDA ≤ 0)")')
        pbc.number_format = FMT_DEC1; pbc.border = BORDER
        r += 1
        _nn = R("project_life")
        ws.cell(r, 1, "→ IRR (annuity)").border = BORDER
        irc = ws.cell(r, 2,
                      f'=IF(B{ebt_row}>0,IFERROR(MAX(-0.99,'
                      f'RATE({_nn},B{ebt_row},-B{b_row},0,0,B{ebt_row}/B{b_row})),-1),"n/a")')
        irc.number_format = "0.0%"; irc.border = BORDER
        r += 2

    # ----------------------------------------------------------------------
    # RECOMMENDED DEPLOYMENT callout (big band) — prose from the Python mirror,
    # live value cells beneath so it stays correct if inputs change.
    # ----------------------------------------------------------------------
    def _fmt_q(v):  # output qty with unit — keep 2 dp for small/fractional outputs
        if v is None:
            return "—"
        prec = 0 if abs(v) >= 10 else 2  # 0.5 t/day must not round to "0"
        return f"{v:,.{prec}f} {out_unit}"

    def _fmt_gbp(v):
        return f"£{v:,.0f}" if v is not None else "—"

    def _fmt_pct(v):
        return f"{v*100:.1f}%" if v is not None else "—"

    def _fmt_yr(v):  # payback years — None/≤0 -> honest text, never "nan"
        return f"{v:.1f} yr" if (v is not None and v > 0) else "n/a (EBITDA ≤ 0)"

    rec = ss["recommend"] if ss else None
    inv = ss["investable"] if ss else None
    five = ss["five"] if ss else None
    npvmax = ss["npv_max"] if ss else None
    callout_lines: List[str] = []
    if ss and rec:
        if inv:
            callout_lines.append(
                f"DEPLOY {_fmt_gbp(rec['capex'])} for {_fmt_q(rec['out'])} "
                f"→ IRR {_fmt_pct(rec['irr'])}, payback "
                f"{_fmt_yr(rec['payback'])} — the sweet spot (first scale clearing the "
                f"{_ECON_DEFAULTS['hurdle_rate']:.0f}% investor hurdle).")
        else:
            callout_lines.append(
                f"DEPLOY {_fmt_gbp(rec['capex'])} for {_fmt_q(rec['out'])} "
                f"→ IRR {_fmt_pct(rec['irr'])}, payback "
                f"{_fmt_yr(rec['payback'])} — the best NPV in range. The "
                f"{_ECON_DEFAULTS['hurdle_rate']:.0f}% investor hurdle is not reached "
                f"anywhere in the 0.2x-5x sweep at the current inputs"
                + (" (economics are UNVERIFIED — no per-unit sale price is "
                   "derivable for this class; set a real sale price on the Inputs "
                   "tab to get a meaningful model)." if _econ_sale_unverified()
                   else ".") )
        # £5M ceiling line
        if five:
            if five["irr"] is None or five["irr"] <= 0.005:
                callout_lines.append(
                    f"At the £5.0M capex ceiling the plant makes only "
                    f"{_fmt_q(five['out'])} and is ~break-even (IRR ≈ 0) — economics "
                    f"turn investable only at larger scale.")
            else:
                callout_lines.append(
                    f"At the £5.0M capex ceiling the plant makes {_fmt_q(five['out'])} "
                    f"at IRR {_fmt_pct(five['irr'])} — "
                    + (f"already above the hurdle." if five['irr'] >= _ECON_DEFAULTS['hurdle_rate']/100
                       else f"below the {_ECON_DEFAULTS['hurdle_rate']:.0f}% hurdle."))
        if inv:
            callout_lines.append(
                f"Economics turn investable above {_fmt_gbp(inv['capex'])} / "
                f"{_fmt_q(inv['out'])}.")
        if ss["npv_monotonic"]:
            callout_lines.append(
                "NOTE: NPV improves monotonically with scale to the top of the "
                "range — the binding constraint here is capital/site, not economics. "
                "Build as large as the site and balance sheet allow.")
    else:
        callout_lines.append("No usable output metric to analyse — sweet-spot "
                             "finder unavailable for this run.")

    # render the callout band
    callout_text = "  ".join(callout_lines)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    cc = ws.cell(r, 1, clean_cell("★ RECOMMENDED DEPLOYMENT"))
    cc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    cc.fill = FILL_TITLE
    cc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    r += 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)
    bc = ws.cell(r, 1, clean_cell(callout_text))
    bc.fill = FILL_RESULT
    bc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    bc.font = Font(name="Calibri", size=11, bold=True, color="1F3A5F")
    for rr in range(r, r + 4):
        ws.row_dimensions[rr].height = 22
    r += 5

    # ----------------------------------------------------------------------
    # SWEET-SPOT THRESHOLDS — LIVE over the sweep (INDEX/MATCH).
    # IRR & EBITDA increase monotonically with output, so the "smallest output
    # meeting a threshold" = the first row above where the threshold is crossed:
    #   MATCH(threshold, ascending_range, 1) returns the LAST row ≤ threshold;
    #   +1 indexes the first row ABOVE it. Guard the edge cases with IFERROR.
    # ----------------------------------------------------------------------
    sub_banner(ws, r, "Sweet-spot thresholds — live over the scale sweep "
                      "(Scenarios tab)", 6)
    r += 1
    header(ws, r, ["What", f"Output ({out_unit})", "Capex £", "IRR", "Payback yr",
                   "How it's found"])
    r += 1
    HURDLE = R("hurdle_rate")
    DISC = R("discount_rate")

    def thr_row(label: str, out_formula: str, how: str, pyrow) -> None:
        """One threshold row: live output cell + capex/IRR/payback looked up at
        that same output via INDEX/MATCH-exact, coloured from the Python mirror."""
        nonlocal r
        ws.cell(r, 1, clean_cell(label)).font = FONT_SUB
        ws.cell(r, 1).border = BORDER
        oc = ws.cell(r, 2, out_formula)
        oc.fill = FILL_RESULT
        oc.border = BORDER
        oc.number_format = FMT_DEC1
        # capex / IRR / payback AT that output: INDEX(col, MATCH(thisOutput, OUT, 0)).
        # The matched output is exactly a sweep value (we INDEX the sweep's own out),
        # so an exact MATCH always resolves.
        mref = f"MATCH(B{r},{OUT},0)"
        cap = ws.cell(r, 3, f"=IFERROR(INDEX({CAP},{mref}),\"—\")")
        cap.border = BORDER
        cap.number_format = FMT_GBP
        ic = ws.cell(r, 4, f"=IFERROR(INDEX({IRRr},{mref}),\"—\")")
        ic.border = BORDER
        ic.number_format = "0.0%"
        pc = ws.cell(r, 5, f"=IFERROR(INDEX({PAY},{mref}),\"—\")")
        pc.border = BORDER
        pc.number_format = FMT_DEC1
        nt = ws.cell(r, 6, clean_cell(how))
        nt.font = FONT_NOTE
        nt.alignment = WRAP_TOP
        nt.border = BORDER
        # colour the output cell green if the Python mirror found a real point
        if pyrow is not None:
            oc.fill = FILL_RESULT
        else:
            oc.fill = FILL_CONST
        r += 1

    # break-even scale: smallest output where EBITDA ≥ 0 (EBITDA ascending)
    be_out = (f"=IFERROR(INDEX({OUT},MATCH(0,{EBI},1)+1),"
              f"INDEX({OUT},1))")
    thr_row("Break-even scale (EBITDA ≥ 0)", be_out,
            "first sweep row with EBITDA ≥ 0 (EBITDA rises with scale)",
            ss["break_even"] if ss else None)
    # viability: smallest output where IRR ≥ discount rate
    via_out = (f"=IFERROR(INDEX({OUT},MATCH({DISC}/100,{IRRr},1)+1),\"> range\")")
    thr_row("Viability threshold (IRR ≥ discount rate)", via_out,
            "first row whose IRR clears the discount rate — capital just covers its "
            "cost", ss["viability"] if ss else None)
    # investable: smallest output where IRR ≥ hurdle
    invv_out = (f"=IFERROR(INDEX({OUT},MATCH({HURDLE}/100,{IRRr},1)+1),\"> range\")")
    thr_row("Investable scale (IRR ≥ hurdle)", invv_out,
            "first row whose IRR clears the editable investor hurdle "
            "(Inputs tab) — the SWEET SPOT", ss["investable"] if ss else None)
    # NPV-max scale in range: output at MAX(NPV)
    npvmax_out = f"=INDEX({OUT},MATCH(MAX({NPV}),{NPV},0))"
    thr_row("NPV-max scale (in range)", npvmax_out,
            "output that maximises NPV across the sweep (trends to the top of the "
            "range when economics improve with scale)", ss["npv_max"] if ss else None)
    r += 1

    # ----------------------------------------------------------------------
    # THE £5M ANCHOR — output affordable at £5M (inverse six-tenths) + its metrics.
    # out_5M = out_ref × (5e6 / capex_ref)^(1/0.6). Live off the Inputs capex cell.
    # ----------------------------------------------------------------------
    sub_banner(ws, r, "The £5M capex anchor — what £5.0M buys & how far below "
                      "investable it sits", 6)
    r += 1
    header(ws, r, ["Metric", "Value", "", "", "", ""])
    r += 1
    five_out_row = r
    # output at £5M (live, inverse scaling off the as-built capex+output; exponent is
    # the tunable scale_exp cell — at n=1.0 this is simply out_ref × 5M / capex_ref)
    ws.cell(r, 1, clean_cell(f"Output affordable at £5.0M ({out_unit})")).font = FONT_SUB
    fo = ws.cell(r, 2, f"={R('out_qty')}*(5000000/{R('capex')})^(1/{R('scale_exp')})")
    fo.fill = FILL_RESULT
    fo.border = BORDER
    fo.number_format = FMT_DEC1
    ws.cell(r, 1).border = BORDER
    r += 1
    # the £5M EBITDA / IRR / payback — computed directly (not via the sweep, since
    # the £5M output isn't one of the sweep grid points). All live off Inputs.
    five_q = f"B{five_out_row}"
    qref = R("out_qty")
    sale_mult = "*1000" if is_ras else ""
    feed5 = f"{five_q}{sale_mult}*{R('fcr')}*{R('feed_price')}"
    energy5 = f"{R('load_kw')}*({five_q}/{qref})*{R('hours')}*{R('load_factor')}*{R('energy_price')}"
    maint5 = f"5000000*{R('maint_pct')}/100"
    rev5 = f"{five_q}{sale_mult}*{R('sale_price')}"
    eb5 = f"({rev5})-({feed5}+{energy5}+{R('labour')}+{maint5}+{R('other_opex')})"
    eb5_row = r
    ws.cell(r, 1, clean_cell("EBITDA at £5.0M")).font = FONT_SUB
    ec5 = ws.cell(r, 2, f"={eb5}")
    ec5.fill = FILL_RESULT
    ec5.border = BORDER
    ec5.number_format = FMT_GBP
    ws.cell(r, 1).border = BORDER
    r += 1
    ws.cell(r, 1, clean_cell("IRR at £5.0M")).font = FONT_SUB
    irr5 = ws.cell(r, 2,
                   f'=IF(B{eb5_row}>0,IFERROR(MAX(-0.99,RATE({R("project_life")},'
                   f'B{eb5_row},-5000000,0,0,B{eb5_row}/5000000)),-1),"n/a")')
    irr5.fill = FILL_RESULT
    irr5.border = BORDER
    irr5.number_format = "0.0%"
    ws.cell(r, 1).border = BORDER
    r += 1
    ws.cell(r, 1, clean_cell("Payback at £5.0M")).font = FONT_SUB
    pay5 = ws.cell(r, 2, f'=IF(B{eb5_row}>0,5000000/B{eb5_row},"n/a (EBITDA≤0)")')
    pay5.fill = FILL_RESULT
    pay5.border = BORDER
    pay5.number_format = FMT_DEC1
    ws.cell(r, 1).border = BORDER
    r += 1
    # gap below investable scale (live: investable output − £5M output)
    ws.cell(r, 1, clean_cell(f"Shortfall vs investable scale ({out_unit})")).font = FONT_SUB
    gap = ws.cell(r, 2,
                  f'=IFERROR(INDEX({OUT},MATCH({HURDLE}/100,{IRRr},1)+1)-B{five_out_row},"—")')
    gap.fill = FILL_CONST
    gap.border = BORDER
    gap.number_format = FMT_DEC1
    ws.cell(r, 3, clean_cell("how far the £5M plant sits below the investable size")
            ).font = FONT_NOTE
    ws.cell(r, 1).border = BORDER
    r += 2

    # ----------------------------------------------------------------------
    # "What moves the sweet spot" note — the 2-3 most sensitive inputs.
    # ----------------------------------------------------------------------
    sub_banner(ws, r, "What moves the sweet spot", 6)
    r += 1
    movers = (
        "Revenue scales linearly with the sale price and 1:1 with output, so the "
        "SALE PRICE is the dominant lever — a 25% lift pulls the investable scale "
        "sharply DOWN (smaller plant clears the hurdle). ENERGY price + connected "
        "load set the largest variable-opex line and bite hardest at small scale "
        "where fixed labour dominates margin. CAPEX (and its six-tenths exponent) "
        "sets the absolute investment and the maintenance line. Flex these three on "
        "the Inputs tab and watch every threshold + curve below move."
    ) if is_ras else (
        "The SALE PRICE per output unit is the dominant lever (revenue is linear in "
        "it and in output). ENERGY price + connected load set the largest variable-"
        "opex line. CAPEX and its six-tenths exponent set the absolute investment "
        "and the maintenance line. Flex these on the Inputs tab — every threshold + "
        "curve moves live."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    mc = ws.cell(r, 1, clean_cell(movers))
    mc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    mc.font = FONT_NOTE
    r += 4

    # ----------------------------------------------------------------------
    # CHART HELPER BLOCK (cols H..L) — a contiguous live mirror of the sweep plus
    # two constant threshold series (discount-rate line + hurdle line) so the
    # IRR-vs-capex chart shows the crossings AS the viability + investable points.
    # Built on THIS tab (cols H+) so every chart series is contiguous & live.
    # ----------------------------------------------------------------------
    hh = r  # helper header row
    ws.cell(hh, 8, clean_cell("Capex £")).font = FONT_SUB
    ws.cell(hh, 9, clean_cell(f"Output {out_unit}")).font = FONT_SUB
    ws.cell(hh, 10, clean_cell("IRR")).font = FONT_SUB
    ws.cell(hh, 11, clean_cell("Discount-rate line")).font = FONT_SUB
    ws.cell(hh, 12, clean_cell("Hurdle line")).font = FONT_SUB
    ws.cell(hh, 13, clean_cell("NPV £")).font = FONT_SUB
    ws.cell(hh, 14, clean_cell("Payback yr (cap 40)")).font = FONT_SUB
    ws.cell(hh, 15, clean_cell("EBITDA margin")).font = FONT_SUB
    hfirst = hh + 1
    n_rows = l - f + 1
    for i in range(n_rows):
        src = f + i  # the sweep source row on the Scenarios tab
        rr = hfirst + i
        # capex (x-axis) + output, mirrored live from the sweep
        ws.cell(rr, 8, f"={sh}!{sw['col_capex']}{src}").number_format = FMT_GBP
        ws.cell(rr, 9, f"={sh}!{sw['col_out']}{src}").number_format = FMT_DEC1
        # IRR — coerce a non-numeric ("n/a") to a plottable NA() so the line breaks
        ws.cell(rr, 10,
                f'=IFERROR(N({sh}!{sw["col_irr"]}{src})+0,NA())').number_format = "0.0%"
        # threshold constant series across the same x-range
        ws.cell(rr, 11, f"={DISC}/100").number_format = "0.0%"
        ws.cell(rr, 12, f"={HURDLE}/100").number_format = "0.0%"
        # NPV mirror
        ws.cell(rr, 13, f"={sh}!{sw['col_npv']}{src}").number_format = FMT_GBP
        # payback capped (text n/a -> 40 so the series plots)
        ws.cell(rr, 14,
                f'=IF({sh}!{sw["col_ebitda"]}{src}>0,'
                f'MIN(40,{sh}!{sw["col_capex"]}{src}/{sh}!{sw["col_ebitda"]}{src}),40)'
                ).number_format = FMT_DEC1
        # EBITDA margin = EBITDA / revenue
        ws.cell(rr, 15,
                f'=IFERROR({sh}!{sw["col_ebitda"]}{src}/{sh}!{sw["col_rev"]}{src},0)'
                ).number_format = "0.0%"
    hlast = hfirst + n_rows - 1

    # ----------------------------------------------------------------------
    # THE CURVES — ScatterChart on a capex / scale x-axis. Threshold lines are
    # constant-value series; their crossing with the IRR curve IS the viability
    # + investable point.
    # ----------------------------------------------------------------------
    from openpyxl.chart import ScatterChart, LineChart, Reference, Series

    def _scatter(title, x_title, y_title, x_col, y_cols_fills, num_fmt="0.0%"):
        ch = ScatterChart()
        ch.title = title
        ch.height, ch.width = 8.5, 15
        ch.x_axis.title = x_title
        ch.y_axis.title = y_title
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        xref = Reference(ws, min_col=x_col, min_row=hfirst, max_row=hlast)
        for (ycol, name) in y_cols_fills:
            yref = Reference(ws, min_col=ycol, min_row=hfirst, max_row=hlast)
            s = Series(yref, xref, title=name)
            ch.series.append(s)
        return ch

    # 1) IRR vs capex with the two threshold lines (the crossings)
    c1 = _scatter("IRR vs capex — crossings = viability & investable points",
                  "Capex £", "IRR",
                  8, [(10, "IRR"), (11, "Discount rate"), (12, "Investor hurdle")])
    # straight-line connectors so the threshold series read as horizontal lines
    for s in c1.series:
        s.smooth = False
        s.marker.symbol = "none"
    style_chart(c1, legend=True)  # 3 genuine series (IRR / discount / hurdle) — keep legend, kill per-point colours
    ws.add_chart(c1, f"A{hh}")

    # 2) NPV vs capex (marks the NPV-max)
    c2 = _scatter("NPV vs capex — peak = NPV-max scale", "Capex £", "NPV £",
                  8, [(13, "NPV")], num_fmt=FMT_GBP)
    for s in c2.series:
        s.smooth = False
    style_chart(c2, legend=None)  # single NPV curve — no per-point legend (was dumping every capex value)
    ws.add_chart(c2, f"A{hh+18}")

    # 3) Payback vs capex
    c3 = _scatter("Payback vs capex", "Capex £", "Payback (yr, cap 40)",
                  8, [(14, "Payback")], num_fmt=FMT_DEC1)
    for s in c3.series:
        s.smooth = False
    style_chart(c3, legend=None)
    ws.add_chart(c3, f"A{hh+36}")

    # 4) EBITDA margin vs scale (output on x)
    c4 = _scatter("EBITDA margin vs scale", f"Output ({out_unit})", "EBITDA margin",
                  9, [(15, "EBITDA margin")])
    for s in c4.series:
        s.smooth = False
    style_chart(c4, legend=None)
    ws.add_chart(c4, f"A{hh+54}")

    # the four scatter charts (anchored in col A) span hh .. hh+72; return below them.
    return max(r, hh + 74)


def _render_sweet_spot_section(ws: Worksheet, state: dict, start_row: int) -> Optional[int]:
    """Render the "Sweet spot & brief reconciliation" block (Phase 1d, 2026-06-24)
    from state['sweetSpot'] (the TS reconcile output — NOT the Python _sweet_spot
    economics mirror). Shows the verdict, the human trade-off statement, the
    recommended operating point (output / capex / £-per-unit), and the cost-output
    frontier as a small table + a tiny £/unit-vs-output line chart. FORMULA-FREE
    (every cell a literal value — zero #REF! risk). Returns the next free row, or
    None when state['sweetSpot'] is absent (older dossiers) so the caller skips
    cleanly."""
    ss = state.get("sweetSpot")
    if not isinstance(ss, dict):
        return None  # older dossier — no reconciliation recorded; skip cleanly

    from openpyxl.chart import LineChart, Reference

    verdict = str(ss.get("verdict", "")).strip() or "unknown"
    objective = str(ss.get("objective", "balanced")).strip() or "balanced"
    unit = str(ss.get("output_unit_label") or "unit")
    stmt = str(ss.get("trade_off_statement", "")).strip()
    rec_out = num(ss.get("recommended_output"))
    rec_cap = num(ss.get("recommended_capex_gbp"))
    rec_cpu = num(ss.get("recommended_cost_per_unit"))
    rescale = num(ss.get("rescale_factor"))
    within_ceiling = ss.get("within_cost_ceiling")
    meets_floor = ss.get("meets_output_floor")
    notes = ss.get("notes") if isinstance(ss.get("notes"), list) else []

    r = start_row

    # ── verdict banner ──
    verdict_label = {
        "compatible": "COMPATIBLE — the brief's cost ceiling and output target fit together",
        "incompatible": "INCOMPATIBLE (reconciled) — cost ceiling & output target conflict; resolved by the stated objective",
        "unconstrained": "UNCONSTRAINED — no cost ceiling stated, so no cost/output tension",
    }.get(verdict, f"verdict: {verdict}")
    sub_banner(ws, r, f"Reconciliation verdict — {verdict_label}", 8)
    vcell = ws.cell(r, 1)
    if verdict == "incompatible":
        vcell.fill = FILL_FAIL; vcell.font = FONT_FAIL
    elif verdict == "compatible":
        vcell.fill = FILL_PASS; vcell.font = FONT_PASS
    r += 1

    # ── the human trade-off statement (wrapped, spanning the width) ──
    if stmt:
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)
        sc = ws.cell(r, 1, clean_cell(stmt))
        sc.alignment = WRAP_TOP
        r += 2
    obj_label = {
        "cost_min": "COST — spend the least to meet the output",
        "output_max": "OUTPUT — deliver the most within budget",
        "balanced": "BALANCED — the best cost/output compromise",
    }.get(objective, objective)
    objc = ws.cell(r, 1, f"Primary objective followed: {obj_label}"
                   + ("  (defaulted — brief did not state a priority)" if ss.get("objective_defaulted") else ""))
    objc.font = FONT_NOTE
    r += 2

    # ── recommended operating point ──
    header(ws, r, ["Recommended operating point", "Value", "", "", "", "", "", ""])
    r += 1
    def _kv(label: str, value: Any, fmt: Optional[str] = None) -> None:
        nonlocal r
        ws.cell(r, 1, label).font = FONT_SUB
        c = ws.cell(r, 2, value)
        if fmt:
            c.number_format = fmt
        c.fill = FILL_RESULT
        r += 1
    if rec_out is not None:
        _kv(f"Recommended output ({unit})", rec_out, FMT_DEC1)
    if rec_cap is not None:
        _kv("Recommended capex (£)", rec_cap, FMT_GBP)
    if rec_cpu is not None:
        _kv(f"Recommended £/{unit}", rec_cpu, FMT_GBP2 if rec_cpu < 100 else FMT_GBP)
    if rescale is not None:
        _kv("Recommended scale vs brief (×)", rescale, FMT_DEC1)
    _kv("Within cost ceiling?", "Yes" if within_ceiling else ("No" if within_ceiling is False else "—"))
    _kv("Meets hard output floor?", "Yes" if meets_floor else ("No" if meets_floor is False else "—"))
    r += 1

    # ── the cost-output frontier table (output | capex | £/unit) ──
    frontier = ss.get("frontier") if isinstance(ss.get("frontier"), list) else []
    if frontier:
        sub_banner(ws, r, "Cost–output frontier (capacity-scaling law) — £/unit falls with scale (economies of scale)", 8)
        r += 1
        header(ws, r, [f"Output ({unit})", "Capex (£)", f"£/{unit}", "", "", "", "", ""])
        r += 1
        chart_first = r
        for pt in frontier:
            if not isinstance(pt, dict):
                continue
            o = num(pt.get("output")); cap = num(pt.get("capex_gbp")); cpu = num(pt.get("cost_per_unit"))
            ws.cell(r, 1, o).number_format = FMT_DEC1
            ws.cell(r, 2, cap).number_format = FMT_GBP
            cc = ws.cell(r, 3, cpu)
            cc.number_format = FMT_GBP2 if (cpu is not None and cpu < 100) else FMT_GBP
            r += 1
        chart_last = r - 1

        # ── tiny line chart: £/unit (y) vs output (x) ──
        if chart_last >= chart_first:
            ch = LineChart()
            ch.title = f"£/{unit} vs output — the sweet-spot curve"
            ch.height, ch.width = 7.5, 14
            ch.y_axis.title = f"£/{unit}"
            ch.x_axis.title = f"Output ({unit})"
            ch.x_axis.delete = False
            ch.y_axis.delete = False
            data = Reference(ws, min_col=3, min_row=chart_first, max_row=chart_last)
            cats = Reference(ws, min_col=1, min_row=chart_first, max_row=chart_last)
            ch.add_data(data, titles_from_data=False)
            ch.set_categories(cats)
            style_chart(ch, legend=None)
            ws.add_chart(ch, f"E{chart_first}")
            r = max(r, chart_last + 16)
        r += 1

    # ── notes (no-auto-rescale disclosure etc.) ──
    for n in notes:
        nc = ws.cell(r, 1, clean_cell(str(n)))
        nc.font = FONT_NOTE
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    return r


def tab_financial_model(wb: Workbook, state: dict) -> bool:
    """A + B (consolidation 2026-06-24) — "Financial model": Economics (base case +
    its opex pie / revenue-vs-EBITDA bar) on top, then the Scenarios scale sweep +
    price sensitivity + its charts, then the Investment-analysis sweet-spot finder +
    its curves — all on ONE sheet. "Inputs & Assumptions" stays a SEPARATE tab (it is
    the editable driver surface every section references).

    Reference safety: every cross-sheet formula either targets 'Inputs & Assumptions'
    (unchanged, still its own tab) or is a SAME-SHEET ref built from `ws.title`/the
    stashed sweep (the Scenarios section stashes `ws._forge_sweep` keyed off THIS sheet,
    which the Investment section reads back) — so no formula carries a now-dead
    "Economics"/"Scenarios" sheet name. Universal; skips only when the Inputs tab
    didn't build (no economic model possible)."""
    if not _ECON_INPUT_ADDR:
        return False  # Inputs tab didn't build -> no economic model

    ws = wb.create_sheet("Financial model")
    # widest of the three sections (Scenarios uses A..M = 13 cols) sets the widths.
    set_widths(ws, {"A": 30, "B": 18, "C": 16, "D": 16, "E": 16, "F": 12,
                    "G": 14, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16})
    title_row(
        ws, "Financial model — economics · scenarios · investment analysis", 8,
        "The whole commercial model on one sheet, every cell a LIVE formula over the "
        "yellow 'Inputs & Assumptions' tab: (1) the base-case Economics (revenue / opex / "
        "EBITDA / NPV / IRR + charts); (2) a fine scale-sweep + Low/Central/High price "
        "sensitivity; (3) the Investment-analysis sweet-spot finder. Edit any input and "
        "everything here recomputes. Money £#,##0; margins 0.0%; years 0.0.",
    )
    r = 4

    def _section(title: str) -> None:
        nonlocal r
        big = ws.cell(r, 1, title)
        big.font = FONT_TITLE; big.fill = FILL_TITLE
        big.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 24
        r += 2

    # ── Economics ──
    _section("Economics — base case (revenue / opex / EBITDA / NPV / IRR)")
    nxt = _render_economics_section(ws, state, r)
    r = (nxt if nxt is not None else r) + 2
    # ── Scenarios (stashes ws._forge_sweep for the section below) ──
    _section("Scenarios — scale sweep & price sensitivity")
    nxt = _render_scenarios_section(ws, state, r)
    r = (nxt if nxt is not None else r) + 2
    # ── Investment Analysis (reads ws._forge_sweep) ──
    _section("Investment analysis — the sweet-spot finder")
    nxt = _render_investment_section(ws, state, r)
    r = (nxt if nxt is not None else r) + 1

    # ── Sweet spot & brief reconciliation (Phase 1d, 2026-06-24) ──
    # Reads state['sweetSpot'] (the TS reconcile output). Skips cleanly (renders
    # nothing) when absent — older dossiers have no reconciliation recorded.
    if state.get("sweetSpot"):
        _section("Sweet spot & brief reconciliation")
        nxt = _render_sweet_spot_section(ws, state, r)
        r = (nxt if nxt is not None else r) + 1

    ws.freeze_panes = "A5"
    back_link(ws, 8)
    return True


# ============================================================================
# MARKDOWN-TABLE PARSER (shared by the schedule tabs #20–22)
# ============================================================================
_MD_SEP = re.compile(r"^\s*\|?[\s:|-]+\|?\s*$")


def parse_md_tables(text: str) -> List[Tuple[str, List[str], List[List[str]]]]:
    """Parse every GitHub-flavoured markdown table in ``text``.
    Returns a list of (preceding_heading, header_cells, [row_cells, ...]).
    A table = a pipe row immediately followed by a |---|---| separator row, then
    consecutive pipe rows. The nearest preceding '#'-heading is attached."""
    lines = text.splitlines()
    tables: List[Tuple[str, List[str], List[List[str]]]] = []
    heading = ""
    i, n = 0, len(lines)

    def cells(s: str) -> List[str]:
        s = s.strip()
        if s.startswith("|"):
            s = s[1:]
        if s.endswith("|"):
            s = s[:-1]
        return [c.strip() for c in s.split("|")]

    while i < n:
        st = lines[i].strip()
        if st.startswith("#"):
            heading = st.lstrip("#").strip()
        if (st.startswith("|") and i + 1 < n
                and _MD_SEP.match(lines[i + 1].strip())
                and "-" in lines[i + 1]):
            hdr = cells(st)
            rows: List[List[str]] = []
            j = i + 2
            while j < n and lines[j].strip().startswith("|"):
                rows.append(cells(lines[j].strip()))
                j += 1
            tables.append((heading, hdr, rows))
            i = j
            continue
        i += 1
    return tables


def _render_md_table(ws: Worksheet, start_row: int, heading: str,
                     hdr: List[str], rows: List[List[str]],
                     spec_col: Optional[int] = None) -> int:
    """Render one parsed md table as a real sortable sheet table starting at
    ``start_row``. Returns (next_free_row, first_body_row). Numeric-looking cells
    are coerced to numbers (so they sort + format correctly); the rest stay as
    text. If ``spec_col`` is given, an 'In spec' ✓/✗ column is coloured."""
    ncol = max(len(hdr), max((len(rw) for rw in rows), default=0))
    if heading:
        sub_banner(ws, start_row, heading, max(ncol, 2))
        start_row += 1
    header(ws, start_row, hdr + [""] * (ncol - len(hdr)))
    r = start_row + 1
    body_first = r
    for rw in rows:
        for ci in range(ncol):
            raw = rw[ci] if ci < len(rw) else ""
            txt = clean_cell(raw)
            cell = ws.cell(r, ci + 1)
            # coerce a clean numeric token (no stray units/parentheses) to a number
            n_only = re.fullmatch(r"-?[\d,]+(?:\.\d+)?", str(txt).replace("£", "").strip())
            if n_only:
                cell.value = num(txt)
            else:
                # Referencify: a cell that IS a master part name (or its tag) becomes a
                # cell-reference to "Part names" — so md-rendered schedules (instrument index,
                # valve list, panel schedule) also point at the one master, not a repeat.
                # Skips the spec ✓/✗ column (left for the live in-spec formula). Literal kept
                # for anything not a registered principal.
                _ref = None
                if not (spec_col is not None and ci + 1 == spec_col) and isinstance(txt, str) and txt.strip():
                    _ref = name_ref(txt) or tag_ref(txt)
                cell.value = _ref if _ref else txt
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            # colour an in-spec ✓/✗ cell
            if spec_col is not None and ci + 1 == spec_col:
                s = str(txt)
                if "✗" in s or s.lower() in ("no", "false", "fail"):
                    cell.fill = FILL_FAIL
                    cell.font = FONT_FAIL
                elif "✓" in s or s.lower() in ("yes", "true", "pass"):
                    cell.fill = FILL_PASS
                    cell.font = FONT_PASS
        r += 1
    # auto-fit row heights so long wrapped text (Service / Description) is not
    # clipped — cells already carry wrap_text (WRAP_TOP) but openpyxl never grows
    # the row to fit it, which read as truncation ("Dual Cornell drain tank…").
    # Estimate wrapped lines per cell from the column width. Universal — keys off
    # column width + content length, no class-specific assumptions.
    for rr in range(body_first, r):
        max_lines = 1
        for ci in range(ncol):
            v = ws.cell(rr, ci + 1).value
            if v is None:
                continue
            cw = ws.column_dimensions[get_column_letter(ci + 1)].width or 12
            approx = max(6, int((cw - 1) * 1.05))   # ~chars per line at this width
            lines = max(1 + str(v).count("\n"), -(-len(str(v)) // approx))
            if lines > max_lines:
                max_lines = lines
        if max_lines > 1:
            ws.row_dimensions[rr].height = min(max_lines, 6) * 14.5
    return r, body_first


# Universal engineering-nomenclature glossary. Standard across EVERY archetype
# (it is industry abbreviation, not RAS-specific) — Tristan: "what does DN stand
# for?", "I don't know what ISA stands for", "abbreviations glossary". Grouped so
# related terms sit together; British spelling; every term spelled out in full.
_GLOSSARY: List[tuple] = [
    ("Pipe sizes & materials", [
        ("DN", "Diameter Nominal — the nominal bore size of a pipe / valve in millimetres (e.g. DN200 ≈ 200 mm bore). A size label, not the exact internal diameter."),
        ("NB", "Nominal Bore — same idea as DN; the nominal internal pipe size."),
        ("CS", "Carbon steel — the default structural / pipe steel where no corrosion or hygiene driver applies."),
        ("316L", "316L stainless steel — low-carbon austenitic stainless; the corrosion-resistant grade for process-water, oxidiser and hygienic duties."),
        ("304 / 304L", "304 / 304L stainless steel — general-purpose austenitic stainless."),
        ("DUPLEX", "Duplex stainless steel — high-strength, chloride-corrosion-resistant stainless for seawater / brackish duties."),
        ("HDPE / PE100", "High-density polyethylene (PE100 grade) — tough, corrosion-free thermoplastic pipe for water / effluent at low pressure."),
        ("PP", "Polypropylene — chemical-resistant thermoplastic, used for ducting, linings and pipe-insulation lagging."),
        ("FRP / GRP", "Fibre-reinforced plastic (glass-reinforced plastic) — moulded composite for tanks and vessels."),
        ("PVC / uPVC", "Poly(vinyl chloride) pipe grades — low-cost drainage / cold-water pipe."),
    ]),
    ("Instrumentation (ISA tags)", [
        ("ISA", "The instrument-tag convention to ISA-5.1 (International Society of Automation) — the letter code naming what an instrument measures and does."),
        ("LT", "Level Transmitter — measures liquid level."),
        ("TT", "Temperature Transmitter — measures temperature."),
        ("PT", "Pressure Transmitter — measures pressure."),
        ("FT", "Flow Transmitter — measures flow rate."),
        ("AT", "Analyser Transmitter — measures a composition variable, e.g. dissolved oxygen, pH or redox."),
        ("AT (pH) / AT (ORP)", "Analyser variant: pH, or oxidation-reduction (redox) potential."),
        ("LSL / LSH", "Level Switch Low / Level Switch High — a discrete level alarm/trip."),
        ("4–20 mA", "The standard analogue field-signal current loop (4 mA = zero, 20 mA = full scale)."),
        ("HART", "Highway Addressable Remote Transducer — a digital signal superimposed on the 4–20 mA loop."),
        ("PLC / SCADA", "Programmable Logic Controller / Supervisory Control And Data Acquisition — the plant control system."),
    ]),
    ("Valves & safety", [
        ("FC", "Fail Closed — the valve drives to the CLOSED position on loss of signal, air or power."),
        ("FO", "Fail Open — the valve drives to the OPEN position on failure."),
        ("PSV / PRV", "Pressure Safety / Relief Valve — mechanical over-pressure protection."),
        ("XV / SDV", "Shutdown / emergency-shutdown Valve — fast on/off safety isolation."),
        ("PCV", "Pressure / Process Control Valve — a modulating control valve."),
        ("HV", "Hand Valve — manual isolation."),
        ("NRV / CV", "Non-Return Valve / Check Valve — prevents reverse flow."),
    ]),
    ("Electrical", [
        ("MCB", "Miniature Circuit Breaker."),
        ("MCCB", "Moulded-Case Circuit Breaker."),
        ("MCC", "Motor Control Centre — the cabinet housing motor starters / drives."),
        ("ATS", "Automatic Transfer Switch — switches the load to standby supply on mains loss."),
        ("UPS", "Uninterruptible Power Supply — battery-backed supply for control / instruments."),
        ("VSD / VFD", "Variable-Speed / Variable-Frequency Drive — controls motor speed."),
        ("LV / HV", "Low Voltage / High Voltage."),
        ("ΔU %", "Volt-drop percentage — the voltage lost along a cable run as a fraction of the supply voltage (kept within spec, typically ≤ 5 %)."),
    ]),
    ("HVAC", [
        ("HVAC", "Heating, Ventilation & Air-Conditioning."),
        ("DX", "Direct-Expansion — a refrigerant-coil cooling / dehumidification system."),
        ("AHU", "Air-Handling Unit — the fan/coil/filter box serving a zone."),
        ("Z-01, Z-02 …", "HVAC zone identifiers — each Z-number is a separately-served air zone (supply / return)."),
        ("ACH", "Air Changes per Hour — room-volume ventilation rate."),
    ]),
    ("Bill-of-materials status codes", [
        ("BESPOKE", "Made-to-order item — fabricated / engineered to spec; no off-the-shelf part number."),
        ("UTILITY", "A utility / commodity item (pipework, cable, civils) priced parametrically."),
        ("SYSTEM", "A multi-part packaged system priced as one unit."),
        ("IDENTIFIED", "A specific real product has been matched (manufacturer + part number)."),
        ("NOT FOUND", "No catalogue match yet — the price is an engineering estimate."),
        ("INSTRUMENT", "A field-instrument line."),
    ]),
    ("Units", [
        ("m³/h", "Cubic metres per hour — volumetric flow."),
        ("m³/s", "Cubic metres per second — volumetric flow."),
        ("kW / kWh", "Kilowatt (power) / kilowatt-hour (energy)."),
        ("mg/L", "Milligrams per litre — concentration (e.g. dissolved oxygen)."),
        ("bar / barg", "Pressure / pressure gauge (above atmospheric)."),
        ("t/yr", "Tonnes per year — annual throughput."),
        ("°C", "Degrees Celsius — temperature."),
    ]),
]


def tab_glossary(wb: Workbook, state: dict) -> bool:
    """Reference glossary of every abbreviation used across the dossier tabs and
    drawings. Universal (standard engineering nomenclature, identical for any
    archetype). Tristan flagged DN / ISA / PT / FC-FO / status-codes / Z-zones as
    undefined."""
    ws = wb.create_sheet("Glossary")
    set_widths(ws, {"A": 20, "B": 104})
    title_row(ws, "Glossary — abbreviations & symbols", 2,
              "Plain-English meaning of every abbreviation used on the schedule, "
              "drawing, bill-of-materials and cost tabs. Standard engineering "
              "nomenclature (British spelling); the same reference applies to any plant.")
    r = 4
    for category, entries in _GLOSSARY:
        sub_banner(ws, r, category, 2)
        r += 1
        for term, meaning in entries:
            tc = ws.cell(r, 1, term)
            tc.font = FONT_SUB
            tc.alignment = WRAP_TOP
            tc.border = BORDER
            mc = ws.cell(r, 2, meaning)
            mc.alignment = WRAP_TOP
            mc.border = BORDER
            # grow the row when the meaning wraps (col B ≈ 102 chars/line)
            lines = max(1, -(-len(meaning) // 102))
            if lines > 1:
                ws.row_dimensions[r].height = lines * 14.5
            r += 1
        r += 1  # spacer between categories
    ws.freeze_panes = "A4"
    back_link(ws, 2)
    return True


_FILL_RISK_HI = PatternFill("solid", fgColor="F4CCCC")   # red-ish
_FILL_RISK_MED = PatternFill("solid", fgColor="FCE5CD")  # amber
_FILL_RISK_LO = PatternFill("solid", fgColor="D9EAD3")   # green

# Universal hazard library. Each entry keys off EQUIPMENT/SERVICE tokens that
# actually appear in the bill of materials (so it fires for whatever plant the
# engine built — no per-class table), and carries the inherent severity ×
# likelihood plus the statutory regulations the hazard triggers. SHARED by the
# Risk Register (X) and the Regulatory & Compliance tab (Y). British spelling;
# regulations default to the UK family (the matrix maps to US/EU equivalents).
#   (token_regex, hazard, cause, severity 1-5, likelihood 1-5, mitigation, [regulations])
_HAZARD_LIB = [
    (r"liquid oxygen|\blox\b|\boxygen\b|oxygenat|\bpsa\b|\bo₂\b",
     "Oxygen enrichment — fire / accelerated combustion",
     "Stored or generated O₂ raises local oxygen concentration; oils and materials ignite readily.",
     4, 2, "O₂-clean materials, exclusion zones, gas detection + ventilation interlocks, no hydrocarbons near O₂.",
     ["Dangerous Substances and Explosive Atmospheres Regulations 2002 (DSEAR)",
      "BCGA / EIGA oxygen-handling codes of practice"]),
    (r"ammonia|\bnh3\b|biofilter|nitrif|\bh2s\b|hydrogen sulphide|chlorine|\bcl2\b",
     "Toxic / asphyxiant gas release",
     "Biological or chemical processes can release ammonia, hydrogen sulphide or chlorine.",
     4, 2, "Gas detection, forced ventilation, scrubbing, respiratory PPE, confined-space controls.",
     ["Control of Substances Hazardous to Health Regulations 2002 (COSHH)"]),
    (r"pressure vessel|pressuris|relief valve|\bpsv\b|compressor|air receiver|autoclave",
     "Stored pressure energy",
     "Pressurised vessels and lines can fail explosively.",
     4, 2, "Design + periodic inspection to pressure code, relief devices, written scheme of examination.",
     ["Pressure Systems Safety Regulations 2000 (PSSR)",
      "Pressure Equipment (Safety) Regulations 2016"]),
    (r"\bpump\b|blower|\bmotor\b|rotat|\bfan\b|mixer|agitator|\bvsd\b|\bdrive\b",
     "Rotating machinery — entanglement / mechanical failure",
     "Exposed rotating parts; stored mechanical energy; loss of a duty machine upsets the process.",
     3, 3, "Fixed guarding, lock-off / isolation (LOTO), N+1 standby on critical duties, vibration monitoring.",
     ["Provision and Use of Work Equipment Regulations 1998 (PUWER)",
      "Supply of Machinery (Safety) Regulations 2008 (UKCA)"]),
    (r"switchgear|transformer|\bmcc\b|\bhv\b|\bkv\b|400 v|busbar|\bups\b|generator|genset|distribution board",
     "Electrical — shock / arc-flash",
     "High fault energy at switchgear and transformers; shock risk from LV / HV.",
     4, 2, "BS 7671 installation, arc-flash assessment, lock-off, IP-rated enclosures, RCD + earthing.",
     ["Electricity at Work Regulations 1989",
      "BS 7671 Requirements for Electrical Installations"]),
    (r"\btank\b|\bvessel\b|\bcolumn\b|\bsump\b|\bbasin\b|\bsilo\b|\bpit\b|reactor",
     "Working at height / confined space (tanks & vessels)",
     "Tank tops present a fall risk; tank interiors are confined spaces (atmosphere / drowning).",
     4, 2, "Edge protection, confined-space permit-to-work, atmosphere testing, rescue plan.",
     ["Work at Height Regulations 2005", "Confined Spaces Regulations 1997"]),
    (r"\bfish\b|aquacultur|livestock|broodstock|larvae|hatchery|\bras\b",
     "Livestock welfare — mass mortality on plant failure",
     "Loss of recirculation, oxygen or temperature control can cause rapid stock mortality.",
     3, 3, "N+1 on critical plant, alarmed O₂ / level / temperature, standby power, automatic failover.",
     ["Welfare of Farmed Animals (England) Regulations 2007", "Animal Welfare Act 2006"]),
    (r"discharge|effluent|sludge|waste ?water|blowdown|\bbleed\b|drain to",
     "Environmental discharge to controlled waters",
     "Process effluent or sludge discharged to controlled waters requires consent.",
     3, 3, "Discharge permit, effluent treatment to consent limits, monitoring + reporting.",
     ["Environmental Permitting (England and Wales) Regulations 2016"]),
    (r"boiler|heater|immersion|heat pump|\bsteam\b|hot water|chiller|refriger",
     "Thermal — burns / scald / refrigerant",
     "Hot surfaces and fluids (scald / burn); refrigerant leak (asphyxiation).",
     3, 2, "Insulation + guarding, thermal relief, refrigerant leak detection + ventilation.",
     ["Provision and Use of Work Equipment Regulations 1998 (PUWER)",
      "Fluorinated Greenhouse Gases Regulations (F-Gas)"]),
    (r"dosing|chemical|\bacid\b|alkali|caustic|bicarb|coagulant|disinfect|biocide",
     "Chemical handling — corrosive / reactive",
     "Dosing chemicals are corrosive or reactive; incompatible mixing is hazardous.",
     3, 2, "Bunding, segregated storage, COSHH assessment, PPE, eyewash / safety shower.",
     ["Control of Substances Hazardous to Health Regulations 2002 (COSHH)",
      "Dangerous Substances and Explosive Atmospheres Regulations 2002 (DSEAR)"]),
    (r"\buv\b|ultraviolet|\bozone\b",
     "Radiation — ultraviolet / ozone exposure",
     "Ultraviolet reactors emit harmful UV; ozone is a respiratory irritant.",
     2, 2, "Interlocked enclosures, UV-opaque shielding, ozone detection + destruct.",
     ["Control of Artificial Optical Radiation at Work Regulations 2010"]),
]

# Statutory duties that apply to ANY industrial plant, by jurisdiction (independent
# of the specific hazards) — the always-on baseline of the regulatory tab.
_BASE_REGS_BY_JURIS = {
    "UK": [("Health and Safety at Work etc. Act 1974", "Overarching duty of care to workers and others."),
           ("Management of Health and Safety at Work Regulations 1999", "Suitable & sufficient risk assessment."),
           ("Construction (Design and Management) Regulations 2015 (CDM)", "Design + construction safety duties."),
           ("Supply of Machinery (Safety) Regulations 2008 (UKCA)", "Machinery conformity + UKCA marking.")],
    "US": [("OSHA General Duty Clause (29 U.S.C. § 654)", "Overarching duty of care."),
           ("OSHA 29 CFR 1910", "General-industry safety standards."),
           ("National Electrical Code (NFPA 70)", "Electrical installation."),
           ("EPA Clean Water Act (NPDES)", "Discharge permitting.")],
    "EU": [("Directive 89/391/EEC (OSH Framework)", "Overarching duty of care."),
           ("Machinery Directive 2006/42/EC (CE)", "Machinery conformity + CE marking."),
           ("ATEX 2014/34/EU", "Equipment in explosive atmospheres."),
           ("Industrial Emissions Directive 2010/75/EU", "Emissions + discharge permitting.")],
}

# Map a UK statute to its rough US/EU equivalent so the hazard regulations follow
# the detected jurisdiction (universal — no per-class content).
_REG_JURIS_MAP = {
    "Dangerous Substances and Explosive Atmospheres Regulations 2002 (DSEAR)":
        {"US": "OSHA 29 CFR 1910.119 (Process Safety Management)", "EU": "ATEX 1999/92/EC"},
    "Control of Substances Hazardous to Health Regulations 2002 (COSHH)":
        {"US": "OSHA Hazard Communication 29 CFR 1910.1200", "EU": "Chemical Agents Directive 98/24/EC"},
    "Pressure Systems Safety Regulations 2000 (PSSR)":
        {"US": "ASME Boiler & Pressure Vessel Code", "EU": "Pressure Equipment Directive 2014/68/EU"},
    "Pressure Equipment (Safety) Regulations 2016":
        {"US": "ASME BPVC", "EU": "Pressure Equipment Directive 2014/68/EU"},
    "Electricity at Work Regulations 1989":
        {"US": "NFPA 70E", "EU": "Low Voltage Directive 2014/35/EU"},
    "BS 7671 Requirements for Electrical Installations":
        {"US": "National Electrical Code (NFPA 70)", "EU": "HD 60364 / IEC 60364"},
    "Environmental Permitting (England and Wales) Regulations 2016":
        {"US": "EPA Clean Water Act (NPDES)", "EU": "Industrial Emissions Directive 2010/75/EU"},
    "Provision and Use of Work Equipment Regulations 1998 (PUWER)":
        {"US": "OSHA 29 CFR 1910 Subpart O", "EU": "Use of Work Equipment Directive 2009/104/EC"},
    "Supply of Machinery (Safety) Regulations 2008 (UKCA)":
        {"US": "ANSI B11 machine safety", "EU": "Machinery Directive 2006/42/EC"},
}


def _derive_hazards(state: dict):
    """Hazards PRESENT in this design, derived from the equipment tokens in the
    bill of materials. Universal — fires for whatever the engine built."""
    bom = state.get("requirementsBom") or []
    corpus = " ".join(
        " ".join(str(b.get(k, "")) for k in ("requirement", "part", "name_human", "status", "tag"))
        for b in bom if isinstance(b, dict)
    ).lower()
    present = []
    for rx, name, cause, sev, lik, mit, regs in _HAZARD_LIB:
        if re.search(rx, corpus):
            present.append(dict(name=name, cause=cause, sev=sev, lik=lik, mit=mit, regs=regs))
    return present


def _rag(score: int):
    if score >= 15:
        return ("High", _FILL_RISK_HI)
    if score >= 8:
        return ("Medium", _FILL_RISK_MED)
    return ("Low", _FILL_RISK_LO)


def _detect_jurisdiction(state: dict) -> str:
    cg = state.get("complianceGate") or {}
    js = cg.get("jurisdictions_detected") or []
    if js:
        j = str(js[0]).upper()
        if j in _BASE_REGS_BY_JURIS:
            return j
        if j in ("GB", "UK", "ENGLAND", "WALES", "SCOTLAND"):
            return "UK"
        if j in ("USA", "US"):
            return "US"
    return "UK"


def _render_risk_section(ws: Worksheet, state: dict, start_row: int) -> Optional[int]:
    """Render the Risk register TABLE onto `ws` starting at `start_row` (header row).
    Returns the next free row, or None when there are no rows to show. The header is
    written by the caller as a section banner. (Was tab_risk_register; refactored into
    a section for the "Risk & Regulatory" merge, 2026-06-24.)"""
    rows = []   # (category, hazard/finding, cause, sev, lik, mitigation, source)

    # 1) LIVE engineering findings — the physics critic's issues
    pc = state.get("physicsCritique") or {}
    for iss in (pc.get("issues") or []):
        if not isinstance(iss, dict):
            continue
        sevtxt = str(iss.get("severity", "")).lower()
        sev = 5 if sevtxt == "high" else 3 if sevtxt in ("medium", "med") else 2
        rows.append(("Engineering — design", iss.get("issue", "").strip() or "Design concern",
                     f"Physics critic ({iss.get('dimension', 'engineering')}), at {iss.get('where', 'design')}.",
                     sev, 3, "Resolve / re-spec before procurement; re-run the physics check to confirm closure.",
                     "state.physicsCritique.issues"))

    # 2) LIVE residual gate flags (QA / commercial)
    for ri in (state.get("residualIssues") or []):
        if not isinstance(ri, dict):
            continue
        gate = str(ri.get("gate", "")).lower()
        cat = ("Commercial — pricing" if "price" in gate else
               "Procurement — part data" if "pn" in gate or "fictional" in gate else
               "Documentation — drawing/layout" if "layout" in gate or "overlap" in gate else
               "Quality assurance")
        rows.append((cat, ri.get("summary", "").strip() or ri.get("gate", "Gate flag"),
                     f"Deterministic {ri.get('gate', 'gate')} raised a flag.",
                     3, 2, "Review the named audit, correct the source line, and re-run the gate.",
                     "state.residualIssues"))

    # 3) LIVE cost-sanity (commercial) when outside / borderline the industry band
    cstat = state.get("costSanity") or {}
    if isinstance(cstat, dict) and cstat.get("ratio_to_nearest_edge"):
        ratio = num(cstat.get("ratio_to_nearest_edge")) or 1.0
        if ratio and ratio > 1.05:
            sev = 4 if ratio > 2 else 3 if ratio > 1.5 else 2
            rows.append(("Commercial — capex", "Capex per output unit outside the typical industry band",
                         clean_cell(cstat.get("message", "")) or "Cost-per-output outside band.",
                         sev, 3, "Value-engineer the high-cost items or confirm the premium is justified.",
                         "state.costSanity"))

    # 4) PROCESS-HAZARD rows from the equipment present (universal hazard library)
    for hz in _derive_hazards(state):
        rows.append(("Process safety / HSE", hz["name"], hz["cause"],
                     hz["sev"], hz["lik"], hz["mit"], "Equipment present in the bill of materials"))

    if not rows:
        return None

    header(ws, start_row, ["#", "Category", "Hazard / finding", "Cause", "S", "L", "Score",
                           "Rating", "Mitigation", "Residual", "Source"])
    r = start_row + 1
    body_first = r
    for i, (cat, hz, cause, sev, lik, mit, src) in enumerate(rows, start=1):
        score = sev * lik
        rating, fill = _rag(score)
        ws.cell(r, 1, i).border = BORDER
        ws.cell(r, 2, clean_cell(cat)).border = BORDER
        c3 = ws.cell(r, 3, clean_cell(hz)); c3.alignment = WRAP_TOP; c3.border = BORDER
        c4 = ws.cell(r, 4, clean_cell(cause)); c4.alignment = WRAP_TOP; c4.border = BORDER
        ws.cell(r, 5, sev).border = BORDER
        ws.cell(r, 6, lik).border = BORDER
        sc = ws.cell(r, 7, score); sc.border = BORDER
        # live recompute so an edited S or L re-scores: =E*F
        sc.value = f"=E{r}*F{r}"
        rc = ws.cell(r, 8, rating); rc.border = BORDER; rc.fill = fill
        rc.alignment = Alignment(horizontal="center")
        c9 = ws.cell(r, 9, clean_cell(mit)); c9.alignment = WRAP_TOP; c9.border = BORDER
        ws.cell(r, 10, "Tolerable (with mitigation)").border = BORDER
        c11 = ws.cell(r, 11, clean_cell(src)); c11.alignment = WRAP_TOP; c11.border = BORDER
        # grow row for the longest wrapped cell
        longest = max(len(str(hz)), len(str(cause)), len(str(mit)))
        if longest > 40:
            ws.row_dimensions[r].height = min(-(-longest // 40), 5) * 14.5
        r += 1
    ws.auto_filter.ref = f"A{start_row}:K{r - 1}"
    return r


def tab_risk_regulatory(wb: Workbook, state: dict) -> bool:
    """E (consolidation 2026-06-24) — "Risk & Regulatory": the Risk register section
    on top, the Regulatory & compliance section below, on ONE sheet (both derive from
    the SAME universal hazard library, so they belong together). Each section keeps its
    full table. Universal; skips only if NEITHER section has content."""
    ws = wb.create_sheet("Risk & Regulatory")
    set_widths(ws, {"A": 6, "B": 22, "C": 40, "D": 40, "E": 6, "F": 6, "G": 7,
                    "H": 10, "I": 46, "J": 16, "K": 26})
    title_row(
        ws, "Risk & Regulatory", 11,
        "Preliminary hazard & risk register PLUS the regulatory / compliance duties, on "
        "one sheet (both derive from the same universal hazard library). Risk rows come "
        "LIVE from the physics critic, the deterministic gate flags, the cost-sanity check "
        "and the equipment present; the regulatory section adds the engine's compliance-gate "
        "verdict and the jurisdiction × hazard → regulation matrix. Score = S × L (≤7 Low / "
        "8–14 Medium / ≥15 High). Universal — not plant-specific. Preliminary; a site-specific "
        "HAZID / HAZOP and a regulatory review are required before construction / sale.",
    )
    r = 4
    # ── Risk register section ──
    sub_banner(ws, r, "Risk register", 11)
    r += 1
    nxt = _render_risk_section(ws, state, r)
    if nxt is None:
        ws.cell(r, 1, "— no live risk findings or process hazards for this design —").font = FONT_NOTE
        r += 2
    else:
        r = nxt + 1
    # ── Regulatory & compliance section ──
    sub_banner(ws, r, "Regulatory & compliance", 11)
    r += 1
    r = _render_regulatory_section(ws, state, r)
    ws.freeze_panes = "A5"
    back_link(ws, 11)
    return True


def _render_regulatory_section(ws: Worksheet, state: dict, start_row: int) -> int:
    """Render the Regulatory & compliance content onto `ws` starting at `start_row`.
    Returns the next free row. (Was tab_regulatory; refactored into a section for the
    "Risk & Regulatory" merge, 2026-06-24.) Universal: the engine's compliance-gate
    verdict + a jurisdiction × hazard → regulation matrix; honest when the gate skipped."""
    cg = state.get("complianceGate") or {}
    juris = _detect_jurisdiction(state)
    hazards = _derive_hazards(state)
    verdict = str(cg.get("verdict", "—")).upper()
    # one-line jurisdiction / verdict note under the section banner
    sub_banner(ws, start_row,
               f"Jurisdiction: {juris}   ·   compliance-gate verdict: {verdict} "
               f"({cg.get('mandatory_covered', 0)}/{cg.get('mandatory_total', 0)} class standards covered). "
               "Statutory duties below are derived from the jurisdiction + the hazards present "
               "(universal matrix), not a per-class list — verify before sale / construction.", 4)
    r = start_row + 1
    # verdict banner row
    if cg.get("reason"):
        sub_banner(ws, r, "Engine compliance-gate verdict", 4)
        r += 1
        vc = ws.cell(r, 1, verdict)
        vc.font = FONT_SUB
        vc.fill = FILL_PASS if verdict == "PASS" else _FILL_RISK_MED
        m = ws.cell(r, 2, clean_cell(cg.get("reason", "")))
        m.alignment = WRAP_TOP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 30
        r += 2

    # 1) always-on statutory duties for the jurisdiction
    sub_banner(ws, r, f"Statutory duties that always apply ({juris})", 4)
    r += 1
    header(ws, r, ["#", "Regulation / standard", "Applies because", "Status"])
    r += 1
    n = 1
    for reg, why in _BASE_REGS_BY_JURIS.get(juris, _BASE_REGS_BY_JURIS["UK"]):
        ws.cell(r, 1, n).border = BORDER
        ws.cell(r, 2, clean_cell(reg)).border = BORDER
        c3 = ws.cell(r, 3, clean_cell(why)); c3.alignment = WRAP_TOP; c3.border = BORDER
        ws.cell(r, 4, "Mandatory — verify").border = BORDER
        r += 1
        n += 1

    # 2) hazard-driven regulations (deduped), mapped to the detected jurisdiction
    r += 1
    sub_banner(ws, r, "Hazard-driven regulations (from the equipment present)", 4)
    r += 1
    header(ws, r, ["#", "Regulation / standard", "Triggered by (hazard present)", "Status"])
    r += 1
    seen = {}
    for hz in hazards:
        for reg in hz["regs"]:
            mapped = reg
            if juris != "UK" and reg in _REG_JURIS_MAP:
                mapped = _REG_JURIS_MAP[reg].get(juris, reg)
            seen.setdefault(mapped, []).append(hz["name"])
    n = 1
    for reg, triggers in seen.items():
        ws.cell(r, 1, n).border = BORDER
        ws.cell(r, 2, clean_cell(reg)).border = BORDER
        # dedupe trigger names, keep readable
        trig = "; ".join(dict.fromkeys(triggers))
        c3 = ws.cell(r, 3, clean_cell(trig)); c3.alignment = WRAP_TOP; c3.border = BORDER
        if len(trig) > 48:
            ws.row_dimensions[r].height = min(-(-len(trig) // 48), 4) * 14.5
        ws.cell(r, 4, "Mandatory — verify").border = BORDER
        r += 1
        n += 1

    # 3) any conflicts the gate found
    conflicts = cg.get("conflicts") or []
    if conflicts:
        r += 1
        sub_banner(ws, r, "Standards conflicts flagged by the engine", 4)
        r += 1
        for c in conflicts:
            ws.cell(r, 1, "⚠").font = FONT_FAIL
            cc = ws.cell(r, 2, clean_cell(str(c))); cc.alignment = WRAP_TOP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            r += 1

    return r


# Universal construction / erection sequence — the standard order of works for a
# process plant. Each phase keys off equipment tokens in the bill of materials, so
# the equipment list populates from whatever the engine built (no per-class table).
#   (phase, token_regex_or_None, predecessor, plant/lifting, hold/witness point)
_ERECTION_PHASES = [
    ("Site establishment & civils",
     r"foundation|slab|\bbund\b|civil|groundwork|plinth|concrete|earthwork|drainage channel",
     "Site handover & set-out", "Excavator, concrete pump",
     "Foundation survey + concrete cube tests signed off"),
    ("Tankage & major static equipment erection",
     r"\btank\b|\bvessel\b|\bcolumn\b|biofilter|degas|\bsilo\b|\bbasin\b|\bsump\b|reactor|skimmer|clarifier|\bmbbr\b|media",
     "Civils complete & cured", "Mobile crane",
     "Tank / vessel hydrostatic (leak) test witnessed"),
    ("Mechanical equipment installation",
     r"\bpump\b|blower|\bfilter\b|heat pump|compressor|screen|mixer|\bfan\b|dehumidif|\buv\b|chiller|\bhrv\b|skid|aerat|exchanger",
     "Tankage set & grouted", "Forklift / overhead crane",
     "Alignment + rotation (bump) test recorded"),
    ("Pipework & ductwork",
     r"\bpipe\b|\bvalve\b|\bline\b|\bduct\b|manifold|header|\bspool\b|fitting|flange|\bdn\d",
     "Equipment installed", "Pipe trolleys, chain hoists",
     "Pressure / leak test certificate per line"),
    ("Electrical installation",
     r"switchgear|transformer|\bmcc\b|\bpanel\b|\bcable\b|\bups\b|generator|genset|distribution|busbar|\bats\b|feeder|breaker|\bmccb\b|\bmcb\b",
     "Equipment set, cable routes ready", "Cable drum jacks",
     "Insulation-resistance + earth-continuity tests"),
    ("Instrumentation, controls & SCADA",
     r"transmitter|\bsensor\b|\bprobe\b|analys|instrument|controller|\bscada\b|\bplc\b|flow meter|\bgauge\b|\bhmi\b|level switch",
     "Electrical energised (LV)", "Hand tools",
     "Loop checks + calibration certificates"),
    ("Pre-commissioning & commissioning",
     None, "All systems installed & tested", "—",
     "Water-on, functional + performance test; client witness"),
]


def tab_assembly_sequence(wb: Workbook, state: dict) -> bool:
    """Z — native Assembly & Erection Sequence (Tristan 2026-06-21: bring the PDF's
    assembly/erection sequence into the Excel). Universal: a standard order-of-works
    whose per-phase equipment list is derived from the bill of materials by discipline
    keyword (no per-class table). Each phase carries its predecessor, lifting plant and
    a hold / witness point."""
    bom = state.get("requirementsBom") or []
    principals = [b for b in bom if isinstance(b, dict) and not b.get("sub_of") and num(b.get("line_gbp"))]
    if not principals:
        principals = [b for b in bom if isinstance(b, dict) and b.get("requirement")]
    if not principals:
        return False

    # assign each principal to the FIRST matching phase (commissioning takes none)
    buckets = {ph[0]: [] for ph in _ERECTION_PHASES}
    for b in principals:
        blob = f"{b.get('requirement', '')} {b.get('part', '')} {b.get('tag', '')}".lower()
        placed = False
        for ph, rx, *_ in _ERECTION_PHASES:
            if rx and re.search(rx, blob):
                tag = clean_cell(b.get("tag", ""))
                nm = clean_cell(b.get("requirement", "")).lstrip("↳ ").split("·")[0].strip()
                buckets[ph].append((nm, tag))      # (name, tag) — referenced at render
                placed = True
                break
        if not placed:
            buckets["Mechanical equipment installation"].append(
                (clean_cell(b.get("requirement", "")).split("·")[0].strip(), ""))

    ws = wb.create_sheet("Assembly sequence")
    set_widths(ws, {"A": 6, "B": 34, "C": 60, "D": 26, "E": 22, "F": 40})
    title_row(
        ws, "Assembly & erection sequence", 6,
        "Standard order of works for the plant, derived from the bill of materials by "
        "discipline. Each step lists the principal equipment installed in it (with tags), "
        "its predecessor, the lifting plant, and the hold / witness point that gates the "
        "next step. Universal sequence — not specific to this plant type. Indicative; a "
        "site-specific construction phase plan (CDM) is required before works begin.",
    )
    header(ws, 4, ["Step", "Phase / activity", "Principal equipment installed (tag)",
                   "Predecessor", "Lifting plant", "Hold / witness point"])
    r = 5
    step = 1
    for ph, rx, pred, plant, hold in _ERECTION_PHASES:
        items = list(dict.fromkeys(buckets.get(ph, [])))
        if ph.startswith("Pre-commissioning"):
            equip = "Whole plant — flush, fill, leak-check, energise, wet-commission, performance test."
        elif items:
            shown = items[:14]
            tail = len(items) - 14
            # Build a concat formula REFERENCING the master "Part names" for each principal's
            # name + tag (one identity; click a precedent). Literal fallback per item; if no
            # item resolves, emit the plain string (no pointless all-literal formula).
            exprs, any_ref = [], False
            for nm, tag in shown:
                nref = name_ref(nm)
                tref = tag_ref(tag) if tag else None
                any_ref = any_ref or bool(nref) or bool(tref)
                nexpr = nref[1:] if nref else '"' + str(nm).replace('"', '""') + '"'
                if tag:
                    texpr = tref[1:] if tref else '"' + str(tag).replace('"', '""') + '"'
                    exprs.append(f'{nexpr} & " [" & {texpr} & "]"')
                else:
                    exprs.append(nexpr)
            if any_ref:
                joined = ' & "; " & '.join(exprs)
                if tail > 0:
                    joined += f' & "  (+{tail} more)"'
                equip = "=" + joined
            else:
                equip = "; ".join(f"{nm}{(' [' + tag + ']') if tag else ''}" for nm, tag in shown) \
                    + (f"  (+{tail} more)" if tail > 0 else "")
        else:
            continue  # skip a phase with no equipment (other than commissioning)
        ws.cell(r, 1, step).border = BORDER
        ws.cell(r, 2, clean_cell(ph)).border = BORDER
        c3 = ws.cell(r, 3, equip); c3.alignment = WRAP_TOP; c3.border = BORDER
        ws.cell(r, 4, clean_cell(pred if step > 1 else "Site handover & set-out")).border = BORDER
        ws.cell(r, 5, clean_cell(plant)).border = BORDER
        c6 = ws.cell(r, 6, clean_cell(hold)); c6.alignment = WRAP_TOP; c6.border = BORDER
        lines = max(1, -(-len(equip) // 58))
        if lines > 1:
            ws.row_dimensions[r].height = min(lines, 6) * 14.5
        r += 1
        step += 1
    ws.freeze_panes = "A5"
    back_link(ws, 6)
    return True


def tab_panel_schedule(wb: Workbook, run_dir: str) -> bool:
    """#20 — Panel / load schedule as a real table (from panel-schedule.md)."""
    path = os.path.join(run_dir, "drawings", "panel-schedule.md")
    if not os.path.exists(path):
        return False
    try:
        text = open(path, "r").read()
    except Exception:  # noqa: BLE001
        return False
    tables = parse_md_tables(text)
    if not tables:
        return False
    ws = wb.create_sheet("Panel schedule")
    set_widths(ws, {"A": 14, "B": 40, "C": 8, "D": 18, "E": 12, "F": 22,
                    "G": 22, "H": 12, "I": 10, "J": 8})
    title_row(
        ws, "Electrical panel / load schedule", 10,
        "Real sortable rows (circuit · load · device · cable · volt-drop). The "
        "engine computes Design I = P·1000 / (√3·V·pf·η) — pf 0.85 / η 0.90 for "
        "motor circuits, pf 1.0 / η 1.0 for resistive loads — from each circuit's "
        "INSTALLED motor frame (so the amps can exceed the duty-kW shown). ΔU % is "
        "the cable volt-drop over its length / CSA at Design I. 'In spec' is a LIVE "
        "formula = (ΔU ≤ 5 %): edit a ΔU cell and the verdict + red flag recompute. "
        "Auto-generated; not for construction.",
    )
    r = 4
    last_circuit_first = None
    for heading, hdr, rows in tables:
        # locate the 'in spec' + 'ΔU' columns for live verdicts + conditional colour
        spec_col = None
        du_col = None
        for idx, h in enumerate(hdr, start=1):
            hl = h.lower()
            if "spec" in hl:
                spec_col = idx
            if "δu" in hl or "volt" in hl or "drop" in hl or "δ" in hl:
                du_col = idx
        r, body_first = _render_md_table(ws, r, heading, hdr, rows, spec_col)
        body_last = r - 1
        # MAKE 'In spec' LIVE: verdict computed from the ΔU cell vs the 5% limit,
        # not the static '✓' parsed from the markdown (Tristan: "doesn't seem to be
        # using real formulas, which makes me suspicious"). Universal — any panel md
        # with a volt-drop + in-spec column gets a live check.
        if spec_col and du_col and len(hdr) >= 6 and body_last >= body_first:
            du_L = get_column_letter(du_col)
            sp_L = get_column_letter(spec_col)
            for rr in range(body_first, body_last + 1):
                cell = ws.cell(rr, spec_col)
                # ELSE branch must be a STATIC value — referencing the spec cell itself
                # ({sp_L}{rr}) is a CIRCULAR reference (the cell is THIS formula). When
                # ΔU is non-numeric there is no volt-drop to check, so show "—".
                cell.value = (f'=IF(ISNUMBER({du_L}{rr}),'
                              f'IF({du_L}{rr}<=5,"✓","✗"),"—")')
            from openpyxl.formatting.rule import CellIsRule
            ws.conditional_formatting.add(
                f"{sp_L}{body_first}:{sp_L}{body_last}",
                CellIsRule(operator="equal", formula=['"✗"'],
                           fill=FILL_FAIL, font=FONT_FAIL))
            ws.conditional_formatting.add(
                f"{sp_L}{body_first}:{sp_L}{body_last}",
                CellIsRule(operator="equal", formula=['"✓"'],
                           fill=FILL_PASS, font=FONT_PASS))
        # autofilter + format the wide circuit table (the one with many columns)
        if len(hdr) >= 6:
            last_circuit_first = (body_first, r - 1, len(hdr))
        r += 1
    if last_circuit_first:
        bf, bl, nc = last_circuit_first
        ws.auto_filter.ref = f"A{bf - 1}:{get_column_letter(nc)}{bl}"
    ws.freeze_panes = "A5"
    back_link(ws, 10)
    return True


def tab_process_schedules(wb: Workbook, run_dir: str) -> int:
    """#21 — Process line list / valve list / instrument index as ONE sheet
    "Process schedules" with three section bands (Tristan 2026-06-24 consolidation:
    they all come from the same drawings/process-schedules.md, so one sheet with
    three banded sections is cleaner than three near-identical tabs). Each section is
    still its own sortable table. Returns the number of sections rendered (>0 => tab
    kept)."""
    path = os.path.join(run_dir, "drawings", "process-schedules.md")
    if not os.path.exists(path):
        return 0
    try:
        text = open(path, "r").read()
    except Exception:  # noqa: BLE001
        return 0
    tables = parse_md_tables(text)
    if not tables:
        return 0

    # a friendly section title per source heading
    name_for = {
        "line": "Process line list",
        "valve": "Process valve list",
        "instrument": "Process instruments",
    }
    # pre-filter to non-empty tables + widest column count across them all (so the
    # title band + back-link span the whole sheet)
    secs = [(h, hd, rw) for (h, hd, rw) in tables if rw]
    if not secs:
        return 0
    span = 2
    for _h, hd, rw in secs:
        span = max(span, len(hd), max((len(r) for r in rw), default=0))

    ws = wb.create_sheet("Process schedules")
    widths = {get_column_letter(i): (14 if i > 1 else 16) for i in range(1, span + 1)}
    set_widths(ws, widths)
    title_row(ws, "Process schedules", span,
              "Process line list, valve list and instrument index — each a sortable "
              "section parsed from process-schedules.md and cross-referenced to the P&ID.")
    made = 0
    r = 4
    used = set()
    for heading, hdr, rows in secs:
        hl = heading.lower()
        sect = next((v for k, v in name_for.items() if k in hl), None)
        if sect is None:
            sect = "Process " + re.sub(r"[^a-z0-9 ]", "", hl).strip()[:24]
        # disambiguate a repeated heading
        base = sect
        n = 1
        while sect in used:
            n += 1
            sect = f"{base} ({n})"
        used.add(sect)
        # widen any 'service'/'description'-like column for THIS section
        for idx, h in enumerate(hdr, start=1):
            if h.lower() in ("service", "description", "measured", "notes", "remarks"):
                ws.column_dimensions[get_column_letter(idx)].width = 54
        # banded section heading + its table; _render_md_table writes the section
        # banner itself when passed a heading.
        nxt, body_first = _render_md_table(ws, r, sect, hdr, rows)
        ncol = max(len(hdr), max((len(rw) for rw in rows), default=0))
        # a per-section auto-filter on the LAST section is most useful; set it each
        # time so the final survivor covers that section's rows.
        ws.auto_filter.ref = f"A{body_first - 1}:{get_column_letter(ncol)}{nxt - 1}"
        r = nxt + 1  # blank spacer row between sections
        made += 1
    ws.freeze_panes = "A4"
    back_link(ws, span)
    return made


def _line_section(row: dict, spec: dict) -> str:
    """Classify a sized run so flow lines and electrical cables never share a units
    column (Tristan: 'water and electrics mixed... amps and cubes'). Universal —
    keys off the connection-schedule's own kind / mechanism, no class assumptions."""
    kind = (spec.get("kind") or "").lower()
    mech = (row.get("mechanism") or "").lower()
    if kind == "cable" or "electric" in mech or mech == "signal" or "power" in mech:
        return "electrical"
    if kind == "pipe" and ("loop" in mech or "fluid" in mech):
        return "pipe"
    return "other"


def _line_basis(row: dict) -> str:
    """How a Line £ is built (Tristan: 'how or where the line cost number comes from')
    — the material take-off + £/m × length + install, from the row's own fields."""
    bits = []
    q = clean_cell(row.get("qty", ""))
    if q:
        bits.append(q)
    uc = num(row.get("unit_cost_gbp"))
    ln = num(row.get("length_m"))
    if uc is not None and ln:
        bits.append(f"£{uc:,.0f}/m × {ln:,.1f} m")
    elif uc is not None:
        bits.append(f"£{uc:,.0f}/unit")
    inst = num(row.get("install_gbp"))
    if inst:
        bits.append(f"install £{inst:,.0f}")
    return " · ".join(bits)


def tab_line_velocity(wb: Workbook, run_dir: str) -> bool:
    """#22 — Line & velocity schedule from connection-schedule.json, SPLIT into
    process-pipe lines (velocity vs ≤3 m/s) and electrical/signal cables (volt-drop
    vs ≤5 %) so flow, current and power units never mix in one column. Adds a Basis
    column showing how each Line £ is derived. within_spec ✗ conditional-formatted RED."""
    cs = load_json(os.path.join(run_dir, "connection-schedule.json"))
    if not cs or not isinstance(cs, dict):
        return False
    rows = cs.get("rows")
    if not rows or not isinstance(rows, list):
        return False
    specs = cs.get("specs") or []  # parallel, carries the spec_limit text

    ws = wb.create_sheet("Line & velocity")
    set_widths(ws, {"A": 6, "B": 22, "C": 22, "D": 16, "E": 16, "F": 14,
                    "G": 16, "H": 10, "I": 9, "J": 13, "K": 62})
    title_row(
        ws, "Line & velocity schedule", 11,
        "Every sized run, SPLIT so units never mix: process-pipe lines (flow m³/s, "
        "velocity vs ≤ 3 m/s) · electrical & signal cables (current A, volt-drop vs "
        "≤ 5 %) · other services. Velocity / ΔU is the AS-SIZED value (post-upsizing); "
        "'In spec' ✗ = RED. Basis shows how each Line £ is built (take-off × £/m + install).",
    )

    # bucket rows by section, preserving the original index for the '#'
    groups: Dict[str, list] = {"pipe": [], "electrical": [], "other": []}
    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        spec = specs[idx] if idx < len(specs) and isinstance(specs[idx], dict) else {}
        groups[_line_section(row, spec)].append((idx, row, spec))

    SECTIONS = [
        ("pipe", "Process pipe lines — flow (m³/s) · velocity limit ≤ 3 m/s"),
        ("electrical", "Electrical & signal cables — current (A) · volt-drop limit ≤ 5 %"),
        ("other", "Other services (air / misc)"),
    ]
    HDR = ["#", "From", "To", "Size", "Rating", "Velocity / ΔU", "Spec limit",
           "Length (m)", "In spec", "Line £", "Basis"]
    r = 4
    grand_fail = 0
    spec_cells = []   # In-spec column cells for the conditional-format range
    for key, label in SECTIONS:
        g = groups[key]
        if not g:
            continue
        sub_banner(ws, r, label, 11)
        r += 1
        header(ws, r, HDR)
        r += 1
        sec_first = r
        for idx, row, spec in g:
            in_spec = row.get("within_spec")
            ws.cell(r, 1, idx + 1).border = BORDER
            # From / To endpoints REFERENCE the master "Part names" row where the endpoint
            # is a registered principal (one identity; click through to the part). The
            # connection-schedule labels endpoints by NAME ("Standby Diesel Generator"), so
            # resolve by name first, tag second; literal kept for boundary endpoints (e.g.
            # "Utility Incomer") the master doesn't carry.
            _fr = name_ref(row.get("from", "")) or tag_ref(row.get("from", ""))
            ws.cell(r, 2, _fr if _fr else clean_cell(row.get("from", ""))).border = BORDER
            _to = name_ref(row.get("to", "")) or tag_ref(row.get("to", ""))
            ws.cell(r, 3, _to if _to else clean_cell(row.get("to", ""))).border = BORDER
            ws.cell(r, 4, clean_cell(row.get("size", ""))).border = BORDER
            ws.cell(r, 5, clean_cell(row.get("rating", ""))).border = BORDER
            # AS-SIZED velocity / volt-drop (post-upsizing) from specs — never the
            # stale pre-upsize row['drop'].
            velnum = num(spec.get("drop_pct_or_velocity"))
            if velnum is not None:
                vc = ws.cell(r, 6, velnum)
                vc.number_format = FMT_DEC2
            else:
                vc = ws.cell(r, 6, clean_cell(row.get("drop", "")))
            vc.border = BORDER
            ws.cell(r, 7, clean_cell(spec.get("spec_limit", ""))).border = BORDER
            lc = ws.cell(r, 8, num(row.get("length_m")))
            lc.number_format = FMT_DEC1
            lc.border = BORDER
            sc = ws.cell(r, 9, "✓" if in_spec is not False else "✗")
            sc.border = BORDER
            sc.alignment = Alignment(horizontal="center")
            spec_cells.append(sc.coordinate)
            if in_spec is False:
                sc.font = FONT_FAIL
                grand_fail += 1
                for col in range(1, 12):
                    ws.cell(r, col).fill = FILL_FAIL
            else:
                sc.fill = FILL_PASS
                sc.font = FONT_PASS
            lt = ws.cell(r, 10, num(row.get("line_total_gbp")))
            lt.number_format = FMT_GBP
            lt.border = BORDER
            basis = _line_basis(row)
            bc = ws.cell(r, 11, basis)
            bc.alignment = WRAP_TOP
            bc.border = BORDER
            if len(basis) > 60:
                ws.row_dimensions[r].height = min(-(-len(basis) // 60), 4) * 14.5
            r += 1
        # per-section Σ line £
        ws.cell(r, 2, f"Σ {label.split('—')[0].strip()}").font = FONT_SUB
        st = ws.cell(r, 10, f"=SUM(J{sec_first}:J{r - 1})")
        st.font = Font(bold=True)
        st.fill = FILL_RESULT
        st.number_format = FMT_GBP
        r += 2

    # grand foot: out-of-spec tally across all sections
    ws.cell(r, 2, "Runs out-of-spec (all sections, as-sized)").font = FONT_SUB
    tc = ws.cell(r, 9, grand_fail)
    tc.font = FONT_FAIL if grand_fail else FONT_PASS
    tc.fill = FILL_FAIL if grand_fail else FILL_PASS
    tc.alignment = Alignment(horizontal="center")

    # any '✗' in an In-spec cell goes red live
    if spec_cells:
        from openpyxl.formatting.rule import CellIsRule
        ws.conditional_formatting.add(
            f"I5:I{r}",
            CellIsRule(operator="equal", formula=['"✗"'], fill=FILL_FAIL, font=FONT_FAIL))
    ws.freeze_panes = "A5"
    back_link(ws, 11)
    return True


# ============================================================================
# TAB — CONTENTS / INDEX (#26)  — built LAST, moved to sheet #1
# ============================================================================
def tab_contents(wb: Workbook, descriptions: Dict[str, str]) -> None:
    """Create the Contents sheet (a hyperlink index to every other tab + a
    one-line description) and MOVE it to position #1. Called after all other
    sheets exist so the full ordered sheet list is known."""
    ws = wb.create_sheet(CONTENTS_SHEET)
    set_widths(ws, {"A": 4, "B": 34, "C": 78})
    title_row(
        ws, "Contents", 3,
        "Click any tab name to jump to it. Each data tab has a '↑ Contents' "
        "link at its top-right to come back. Yellow cells anywhere are editable "
        "inputs; green cells are live formulas.",
    )
    header(ws, 4, ["#", "Tab", "What's on it"])
    r = 5
    n = 1
    for name in wb.sheetnames:
        if name == CONTENTS_SHEET:
            continue
        ws.cell(r, 1, n).border = BORDER
        c = ws.cell(r, 2, name)
        c.hyperlink = f"#'{name}'!A1"
        c.font = FONT_LINK
        c.border = BORDER
        d = ws.cell(r, 3, descriptions.get(name, _default_desc(name)))
        d.alignment = WRAP_TOP
        d.font = FONT_NOTE
        d.border = BORDER
        r += 1
        n += 1
    ws.freeze_panes = "A5"
    # move to the very front
    idx = wb.sheetnames.index(CONTENTS_SHEET)
    if idx != 0:
        wb.move_sheet(CONTENTS_SHEET, -idx)


def _default_desc(name: str) -> str:
    """Fallback one-line description for image / module tabs not in the static map.
    Distinct lines per drawing TYPE (matched on the title) so GA/P&ID/BFD/Single-line/
    HVAC never share an identical generic string."""
    low = name.lower()
    if low.startswith("module —"):
        return "Per-module Blender render."
    if low.startswith("isometric"):
        return "Representative pipe isometric drawing."
    if low.startswith("render"):
        return "Photoreal Blender render."
    if low.startswith("ga") or "general arrangement" in low:
        return "General arrangement — equipment layout & footprint."
    if "p&id" in low or low.startswith("pid"):
        return "Piping & instrumentation diagram — process flow, valves, instruments."
    if low.startswith("bfd") or "block flow" in low:
        return "Block flow diagram — major process blocks & streams."
    if "single-line" in low or "single line" in low:
        return "Single-line electrical distribution diagram."
    if "hvac" in low:
        return "Ventilation & climate-control schematic."
    return "Engineering drawing / view."


# ============================================================================
# IMAGE TABS
# ============================================================================
def ensure_png(src_path: str, run_dir: str) -> Optional[str]:
    """
    Return a PNG path for an image reference. If only an SVG exists, convert it to
    PNG (cairosvg preferred, then rsvg-convert). Returns None if nothing usable.
    """
    if src_path.lower().endswith(".png") and os.path.exists(src_path):
        return src_path
    # if given a .svg, try png sibling first
    base, ext = os.path.splitext(src_path)
    png_sib = base + ".png"
    if os.path.exists(png_sib):
        return png_sib
    if ext.lower() == ".svg" and os.path.exists(src_path):
        out_png = os.path.join(run_dir, ".excel-tmp",
                               os.path.basename(base) + ".png")
        os.makedirs(os.path.dirname(out_png), exist_ok=True)
        # try cairosvg
        try:
            import cairosvg  # noqa: WPS433
            cairosvg.svg2png(url=src_path, write_to=out_png, output_width=3000)
            return out_png
        except Exception:  # noqa: BLE001
            pass
        # try rsvg-convert
        try:
            subprocess.run(["rsvg-convert", "-w", "3000", "-o", out_png, src_path],
                           check=True, capture_output=True, timeout=60)
            if os.path.exists(out_png):
                return out_png
        except Exception:  # noqa: BLE001
            pass
    return None


def downscale_png(src_png: str, run_dir: str, max_px: int = 3400) -> str:
    """
    Downscale a PNG so the whole workbook stays manageable. Raised 1400→2600→3400
    (Tristan: "make all of the diagrams higher resolution" + "A2 minimum"): native
    drawings are 2160–6560 px and the densest (single-line 6140, P&ID 4380, block-
    flow 4526, process-schedule 4356) were still being crushed to 2600. 3400 keeps
    the dense P&ID / single-line / process-schedule legible on zoom (≈ A2 at ~150 dpi)
    while bounding workbook size; smaller renders (hero, isometrics ~2360 px) pass
    through untouched. Universal.
    Returns a path to the (possibly) downscaled PNG. Keeps aspect ratio. JPEG-quality
    PNG compression via Pillow optimisation.
    """
    try:
        from PIL import Image
    except Exception:  # noqa: BLE001
        return src_png
    try:
        with Image.open(src_png) as im:
            im = im.convert("RGB") if im.mode in ("RGBA", "P") else im
            w, h = im.size
            scale = min(1.0, max_px / float(max(w, h)))
            if scale < 1.0:
                im = im.resize((max(1, int(w * scale)), max(1, int(h * scale))))
            out_dir = os.path.join(run_dir, ".excel-tmp")
            os.makedirs(out_dir, exist_ok=True)
            # Name the downscaled file by its PARENT-QUALIFIED relative path, NOT the bare
            # basename — interior 00-hero.png and exterior/00-hero.png share a basename, so a
            # basename-keyed temp made the exterior OVERWRITE the interior and both tabs embedded
            # the SAME image (Tristan 2026-06-22: "interior and exterior layouts are identical").
            try:
                _rel = os.path.relpath(src_png, run_dir)
            except ValueError:
                _rel = os.path.basename(src_png)
            _safe = _rel.replace(os.sep, "_").replace("..", "_")
            out = os.path.join(out_dir, "ds_" + _safe)
            # save as compressed JPEG inside a .png-ext wrapper would confuse openpyxl;
            # keep PNG but optimise.
            im.save(out, format="PNG", optimize=True)
            return out
    except Exception as exc:  # noqa: BLE001
        print(f"    ! downscale failed for {src_png}: {exc}")
        return src_png


_SHEET_BAD = re.compile(r"[:\\/?*\[\]]")


def safe_sheet_title(name: str, used: set) -> str:
    t = _SHEET_BAD.sub("-", name)[:31].strip() or "sheet"
    base = t
    i = 2
    while t.lower() in used:
        suffix = f"-{i}"
        t = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(t.lower())
    return t


def add_image_tab(wb: Workbook, run_dir: str, png_path: str, title: str,
                  caption: str, used_titles: set) -> bool:
    from openpyxl.drawing.image import Image as XLImage
    ds = downscale_png(png_path, run_dir)
    try:
        sheet_title = safe_sheet_title(title, used_titles)
        ws = wb.create_sheet(sheet_title)
        # Shared navy title band (matches the data tabs) + a ↑ Contents back-link, instead
        # of a bare bold cell. title_row returns the next free row; the image goes below it.
        img_row = title_row(ws, title, 6, caption)
        back_link(ws, 6)
        img = XLImage(ds)
        # cap on-sheet display size (keep aspect) so the tab is readable. Raised
        # 1100→1700 to match the higher-res embed: the underlying PNG is up to
        # 2600 px so zooming stays sharp. Universal.
        max_w = 1700
        if img.width and img.width > max_w:
            ratio = max_w / float(img.width)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, f"A{img_row}")
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"    ! could not embed {png_path}: {exc}")
        return False


def collect_image_specs(run_dir: str) -> List[Tuple[str, str, str]]:
    """
    Return ordered list of (file_path, tab_title, caption). UNIVERSAL discovery:
      1. Blender hero (00-hero.png or blender-cover.png)
      2. each module-*.png at the run root (NOT the *-top-front variants)
      3. the 8 named engineering drawings (PNG, or SVG->PNG) from drawings/
         + a few representative isometrics if present
    Missing files are simply skipped.
    """
    specs: List[Tuple[str, str, str]] = []
    draw = os.path.join(run_dir, "drawings")

    def first_existing(*cands: str) -> Optional[str]:
        for c in cands:
            if c and os.path.exists(c):
                return c
        return None

    # 1. hero (interior layout — the materialed render's 00-hero)
    hero = first_existing(os.path.join(run_dir, "00-hero.png"),
                          os.path.join(run_dir, "blender-cover.png"))
    if hero:
        specs.append((hero, "Render — Interior layout",
                      "Blender 3D render — interior plant layout (no walls/roof), view 1."))
    # 1b. SECOND interior angle + the EXTERIOR building views (Tristan 2026-06-22: the Excel was
    #     missing the updated renders + the external building). The exterior set is rendered to
    #     <run>/exterior/ from the SAME scene build (BLENDER_PLANT_SHELL=1).
    _ext = os.path.join(run_dir, "exterior")
    # The full set for BOTH surfaces (Tristan 2026-06-25): hero + front-corner + top-down + back-
    # corner, interior then exterior. Each is guarded by os.path.exists, so a class that renders
    # fewer views simply gets fewer tabs (universal — no per-class assumption).
    for _p, _ttl, _cap in (
        (os.path.join(run_dir, "02-corner-FR.png"), "Render — Interior layout (view 2)",
         "Interior plant layout — front-corner angle."),
        (os.path.join(run_dir, "01-top.png"), "Render — Interior layout (top)",
         "Interior plant layout — top-down plan view."),
        (os.path.join(run_dir, "03-corner-BL.png"), "Render — Interior layout (view 3)",
         "Interior plant layout — back-corner angle."),
        (os.path.join(_ext, "00-hero.png"), "Render — Building exterior",
         "Architectural exterior of the container enclosure, view 1."),
        (os.path.join(_ext, "02-corner-FR.png"), "Render — Building exterior (view 2)",
         "Architectural exterior — front-corner angle."),
        (os.path.join(_ext, "01-top.png"), "Render — Building exterior (top)",
         "Architectural exterior — top-down view."),
        (os.path.join(_ext, "03-corner-BL.png"), "Render — Building exterior (view 3)",
         "Architectural exterior — back-corner angle."),
    ):
        if os.path.exists(_p):
            specs.append((_p, _ttl, _cap))

    # 2. module renders — REMOVED (Tristan 2026-06-20): the per-module Blender
    # highlight renders (module-*.png) read as poor quality and their provenance was
    # unclear, so they no longer get their own tabs. Universal across all classes.

    # 3. the 8 engineering drawings (canonical names + aliases)
    eng = [
        ("general-arrangement", "GA — General Arrangement",
         "General arrangement / plant layout."),
        ("pid", "P&ID", "Piping & instrumentation diagram."),
        ("block-flow-diagram", "BFD — Block Flow",
         "Block flow diagram of the process."),
        ("single-line-diagram", "Single-line",
         "Electrical single-line diagram."),
        ("hvac-layout", "HVAC", "HVAC / ventilation layout."),
        # NOTE: panel-schedule + process-schedules are deliberately NOT embedded
        # as PDF-page images — they are rendered as NATIVE, sortable Excel rows by
        # tab_panel_schedule + tab_process_schedules (Tristan 2026-06-20: "panel /
        # process schedule should be in excel, not a pdf"). Universal for any class
        # that emits those schedules.
    ]
    for stem, ttl, cap in eng:
        # try drawings/<stem>.png, drawings/<stem>.svg, then run-root variants
        cand_png = os.path.join(draw, stem + ".png")
        cand_svg = os.path.join(draw, stem + ".svg")
        root_png = os.path.join(run_dir, stem + ".png")
        path = first_existing(cand_png, root_png)
        if not path:
            path = ensure_png(cand_svg, run_dir)
        if path:
            specs.append((path, ttl, cap))

    # representative isometrics (first up to 3) so the iso family is shown without
    # exploding the file size with all ~60.
    isos = sorted(fn for fn in os.listdir(draw)
                  if fn.startswith("isometric-") and fn.endswith(".png")) if os.path.isdir(draw) else []
    for fn in isos[:3]:
        tag = fn[len("isometric-"):-len(".png")]
        specs.append((os.path.join(draw, fn), f"Isometric — {tag}",
                      f"Representative pipe isometric ({tag})."))

    return specs


# ============================================================================
# MAIN
# ============================================================================
def _render_cabinet_section(ws: Worksheet, run_dir: str, start_row: int) -> Optional[int]:
    """Render the Cabinet schedule TABLE onto `ws` starting at `start_row`. Returns the
    next free row, or None when there is no `cabinets` block to show. (Was
    tab_cabinet_schedule; refactored into a section appended to "Connection trace",
    2026-06-24 — both read the same parts-ledger connectivity verdict, so they cannot
    disagree.) The cabinet rows prove every small electrical / control device is HOUSED
    and that the cabinet + its contents show their IN/OUT connectors, all connected."""
    pl = load_json(os.path.join(run_dir, "parts-ledger.json"))
    cab = (pl.get("cabinets") if isinstance(pl, dict) else None) or {}
    cabinets = cab.get("cabinets") or []
    if not cabinets:
        return None

    proven = cab.get("all_cabinets_proven")
    sub_banner(ws, start_row,
               f"{cab.get('n_cabinets', len(cabinets))} cabinet(s)   ·   "
               f"{cab.get('n_housed', 0)} housed device(s)   ·   "
               f"{cab.get('n_cabinets_all_connected', 0)}/{cab.get('n_cabinets', len(cabinets))} "
               f"cabinets fully connected   ·   "
               f"ALL CONNECTORS PROVEN: {'YES ✓' if proven else 'NO — see ✗ rows'}", 6)
    r = start_row + 1

    def _join(items):
        out = []
        for it in (items or []):
            s = str(it)
            # ledger strings look like 'Main Breaker (X-123) via cable [electrical_bus]'
            out.append(s)
        return "\n".join(out) if out else "—"

    for c in cabinets:
        # ── cabinet header band ──
        dom = (c.get("domain") or "").upper()
        verdict = "✓ ALL CONNECTED" if c.get("all_connected") else "✗ INCOMPLETE"
        hdr_txt = (f"{clean_cell(c.get('name'))}  [{dom} CABINET · {clean_cell(c.get('tag'))}]   "
                   f"— {c.get('n_contents', 0)} device(s) housed   ·   "
                   f"cabinet feeds: {c.get('n_in', 0)} in / {c.get('n_out', 0)} out   ·   {verdict}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        hc = ws.cell(r, 1, clean_cell(hdr_txt))
        hc.font = Font(bold=True, color="FFFFFF")
        hc.fill = PatternFill("solid", fgColor="107C10" if c.get("all_connected") else "C00000")
        hc.alignment = LEFT_TOP
        r += 1
        header(ws, r, ["Housed device", "Tag", "Function", "Inputs ← (from)",
                       "Outputs → (to)", "Connectors"])
        r += 1
        contents = c.get("contents") or []
        if not contents:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(r, 1, "— no discrete devices itemised (cabinet feeds the loads directly) —").font = FONT_NOTE
            r += 1
        for d in contents:
            ws.cell(r, 1, clean_cell(d.get("name"))).border = BORDER
            ws.cell(r, 2, clean_cell(d.get("tag"))).border = BORDER
            ws.cell(r, 3, clean_cell(d.get("function"))).border = BORDER
            ci = ws.cell(r, 4, _join(d.get("inputs"))); ci.alignment = WRAP_TOP; ci.border = BORDER
            co = ws.cell(r, 5, _join(d.get("outputs"))); co.alignment = WRAP_TOP; co.border = BORDER
            ok = d.get("connected")
            sc = ws.cell(r, 6, f"✓ {d.get('n_in',0)} in / {d.get('n_out',0)} out" if ok
                         else f"✗ unconnected ({d.get('n_in',0)} in / {d.get('n_out',0)} out)")
            sc.border = BORDER
            sc.font = Font(color="107C10", bold=True) if ok else Font(color="C00000", bold=True)
            r += 1
        r += 1  # spacer between cabinets
    return r


def _render_tool_io_section(ws: Worksheet, state: dict, run_dir: str, start_row: int) -> Optional[int]:
    """Render the Tool-provenance table onto `ws` starting at `start_row` (header row).
    Returns the next free row, or None when no tools ran. (Was tab_tool_io; refactored
    into a section folded under "⚠ Checks", 2026-06-24.) Every engineering tool that ran,
    with where each output's INPUT came from and where its OUTPUT goes, plus a
    USED/STALE/ORPHANED/TRACED verdict — the 'computed by verified tools' proof."""
    import re as _re
    tu = load_json(os.path.join(run_dir, "4-orchestrator-tools-used.json"))
    if not isinstance(tu, dict) or not tu.get("tools"):
        return None
    q = (state.get("orchestratorContract") or {}).get("quantities") or {}
    qval = {k.lower(): num(v.get("value") if isinstance(v, dict) else v) for k, v in q.items()}
    consumed = {v for v in qval.values() if v is not None}
    for r in (state.get("requirementsBom") or []):
        for f in ("unit_gbp", "line_gbp", "qty"):
            n = num(r.get(f))
            if n is not None:
                consumed.add(round(n, 3))
    for v in (state.get("costStack") or {}).values():
        n = num(v)
        if n is not None:
            consumed.add(round(n, 3))

    def _present(x: float) -> bool:
        for cand in (x, x * 1000.0, x / 1000.0):
            for c in consumed:
                if abs(cand - c) <= max(abs(c) * 0.02, 0.01):
                    return True
        return False

    # RIGOROUS field-level consumption proof (Tristan 2026-06-24: prove the output GOES somewhere,
    # not just that its value coincidentally appears): the set of every token any tool declares as an
    # INPUT. An output field is genuinely CONSUMED if its name (or a distinctive token of it) is some
    # other tool's input — a real edge in the tool DAG, universal across archetypes.
    _input_tokens: set = set()
    for _t in tu["tools"]:
        for _c in (_t.get("claims") or []):
            for _tok in _re.findall(r"[a-z][a-z0-9_]{3,}", str(_c.get("input_summary", "")).lower()):
                _input_tokens.add(_tok)
    _STOP = {"from", "each", "total", "rated", "with", "into", "flow", "load", "duty", "this", "that"}
    def _consumed_by_tool(field: str) -> bool:
        f = str(field).lower().strip()
        if not f:
            return False
        if f in _input_tokens:
            return True
        toks = [t for t in _re.findall(r"[a-z][a-z0-9]{3,}", f) if t not in _STOP]
        return bool(toks) and any(t in _input_tokens for t in toks)

    header(ws, start_row, ["Tool", "Output field", "Value", "Unit", "Input — from",
                           "→ Consumed by", "Status"])
    r = start_row + 1
    body_first = r
    for t in tu["tools"]:
        tid = str(t.get("tool_id", "?"))
        for c in (t.get("claims") or []):
            field = str(c.get("field", ""))
            val = num(c.get("value"))
            of = str(c.get("output_field", "")).lower()
            f = _re.sub(r"^calc_", "", field).lower()
            qk = of if of in qval else (f if f in qval else None)
            if qk is not None and val is not None and qval[qk] is not None:
                qv = qval[qk]
                if abs(val - qv) <= max(abs(qv) * 0.02, 0.01):
                    status, cons = "USED", f"{qk} = {qv:g}"
                else:
                    status, cons = "STALE", f"{qk} = {qv:g}  (≠ tool {val:g})"
            elif _consumed_by_tool(c.get("output_field") or field):
                status, cons = "USED", "feeds another tool's input (DAG edge)"
            elif val is not None and abs(val) in (0.0, 1.0):
                status, cons = "—", "(zero/unit — not checked)"
            elif val is not None and _present(val):
                status, cons = "TRACED", "value used in the design"
            else:
                status, cons = "ORPHANED", "appears nowhere downstream"
            ws.cell(r, 1, tid).border = BORDER
            ws.cell(r, 2, clean_cell(c.get("output_field") or field)).border = BORDER
            ws.cell(r, 3, val).border = BORDER
            ws.cell(r, 4, clean_cell(c.get("unit", ""))).border = BORDER
            e = ws.cell(r, 5, clean_cell(c.get("input_summary", "")))
            e.alignment = WRAP_TOP
            e.font = FONT_NOTE
            e.border = BORDER
            ws.cell(r, 6, clean_cell(cons)).border = BORDER
            sc = ws.cell(r, 7, status)
            sc.border = BORDER
            if status in ("USED", "TRACED"):
                sc.fill, sc.font = FILL_PASS, FONT_PASS
            elif status in ("STALE", "ORPHANED"):
                sc.fill, sc.font = FILL_FAIL, FONT_FAIL
            else:
                sc.fill, sc.font = FILL_ADVISORY, FONT_ADVISORY
            r += 1
    apply_col_formats(ws, body_first, {3: FMT_NUM}, r - 1)
    return r


def _reorder_tabs(wb: Workbook) -> None:
    """Reorder sheets into a READER-NARRATIVE sequence (Tristan 2026-06-23): Story (hero render
    early) → Commercial (economics early) → Engineering → Drawings → Reference/Audit (the plumbing
    last). PURE presentation — Excel formula references are by sheet NAME, so tab order never breaks
    a formula. Universal — keys on tab name/prefix, no archetype logic; unknown tabs take a sensible
    middle rank."""
    _RANK = {
        "Executive Summary": -1, "Contents": 0, "⭐ Scorecard": 0.5, "Overview": 1,
        "Render — Interior layout": 2,                 # HERO render — early (exact name)
        "Brief": 3,
        # Calcs + Quantities are FOUNDATIONAL — they feed the BoM, cost waterfall and financial
        # model, so they come BEFORE them (Tristan: "calcs must come before the BoM and the cost
        # waterfall — they choose the categories and what happens downstream").
        "Quantities": 4, "Calculations": 5,
        "Financial model": 10,
        "Cost waterfall": 14, "Inputs & Assumptions": 15,
        "Bill of Materials (Ledger)": 20, "Sense-check": 21,
        "Line & velocity": 24, "Panel schedule": 25, "Process schedules": 26,
        "Assembly sequence": 29,
        "Risk & Regulatory": 30,
        "⚠ Checks": 90, "⚠ Audit": 91, "Connection trace": 92, "Part names": 93, "Glossary": 94,
    }

    def _rank(title: str) -> int:
        if title in _RANK:
            return _RANK[title]
        t = title.lower()
        if t.startswith("render"):
            return 50                                   # non-hero renders → Drawings
        for i, p in enumerate(("ga", "p&id", "bfd", "single", "hvac")):
            if t.startswith(p):
                return 55 + i
        if t.startswith("isometric"):
            return 70
        if t.startswith("module"):
            return 72
        return 40                                       # unknown data tab: end of engineering

    orig = {id(ws): i for i, ws in enumerate(wb._sheets)}
    wb._sheets.sort(key=lambda ws: (_rank(ws.title), orig[id(ws)]))


def tab_benchmark(wb: Workbook, state: dict) -> None:
    """The INDEPENDENT SENSE-CHECK (Tristan 2026-06-25: a generative LLM gives a top-down expected
    BoM/process/cost; the deterministic engine is diffed against it; each line out by more than the
    band is flagged). The benchmark net (gate 36) already computes this every run — it was just
    never shown. Renders state.benchmarkExpectation (LLM top-down) vs the engine, the per-dimension
    divergence, and the per-line faults the LLM caught in the deterministic bill."""
    be = state.get("benchmarkExpectation") or {}
    bd = state.get("benchmarkDivergence") or {}
    bf = state.get("benchmarkFaults") or []
    if not be and not bd and not bf:
        return
    ws = wb.create_sheet("Sense-check")
    set_widths(ws, {"A": 30, "B": 30, "C": 26, "D": 18, "E": 60})
    title_row(ws, "Independent sense-check — generative LLM expectation vs the deterministic engine", 5,
              "An LLM independently estimates, TOP-DOWN from the market, what a plant like this should "
              "cost, output and be built from. The deterministic engine builds BOTTOM-UP from parts. The "
              "two methods are diffed: a large gap means one is wrong. This is the engine's own sanity "
              "check — shown here so you can see it, not just trust it.")
    r = 4
    worst = str(bd.get("worst", "?")).upper()
    wfill = FILL_FAIL if worst in ("RADICAL", "FAIL") else (FILL_ADVISORY if worst in ("WARN", "REVIEW") else FILL_PASS)
    vc = ws.cell(r, 1, f"VERDICT: {worst}")
    vc.font = FONT_TITLE
    vc.fill = wfill
    ws.cell(r, 2).fill = wfill
    ws.cell(r, 5, clean_cell(bd.get("summary") or "")).font = FONT_NOTE
    r += 2

    # ---- cost: LLM envelope vs engine ----
    ec = be.get("expected_cost") or {}
    if ec:
        sub_banner(ws, r, "All-in cost — LLM top-down envelope vs the engine", 5)
        r += 1
        header(ws, r, ["", "LLM expected (top-down)", "Engine (bottom-up)", "Ratio", "Basis"])
        r += 1
        eng = (num((state.get("costStack") or {}).get("oem_transfer_price_gbp"))
               or num((state.get("costStack") or {}).get("installed_asp_gbp")) or 0)
        exp = num(ec.get("expected_gbp")) or 0
        ratio = (eng / exp) if exp else None
        ws.cell(r, 1, "All-in cost").font = FONT_SUB
        ws.cell(r, 2, f"£{round(num(ec.get('low_gbp')) or 0):,}–£{round(num(ec.get('high_gbp')) or 0):,} (mid £{round(exp):,})")
        ws.cell(r, 3, f"£{round(eng):,}")
        rc = ws.cell(r, 4, f"{ratio:.1f}×" if ratio else "—")
        rc.fill = FILL_FAIL if (ratio and (ratio > 1.5 or ratio < 0.67)) else FILL_PASS
        e = ws.cell(r, 5, clean_cell(ec.get("basis") or ""))
        e.alignment = WRAP_TOP
        e.font = FONT_NOTE
        r += 2

    # ---- per-dimension divergence (expected vs deterministic) ----
    findings = bd.get("findings") or []
    if findings:
        sub_banner(ws, r, f"Per-dimension divergence — {len(findings)} checked (cost · output · sizing · components)", 5)
        r += 1
        header(ws, r, ["Dimension", "LLM expected", "Engine (deterministic)", "Ratio", "Verdict"])
        r += 1
        for f in findings:
            if not isinstance(f, dict):
                continue
            ws.cell(r, 1, clean_cell(f.get("dimension", ""))).font = FONT_SUB
            ws.cell(r, 2, clean_cell(str(f.get("expected", "")))).alignment = WRAP_TOP
            ws.cell(r, 3, clean_cell(str(f.get("deterministic", "")))).alignment = WRAP_TOP
            rr = f.get("ratio")
            ws.cell(r, 4, f"{num(rr):.1f}×" if num(rr) else "—")
            verd = str(f.get("verdict", ""))
            vcell = ws.cell(r, 5, verd)
            vcell.fill = FILL_FAIL if verd in ("radical", "fail") else (FILL_ADVISORY if verd in ("warn", "review") else FILL_PASS)
            r += 1
        r += 1

    # ---- per-LINE faults: the sense-check flags (the LLM caught these in the bill) ----
    if bf:
        sub_banner(ws, r, f"Per-line-item faults — {len(bf)} lines the LLM flags as wrong in the engine's bill", 5)
        r += 1
        header(ws, r, ["Bill line", "What's wrong", "Magnitude", "LLM-suggested", "Likely cause"])
        r += 1
        for f in bf:
            if not isinstance(f, dict):
                continue
            ws.cell(r, 1, clean_cell(f.get("line", ""))).font = FONT_SUB
            iss = ws.cell(r, 2, clean_cell(f.get("issue", "")))
            iss.alignment = WRAP_TOP
            mc = ws.cell(r, 3, clean_cell(f.get("magnitude", "")))
            mc.fill = FILL_FAIL
            ws.cell(r, 4, clean_cell(f.get("suggested", "")))
            lc = ws.cell(r, 5, clean_cell(f.get("likely_cause", "")))
            lc.alignment = WRAP_TOP
            lc.font = FONT_NOTE
            r += 1
        r += 1

    # ---- LLM-expected bill structure (top-down % of cost) ----
    ebom = be.get("expected_bom") or []
    if ebom:
        sub_banner(ws, r, "LLM-expected bill structure (top-down — what the cost SHOULD break into)", 5)
        r += 1
        header(ws, r, ["Expected line / category", "% of cost", "Note", "", ""])
        r += 1
        for b in ebom:
            if not isinstance(b, dict):
                continue
            ws.cell(r, 1, clean_cell(b.get("item", ""))).font = FONT_SUB
            ws.cell(r, 2, f"{num(b.get('typical_pct_of_cost')) or 0:g}%")
            nt = ws.cell(r, 3, clean_cell(b.get("note", "")))
            nt.alignment = WRAP_TOP
            nt.font = FONT_NOTE
            r += 1
        r += 1
    reasoning = be.get("reasoning")
    if reasoning:
        sub_banner(ws, r, "LLM reasoning (how the top-down expectation was formed)", 5)
        r += 1
        rc2 = ws.cell(r, 1, clean_cell(reasoning))
        rc2.alignment = WRAP_TOP
        rc2.font = FONT_NOTE
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=5)
        except Exception:  # noqa: BLE001
            pass


def tab_scorecard(wb: Workbook, state: dict) -> None:
    """The ≥8-every-section self-audit, surfaced for the reader (Tristan 2026-06-25: the dossier
    must SHOW its own quality, not bury it). Renders state.selfAudit — the MIN section score vs the
    ≥8 floor, the mean, and every section's score + its top defect — so the reader sees exactly
    where the design is solid and where it isn't. This is the headline quality measure."""
    sa = state.get("selfAudit") or {}
    secs = sa.get("sections") or []
    if not isinstance(secs, list) or not secs:
        return
    ws = wb.create_sheet("⭐ Scorecard")
    set_widths(ws, {"A": 28, "B": 9, "C": 16, "D": 82})
    title_row(ws, "Quality scorecard — every section against the ≥8 floor", 4,
              "The engine's own self-audit. The AIM is ≥8 on EVERY section — the floor, not the "
              "average. A section below 8 is flagged; a BLOCKING defect means the dossier is not "
              "yet shippable. This is the headline quality measure for the whole dossier.")
    r = 4
    mn = sa.get("min_score")
    mean = sa.get("mean_score")
    n_ok = sum(1 for s in secs if isinstance(s.get("score"), (int, float)) and s.get("score") >= 8)
    min_fill = FILL_PASS if (isinstance(mn, (int, float)) and mn >= 8) else FILL_FAIL
    mc = ws.cell(r, 1, f"MIN section score  {mn}/10")
    mc.font = FONT_TITLE
    mc.fill = min_fill
    ws.cell(r, 2).fill = min_fill
    ws.cell(r, 3, f"{n_ok}/{len(secs)} sections ≥8").font = FONT_SUB
    ws.cell(r, 4, f"mean {mean}/10   ·   AIM: every section ≥8").font = FONT_NOTE
    r += 2
    header(ws, r, ["Section", "Score", "vs ≥8 floor", "Top defect (why it's below 8)"])
    r += 1
    for s in secs:
        if not isinstance(s, dict):
            continue
        score = s.get("score")
        ok = isinstance(score, (int, float)) and score >= 8
        ws.cell(r, 1, clean_cell(s.get("name", ""))).font = FONT_SUB
        scell = ws.cell(r, 2, score)
        scell.fill = FILL_PASS if ok else (FILL_FAIL if s.get("blocking") else FILL_ADVISORY)
        ws.cell(r, 3, "PASS ✓" if ok else ("⛔ BLOCKING" if s.get("blocking") else "below 8"))
        defect = s.get("defects") or [""]
        dcell = ws.cell(r, 4, clean_cell(defect[0] if defect else ""))
        dcell.alignment = WRAP_TOP
        dcell.font = FONT_NOTE
        r += 1

    # ---- engine quality metrics (#92): the dossier carries its own confidence rating ----
    r += 1
    sub_banner(ws, r, "Engine quality metrics — the dossier's own confidence rating", 4)
    r += 1
    try:
        from provenance import audit_provenance
        _trace = round((audit_provenance(state).scorecard().get("traceable_fraction") or 0) * 100)
    except Exception:  # noqa: BLE001
        _trace = None
    # calc-coverage: % of DERIVED numbers whose formula is shown (mirrors check_calc_coverage)
    _q = ((state.get("orchestratorContract") or {}).get("quantities") or {})
    _roots = {"brief", "physics_constant", "constant", "standard", "anchor", "datasheet", "spec"}
    _wc = (state.get("worked_calculations")
           or ((state.get("orchestratorContract") or {}).get("worked_calculations")) or {})
    _worked = set()
    if isinstance(_wc, dict):
        for _cs in _wc.values():
            for _c in (_cs or []):
                if isinstance(_c, dict):
                    _f = _c.get("output_field") or _c.get("field") or _c.get("label")
                    if _f:
                        _worked.add(str(_f).lower())
    _shown = _tot = 0
    for _k, _v in _q.items():
        if not isinstance(_v, dict) or str(_v.get("source", "")).lower() in _roots:
            continue
        _tot += 1
        _sd = str(_v.get("source_detail") or "")
        if _k.lower() in _worked or (len(_sd) > 3 and any(o in _sd for o in ("=", "×", "*", "/", "+"))):
            _shown += 1
    _cov = round(_shown / _tot * 100) if _tot else 100
    for _label, _val, _aim in [
        ("Sections at ≥8 (AIM: all)", f"{n_ok} / {len(secs)}", n_ok == len(secs)),
        ("Min section score", f"{mn}/10", isinstance(mn, (int, float)) and mn >= 8),
        ("Traceability — every number → the brief", f"{_trace}%" if _trace is not None else "—", (_trace or 0) >= 100),
        ("Calc-coverage — every number shows its formula", f"{_cov}%", _cov >= 100),
    ]:
        ws.cell(r, 1, _label).font = FONT_SUB
        vcl = ws.cell(r, 2, _val)
        vcl.fill = FILL_PASS if _aim else FILL_ADVISORY
        ws.cell(r, 4, "AIM: 100%" if "—" not in str(_val) else "").font = FONT_NOTE
        r += 1

    # ---- PER-TAB deterministic scorecard (Tristan 2026-06-26): every tab vs the ≥8 floor ----
    _tabsc = state.get("tabScorecard") or {}
    if isinstance(_tabsc, dict) and _tabsc:
        r += 1
        _su = state.get("tabScorecardSummary") or {}
        sub_banner(ws, r, f"Per-tab quality — every tab against the ≥8 floor "
                          f"(worst: {_su.get('min_tab', '?')} {_su.get('min_score', '?')}/10; "
                          f"{len(_su.get('fail_tabs') or [])} FAIL, {len(_su.get('unscored_tabs') or [])} UNSCORED)", 4)
        r += 1
        header(ws, r, ["Tab", "Score", "vs ≥8 floor", "Top issue / coverage gap (fix at source)"])
        r += 1
        for _tab, _v in _tabsc.items():
            if not isinstance(_v, dict):
                continue
            _st = _v.get("status")
            _scv = _v.get("score")
            ws.cell(r, 1, clean_cell(_tab)).font = FONT_SUB
            _scell = ws.cell(r, 2, _scv if _scv is not None else "—")
            _scell.fill = FILL_PASS if _st == "PASS" else (FILL_ADVISORY if _st == "UNSCORED" else FILL_FAIL)
            ws.cell(r, 3, "PASS ✓" if _st == "PASS" else ("UNSCORED" if _st == "UNSCORED" else "⛔ below 8"))
            _iss = _v.get("issues") or [""]
            _icell = ws.cell(r, 4, clean_cell(_iss[0] if _iss else ""))
            _icell.alignment = WRAP_TOP
            _icell.font = FONT_NOTE
            r += 1


def build(run_dir: str, out_path: str) -> dict:
    state = load_json(os.path.join(run_dir, "state.json"))
    if state is None:
        raise SystemExit(f"No state.json in {run_dir}")
    sha = git_short_sha()
    global _RUN_DIR
    _RUN_DIR = run_dir            # so _tab_quality_banner can score drawing/render/meta sheets

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    # ---- DETERMINISTIC SHIP-GATE AUDIT — run FIRST so the Exec Summary validation card can read
    # the verdict (state["_dossierAudit"]) and the "⚠ Audit" tab can render the findings. `rows` are
    # the assembled BoM (state.requirementsBom); the audit guards every key access itself. ----
    rows = state.get("requirementsBom") or []
    # ---- SELF-CORRECTING REPAIR LOOP (Tristan 2026-06-25: "the engine should look at things,
    # see that it doesn't work, and FIX it — find all the problems and fix them internally,
    # without producing a report until it's fixed"). Don't merely flag defects: run the
    # deterministic audit→FIX→re-audit loop so the auto-fixable ones (untagged principals,
    # duplicate parts, £0 lines a sibling can price) are CORRECTED before any tab is rendered.
    # Every tab is then built from the REPAIRED bill, and the verdict reflects ONLY the genuine
    # remaining gaps (human-input / source-rule), surfaced as questions — never a silent
    # auto-fixable defect. The source-rule fixes (sizing in the contract, the tag rule in the
    # emitter) are the deeper track; this closes the loop for the deliverable surface. ----
    _repair = repair_dossier(state, rows, run_dir)
    state, rows = _repair.state, _repair.rows
    state["requirementsBom"] = rows                       # downstream tabs read the repaired bill
    report = audit_dossier(state, rows, run_dir)          # audit of the REPAIRED dossier
    state["_dossierAudit"] = report.scorecard()
    # DETERMINISTIC PER-TAB SCORECARD (Tristan 2026-06-26): score EVERY tab against the ≥8 floor and
    # stamp the score ON each tab (title_row reads _TAB_SCORES). The minimum scored tab is the
    # dossier's per-tab floor; UNSCORED tabs are coverage gaps to close. Written to state so the
    # chain can gate + route the loop (a tab <8 → its source_rule fix → re-run → tab ≥8).
    global _TAB_SCORES
    _TAB_SCORES = tab_scores(state, rows, run_dir)
    state["tabScorecard"] = _TAB_SCORES
    _ts_summary = tab_scorecard_summary(_TAB_SCORES)
    state["tabScorecardSummary"] = _ts_summary
    print(f"  · per-tab scorecard: min {_ts_summary['min_tab']}={_ts_summary['min_score']}/10 · "
          f"{len(_ts_summary['fail_tabs'])} FAIL, {len(_ts_summary['unscored_tabs'])} UNSCORED")
    # ── PER-TAB ≥8 FLOOR IS THE SHIP GATE (Tristan 2026-06-26: "an 8/10 on every single tab
    #    as a MINIMUM must be part of the code"). The dossier is NOT validated unless EVERY tab
    #    scores a genuine ≥8 — a FAIL tab (<8) OR an UNSCORED tab (no deterministic check exists
    #    to certify it, so it cannot be claimed ≥8) BOTH block. This folds into ship_ok so the
    #    existing SHIP GATE / "⚠ Audit" tab / DRAFT banner all reflect it; the chain's hard-gate
    #    (exit 37) reads tabScorecardSummary.all_pass. UNSCORED counts as a fail by design — a
    #    genuine 8 means a check looked and passed, never "nothing looked". ──
    _aud_sc = state.get("_dossierAudit") or {}
    if not _ts_summary.get("all_pass"):
        _aud_sc["ship_ok"] = False
        if _aud_sc.get("verdict") not in (None, "FAIL"):
            _aud_sc["verdict"] = "FAIL"
        _aud_sc["per_tab_floor"] = {
            "all_pass": False, "min_tab": _ts_summary.get("min_tab"),
            "min_score": _ts_summary.get("min_score"),
            "fail_tabs": _ts_summary.get("fail_tabs"),
            "unscored_tabs": _ts_summary.get("unscored_tabs"),
        }
        state["_dossierAudit"] = _aud_sc
        print(f"  · SHIP GATE: per-tab ≥8 floor NOT met — "
              f"{len(_ts_summary['fail_tabs'])} tab(s) <8, {len(_ts_summary['unscored_tabs'])} UNSCORED "
              f"→ ship_ok=False (every tab must be a genuine ≥8)")
    # Persist the per-tab scorecard + a ROUTED punch-list — the loop signal. Every tab <8 or UNSCORED,
    # its issue + the source rule to fix at. The chain / operator / next run reads these to drive the
    # ≥8 loop: tab <8 → fix the source rule → re-run → tab re-scores. (Fix the SOURCE, not the symptom.)
    try:
        with open(os.path.join(run_dir, "tab-scorecard.json"), "w", encoding="utf-8") as _fh:
            json.dump({"tabs": _TAB_SCORES, "summary": _ts_summary}, _fh, indent=2, default=str)
        _punch = [
            f"- **{_t}** — {(_v.get('score') if _v.get('score') is not None else 'UNSCORED')}/10 "
            f"{_v.get('status')} — {(_v.get('issues') or ['—'])[0]}\n"
            f"  FIX (at source): {_v.get('fix') or 'write a deterministic check for this tab'}"
            for _t, _v in _TAB_SCORES.items() if _v.get("status") in ("FAIL", "UNSCORED")
        ]
        if _punch:
            with open(os.path.join(run_dir, "tab-scorecard-punchlist.md"), "w", encoding="utf-8") as _fh:
                _fh.write(
                    f"# Per-tab quality punch-list — min {_ts_summary['min_tab']} "
                    f"{_ts_summary['min_score']}/10 ({len(_ts_summary['fail_tabs'])} FAIL, "
                    f"{len(_ts_summary['unscored_tabs'])} UNSCORED)\n\n"
                    "Every tab must score ≥8. Fix each at SOURCE (the rule that produced the defect), "
                    "not the symptom; re-run; the tab re-scores. An UNSCORED tab's fix is to write its "
                    "deterministic check.\n\n" + "\n".join(_punch) + "\n")
    except Exception:  # noqa: BLE001 — persistence must never break the build
        pass
    state["_dossierRepair"] = {
        **_repair.summary(),
        "fixes": list(_repair.fixes_applied),
        "needs_input": [
            {"check": f.check, "severity": f.severity, "tab": f.tab,
             "message": f.message, "why": f.source_rule}
            for f in _repair.remaining()
        ],
    }
    if _repair.fixes_applied:
        print(f"  · self-repair: {len(_repair.fixes_applied)} fix(es) applied over "
              f"{_repair.iterations} pass(es); {len(_repair.remaining())} gap(s) need input")

    print("  · Executive Summary")
    tab_executive_summary(wb, state, run_dir, sha)
    print("  · Overview")
    tab_overview(wb, state, run_dir, sha)
    print("  · ⭐ Scorecard")
    tab_scorecard(wb, state)
    print("  · Sense-check")
    tab_benchmark(wb, state)
    print("  · Brief")
    tab_brief(wb, run_dir)
    # Place Brief immediately AFTER Overview, BEFORE ⚠ Checks (created next).
    # openpyxl appends at the end, so move it up to index 1.
    try:
        _bi = wb.sheetnames.index("Brief")
        if _bi != 1:
            wb.move_sheet("Brief", 1 - _bi)
    except ValueError:
        pass
    print("  · ⚠ Checks")
    fail_count = tab_checks(wb, state, run_dir)
    checks_ws = wb["⚠ Checks"]
    fail_labels = getattr(checks_ws, "_forge_fail_labels", [])
    print("  · ⚠ Audit")
    tab_audit(wb, report)
    print("  · Part names")
    tab_parts_master(wb, state, run_dir)
    print("  · Quantities")
    tab_quantities(wb, state)
    print("  · Calculations")
    live_n, static_n = tab_calculations(wb, state, run_dir)
    print("  · Bill of Materials (Ledger)")
    tab_bom(wb, state, run_dir)   # BoM + Cost + Spec sheets merged into one ledger sheet
    # Cost provenance is now a collapsible group on the ledger; report has_cost from state.
    _cb = state.get("costBasis")
    has_cost = bool(isinstance(_cb, dict) and _cb.get("lines"))

    # ---- NEW high-value engineering + schedule tabs (each self-guards: a tab
    # whose source data is absent is SKIPPED cleanly so any class still builds) --
    skipped: List[str] = []

    def add_tab(name: str, fn) -> None:
        """Run a guarded tab builder; record what it did. Never let one tab kill
        the build — a builder exception is caught + logged as a skip."""
        try:
            ok = fn()
        except Exception as exc:  # noqa: BLE001 — robustness: skip, don't crash
            print(f"  ! {name} raised, skipping: {exc}")
            skipped.append(f"{name} (error: {exc})")
            return
        if ok:
            print(f"  · {name}")
        else:
            print(f"  · {name} — skipped (no source data)")
            skipped.append(f"{name} (no source data)")

    add_tab("Connection trace", lambda: tab_connection_trace(wb, state, run_dir))
    add_tab("Cost waterfall", lambda: tab_cost_waterfall(wb, state))
    # ---- ECONOMICS MODEL: Inputs -> Financial model (Economics + Scenarios +
    # Investment Analysis on ONE sheet, 2026-06-24 consolidation). Inputs is built
    # FIRST + kept SEPARATE so _ECON_INPUT_ADDR is populated before the model
    # references it. Each self-guards (skips cleanly with no usable output metric).
    add_tab(INPUTS_SHEET, lambda: tab_inputs_assumptions(wb, state))
    add_tab("Financial model", lambda: tab_financial_model(wb, state))
    add_tab("Panel schedule", lambda: tab_panel_schedule(wb, run_dir))
    # process schedules creates 0..3 sheets; treat >0 as success
    add_tab("Process schedules", lambda: tab_process_schedules(wb, run_dir) > 0)
    add_tab("Line & velocity", lambda: tab_line_velocity(wb, run_dir))
    add_tab("Risk & Regulatory", lambda: tab_risk_regulatory(wb, state))
    add_tab("Assembly sequence", lambda: tab_assembly_sequence(wb, state))
    add_tab("Glossary", lambda: tab_glossary(wb, state))

    print("  · Image tabs")
    used_titles = {t.lower() for t in wb.sheetnames}
    specs = collect_image_specs(run_dir)
    img_ok = 0
    for path, ttl, cap in specs:
        png = ensure_png(path, run_dir)
        if png and add_image_tab(wb, run_dir, png, ttl, cap, used_titles):
            img_ok += 1
            print(f"      + {ttl}")

    # ---- reorder into the reader-narrative sequence BEFORE Contents is built, so the Contents
    # index lists the tabs in the new order (Story → Commercial → Engineering → Drawings → Audit) ----
    _reorder_tabs(wb)

    # ---- CONTENTS (#26): built LAST so the full ordered tab list is known,
    # then moved to sheet #1 with a one-line description + hyperlink per tab ----
    print("  · Contents (sheet #1)")
    tab_contents(wb, _TAB_DESCRIPTIONS)

    # The Executive Summary is the COVER — force it to the very front, ahead of Contents
    # (tab_contents moves itself to index 0; the wow cover must precede it). Tristan 2026-06-24.
    try:
        _ei = wb.sheetnames.index("Executive Summary")
        if _ei != 0:
            wb.move_sheet("Executive Summary", -_ei)
    except ValueError:
        pass

    # ---- FINAL SANITISATION (Tristan 2026-06-23): defang any cell openpyxl turned into a FORMULA
    # that is actually engine PROSE. openpyxl stores ANY string starting with "=" as a formula;
    # engine notes that start with "=" (e.g. a worked-calc assumption "= 3.91 t/day (g/s x 86.4 /
    # 1000)") then become invalid <f> records → Excel strips them ("Removed Records: Formula" and
    # the file opens broken). Real build-generated formulas NEVER start with "= " (space after =) and
    # never contain a digit directly followed by whitespace+letter — so this net is safe for genuine
    # formulas (incl. "=...*8000/1e9", "='Inputs & Assumptions'!$B$5", "=IF(...)"). ----
    # Detector: a string openpyxl will store as a formula that is NOT a valid Excel formula —
    # engine PROSE ("= 3.91 t/day …"), chemical formulas, OR a live worked-calc formula left with an
    # UNBOUND symbol (a bare letter like "E" in "=B187/(2*B189*E-…)"). Strip every VALID formula
    # token; if ANY letter remains it cannot be a real formula → defang. Safe for genuine formulas
    # (=…*8000/1e9, =IF(…), ='Inputs & Assumptions'!$B$5, =CO2 cell-ref) → they strip to nothing.
    def _is_invalid_formula(_s: str) -> bool:
        _t = re.sub(r'"[^"]*"', "", _s)                     # string literals ("PASS"/"FAIL"/…) — letters here are VALID
        _t = re.sub(r"'[^']*'!", "", _t)                    # sheet refs
        _t = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "", _t)         # cell refs (incl CO2-style)
        _t = re.sub(r"[A-Za-z][A-Za-z0-9.]*\s*\(", "(", _t) # function calls -> (
        _t = re.sub(r"\bin_[a-z0-9_]+\b", "", _t)           # known defined names
        _t = re.sub(r"\d+\.?\d*([eE][+\-]?\d+)?", "", _t)    # numbers incl scientific (1e9)
        _t = re.sub(r"[\s+\-*/^%(),:.=&<>​]", "", _t)  # operators / whitespace / zero-width
        return bool(re.search(r"[A-Za-z]", _t))
    # Excel's FILE LOADER rejects lowercase scientific-notation number literals (e.g. "1e9", "1.5e6")
    # in stored formulas — it silently drops the formula ("Removed Records: Formula") even though the
    # UI accepts 1e9 typed live. (Confirmed via LibreOffice round-trip: it rewrites 1e9 -> 1000000000.)
    # So expand any sci-notation literal in a SURVIVING formula to a plain number. Lookbehind/ahead
    # keep it from touching identifiers/cell-refs (e.g. "B1e9" or a name).
    _sci = re.compile(r"(?<![A-Za-z0-9_])(\d+\.?\d*[eE][+\-]?\d+)(?![A-Za-z0-9_])")
    def _expand_sci(_m: "re.Match") -> str:
        _val = float(_m.group(1))
        return str(int(_val)) if _val == int(_val) else repr(_val)
    _defanged = _normalised = 0
    for _ws in wb.worksheets:
        for _row in _ws.iter_rows():
            for _c in _row:
                _v = _c.value
                if not (isinstance(_v, str) and _v.startswith("=")):
                    continue
                if _is_invalid_formula(_v):
                    _c.value = clean_cell(_v)        # zero-width-space prefix → stored as TEXT, not a formula
                    _defanged += 1
                elif _sci.search(_v):
                    _c.value = _sci.sub(_expand_sci, _v)   # 1e9 -> 1000000000 (Excel-loader-safe)
                    _normalised += 1
    if _defanged or _normalised:
        print(f"  · Excel-corruption guard: defanged {_defanged} prose-as-formula + normalised {_normalised} sci-notation formula(s)")

    wb.save(out_path)

    size_mb = os.path.getsize(out_path) / (1024 * 1024)
    return {
        "out_path": out_path,
        "size_mb": size_mb,
        "tabs": wb.sheetnames,
        "live_calcs": live_n,
        "static_calcs": static_n,
        "fail_count": fail_count,
        "fail_labels": fail_labels,
        "image_tabs": img_ok,
        "has_cost": has_cost,
        "skipped_tabs": skipped,
        # The floor-aware audit (ship_ok already folds in the per-tab ≥8 gate above).
        "audit": state.get("_dossierAudit") or report.scorecard(),
        "tab_floor": _ts_summary,
        "sha": sha,
    }


def _selftest() -> int:
    """Pure guards for the compliance MATCHER + direction + class display — the false-PASS class of
    bug (2026-06-25). Exits non-zero on any failure; wired into verify-engine-guards.sh."""
    bad = 0
    # (1) _match_quantity must match the ACHIEVED quantity by NAME, NOT the target-closest ECHO.
    qs = {
        "nameplate_capacity_kwh": {"value": 2912, "unit": "kWh"},
        "usable_capacity_kwh_requested": {"value": 5000, "unit": "kWh"},   # brief echo — must be ignored
        "continuous_power_kw": {"value": 1000, "unit": "kW"},
        "peak_power_kw": {"value": 1250, "unit": "kW"},
    }
    m = _match_quantity({"key_metric": "nameplate_capacity_kwh", "value": 5000, "unit": "kWh"}, qs)
    if not m or m[0] != "nameplate_capacity_kwh" or m[1] != 2912:
        print(f"  FAIL matcher grabbed the echo, not the achieved nameplate (got {m})"); bad += 1
    m = _match_quantity({"key_metric": "rated_power_kw", "value": 2500, "unit": "kW"}, qs)
    if not m or m[0] != "continuous_power_kw":   # rated→continuous, not peak
        print(f"  FAIL matcher power → {m} (want continuous_power_kw)"); bad += 1
    # (2) _humanize_class display names
    if _humanize_class("bess") != "Battery Energy Storage System":
        print(f"  FAIL humanize bess (got {_humanize_class('bess')})"); bad += 1
    # (3) X1 (Tristan 2026-06-27): EVERY sheet gets an honest quality banner — a drawing scores from
    # parts-ledger coverage, and an EMPTY drawing is a 0/FAIL, NEVER a fake PASS. Universal (keyed on
    # drawing TYPE, not archetype).
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        with open(os.path.join(_td, "parts-ledger.json"), "w") as _fh:
            json.dump({"coverage_by_drawing": {
                "pid": {"present": 0, "expected": 44, "pct": 0.0},
                "general-arrangement": {"present": 73, "expected": 77, "pct": 94.8},
            }}, _fh)
        _COV_CACHE.clear()
        _pid = _aux_tab_score("P&ID", _td)
        if not _pid or _pid.get("status") != "FAIL" or _pid.get("score") != 0:
            print(f"  FAIL X1 aux-score: an EMPTY P&ID must be 0/FAIL, not a fake pass (got {_pid})"); bad += 1
        _ga = _aux_tab_score("GA — General Arrangement", _td)
        if not _ga or _ga.get("status") != "PASS":
            print(f"  FAIL X1 aux-score: a 95%-covered GA must PASS (got {_ga})"); bad += 1
        _COV_CACHE.clear()
    print("build-excel-export selftest:", "OK" if bad == 0 else f"{bad} FAIL")
    return bad


def main() -> None:
    if "--selftest" in sys.argv[1:]:
        raise SystemExit(1 if _selftest() else 0)
    run_dir = sys.argv[1] if len(sys.argv) > 1 else "out/ras-v26-verify"
    run_dir = os.path.normpath(run_dir)
    if not os.path.isdir(run_dir):
        raise SystemExit(f"Run dir not found: {run_dir}")
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(run_dir, "dossier.xlsx")
    print(f"Building Excel review workbook for {run_dir} -> {out_path}")
    res = build(run_dir, out_path)
    print("\nDONE")
    print(f"  path        : {res['out_path']}")
    print(f"  size        : {res['size_mb']:.2f} MB")
    print(f"  tabs ({len(res['tabs'])})  : {res['tabs']}")
    print(f"  live calcs  : {res['live_calcs']}")
    print(f"  static calcs: {res['static_calcs']}")
    print(f"  image tabs  : {res['image_tabs']}")
    if res.get("skipped_tabs"):
        print(f"  skipped     : {res['skipped_tabs']}")
    print(f"  CHECKS FAIL : {res['fail_count']}")
    for fl in res["fail_labels"]:
        print(f"      FAIL -> {fl}")
    _aud = res.get("audit") or {}
    if _aud:
        print(f"  SHIP GATE   : {_aud.get('verdict')}  ·  {_aud.get('high', 0)} HIGH "
              f"· {_aud.get('med', 0)} MED · {_aud.get('low', 0)} LOW "
              f"· ship_ok={_aud.get('ship_ok')}")
    # PER-TAB ≥8 FLOOR — the codified ship requirement (Tristan 2026-06-26): the dossier is
    # NOT validated until EVERY tab is a genuine ≥8 (a <8 tab OR an UNSCORED tab blocks). The
    # workbook is still written (so it is inspectable + carries the DRAFT banner), but the
    # process EXITS NON-ZERO so the chain's excel_deliverable step + CI register the gate.
    _tf = res.get("tab_floor") or {}
    if not _tf.get("all_pass", True):
        print(f"  PER-TAB ≥8 GATE: FAIL — min {_tf.get('min_tab')}={_tf.get('min_score')}/10; "
              f"<8: {_tf.get('fail_tabs')}; UNSCORED: {_tf.get('unscored_tabs')}")
        print("  → dossier is a DRAFT (not validated): every tab must reach a genuine ≥8.")
        raise SystemExit(2)


if __name__ == "__main__":
    main()
