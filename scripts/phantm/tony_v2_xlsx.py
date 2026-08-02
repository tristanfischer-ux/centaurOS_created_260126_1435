"""PHANTM v2 — live Excel workbook that backs the Tony answer PDF, in PDF order.

Sheet order = answer PDF section order (§0…§9 + Summary). Every figure in the
PDF has a cell here; the INDEX sheet maps PDF → sheet!cell so Tony can track
without hunting.

Rules:
  - Amber = editable input; blue = formula; green = FE solver input; red = warn.
  - Derived cells are real formulas (named refs). FE / gap-matrix forces are
    solver inputs, labelled as such — same numbers as the PDF tables.
  - LibreOffice headless verify() recomputes and diffs key cells vs Python.

Run: ~/.venvs/phantm/bin/python tony_v2_xlsx.py
       -> out/PHANTM-TONY-V2-CALC.xlsx
"""

from __future__ import annotations

import datetime
import json
import math
import os
import shutil
import subprocess
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

from tony_v2 import pm_options, supply_ceiling

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "out")
SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E8F1FB")
WARN_FILL = PatternFill("solid", fgColor="FCE4E4")
FE_FILL = PatternFill("solid", fgColor="E2EFDA")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
IDX_FILL = PatternFill("solid", fgColor="DDEBF7")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Sheet titles — keep ≤31 chars; order matches PDF
S_INDEX = "0 INDEX (PDF map)"
S_INPUTS = "1 INPUTS"
S_GEOM = "§0 GEOMETRY"
S_MASS = "§1 MASS + TARGETS"
S_FORCE = "§2 FORCE (coil only)"
S_GAP = "§2b GAP MATRIX"
S_ELEC = "§3 ELECTRICAL"
S_PULSE = "§4 PULSE WIDTH"
S_ENERGY = "§5 ENERGY + RAIL"
S_RES = "§6 RESONANCE"
S_MAG = "§7 MAGNET SCREEN"
S_WIRE = "§8 WIRE LEVERS"
S_SUM = "§9 + SUMMARY"


def header(ws, row, *cells):
    for i, txt in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=txt)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def cols(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def put(ws, row, label, value, unit="", pdf="", note="", kind="calc",
        name=None, fmt=None):
    """label | value | unit | PDF § | note."""
    ws.cell(row=row, column=1, value=label).alignment = Alignment(wrap_text=True)
    c = ws.cell(row=row, column=2, value=value)
    c.fill = {"input": INPUT_FILL, "calc": CALC_FILL,
              "warn": WARN_FILL, "fe": FE_FILL}.get(kind, CALC_FILL)
    c.border = BOX
    c.font = BOLD if kind in ("input", "warn", "fe") else Font()
    if fmt:
        c.number_format = fmt
    ws.cell(row=row, column=3, value=unit)
    ws.cell(row=row, column=4, value=pdf)
    ws.cell(row=row, column=5, value=note).alignment = Alignment(wrap_text=True)
    if name:
        ws.parent.defined_names.add(
            DefinedName(name, attr_text=f"'{ws.title}'!$B${row}"))
    return c


def section_banner(ws, row, text, colour="1F3864"):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color=colour)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    return row + 1


def _interp_i(rows, target_mn):
    i_arr = [r["current_a"] for r in rows]
    f_arr = [r["force_mn"] for r in rows]
    if target_mn > max(f_arr):
        return None
    for a, b, fa, fb in zip(i_arr, i_arr[1:], f_arr, f_arr[1:]):
        if fa <= target_mn <= fb or fb <= target_mn <= fa:
            if fb == fa:
                return a
            return a + (target_mn - fa) * (b - a) / (fb - fa)
    return None


def build():
    num = json.load(open(os.path.join(OUT, "tony-v2-numbers.json")))
    fe = json.load(open(os.path.join(OUT, "tony-v2-fe.json")))
    mx = json.load(open(os.path.join(OUT, "tony-v2-matrix.json")))
    wb = Workbook()

    # ================================================================== INPUTS
    ws = wb.active
    ws.title = S_INPUTS
    cols(ws, {"A": 40, "B": 14, "C": 10, "D": 10, "E": 70})
    header(ws, 1, "Input", "Value", "Unit", "PDF §", "Source / note")
    ws["A2"] = ("Amber = edit. All other sheets are formulas (or green FE "
                "inputs). Sheet order matches the answer PDF.")
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:E2")

    r = 4
    r = section_banner(ws, r, "TRANSLATOR (drawing + worksheet)")
    put(ws, r, "Teeth", 25, "-", "§0", "your drawing", "input", "N_teeth"); r += 1
    put(ws, r, "Slots", 24, "-", "§0", "teeth − 1", "input", "N_slots"); r += 1
    put(ws, r, "Tooth width", 125, "µm", "§0",
        "forced by length closure (swap → 7.675 mm miss)",
        "input", "tooth_w"); r += 1
    put(ws, r, "Slot width", 187, "µm", "§0", "your drawing", "input", "slot_w"); r += 1
    put(ws, r, "Slot depth (each face)", 280, "µm", "§0", "worksheet",
        "input", "slot_d"); r += 1
    put(ws, r, "Central core", 280, "µm", "§0", "worksheet", "input", "core"); r += 1
    put(ws, r, "Transverse width L_y", 1200, "µm", "§0", "worksheet",
        "input", "L_y"); r += 1
    put(ws, r, "Density (Fe-Co)", 8.12, "g/cm³", "§1", "worksheet",
        "input", "rho_FeCo"); r += 1
    put(ws, r, "Detent multiple", 30, "× M_t·g", "§1", "F_d = 30·M_t·g",
        "input", "DETENT_G"); r += 1

    r += 1
    r = section_banner(ws, r, "COIL AND DRIVE")
    put(ws, r, "Turns", 70, "-", "§3",
        "your correction (drawing shows 60); packing grid may show 72 slots",
        "input", "N_turns"); r += 1
    put(ws, r, "Wire diameter (bare)", 40, "µm", "§3", "drawing",
        "input", "d_bare"); r += 1
    put(ws, r, "Wire diameter (enamelled)", 48, "µm", "§3",
        "grade-2 class for 40 µm bare", "input", "d_od"); r += 1
    put(ws, r, "Winding window length", 1521, "µm", "§3/§8", "drawing",
        "input", "window_len"); r += 1
    put(ws, r, "Wound limb width", 400, "µm", "§3",
        "least certain input — drives mean turn / R",
        "input", "limb_w"); r += 1
    put(ws, r, "Supply rail", 5.0, "V", "§2/§5", "worksheet",
        "input", "V_rail"); r += 1
    put(ws, r, "Working gap (as drawn)", 60, "µm", "§2", "drawing",
        "input", "gap"); r += 1
    put(ws, r, "Step rate", 10, "/s", "§5/§6", "worksheet",
        "input", "step_rate"); r += 1
    put(ws, r, "Aligned inductance L_al (FE)", 
        max(r0["L_aligned_uh"] for r0 in fe["force_vs_current"]),
        "µH", "§3", "solver input — max L_aligned from force sweep",
        "fe", "L_al_uh", fmt="0.0"); r += 1

    r += 1
    r = section_banner(ws, r, "CONSTANTS")
    put(ws, r, "g", 9.80665, "m/s²", "", "", "input", "g_accel"); r += 1
    put(ws, r, "Copper resistivity @ 20 °C", 1.72e-8, "Ω·m", "§3", "",
        "input", "rho_Cu"); r += 1
    put(ws, r, "Cu temperature coefficient", 0.00393, "/K", "§5", "",
        "input", "alpha_Cu"); r += 1
    put(ws, r, "Cu volumetric heat capacity", 3.45e6, "J/(m³·K)", "§5", "",
        "input", "C_cu"); r += 1

    # ================================================================== §0
    ws0 = wb.create_sheet(S_GEOM)
    cols(ws0, {"A": 42, "B": 14, "C": 10, "D": 10, "E": 72})
    header(ws0, 1, "Quantity", "Value", "Unit", "PDF §", "Formula / check")
    r = 2
    put(ws0, r, "Tooth pitch", "=tooth_w+slot_w", "µm", "§0",
        "tooth + slot", name="pitch"); r += 1
    put(ws0, r, "Tooth duty", "=tooth_w/pitch", "-", "§0",
        "0.401 ≈ duty-sweep optimum 0.40", name="duty", fmt="0.000"); r += 1
    put(ws0, r, "Length L_z (closure)",
        "=N_teeth*tooth_w+N_slots*slot_w", "µm", "§0",
        "must match stated 7612 µm → 7613", name="L_z"); r += 1
    put(ws0, r, "Across-gap L_x (closure)", "=core+2*slot_d", "µm", "§0",
        "must match stated 840 µm", name="L_x"); r += 1
    put(ws0, r, "Gap / tooth (fringing ratio)", "=gap/tooth_w", "-", "§2",
        "0.48 at as-drawn → heavy fringing", fmt="0.00"); r += 1

    # ================================================================== §1
    ws1 = wb.create_sheet(S_MASS)
    cols(ws1, {"A": 42, "B": 14, "C": 10, "D": 10, "E": 72})
    header(ws1, 1, "Quantity", "Value", "Unit", "PDF §", "Formula / note")
    r = 2
    put(ws1, r, "Envelope volume", "=L_x*L_y*L_z/10^9", "mm³", "§1",
        "µm³ ÷ 10⁹", name="V_env", fmt="0.000"); r += 1
    put(ws1, r, "Volume removed by slots",
        "=2*N_slots*slot_w*slot_d*L_y/10^9", "mm³", "§1",
        "2 faces × slots", name="V_slots", fmt="0.000"); r += 1
    put(ws1, r, "Solid volume", "=V_env-V_slots", "mm³", "§1",
        "PDF table", name="V_solid", fmt="0.000"); r += 1
    put(ws1, r, "Solid fraction", "=V_solid/V_env", "-", "§1", "",
        fmt="0.0%"); r += 1
    put(ws1, r, "MASS M_t", "=V_solid/1000*rho_FeCo", "g", "§1",
        "cm³ × density", name="M_t", fmt="0.0000"); r += 1
    put(ws1, r, "Mass", "=M_t*1000", "mg", "§1",
        "PDF bold: 37.8 mg", name="mass_mg", fmt="0.00"); r += 1
    put(ws1, r, "Weight", "=M_t/1000*g_accel*1000", "mN", "§1", "",
        name="W_t", fmt="0.000"); r += 1
    put(ws1, r, "DETENT TARGET F_d",
        "=DETENT_G*M_t/1000*g_accel*1000", "mN", "§1",
        "30 × M_t × g — PDF bold", name="F_d", fmt="0.00"); r += 1
    put(ws1, r, "Stepping force 1.5 × F_d", "=1.5*F_d", "mN", "§1", "",
        name="F_step15", fmt="0.00"); r += 1
    put(ws1, r, "Stepping force 2 × F_d", "=2*F_d", "mN", "§1", "",
        name="F_step20", fmt="0.00"); r += 1
    put(ws1, r, "Step size (3 phases)", "=pitch/3", "µm", "§1",
        "third of a pitch", name="step_um", fmt="0.0"); r += 1

    # ================================================================== §2 FORCE
    ws2 = wb.create_sheet(S_FORCE)
    cols(ws2, {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
               "G": 14, "H": 14, "I": 36})
    ws2["A1"] = ("PDF §2 — Magnetic performance (60 µm, 70 turns). Coil "
                 "reluctance ONLY (Br=0). Green = FE solver; blue = workbook "
                 "formulas from R_c / N_turns. Do NOT add §7 detent by hand.")
    ws2["A1"].font = Font(bold=True, size=11, color="006100")
    ws2.merge_cells("A1:I1")
    ws2["A2"] = ("Disclosures (same as PDF): ~25% DC end-attraction removed "
                 "before peaking; MST vs ½·i²·dL/dx ~28% median — force from "
                 "stress tensor; back iron / wound limb approximate. "
                 "Convergence at 0.40 A: 1.071→1.070→1.067 mN (0.4%).")
    ws2["A2"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells("A2:I2")
    ws2.row_dimensions[2].height = 48

    header(ws2, 4, "I (A)", "F (mN) FE", "A-turns", "V = I·R", "P = I²R",
           "J (A/mm²)", "dL/dx (H/m)", "L_al (µH)", "Note")
    r = 5
    force_start = r
    for row in fe["force_vs_current"]:
        ws2.cell(row=r, column=1, value=row["current_a"]).fill = FE_FILL
        c = ws2.cell(row=r, column=2, value=row["force_mn"])
        c.fill, c.number_format = FE_FILL, "0.00"
        c3 = ws2.cell(row=r, column=3, value=f"=A{r}*N_turns")
        c3.fill, c3.number_format = CALC_FILL, "0.0"
        c4 = ws2.cell(row=r, column=4, value=f"=A{r}*R_c")
        c4.fill, c4.number_format = CALC_FILL, "0.00"
        c5 = ws2.cell(row=r, column=5, value=f"=A{r}^2*R_c")
        c5.fill, c5.number_format = CALC_FILL, "0.00"
        c6 = ws2.cell(row=r, column=6,
                      value=f"=A{r}/(PI()*(d_bare/2)^2)*10^6")
        c6.fill, c6.number_format = CALC_FILL, "0"
        c7 = ws2.cell(row=r, column=7, value=row["peak_dL_dx_h_per_m"])
        c7.fill, c7.number_format = FE_FILL, "0.0000"
        c8 = ws2.cell(row=r, column=8, value=row["L_aligned_uh"])
        c8.fill, c8.number_format = FE_FILL, "0.0"
        note = ""
        if row["current_a"] in (0.30, 0.35, 0.40):
            note = "worksheet (PDF bold)"
        elif row["volts_across_coil"] > 5.0:
            note = "⚠ over 5 V cold"
        elif abs(row["current_a"] - 1.2) < 1e-9:
            note = "best on cold rail ≈"
        ws2.cell(row=r, column=9, value=note)
        r += 1
    force_end = r - 1

    # 0.30, 0.35, 0.40 are first three rows of typical sweep — find 0.40
    r04 = None
    r03 = None
    r12 = None
    for i, row in enumerate(fe["force_vs_current"]):
        if abs(row["current_a"] - 0.40) < 1e-9:
            r04 = force_start + i
        if abs(row["current_a"] - 0.30) < 1e-9:
            r03 = force_start + i
        if abs(row["current_a"] - 1.20) < 1e-9:
            r12 = force_start + i
    assert r04 and r03 and r12
    wb.defined_names.add(DefinedName("F_040", attr_text=f"'{S_FORCE}'!$B${r04}"))
    wb.defined_names.add(DefinedName("F_030", attr_text=f"'{S_FORCE}'!$B${r03}"))
    wb.defined_names.add(DefinedName("F_120", attr_text=f"'{S_FORCE}'!$B${r12}"))
    wb.defined_names.add(
        DefinedName("J_040", attr_text=f"'{S_FORCE}'!$F${r04}"))

    r += 1
    r = section_banner(ws2, r, "PDF §2 — shortfall & “to reach” (60 µm)")
    # Use 5-col put on a wide sheet — put writes A-E; OK
    put(ws2, r, "F at 0.30 A (FE)", f"=B{r03}", "mN", "§2",
        "PDF headline range low", "fe", fmt="0.00"); r += 1
    put(ws2, r, "F at 0.40 A (FE)", f"=B{r04}", "mN", "§2",
        "PDF headline", "fe", fmt="0.00"); r += 1
    put(ws2, r, "F_d / F(0.40 A)", f"=F_d/B{r04}", "×", "§2",
        "force shortfall at worksheet current", fmt="0.0"); r += 1
    put(ws2, r, "F/i² at 0.40 A (linear check)",
        f"=B{r04}/(0.4^2)", "mN/A²", "§2",
        "should be flat in unsaturated region", fmt="0.00"); r += 1
    put(ws2, r, "Best F on cold rail (~1.2 A)", f"=B{r12}", "mN", "§2",
        "still short of F_d at 60 µm", "fe", fmt="0.00"); r += 1
    put(ws2, r, "J at 0.40 A", f"=F{r04}", "A/mm²", "§2",
        "OPEN — not a cleared thermal design", fmt="0"); r += 1
    mst = next(x["mst_vs_energy_median_pct"] for x in fe["force_vs_current"]
               if abs(x["current_a"] - 0.40) < 1e-9)
    put(ws2, r, "MST vs ½i²dL/dx median @0.4 A", mst, "%", "§2",
        "disclosure — force from stress tensor", "fe", fmt="0"); r += 1

    r += 1
    header(ws2, r, "To reach (PDF §2)", "Target (mN)", "I needed (A)",
           "Volts", "Inside 5 V?", "", "", "", "")
    r += 1
    fd_py = num["translator"]["detent_target_mn"]
    f15_py, f20_py = num["translator"]["stepping_force_mn"]
    rows60 = fe["force_vs_current"]
    for lbl, tgt in (("detent F_d", fd_py),
                     ("stepping 1.5×F_d", f15_py),
                     ("stepping 2×F_d", f20_py)):
        i_need = _interp_i(rows60, tgt)
        ws2.cell(row=r, column=1, value=lbl)
        ws2.cell(row=r, column=2, value=round(tgt, 2)).number_format = "0.00"
        if i_need is None:
            ws2.cell(row=r, column=3, value="not reached <2 A").fill = WARN_FILL
            ws2.cell(row=r, column=4, value="—")
            ws2.cell(row=r, column=5, value="—")
        else:
            c = ws2.cell(row=r, column=3, value=round(i_need, 3))
            c.fill, c.number_format = FE_FILL, "0.00"
            c4 = ws2.cell(row=r, column=4, value=f"=C{r}*R_c")
            c4.fill, c4.number_format = CALC_FILL, "0.00"
            c5 = ws2.cell(row=r, column=5,
                          value=f'=IF(D{r}<=V_rail,"yes","NO — over rail")')
            c5.fill = CALC_FILL
        r += 1

    # ================================================================== §2b
    ws2b = wb.create_sheet(S_GAP)
    cols(ws2b, {"A": 12, "B": 14, "C": 14, "D": 14, "E": 16, "F": 18, "G": 42})
    ws2b["A1"] = ("PDF §2b — Gap × current design space. Same coil-only FE "
                  "model as §2. Closing gap restores modulation.")
    ws2b["A1"].font = Font(bold=True, size=11)
    ws2b.merge_cells("A1:G1")
    ws2b["A2"] = ("±10–20 µm etch variation on 125 µm teeth is FIRST-ORDER "
                  "vs 40/30 µm gap (§9). Green = solver.")
    ws2b["A2"].alignment = Alignment(wrap_text=True)
    ws2b.merge_cells("A2:G2")

    header(ws2b, 4, "Gap (µm)", "Mod% @0.4A", "F@0.4A (mN)", "F@1.2A (mN)",
           "I for F_d (A)", "I for 1.5×F_d (A)", "Reading (PDF §2b)")
    r = 5
    r_ohm_py = num["coil"]["resistance_ohm"]
    gap_i_fd_40 = None
    gap_i_15_40 = None
    gap_i_15_30 = None
    gap_i_fd_30 = None
    for gap in mx["gaps_um"]:
        rows = mx["by_gap"][str(gap)]
        by_i = {row["current_a"]: row for row in rows}
        i_fd = _interp_i(rows, fd_py)
        i_15 = _interp_i(rows, f15_py)
        if gap == 40:
            gap_i_fd_40, gap_i_15_40 = i_fd, i_15
        if gap == 30:
            gap_i_fd_30, gap_i_15_30 = i_fd, i_15
        ws2b.cell(row=r, column=1, value=gap).fill = FE_FILL
        ws2b.cell(row=r, column=2,
                  value=by_i[0.4]["modulation_pct"]).fill = FE_FILL
        ws2b.cell(row=r, column=3, value=by_i[0.4]["force_mn"]).fill = FE_FILL
        ws2b.cell(row=r, column=4, value=by_i[1.2]["force_mn"]).fill = FE_FILL
        if i_fd is None:
            ws2b.cell(row=r, column=5, value="unreachable").fill = WARN_FILL
        else:
            c = ws2b.cell(row=r, column=5, value=round(i_fd, 3))
            c.fill, c.number_format = FE_FILL, "0.00"
        if i_15 is None:
            ws2b.cell(row=r, column=6, value="unreachable").fill = WARN_FILL
            read = "1.5× unreachable on this sweep"
        else:
            c = ws2b.cell(row=r, column=6, value=round(i_15, 3))
            c.fill, c.number_format = FE_FILL, "0.00"
            v = i_15 * r_ohm_py
            if v > 5.0:
                read = f"{v:.2f} V — over cold rail"
            elif i_15 > supply_ceiling(r_ohm_py, 5.0, 3.0):
                read = f"{v:.2f} V cold — fails warm"
            else:
                read = f"{v:.2f} V — inside with margin"
        if i_fd is None:
            read = "detent unreachable on 5 V / 40 µm"
        ws2b.cell(row=r, column=7, value=read)
        r += 1

    r += 1
    put(ws2b, r, "Recommended I @40 µm for F_d (FE)",
        round(gap_i_fd_40, 3) if gap_i_fd_40 else "—", "A", "§2b/Sum",
        "PDF recommended path", "fe", "I_40_fd", fmt="0.00"); r += 1
    put(ws2b, r, "I @40 µm for 1.5×F_d (FE)",
        round(gap_i_15_40, 3) if gap_i_15_40 else "—", "A", "§2b",
        "cold-OK / warm-FAIL with 40 µm wire", "fe", "I_40_15",
        fmt="0.00"); r += 1
    put(ws2b, r, "I @30 µm for F_d / 1.5×",
        (f"{gap_i_fd_30:.2f} / {gap_i_15_30:.2f}"
         if gap_i_fd_30 and gap_i_15_30 else "—"),
        "A", "§2b", "comfortable inside 5 V", "fe"); r += 1

    # ================================================================== §3
    ws3 = wb.create_sheet(S_ELEC)
    cols(ws3, {"A": 44, "B": 14, "C": 10, "D": 10, "E": 72})
    header(ws3, 1, "Quantity", "Value", "Unit", "PDF §", "Formula / note")
    r = 2
    put(ws3, r, "Turns per layer (packing)", "=INT(window_len/d_od)", "-", "§3",
        "", name="turns_layer"); r += 1
    put(ws3, r, "Layers (packing)", "=ROUNDUP(N_turns/turns_layer,0)", "-",
        "§3", "grid may exceed N_turns", name="layers"); r += 1
    put(ws3, r, "Winding build", "=layers*d_od", "µm", "§3", "",
        name="build"); r += 1
    put(ws3, r, "Window capacity (packing slots)", "=turns_layer*layers", "-",
        "§3", "only N_turns wound", name="cap_turns"); r += 1
    put(ws3, r, "Mean turn length", "=2*((limb_w+build)+(L_y+build))", "µm",
        "§3", "", name="l_turn"); r += 1
    put(ws3, r, "Wire length", "=N_turns*l_turn/1000", "mm", "§3", "",
        name="wire_len", fmt="0.0"); r += 1
    put(ws3, r, "Wire cross-section", "=PI()*(d_bare/2)^2/10^12", "m²", "§3",
        "", name="A_wire"); r += 1
    put(ws3, r, "RESISTANCE R_c", "=rho_Cu*(wire_len/1000)/A_wire", "Ω", "§3",
        "PDF bold — ρ·ℓ/A", name="R_c", fmt="0.00"); r += 1
    put(ws3, r, "Aligned inductance L_al", "=L_al_uh", "µH", "§3",
        "FE input from sheet 1", fmt="0.0"); r += 1
    put(ws3, r, "Time constant τ = L/R",
        "=(L_al_uh*1e-6)/R_c*1e6", "µs", "§3", "PDF bold",
        name="tau_us", fmt="0.0"); r += 1
    put(ws3, r, "3τ (electrical rise)", "=3*tau_us", "µs", "§4",
        "pulse set by mechanics, not electronics", fmt="0"); r += 1
    put(ws3, r, "Cold rail ceiling", "=V_rail/R_c", "A", "§3",
        "PDF bold", name="I_max_cold", fmt="0.000"); r += 1
    put(ws3, r, "Volts per ampere-turn",
        "=rho_Cu*(l_turn/10^6)/A_wire", "V/At", "§8",
        "V=(N·i)·ρ·l/A — N cancels at fixed l_turn",
        name="V_per_At", fmt="0.0000"); r += 1

    # ================================================================== §4
    ws4 = wb.create_sheet(S_PULSE)
    cols(ws4, {"A": 48, "B": 14, "C": 10, "D": 10, "E": 72})
    header(ws4, 1, "Quantity", "Value", "Unit", "PDF §", "Formula / note")
    ws4["A2"] = ("PDF §4 — Drive pulse width. Ballistic estimate at planning "
                 "force 1.5×F_d (order-of-magnitude; ignores real F(x)).")
    ws4["A2"].font = Font(italic=True)
    ws4.merge_cells("A2:E2")
    r = 3
    put(ws4, r, "Force used for ballistic step", "=F_step15", "mN", "§4",
        "1.5×F_d — same as PDF planning", name="F_ball", fmt="0.00"); r += 1
    put(ws4, r, "Ballistic step time",
        "=SQRT(2*(step_um/1e6)/((F_ball/1000)/(M_t/1000)))*1000", "ms", "§4",
        "t = √(2s/a) from rest", name="t_step_ms", fmt="0.00"); r += 1
    put(ws4, r, "Recommended pulse (1.5× ballistic)",
        "=MAX(1.5*t_step_ms,1)", "ms", "§4",
        "PDF lower bound of recommended range", name="t_pulse_lo",
        fmt="0.0"); r += 1
    put(ws4, r, "Recommended pulse (2.5× ballistic)",
        "=1.6*t_pulse_lo", "ms", "§4",
        "PDF upper bound (~1.5–2.5× ballistic)", name="t_pulse_hi",
        fmt="0.0"); r += 1
    put(ws4, r, "Pulse width used in §5 energy", 1.0, "ms", "§5",
        "AMBER — PDF uses max(1.5×t_step, 1.0) ≈ 1.0 ms at planning",
        "input", "t_pulse", fmt="0.0"); r += 1

    # ================================================================== §5
    ws5 = wb.create_sheet(S_ENERGY)
    cols(ws5, {"A": 48, "B": 14, "C": 10, "D": 10, "E": 72})
    header(ws5, 1, "Quantity", "Value", "Unit", "PDF §", "Why it matters")
    ws5["A2"] = ("PDF §5 — Energy/power + rail warm margin. THIS is where "
                 "cold-OK / warm-FAIL shows up. Edit amber planning current.")
    ws5["A2"].font = Font(italic=True, bold=True)
    ws5.merge_cells("A2:E2")
    r = 3
    put(ws5, r, "Planning current", 1.00, "A", "§5",
        "AMBER. Try 0.95 (40 µm F_d), 1.37 (40 µm 1.5× — cold OK, warm FAIL)",
        "input", "I_plan", fmt="0.00"); r += 1
    put(ws5, r, "Ampere-turns", "=N_turns*I_plan", "At", "§2/§8",
        "70×0.40=28 At — why as-drawn is short", name="At_plan",
        fmt="0.0"); r += 1
    put(ws5, r, "Volts needed (cold)", "=I_plan*R_c", "V", "§5", "",
        name="V_need", fmt="0.000"); r += 1
    put(ws5, r, "Inside 5 V cold?",
        '=IF(V_need<=V_rail,"yes","NO — over the cold rail")', "", "§5",
        ""); r += 1
    put(ws5, r, "Ohmic energy per pulse",
        "=I_plan^2*R_c*t_pulse/1000", "J", "§5",
        "I²R·t — PDF mJ = ×1000", name="E_ohm", fmt="0.000000"); r += 1
    put(ws5, r, "Ohmic energy", "=E_ohm*1000", "mJ", "§5",
        "PDF figure", fmt="0.00"); r += 1
    put(ws5, r, "Magnetic energy ½LI²",
        "=0.5*(L_al_uh*1e-6)*I_plan^2*1000", "mJ", "§5",
        "negligible vs ohmic", name="E_mag_mJ", fmt="0.000"); r += 1
    put(ws5, r, "Energy per step", "=E_ohm*1000+E_mag_mJ", "mJ", "§5",
        "PDF bold", name="E_step_mJ", fmt="0.00"); r += 1
    put(ws5, r, "Average power @ step rate",
        "=E_step_mJ*step_rate", "mW", "§5", "PDF bold", fmt="0.0"); r += 1
    put(ws5, r, "Copper volume", "=(wire_len/1000)*A_wire", "m³", "§5",
        "", name="V_cu"); r += 1
    put(ws5, r, "ΔT per pulse (adiabatic)", "=E_ohm/(C_cu*V_cu)", "K", "§5",
        "does NOT clear insulation / EM / duty", name="dT", fmt="0.0"); r += 1
    put(ws5, r, "R when warm", "=R_c*(1+alpha_Cu*dT)", "Ω", "§5", "",
        name="R_hot", fmt="0.000"); r += 1
    put(ws5, r, "Rail ceiling, warm", "=V_rail/R_hot", "A", "§5",
        "compare to I_plan", "warn", "I_max_hot", fmt="0.000"); r += 1
    put(ws5, r, "Still enough when warm?",
        '=IF(I_max_hot>=I_plan,"yes","NO — fails once warm")', "", "§5",
        "", "warn"); r += 1
    put(ws5, r, "Margin cold vs I_plan",
        "=(I_max_cold-I_plan)/I_plan", "-", "§5", "", fmt="0.0%"); r += 1
    r += 1
    c = ws5.cell(row=r, column=1, value=(
        "TRY 1.37 A (40 µm-gap 1.5× step): ~4.96 V cold (UNDER 5 V) but warm "
        "ceiling drops below ask → cold-OK / warm-FAIL. Thicker wire (§8) or "
        "30 µm gap (§2b) buys margin."))
    c.alignment = Alignment(wrap_text=True)
    c.font = Font(bold=True)
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws5.row_dimensions[r].height = 48

    # ================================================================== §6
    ws6 = wb.create_sheet(S_RES)
    cols(ws6, {"A": 52, "B": 14, "C": 10, "D": 10, "E": 72})
    header(ws6, 1, "Quantity", "Value", "Unit", "PDF §", "Formula / note")
    r = 2
    put(ws6, r, "Worksheet form k = 2π·F_d/(S/2)",
        "=2*PI()*(F_d/1000)/((step_um/2)/10^6)", "N/m", "§6",
        "S = step — inflates k ×6", name="k_tony", fmt="0.0"); r += 1
    put(ws6, r, "Resonance (worksheet form)",
        "=1/(2*PI())*SQRT(k_tony/(M_t/1000))", "Hz", "§6", "",
        fmt="0"); r += 1
    put(ws6, r, "Sinusoidal k = 2π·F_d/pitch",
        "=2*PI()*(F_d/1000)/(pitch/10^6)", "N/m", "§6",
        "F(x)=F_d·sin(2πx/pitch) — USE THIS", name="k_sin", fmt="0.0"); r += 1
    put(ws6, r, "Resonance (sinusoidal)",
        "=1/(2*PI())*SQRT(k_sin/(M_t/1000))", "Hz", "§6",
        "PDF bold ≈387 Hz", name="f_res", fmt="0"); r += 1
    put(ws6, r, "Ratio (worksheet / sinusoidal)",
        "=SQRT(k_tony/k_sin)", "-", "§6", "√6 ≈ 2.45", fmt="0.00"); r += 1
    put(ws6, r, "Margin over step rate", "=f_res/step_rate", "×", "§6",
        "conclusion unchanged", fmt="0"); r += 1

    # ================================================================== §7
    # At for screen: 70 × I_40_15 (or planning) — PDF uses ~at_needed from 40µm path
    at_needed = 70.0 * (gap_i_15_40 if gap_i_15_40 else 1.0)
    ws7 = wb.create_sheet(S_MAG)
    cols(ws7, {"A": 28, "B": 12, "C": 16, "D": 14, "E": 14, "F": 10, "G": 48})
    ws7["A1"] = ("PDF §7 — Magnet SCREEN only (H_c×t). Not a load-line solve. "
                 f"Required At ≈ {at_needed:.0f} from 40 µm-gap 1.5× path "
                 "(N×I). Ratio column uses At_req below.")
    ws7["A1"].font = Font(bold=True, color="9C0006")
    ws7.merge_cells("A1:G1")
    put(ws7, 3, "Required At (screen)", at_needed, "At", "§7",
        "70 × I(1.5×F_d @40 µm) — edit if path changes",
        "input", "At_req", fmt="0.0")
    header(ws7, 5, "Material", "H_c (kA/m)", "Thinnest (mm)", "MMF (At)",
           "vs At_req", "PDF §", "Note")
    r = 6
    for row in pm_options(at_needed):
        ws7.cell(row=r, column=1, value=row["material"])
        ws7.cell(row=r, column=2, value=row["hc_ka_m"]).fill = INPUT_FILL
        ws7.cell(row=r, column=3,
                 value=row["supplier_min_thickness_mm"]).fill = INPUT_FILL
        c = ws7.cell(row=r, column=4, value=f"=B{r}*C{r}")
        c.fill, c.number_format = CALC_FILL, "0"
        c2 = ws7.cell(row=r, column=5, value=f"=D{r}/At_req")
        c2.fill, c2.number_format = CALC_FILL, "0.00"
        ws7.cell(row=r, column=6, value="§7")
        ws7.cell(row=r, column=7, value=row.get("verdict", row.get("note", "")))
        r += 1

    # ================================================================== §8
    ws8 = wb.create_sheet(S_WIRE)
    cols(ws8, {"A": 18, "B": 14, "C": 16, "D": 18, "E": 14, "F": 48})
    ws8["A1"] = ("PDF §8 — Ampere-turn / wire levers. V = (N·i)·ρ·l/A; "
                 "at fixed mean-turn, N cancels. Thicker wire lowers volts.")
    ws8["A1"].font = Font(bold=True, size=11)
    ws8.merge_cells("A1:F1")
    put(ws8, 3, "At for wire table", "=At_req", "At", "§8",
        "same as §7 screen At", "calc")
    header(ws8, 5, "Bare Ø (µm)", "V per At", "V for At_req",
           "Turns fit (3 layers)", "I needed", "Comment")
    r = 6
    for d in (40, 50, 63):
        ws8.cell(row=r, column=1, value=d).fill = INPUT_FILL
        c = ws8.cell(
            row=r, column=2,
            value=f"=rho_Cu*(l_turn/10^6)/(PI()*({d}/2)^2/10^12)")
        c.fill, c.number_format = CALC_FILL, "0.0000"
        c3 = ws8.cell(row=r, column=3, value=f"=B{r}*At_req")
        c3.fill, c3.number_format = CALC_FILL, "0.00"
        c4 = ws8.cell(row=r, column=4, value=f"=INT(window_len/({d}*1.2))*3")
        c4.fill = CALC_FILL
        c5 = ws8.cell(row=r, column=5, value=f"=At_req/D{r}")
        c5.fill, c5.number_format = CALC_FILL, "0.00"
        ws8.cell(row=r, column=6, value=(
            "present choice — near rail → warm FAIL at 1.5× path"
            if d == 40 else
            "fewer turns, higher I, lower V — the point of thicker wire"))
        r += 1
    r += 1
    ws8.cell(row=r, column=1, value="Levers (PDF order)").font = BOLD
    r += 1
    for line in (
        "1. Close gap (§2b) — 60→40 µm makes detent reachable; 30 µm puts 1.5× inside rail",
        "2. Thicker wire — mandatory near rail (warm margin); keeps NI when turns fall",
        "3. Deepen pole slots — more modulation; tooling geometry only",
        "4. More window — 1521 µm holds ~93 turns of 40 µm; room for copper",
    ):
        ws8.cell(row=r, column=1, value=line)
        ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    # ================================================================== §9 + Summary
    ws9 = wb.create_sheet(S_SUM)
    cols(ws9, {"A": 36, "B": 28, "C": 12, "D": 10, "E": 48})
    header(ws9, 1, "PDF Summary row", "Value / formula", "Unit", "PDF §",
           "Note")
    r = 2
    put(ws9, r, "Translator mass", "=mass_mg", "mg", "Sum", "", fmt="0.0"); r += 1
    put(ws9, r, "As-drawn gap", "=gap", "µm", "Sum",
        "detent unreachable on 5 V / 40 µm"); r += 1
    put(ws9, r, "Force @ 0.40 A / 60 µm", "=F_040", "mN", "Sum", "",
        fmt="0.00"); r += 1
    put(ws9, r, "Target F_d", "=F_d", "mN", "Sum", "", fmt="0.00"); r += 1
    put(ws9, r, "I for F_d @ 40 µm", "=I_40_fd", "A", "Sum",
        "recommended path", fmt="0.00"); r += 1
    put(ws9, r, "Pulse width (planning)",
        '=TEXT(t_pulse_lo,"0.0")&" – "&TEXT(t_pulse_hi,"0.0")', "ms", "Sum",
        ""); r += 1
    put(ws9, r, "Resonant frequency", "=f_res", "Hz", "Sum", "", fmt="0"); r += 1
    r += 1
    ws9.cell(row=r, column=1, value="§9 Manufacturability (qualitative — see PDF)").font = BOLD
    r += 1
    for line in (
        "Translator: YES — photochemical etch + stack Fe-Co (near process edge; ±10–20 µm vs gap)",
        "Precision microstamp: viable at volume; die roll couples to gap choice",
        "Micro-MIM: high risk — prove before assuming",
        "Pole-piece CORE: same as translator — etchable/stampable",
        "WINDING: the real question — prefer pre-wound bonded coil + split core",
        "Planar/deposited coils: unattractive at 70 turns + ampere pulses",
    ):
        ws9.cell(row=r, column=1, value=line)
        ws9.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    # ================================================================== INDEX last-built but moved to front
    # Collect map rows for INDEX
    index_rows = [
        ("Intro / correction", "Force reinstated; MST~28%; DC~25%",
         S_FORCE, "A1:A2", "narrative"),
        ("Headline F @0.30–0.40 A", "FE table", S_FORCE, f"B{r03}:B{r04}", "FE"),
        ("Headline F_d", "30·M_t·g", S_MASS, "B9", "formula"),
        ("§0 L_z closure", "25×125+24×187", S_GEOM, "B3", "formula"),
        ("§0 L_x closure", "core+2·slot_d", S_GEOM, "B4", "formula"),
        ("§0 duty", "tooth/pitch", S_GEOM, "B2", "formula"),
        ("§1 solid volume", "envelope − slots", S_MASS, "B3", "formula"),
        ("§1 mass M_t", "mg", S_MASS, "B6", "formula"),
        ("§1 F_d / 1.5× / 2×", "targets", S_MASS, "B8:B10", "formula"),
        ("§1 step size", "pitch/3", S_MASS, "B11", "formula"),
        ("§2 force table", "I, F, V, P, J, dL/dx", S_FORCE, f"A5:H{force_end}", "FE+formula"),
        ("§2 shortfall F_d/F", "×", S_FORCE, "shortfall block", "formula"),
        ("§2 to-reach table", "I for F_d / 1.5 / 2", S_FORCE, "to-reach block", "FE interp"),
        ("§2b gap matrix", "gap × F × I", S_GAP, "A5:G…", "FE"),
        ("§2b 40/30 µm paths", "recommended I", S_GAP, "named I_40_*", "FE"),
        ("§3 R_c, L_al, τ", "electrical", S_ELEC, "B8:B11", "formula+FE"),
        ("§3 cold ceiling", "V/R", S_ELEC, "B12", "formula"),
        ("§4 ballistic / pulse", "ms", S_PULSE, "B3:B6", "formula"),
        ("§5 E_ohm, E_mag, E_step, P_avg", "mJ / mW", S_ENERGY, "B6:B10", "formula"),
        ("§5 ΔT / warm ceiling", "cold-OK warm-FAIL", S_ENERGY, "B12:B15", "formula"),
        ("§6 k and f_res", "√6 correction", S_RES, "B2:B6", "formula"),
        ("§7 magnet screen", "H_c×t table", S_MAG, "A6:…", "formula"),
        ("§8 V/At wire table", "40/50/63 µm", S_WIRE, "A6:E8", "formula"),
        ("§8 levers list", "gap → wire → slots → window", S_WIRE, "text", "—"),
        ("§9 + Summary", "one-page figures", S_SUM, "all", "formula"),
        ("All amber inputs", "single place", S_INPUTS, "all", "input"),
    ]

    wsi = wb.create_sheet(S_INDEX, 0)
    cols(wsi, {"A": 36, "B": 32, "C": 26, "D": 18, "E": 14})
    wsi["A1"] = ("PHANTM v2 CALC — backs answer PDF V7.2, in PDF section order")
    wsi["A1"].font = Font(bold=True, size=14)
    wsi.merge_cells("A1:E1")
    wsi["A2"] = ("Open the PDF and this workbook side by side. Each PDF figure "
                 "is listed below with its sheet and cells. Amber=edit, "
                 "blue=formula, green=FE solver, red=warning.")
    wsi["A2"].alignment = Alignment(wrap_text=True)
    wsi.merge_cells("A2:E2")
    wsi.row_dimensions[2].height = 36
    header(wsi, 4, "PDF location", "Figure / claim", "Excel sheet",
           "Cells", "Kind")
    for i, (pdf, fig, sheet, cells, kind) in enumerate(index_rows, start=5):
        wsi.cell(row=i, column=1, value=pdf).fill = IDX_FILL
        wsi.cell(row=i, column=2, value=fig)
        wsi.cell(row=i, column=3, value=sheet)
        wsi.cell(row=i, column=4, value=cells)
        wsi.cell(row=i, column=5, value=kind)

    r = 5 + len(index_rows) + 1
    wsi.cell(row=r, column=1, value="SHEET ORDER (= PDF ORDER)").font = BOLD
    r += 1
    for name in (S_INDEX, S_INPUTS, S_GEOM, S_MASS, S_FORCE, S_GAP, S_ELEC,
                 S_PULSE, S_ENERGY, S_RES, S_MAG, S_WIRE, S_SUM):
        wsi.cell(row=r, column=1, value=name)
        r += 1

    for s in wb.worksheets:
        s.freeze_panes = "A2"

    path = os.path.join(OUT, "PHANTM-TONY-V2-CALC.xlsx")
    wb.save(path)
    return path, num, fe, mx, dict(
        r04=r04, force_sheet=S_FORCE, mass_sheet=S_MASS, elec_sheet=S_ELEC,
        energy_sheet=S_ENERGY, res_sheet=S_RES,
    )


def _python_expectations(num):
    g = 9.80665
    rho_cu = 1.72e-8
    lx = 280 + 2 * 280
    ly = 1200
    lz = 25 * 125 + 24 * 187
    v_env = lx * ly * lz / 1e9
    v_slots = 2 * 24 * 187 * 280 * ly / 1e9
    v_solid = v_env - v_slots
    m_g = v_solid / 1000 * 8.12
    fd = 30 * m_g / 1000 * g * 1000
    d_od = 48
    turns_layer = int(1521 // d_od)
    layers = math.ceil(70 / turns_layer)
    build = layers * d_od
    l_turn = 2 * ((400 + build) + (1200 + build))
    wire_len_m = 70 * l_turn * 1e-6
    a_wire = math.pi * (20e-6) ** 2
    r_ohm = rho_cu * wire_len_m / a_wire
    i_cold = 5.0 / r_ohm
    i_plan = 1.0
    e_pulse = i_plan ** 2 * r_ohm * 0.001
    v_cu = wire_len_m * a_wire
    dT = e_pulse / (3.45e6 * v_cu)
    i_hot = 5.0 / (r_ohm * (1 + 0.00393 * dT))
    pitch_m = 312e-6
    step_m = pitch_m / 3
    k_tony = 2 * math.pi * (fd / 1000) / (step_m / 2)
    k_sin = 2 * math.pi * (fd / 1000) / pitch_m
    f_sin = (1 / (2 * math.pi)) * math.sqrt(k_sin / (m_g / 1000))
    t = num["translator"]
    c = num["coil"]
    return dict(
        L_z=lz, L_x=lx, V_solid=v_solid, M_t_mg=m_g * 1000, F_d=fd,
        R_c=r_ohm, I_max_cold=i_cold, dT=dT, I_max_hot=i_hot,
        f_res=f_sin, k_ratio=math.sqrt(k_tony / k_sin),
        mass_json=t["mass_mg"], fd_json=t["detent_target_mn"],
        r_json=c["resistance_ohm"],
    )


def verify(path: str, num: dict, meta: dict) -> None:
    if not os.path.exists(SOFFICE):
        raise SystemExit(f"LibreOffice not found at {SOFFICE}")
    exp = _python_expectations(num)
    assert abs(exp["mass_json"] - exp["M_t_mg"]) < 0.02
    assert abs(exp["fd_json"] - exp["F_d"]) < 0.02
    assert abs(exp["r_json"] - exp["R_c"]) < 0.005

    with tempfile.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, "in.xlsx")
        shutil.copy2(path, src)
        subprocess.run(
            [SOFFICE, "--headless", "--norestore", "--calc",
             "--convert-to", "xlsx", "--outdir", tmp, src],
            check=True, capture_output=True, timeout=120,
        )
        candidates = [f for f in os.listdir(tmp) if f.endswith(".xlsx")]
        assert candidates, "LibreOffice produced no xlsx"
        wb = load_workbook(os.path.join(tmp, candidates[0]), data_only=True)

        def cell(sheet, coord):
            v = wb[sheet][coord].value
            if v is None:
                raise AssertionError(f"{sheet}!{coord} is None after recalc")
            return float(v)

        # Cell map (header row 1; data from row 2 unless noted):
        # §0: pitch2 duty3 L_z4 L_x5 | §1: Venv2 … Vsolid4 mass7 F_d9
        # §3: … R_c9 I_cold13 | §5: … dT13 I_hot15 | §6: f_res5 ratio6
        checks = [
            (S_GEOM, "B4", exp["L_z"], 0.5, "L_z"),
            (S_GEOM, "B5", exp["L_x"], 0.5, "L_x"),
            (S_MASS, "B4", exp["V_solid"], 0.01, "V_solid"),
            (S_MASS, "B7", exp["M_t_mg"], 0.05, "mass_mg"),
            (S_MASS, "B9", exp["F_d"], 0.05, "F_d"),
            (S_ELEC, "B9", exp["R_c"], 0.01, "R_c"),
            (S_ELEC, "B13", exp["I_max_cold"], 0.01, "I_max_cold"),
            (S_ENERGY, "B13", exp["dT"], 0.15, "dT"),
            (S_ENERGY, "B15", exp["I_max_hot"], 0.02, "I_max_hot"),
            (S_RES, "B5", exp["f_res"], 1.0, "f_res"),
            (S_RES, "B6", exp["k_ratio"], 0.02, "k_ratio"),
        ]
        r04 = meta["r04"]
        v040 = cell(S_FORCE, f"D{r04}")
        assert abs(v040 - 0.40 * exp["R_c"]) < 0.02, f"V@0.4A={v040}"

        print("LibreOffice recalc checks:")
        bad = 0
        for sheet, coord, expect, tol, label in checks:
            got = cell(sheet, coord)
            ok = abs(got - expect) <= tol
            print(f"  {'OK' if ok else 'FAIL'} {label:12s}  "
                  f"{sheet}!{coord} = {got:.6g}  (expect {expect:.6g} ±{tol})")
            if not ok:
                bad += 1
        if bad:
            raise SystemExit(f"{bad} formula check(s) FAILED")
        print(f"  OK V@0.40A formula  {S_FORCE}!D{r04} = {v040:.4f}")
        print(f"all {len(checks)+1} checks passed")


def deliverable_name(subject: str = "PHANTM-TONY-V2-CALC") -> str:
    """Tony naming: date → time → subject → version (e.g. 20260728-2215-…-V7.2)."""
    ver = json.load(open(os.path.join(HERE, "version.json")))
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M")
    tag = f"V{ver['major']}.{ver['minor']}"
    return f"{stamp}-{subject}-{tag}.xlsx"


if __name__ == "__main__":
    p, num, fe, mx, meta = build()
    print(f"wrote {p}")
    verify(p, num, meta)
    name = deliverable_name()
    tagged = os.path.join(OUT, name)
    shutil.copy2(p, tagged)
    dl = os.path.expanduser(f"~/Downloads/{name}")
    shutil.copy2(p, dl)
    print(f"tagged {tagged}")
    print(f"downloads {dl}")
