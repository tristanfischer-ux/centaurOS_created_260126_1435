"""PHANTM — live-formula Excel workbook, single clear sheet.

Design rules (2026-07-24 rebuild after Excel repair-prompt + opacity feedback):
- ONE sheet, sectioned with banners; yellow = edit, white = live formula,
  grey = finite-element result (NOT a formula), red = placeholder to set.
- NO defined names (the likely corruption vector in v1) — plain cell refs.
- Every derived row's expected value is recomputed here in Python and printed
  into the note column, so a wrong formula is visible as a mismatch on sight.
- Honeycomb cell-array structural calcs included (Tony's prototype photos,
  24 Jul). Cell sizes are INPUTS from the outline spec — no RF is modelled.

Run: ~/.venvs/phantm/bin/python excel_calc.py  → out/PHANTM-CALC.xlsx
"""
import datetime
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

YEL = PatternFill("solid", fgColor="FFF3B0")   # editable input
GRY = PatternFill("solid", fgColor="E8EAED")   # FE constant — do not edit
RED = PatternFill("solid", fgColor="F8CECC")   # placeholder — must be set
BAN = PatternFill("solid", fgColor="D9E2F1")   # section banner
HDR = Font(bold=True)
SMALL = Font(size=9)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()
ws = wb.active
ws.title = "PHANTM calculator"
R = {}   # symbol -> row number of its value cell (column C)


def C(sym):
    return f"C{R[sym]}"


def _next():
    return ws.max_row + 1


def banner(text):
    r = _next() + 1   # blank row before each banner
    ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=11)
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = BAN


PART = PatternFill("solid", fgColor="9DB8D9")  # part banner (darker than BAN)


def part_banner(text):
    r = _next() + 2   # extra blank row before each PART
    ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=12)
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = PART


def note(text):
    r = _next()
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(italic=True, size=9)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c.alignment = WRAP


def head():
    r = _next()
    for col, t in enumerate(("Quantity", "Symbol", "Value", "Unit",
                             "Formula (in symbols)", "What it is / what to check"), 1):
        cc = ws.cell(row=r, column=col, value=t)
        cc.font = HDR


def inp(sym, label, val, unit, expl, fill=YEL):
    r = _next()
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=sym)
    c = ws.cell(row=r, column=3, value=val)
    c.fill = fill
    ws.cell(row=r, column=4, value=unit)
    ws.cell(row=r, column=5, value="input" if fill is YEL else
            ("SET THIS" if fill is RED else "FE result — not a formula"))
    ws.cell(row=r, column=6, value=expl).font = SMALL
    R[sym] = r


def der(sym, label, formula, unit, symtext, expected, expl):
    """formula: Excel string with {x} placeholders -> cell refs. expected: python float or str."""
    r = _next()
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=sym)
    ws.cell(row=r, column=3, value=formula.format(**{k: C(k) for k in R}))
    ws.cell(row=r, column=4, value=unit)
    ws.cell(row=r, column=5, value=symtext)
    exp = f"expect ≈ {expected:.4g} — " if isinstance(expected, float) else (
        f"expect {expected} — " if expected else "")
    ws.cell(row=r, column=6, value=exp + expl).font = SMALL
    R[sym] = r


# ============================ header + instructions ==========================
ws["A1"] = "PHANTM — CALCULATION WORKBOOK (v6, mirrors the report's de-risking funnel) — CONFIDENTIAL"
ws["A1"].font = Font(bold=True, size=13)
ws["A2"] = (f"Generated {datetime.datetime.now():%Y-%m-%d %H:%M} local. One sheet, six parts matching "
            "the report: 0 state of play · I concept inputs · II hardening (live formulas + FE) · "
            "III design choices · IV manufacture · V glossary + where next.")
ws["A2"].font = SMALL

banner("HOW TO USE THIS SHEET")
note("1. EDIT ONLY THE YELLOW CELLS. Every white cell is a live Excel formula and recomputes "
     "instantly — click any white value to see its formula; column E states the same formula in symbols.")
note("2. GREY CELLS ARE FINITE-ELEMENT RESULTS, NOT FORMULAS. They are real solved numbers from the "
     "2D nonlinear FE model. Changing a yellow input does NOT update them — they must be re-solved "
     "(scripts/phantm/femm/). The 'invalidated by' note on each says which inputs stale it.")
note("3. The waveguide cutoff λc is DERIVED from the measured 3.10 mm interior cell by a validated "
     "eigensolver (scripts/phantm/hexcell.py) — Vlad to confirm; every λg row updates if he edits it.")
note("4. DO NOT add a simple-formula force estimate. Lumped magnetic-circuit formulas over-predict "
     "force ≈15–40× at this gap/tooth scale (fringing at gap/tooth = 1/3 kills the permeance "
     "modulation) — that is the report's central finding. Force numbers live ONLY in the grey FE block.")
note("5. Units are embedded in each formula (column E shows the conversion factors). If you change a "
     "unit, change the factor — the note column states the expected value as a cross-check.")

# ============================ Part 0 · state of play =========================
import json as _json
import os as _os
_reg = _json.load(open(_os.path.join(_os.path.dirname(__file__), "out",
                                     "claims-register.json")))
part_banner("PART 0 — STATE OF PLAY (mirrors report Part 0)")
note(f"The report's claims register holds {_reg['total']} load-bearing claims "
     f"({_reg['counts']['PROVEN-FE']} PROVEN-FE, {_reg['counts']['PROVEN-calc']} PROVEN-calc, "
     f"{_reg['counts']['RED-TEAMED']} RED-TEAMED, {_reg['counts']['MEASURED']} MEASURED, "
     f"{_reg['counts']['TESTED']} TESTED, {_reg['counts']['ESTIMATE']} ESTIMATE) with "
     f"{_reg['open_gates']} OPEN GATES: transient-eddy FE · damper-vent tolerance · "
     "DEMAGNETISATION FE (blocks dual drive) · stamped-tooth edge coupon · driver-EMC layout spec. "
     "Headlines: detent 7.72 mN = 5.0 g (fixed design, as-drawn registration) · BALANCED optimised "
     "set 14.6/11.0 g at 37 mJ/step · cell fc = 53.56 GHz, full 2π from ≈57.4–58.0 GHz · 16.7°/step "
     "at 70 GHz · $0.10–0.25/actuator at ≈10 M/yr. Recommended first hardware: Prototype-A (the "
     "original 5 g design, unipolar drive — zero open gates). Full table: report §7.")

part_banner("PART I — CONCEPT INPUTS: the numbers the design rests on (yellow = fettle)")

# ============================ 1 · constants ==================================
banner("1 · PHYSICAL CONSTANTS (fixed — no reason to edit)")
head()
inp("g", "Standard gravity", 9.80665, "m/s²", "used in Fd = n·g·Mt")
inp("mu0", "Vacuum permeability", 1.2566370614e-6, "H/m", "4π×10⁻⁷")
inp("rhoc", "Copper resistivity (20 °C)", 1.72e-8, "Ω·m", "coil resistance; +0.39 %/K when hot")
inp("dcu", "Copper density", 8960, "kg/m³", "coil mass")
inp("ccu", "Copper heat capacity", 385, "J/(kg·K)", "adiabatic coil heating")
inp("c0", "Speed of light", 299792458, "m/s", "free-space wavelength")

# ============================ 2 · geometry ===================================
banner("2 · TRANSLATOR + STATOR GEOMETRY (Tony's baseline; yellow = fettle)")
head()
inp("w1", "Translator toothed width", 1.549, "mm", "the toothed faces' width")
inp("w2", "Translator transverse width", 1.55, "mm", "the un-toothed width")
inp("L", "Translator length", 12.5, "mm", "")
inp("sd", "Translator slot depth", 0.465, "mm", "")
inp("sw", "Slot width", 0.232, "mm", "slot + land = tooth pitch")
inp("lw", "Land (tooth) width", 0.232, "mm", "")
inp("dsmc", "SMC density", 7400, "kg/m³", "7300–7600 depending on grade; Mt scales linearly")
inp("sep", "Slot-face separation", 1.704, "mm", "sets the working gap: (sep − w1)/2")
inp("sl", "Slot-section bar length", 1.708, "mm", "NOT the separation — two different 1.70x numbers")
inp("sax", "Pole axial length", 1.16, "mm", "")
inp("sdep", "Slot-section radial depth", 0.465, "mm", "")
inp("sslot", "Stator slot depth", 0.155, "mm", "")
inp("spc", "Inter-pole spacing", 0.374, "mm", "brief value; Tony CAD reads 0.400; exact ⅓-pitch row below")
inp("np", "Number of poles", 3, "-", "")
inp("gfix", "Working gap — FIXED design", 0.020, "mm", "baseline was 0.0775; FE says ≤0.020 needed for 5 g")
inp("brax", "Bridge axial thickness (F2, ×1.5)", 0.348, "mm", "baseline 0.232; stays inside pole footprint")
inp("brtr", "Bridge transverse width", 1.162, "mm", "")

# ============================ 3 · magnet =====================================
banner("3 · MAGNET (NdFeB in-series slugs)")
head()
inp("br", "Remanence Br", 1.30, "T", "N42 class; N52 = 1.45 T is the force-ladder lever")
inp("mur", "Recoil permeability", 1.05, "-", "")
inp("pml", "Magnet length Pm (fixed design)", 0.243, "mm", "the TRIM parameter at prototype")
inp("dpm", "NdFeB density", 7500, "kg/m³", "")

# ============================ 4 · coil + drive ===============================
banner("4 · COIL + DRIVE")
head()
inp("N", "Turns per coil", 20, "-", "more turns does NOT raise the 1 V MMF ceiling (see row below)")
inp("dw", "Bare wire diameter", 0.050, "mm", "Ø50 µm bondable magnet wire")
inp("mtl", "Mean turn length", 3.15, "mm", "from coil former geometry")
inp("V", "Supply voltage", 1.0, "V", "1 V caps MMF at ≈36 A-turns; full drive point needs ≈1.9 V")
inp("Id", "Drive current (pulse)", 1.8, "A", "set 3.35 to see the full-drive energy/ΔT")
inp("Ix", "Full-drive current Ic* (FE)", 3.35, "A", "the FE-solved 2·Fd drive point of the fixed design")
inp("tp", "Pulse width", 1.5, "ms", "")
inp("lc", "Coil inductance (FE)", 0.60, "µH", "FE at the drive point; lumped model said 2.2", fill=GRY)

# ============================ 5 · RF + dynamics inputs =======================
banner("5 · OPERATING BAND + DYNAMICS INPUTS")
head()
inp("f", "Operating frequency", 70, "GHz", "Tony 24 Jul: outline spec was the ≈70 GHz band")
inp("lamc", "Waveguide CUTOFF wavelength λc", 5.598, "mm",
    "DERIVED: validated FD eigensolver on the 3.10 mm interior hex (hexcell.json; "
    "fc = 53.56 GHz) — Vlad to confirm against the metallised cell")
inp("kdet", "Detent stiffness (FE)", 200, "N/m", "slope of the FE detent curve at zero", fill=GRY)

# ============================ derived: mass + force ==========================
p_ = 0.232 + 0.232
nslots_ = math.floor(12.5 / p_)
mt_ = (1.549 * 1.55 * 12.5 - 2 * nslots_ * 0.465 * 0.232 * 1.55) * 7400 / 1000  # mg
fd_ = 5 * 9.80665 * mt_ / 1000
part_banner("PART II — HARDENING: live formulas + finite-element facts (mirrors report Part II)")

banner("6 · MASS + FORCE TARGETS (live formulas)")
head()
der("p", "Tooth pitch", "={sw}+{lw}", "mm", "p = slot + land", 0.464, "")
der("ns", "Slots per face", "=FLOOR({L}/{p},1)", "-", "⌊L/p⌋", float(nslots_), "")
der("mt", "Translator mass Mt",
    "=({w1}*{w2}*{L}-2*{ns}*{sd}*{sw}*{w2})*{dsmc}/1000", "mg",
    "(bar − 52 slots) × ρ; mm³·kg/m³ ÷1000 → mg", mt_,
    "matches Tony's ≈0.16 g. THE spec-defining number: Fd scales with it")
der("wm", "Baseline working gap Wm", "=({sep}-{w1})/2*1000", "µm", "(sep − w1)/2", 77.5,
    "the as-specified 77.5 µm; the FIXED design closes it to the gap in section 2")
der("clo", "Magnetic loop closure (radial envelope)", "=2*{sdep}+{sep}", "mm",
    "2·slot-section depth + separation", 2.634, "must equal the bridge span — and it does")
der("ext", "Stator extent", "={np}*{sax}+({np}-1)*{spc}", "mm", "3 poles + 2 spacings", 4.228, "")
der("str", "Usable stroke", "={L}-{ext}", "mm", "L − stator extent", 12.5 - 4.228,
    "compare with the λg/2 stroke-need row in section 9")
der("fd", "Detent target Fd = 5·g·Mt", "=5*{g}*{mt}/1000", "mN", "5·g·Mt; mg ÷1000 → mN", fd_,
    "the brief's 5 g MINIMUM (Tony 24 Jul)")
der("fd2", "Drive target 2·Fd", "=2*{fd}", "mN", "2·Fd", 2 * fd_, "")
der("f20", "20 g ambition", "=20*{g}*{mt}/1000", "mN", "20·g·Mt", 4 * fd_, "Tony's ladder lower bound")
der("f30", "30 g ambition", "=30*{g}*{mt}/1000", "mN", "30·g·Mt", 6 * fd_,
    "needs gap ≤≈12 µm or a wider magnetic section — report §A.3")

# ============================ derived: magnetics =============================
hc_ = 1.30 / (1.2566370614e-6 * 1.05) / 1000
pma_ = 0.348 * 1.162
banner("7 · MAGNET CIRCUIT (live formulas — sanity only; forces are FE)")
head()
der("hc", "Coercive field Hc", "={br}/({mu0}*{mur})/1000", "kA/m", "Br/(µ0·µr)", hc_, "")
der("pma", "PM cross-section", "={brax}*{brtr}", "mm²", "bridge axial × transverse", pma_,
    "the F2 ×1.5 section; baseline was 0.232×1.162")
der("mmf", "PM MMF", "={br}/({mu0}*{mur})*{pml}/1000", "A-turns", "Hc·Pm; mm ÷1000 → m",
    1.30 / (1.2566370614e-6 * 1.05) * 0.243e-3, "")
der("phi", "Flux ceiling Br·A", "={br}*{pma}", "µWb", "T·mm² → µWb directly", 1.30 * pma_,
    "the hard PM ceiling no geometry change beats")

# ============================ derived: coil ==================================
aw_ = math.pi * 0.025 ** 2
rc_ = 1.72e-8 * 0.063 / (aw_ * 1e-6)
iinf_ = 1.0 / rc_
es_ = 1.8 ** 2 * rc_ * 1.5
mcu_ = 8.96 * aw_ * 63
banner("8 · COIL ELECTRICAL (live formulas)")
head()
der("aw", "Wire cross-section", "=PI()*({dw}/2)^2", "mm²", "π(d/2)²", aw_, "")
der("lwr", "Wire length", "={N}*{mtl}", "mm", "N × mean turn", 63.0, "")
der("rc", "Coil resistance Rc", "={rhoc}*({lwr}/1000)/({aw}*0.000001)", "Ω",
    "ρ·L/A; mm→m, mm²→m²", rc_, "grows ≈9 % at +22 K — headroom check")
der("iinf", "Steady current on V", "={V}/{rc}", "A", "V/R", iinf_, "")
der("mmfc", "MMF ceiling at V", "={N}*{iinf}", "A-turns", "N·V/R", 20 * iinf_,
    "fixed wire gauge: R∝N so N·I∞ is TURNS-INDEPENDENT — more turns alone can't fix a 1 V limit")
der("vx", "Voltage for Ic*", "={Ix}*{rc}", "V", "Ic*·R", 3.35 * rc_,
    "≈1.9 V incl. driver drop — the full-drive point is out of reach of 1 V")
der("jx", "Current density at Ic*", "={Ix}/{aw}", "A/mm²", "I/A", 3.35 / aw_,
    "PULSE ONLY — never continuous")
der("tau", "Electrical time constant", "={lc}/{rc}", "µs", "L/R; µH/Ω → µs", 0.60 / rc_,
    "µs-scale: the coil current is never the step-time limit")
der("es", "Energy per step", "={Id}^2*{rc}*{tp}", "mJ", "I²·R·t; A²·Ω·ms → mJ", es_,
    "at the default 1.8 A; set Id = 3.35 → ≈9.3 mJ")
der("mcu", "Coil copper mass", "={dcu}/1000*{aw}*{lwr}", "mg", "ρ·A·L; kg/m³→mg/mm³ is ÷1000",
    mcu_, "")
der("dT", "Adiabatic ΔT per step", "={es}*1000/({mcu}*{ccu})", "K", "E/(m·c); mJ,mg → ×1000",
    es_ * 1000 / (mcu_ * 385), "6.3 K at 1.8 A / ≈22 K at 3.35 A — duty-limit the full-drive pulse")

# ============================ derived: RF + steps ============================
lam0_ = 299792458 / 70e9 * 1000
lamg_ = lam0_ / math.sqrt(1 - (lam0_ / 5.598) ** 2)
step_ = p_ / 3 * 1000
banner("9 · WAVELENGTH, STROKE NEED + STEP GEOMETRY")
note("The reflection phase per mechanical step is 4π·Δd/λ. INSIDE the cell the guide wavelength λg "
     "rules (Tony 24 Jul), and λg > λ0 always — so free-space figures are UPPER bounds on phase per "
     "step and UNDER-estimates of stroke need. No RF is modelled here: λc is Tony/Vlad's number.")
head()
der("lam0", "Free-space wavelength λ0", "={c0}/({f}*1000000000)*1000", "mm", "c/f → mm", lam0_, "")
der("lamg", "Guide wavelength λg",
    "=IF({lam0}>={lamc},\"set real cutoff — below cutoff\",{lam0}/SQRT(1-({lam0}/{lamc})^2))",
    "mm", "λ0/√(1−(λ0/λc)²)", lamg_,
    "below-cutoff guard shows a message instead of #NUM")
der("stp", "Nominal step (pitch/3)", "={p}/3*1000", "µm", "p/3", step_,
    "FE actual at 374 µm registration: 172.6/146.1/145.3 (grey block)")
der("phg", "Phase per step IN GUIDE", "=IF(ISNUMBER({lamg}),DEGREES(4*PI()*{stp}/1000/{lamg}),\"needs λc\")",
    "°", "4π·Δd/λg", math.degrees(4 * math.pi * step_ / 1000 / lamg_), "the honest per-step quantisation")
der("phf", "Phase per step free-space", "=DEGREES(4*PI()*{stp}/1000/{lam0})", "°", "4π·Δd/λ0",
    math.degrees(4 * math.pi * step_ / 1000 / lam0_), "upper bound only")
der("sn", "Stroke needed = λg/2", "=IF(ISNUMBER({lamg}),{lamg}/2,\"needs λc\")", "mm", "λg/2",
    lamg_ / 2, "compare with the usable-stroke row in section 6 — margin is large")
der("off", "Registration offset (as drawn)", "=MOD({sax}+{spc},{p})*1000", "µm",
    "(pole + spacing) mod p", 142.0,
    "vs ideal p/3 = 154.7 — the offset is the detent-vs-uniformity knob (report §9)")
der("exsp", "Exact ⅓-pitch spacing", "=(3*{p}+{p}/3-{sax})*1000", "µm", "3p + p/3 − pole",
    (3 * p_ + p_ / 3 - 1.16) * 1000, "386.7 if p = 464 µm rules; 390 if 465")
der("fr", "Detent ring frequency", "=SQRT({kdet}/({mt}/1000000))/(2*PI())", "Hz",
    "√(k/m)/2π; mg ÷10⁶ → kg", math.sqrt(200 / (mt_ / 1e6)) / (2 * math.pi), "")

# ============================ FE results =====================================
banner("10 · FINITE-ELEMENT RESULTS — GREY = NOT FORMULAS, DO NOT EDIT")
note("These come from the 2D nonlinear FE solves (native xfemm; C-core validation gate passed). "
     "They CANNOT be spreadsheet formulas — rerun scripts/phantm/femm/ after changing any input "
     "named in the last column. Treat them as the measured truth the yellow cells must be judged against.")
head()
fe_rows = [
    ("Baseline net detent plateau", "0.47–0.52", "mN", "pm-ic-sweeps.json", "gap, tooth geometry — ×15 short of Fd for ANY magnet length"),
    ("Baseline drive plateau (8 A)", 2.5, "mN", "pm-ic-sweeps.json", "gap, tooth geometry — ×6 short of 2·Fd"),
    ("Fixed design Pm*", 243, "µm", "fixed-design.json", "gap, PM section, registration"),
    ("Fixed design breakaway detent", 7.72, "mN", "fixed-design.json", "= 5.0 g — meets the brief minimum"),
    ("Fixed design Ic*", 3.35, "A", "fixed-design.json", "turns, gap; needs ≈1.9 V"),
    ("Fixed drive peak / path-min", "15.4 / 4.7", "mN", "fixed-design.json", "gap, PM section"),
    ("Coil inductance at drive point", 0.60, "µH", "fixed-design.json", "gap, turns"),
    ("Detent at EXACT ⅓ registration", 5.95, "mN", "f3-registration-check.json", "registration — the detent-vs-uniformity trade"),
    ("Steps at 374 µm registration", "172.6 / 146.1 / 145.3", "µm", "fix-alternatives.json", "pole spacing"),
    ("Detent at gap 100 / 150 µm", "0.67 / 0.48", "mN", "tony-gap-check.json", "gap — Tony's manufacturable direction FAILS the 5 g test"),
    ("Gap sensitivity", "≈ −8 % per µm", "-", "femm-variants.json", "gap (20–40 µm band) — ±5 µm scatter = ±40 % force"),
    ("Gap-flux modulation (baseline)", "≈8 %", "-", "femm sweep diagnostics", "gap/tooth ratio = 1/3 — why lumped formulas fail here"),
]
for label, val, unit, art, inval in fe_rows:
    r = _next()
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=3, value=val)
    ws.cell(row=r, column=4, value=unit)
    ws.cell(row=r, column=5, value=art).font = SMALL
    ws.cell(row=r, column=6, value="invalidated by: " + inval).font = SMALL
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = GRY

# ============================ honeycomb ======================================
af_, t_, dep_, apw_, aph_ = 3.1, 0.15, 7.75, 100.0, 100.0
cella_ = math.sqrt(3) / 2 * (af_ + t_) ** 2
ncell_ = math.floor(apw_ * aph_ / cella_)
wall_ = math.sqrt(3) / 2 * ((af_ + t_) ** 2 - af_**2) * dep_
cellm_ = wall_ * 1240 / 1000
hcm_ = ncell_ * cellm_ / 1000
ssv_ = 2 * (1.16 * 1.708 * 0.465 - 3 * 0.155 * 0.232 * 1.708)
brv_ = 0.348 * 1.162 * 2.634
pole_ = (ssv_ + brv_) * 7400 / 1000
pmm_ = pma_ * 0.243 * 7500 / 1000
act_ = mt_ + 3 * (pole_ + pmm_ + mcu_)
banner("11 · HONEYCOMB CELL ARRAY — STRUCTURE ONLY (Tony's CAD, 24 Jul)")
note("Structural/packing calculations for the cell lattice that houses one actuator per cell. "
     "Dimensions from Tony's SketchUp file ('24-hex sub array, 3.1 mm between flats, 150 micron "
     "wall'): 3.1 mm across-flats, 150 µm walls, cells 7.7 mm deep (earlier CAD), tiled as "
     "24-hex sub-arrays. The cell-size ↔ band mapping is RF and stays with Tony/Vlad — cell "
     "sizes here are INPUTS. The large-cell photos are the demo-scale prototypes (≈25–30 mm).")
head()
inp("af", "Hex cell INTERIOR across-flats", 3.1, "mm",
    "the RF aperture — STL wall-plane forensics: walls at 3.25 mm pitch, opening 3.10")
inp("thc", "Cell wall thickness", 0.15, "mm", "Tony .skp 24 Jul: '150micron wall'")
inp("dhc", "Cell depth", 7.75, "mm", "MEASURED from Tony's STL bounding box (drawing reads 7.7)")
inp("apw", "Aperture width", 100.0, "mm", "SET FROM CAD — round aperture: use its bounding square")
inp("aph", "Aperture height", 100.0, "mm", "SET FROM CAD")
inp("csub", "Cells per sub-array", 24, "-", "Tony .skp: tileable 24-hex sub-array")
inp("nsub", "Number of sub-arrays", 1, "-", "SET FROM CAD — how many 24-hex tiles the aperture uses")
inp("dhcm", "Wall material density", 1240, "kg/m³",
    "3D-print resin/PLA ≈1240 (as prototyped); aluminium 2700; metallised print adds ≈5 %")
inp("sqs", "Square-grid pitch (alt. prototype)", 3.0, "mm", "the square egg-crate array in photos 3–4")
der("hexa", "Hex side length (interior)", "={af}/SQRT(3)", "mm", "a/√3", af_ / math.sqrt(3), "")
der("pitch", "Tiling pitch", "={af}+{thc}", "mm", "a + t (one shared wall)", af_ + t_, "")
der("cella", "Tiling cell area", "=SQRT(3)/2*({af}+{thc})^2", "mm²", "(√3/2)·p²",
    math.sqrt(3) / 2 * (af_ + t_) ** 2, "area each cell occupies in the lattice")
der("ncell", "Cells per aperture (by area)", "=FLOOR({apw}*{aph}/{cella},1)", "-",
    "⌊A_aperture/A_cell⌋", float(ncell_),
    "edge effects ignored — the cluster count below is the exact method")
der("ncl", "Cells from sub-array tiling", "={csub}*{nsub}", "-", "cells/sub-array × sub-arrays",
    24.0, "Tony's .skp tiling — USE THIS count once the sub-array number is known")
der("rdhc", "Honeycomb relative density", "=1-({af}/({af}+{thc}))^2", "-", "1 − (a/p)²",
    1 - (af_ / (af_ + t_)) ** 2, "fraction of the slab that is wall material")
der("open", "Open-area fraction", "=({af}/({af}+{thc}))^2", "-", "(a/p)²",
    (af_ / (af_ + t_)) ** 2, "")
der("wall", "Wall volume per cell", "=SQRT(3)/2*(({af}+{thc})^2-{af}^2)*{dhc}", "mm³",
    "(√3/2)(p² − a²)·L (walls shared)", wall_, "")
der("cellm", "Wall mass per cell", "={wall}*{dhcm}/1000", "mg", "V·ρ ÷1000 → mg", cellm_, "")
der("hcm", "Honeycomb lattice mass", "={ncl}*{cellm}/1000", "g", "N·m_cell ÷1000 → g",
    24 * cellm_ / 1000, "per 24-hex sub-array at the default tile count")
der("areal", "Areal density", "={dhcm}*{dhc}/1000*{rdhc}", "kg/m²", "ρ·depth·relative density",
    1240 * dep_ / 1000 * 2 * t_ / af_, "printed plastic; ×2.18 if aluminium")
der("wfit", "Widest actuator that fits at full height",
    "=(2*{af}-{clo})/SQRT(3)", "mm", "(2·AF − h)/√3, h = radial envelope 2.634",
    (2 * af_ - 2.634) / math.sqrt(3),
    "hex geometry: the corner cuts width as height grows")
der("fitq", "Does the actuator fit THIS cell?",
    "=IF(AND({clo}<={af},{sl}<={wfit}),\"FITS\",\"DOES NOT FIT\")", "-",
    "needs h ≤ AF and w ≤ (2·AF − h)/√3", "FITS",
    "envelope 1.708 wide × 2.634 tall vs the across-flats above")
der("fit50", "Fit check @ 50 GHz cell (AF = 3.0)",
    "=IF(AND({clo}<=3,{sl}<=(2*3-{clo})/SQRT(3)),\"FITS\",\"DOES NOT FIT\")", "-",
    "same test at AF = 3.0", "FITS", "w_max = 1.944 ≥ 1.708")
der("fit80", "Fit check @ 80 GHz cell (AF = 1.9)",
    "=IF(AND({clo}<=1.9,{sl}<=(2*1.9-{clo})/SQRT(3)),\"FITS\",\"DOES NOT FIT\")", "-",
    "same test at AF = 1.9", "DOES NOT FIT",
    "radial 2.634 > 1.9 — needs ≥2-deep axial staggering (report §9.5)")
der("fit160", "Fit check @ 160 GHz cell (AF = 0.94)",
    "=IF(AND({clo}<=0.94,{sl}<=(2*0.94-{clo})/SQRT(3)),\"FITS\",\"DOES NOT FIT\")", "-",
    "same test at AF = 0.94", "DOES NOT FIT", "5× over — per-band redesign territory")
der("ssv", "Slot-section steel per pole (2 bars)",
    "=2*({sax}*{sl}*{sdep}-3*{sslot}*{sw}*{sl})", "mm³", "2·(bar − 3 stator slots)", ssv_, "")
der("brv", "Bridge steel per pole", "={brax}*{brtr}*{clo}", "mm³", "axial × transverse × span",
    brv_, "wrap approximated as one prism")
der("polem", "Steel mass per pole", "=({ssv}+{brv})*{dsmc}/1000", "mg", "V·ρ ÷1000", pole_, "")
der("pmm", "Magnet mass per pole", "={pma}*{pml}*{dpm}/1000", "mg", "A·Pm·ρ ÷1000", pmm_, "")
der("actm", "ONE actuator, total mass",
    "={mt}+{np}*({polem}+{pmm}+{mcu})", "mg", "translator + 3·(steel + magnet + coil)", act_,
    "excludes reflector (<2 % — Tony 24 Jul), frame and bearing")
der("apact", "Actuators, total mass", "={ncl}*{actm}/1000", "g", "N·m_act ÷1000",
    24 * act_ / 1000, "one actuator per cell — DOMINATES the lattice mass ≈29×")
der("aptot", "Moving hardware total", "={hcm}+{apact}", "g", "lattice + actuators",
    24 * (cellm_ / 1000 + act_ / 1000), "per sub-array; frame, feed, reflectors excluded")
inp("stlv", "24-hex sub-array solid volume (STL, measured)", 192.25, "mm³",
    "mesh integral of Tony's 1,640-triangle STL — ground truth incl. boundary walls", fill=GRY)
der("stlm", "24-hex lattice mass (measured)", "={stlv}*{dhcm}/1000", "mg", "V·ρ ÷1000",
    192.25 * 1.24, "= 0.238 g printed / 0.519 g Al; +25% over the shared-wall asymptote "
    "(an isolated tile owns its outer walls)")
inp("stlv7", "7-hex sub-array solid volume (STL, measured)", 64.53, "mm³",
    "mesh integral of Tony's 264-triangle STL (9.9 × 9.6 × 7.75 mm tile)", fill=GRY)
der("stlm7", "7-hex lattice mass (measured)", "={stlv7}*{dhcm}/1000", "mg", "V·ρ ÷1000",
    64.53 * 1.24, "= 0.080 g printed / 0.174 g Al; +44% over the asymptote — the smaller "
    "the tile, the bigger the boundary-wall share")
der("sqa", "Square-grid cell area", "={sqs}^2", "mm²", "s²", 9.0, "")
der("sqrd", "Square-grid relative density", "=2*{thc}/{sqs}", "-", "2t/s — same form as hex",
    2 * t_ / 3.0, "")
der("sqfit", "Square-grid actuator fit",
    "=IF(AND({clo}<={sqs},{sl}<={sqs}),\"FITS\",\"DOES NOT FIT\")", "-",
    "rectangle needs s ≥ 2.634 only", "FITS",
    "square cells take the rectangular actuator with no corner clash — hex needs the width check; "
    "hex packs ≈15 % more cells per area at AF = s")

note("Manufacture note: at demo scale the lattice is conventional (bonded/folded sheet, printing). "
     "At 3.0 mm cells and below it becomes precision fabrication — wire-EDM'd or electroformed — "
     "and belongs in the same supplier conversations as the actuator (report §23).")

# ============================ 12 · wall-thickness proof ======================
banner("12 · WALL-THICKNESS PROOF — thicker walls are RF-free (report §12; wall-proof.json)")
note("The cell's cutoff is the eigenvalue of the INTERIOR cross-section — wall thickness is not a "
     "parameter (∂fc/∂wall = 0 at fixed interior). The wall's job is only to be ≥ ≈10 skin depths of "
     "conductor. MANDATORY RULE: thicken OUTWARD — hold the interior at 3.10 mm and let the pitch "
     "grow; thickening INWARD at fixed pitch moves fc by 17.3 MHz/µm and loses the band bottom. "
     "Change the wall-thickness input in section 11 and watch: pitch/mass/open-area move, fc does not.")
head()
inp("fghz", "Frequency for skin-depth check", 70.0, "GHz", "57.5 = band bottom (worst case)")
inp("rhop", "Plated-Cu resistivity (worst case)", 3.0e-8, "Ω·m",
    "rough electroless Cu; bulk Cu is 1.72e-8 — plated is the honest case")
inp("tpl", "Plating thickness", 3.0, "µm", "the §21 spec floor (3–5 µm + flash)")
inp("itol", "Interior tolerance (fabrication route)", 25.0, "µm",
    "±10 electroform/EDM · ±25 good print · ±50 loose print (KILLED)")
der("skin", "Skin depth at f", "=SQRT({rhop}/(PI()*{fghz}*1000000000*4*PI()*0.0000001))*1000000",
    "µm", "√(ρ/(π·f·µ0))", math.sqrt(3.0e-8 / (math.pi * 70e9 * 4e-7 * math.pi)) * 1e6,
    "0.36 µm at 57.5 GHz, 0.29 at 90 — fields die in a fraction of a micron")
der("att3", "Field attenuation through the plating alone", "=8.686*{tpl}/{skin}", "dB",
    "8.686·t/δ", 8.686 * 3.0 / (math.sqrt(3.0e-8 / (math.pi * 70e9 * 4e-7 * math.pi)) * 1e6),
    "72–93 dB across the band — the substrate behind it is invisible")
der("attw", "Field attenuation through the full wall", "=8.686*{thc}*1000/{skin}", "dB",
    "8.686·t_wall/δ", 8.686 * 150.0 / (math.sqrt(3.0e-8 / (math.pi * 70e9 * 4e-7 * math.pi)) * 1e6),
    "≈4,000 dB at 150 µm — cell-to-cell through-wall coupling does not exist; thicker only lowers it")
der("fcsc", "Cutoff scatter from interior tolerance", "=17.3*{itol}/1000", "GHz",
    "17.3 MHz/µm × tolerance", 17.3 * 25.0 / 1000,
    "±0.43 GHz at ±25 µm — comfortably inside the ≈57.5 GHz band-edge margin; ±50 µm is the kill line")
note("The one real cost of thicker walls is ARRAY pitch (section 11's pitch row: interior + wall) — "
     "grating-lobe/scan headroom is Tony/Vlad's array budget: broadside limit = 299.79/pitch GHz "
     "(92.2 at 3.25 → 88.2 at 3.40). Red-teamed by grok-4.5 (8 attacks, none landed): report §12.")

part_banner("PART III — DESIGN CHOICES: the campaign Pareto and the drive that survived "
            "(mirrors report Part III)")

# ============================ 13 · optimisation campaign =====================
banner("13 · OPTIMISATION CAMPAIGN — the Pareto (report §14; grey = FE campaign results)")
note("≈250 FE solutions, 24 Jul, council-audited. These are MEASURED campaign numbers, not "
     "formulas: the geometry stack is duty 0.40 teeth (186 µm) + translator slots 1.5× + stator "
     "slots 2× at gap 20 µm. Detent strength TRADES against stepping cost — both optimised sets "
     "need DUAL-coil drive (pull target pole, CANCEL holding pole with reverse current). "
     "Recommendation: the BALANCED set. Artefacts: out/opt/*.json; decisions: OPTIMISATION-PLAN.md.")
head()
opt_rows = [
    ("Original FIXED design", "5.0 / 3.8 g", "-", "k 200 N/m · single 1.8 A · 2.7 mJ · settle 2.5-4 ms", "§9 baseline", "opt-sweeps-1 baseline row"),
    ("BALANCED set (N42, Pm 0.4) — RECOMMENDED", "15.2/11.4 mN = 14.6/11.0 g", "-", "k 415 N/m · DUAL ±3.35 A hold 3 ms · 37 mJ · settle 5.5 ms · ΔT ≈44 K", "2.2× spec worst-reg (2.9× drawn)", "opt-sweeps-3 + damper-balanced.json"),
    ("MAX-FORCE set (N52, Pm 0.5)", "19.9/14.9 mN = 19.1/14.3 g", "-", "k 544 N/m · DUAL ±5 A hold 3 ms · 83 mJ · settle 8 ms · ΔT ≈97 K", "if vibration demands", "opt-sweeps-4 + damper.json"),
    ("Gap-30 full stack (the tolerance trade)", "11.3 / 7.7 g", "-", "3 basins both registrations; assembly tolerance 1.5× looser", "Tony's margin-vs-cost call", "opt-sweeps-3"),
    ("Gap-40 (KILLED)", "1 basin as-drawn", "-", "breakaway passed 5 g but not a stepper — basin-check mirage", "dead", "opt-recentre.json"),
    ("Damper spec (both sets)", "Ø0.15 mm vent", "-", "captures at EVERY hold ≥3 ms; larger vents escape; sealed mis-parks", "Tony item 10 solved", "damper*.json"),
]
for label, val, unit, detail, verdict, art in opt_rows:
    r = _next()
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=3, value=val)
    ws.cell(row=r, column=4, value=unit)
    ws.cell(row=r, column=5, value=art).font = SMALL
    ws.cell(row=r, column=6, value=detail + " — " + verdict).font = SMALL
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = GRY

# ============================ 14 · drive electronics =========================
banner("14 · DRIVE ELECTRONICS (per report §15; artefact drive-electronics.json)")
note("UNIPOLAR drive (council-corrected): direction comes from the phase SEQUENCE, so each coil "
     "needs ONE low-side FET + clamp (72/tile) — no H-bridges. The buck RAIL VOLTAGE is the "
     "current control (resistive coil, τ ≈ 1.1 µs): 1.07 V → 1.8 A step, 2.0 V → 3.35 A full "
     "drive (total path 0.60 Ω). TWO REGIMES — never mix. Idle power is ZERO. "
     "Stack: aperture PCB (holes + pads + per-coil FETs) + driver PCB (burst-rated buck, MCU).")
head()
inp("vrail", "Buck rail voltage", 2.0, "V", "DAC-set 0.8–2.1 V; current = V/(rc+rext)")
inp("rext", "FET + trace path resistance", 0.045, "Ω", "15 mΩ FET + 30 mΩ trace/web (sol-consistent)")
inp("npar", "Cells stepped in parallel", 8, "-", "driver-board sizing knob")
inp("nstp", "Average steps per re-point", 10, "-", "assumption — depends on the phase map")
inp("tstp", "Time per completed step", 4, "ms", "1.5 ms pulse + settle (report §9.4 upper)")
inp("napc", "Cells in a 10 cm aperture", 1093, "-", "100×100 mm / 9.15 mm² tiling area")
der("irail", "Rail current, FULL-drive burst", "={npar}*{Ix}", "A", "n·Ic*", 8 * 3.35, "")
der("burst", "Pulse power, FULL-drive (worst case)", "={npar}*{Ix}^2*({rc}+{rext})", "W",
    "n·Ic*²·(Rc+Rext)", 8 * 3.35**2 * 0.5968, "53.6 W at 8-parallel; ×0.375 duty ⇒ ≈20 W avg")
der("bstep", "Pulse power, STEPPING", "={npar}*{Id}^2*({rc}+{rext})", "W",
    "n·Id²·(Rc+Rext)", 8 * 1.8**2 * 0.5968, "15.5 W at 8-parallel; ≈5.8 W avg over a re-point")
der("tilet", "24-cell tile re-point time", "=CEILING({csub}/{npar},1)*{nstp}*{tstp}/1000", "s",
    "⌈cells/n⌉·steps·t", math.ceil(24 / 8) * 10 * 4 / 1000, "")
der("tilee", "24-cell tile re-point energy", "={csub}*{nstp}*{es}/1000", "J", "cells·steps·E_step",
    24 * 10 * 2.682 / 1000, "at the 1.8 A step energy")
der("apt", "10 cm aperture re-point time", "=CEILING({napc}/{npar},1)*{nstp}*{tstp}/1000", "s",
    "⌈N/n⌉·steps·t", math.ceil(1093 / 8) * 10 * 4 / 1000, "5.5 s @8-par; set npar=64 → 0.69 s")
der("ape", "Panel re-point energy, STEPPING", "={napc}*{nstp}*{es}/1000", "J", "N·steps·E_step",
    1093 * 10 * 2.682 / 1000, "idle power between re-points is ZERO")
der("apef", "Panel re-point energy, FULL-drive", "={napc}*{nstp}*{Ix}^2*{rc}*{tp}/1000", "J",
    "N·steps·Ic*²·R·t", 1093 * 10 * 3.35**2 * 0.5518 * 1.5 / 1000, "the worst-case bound")

part_banner("PART IV — HOW TO MANUFACTURE (mirrors report Part IV)")

# ============================ 15 · manufacture numbers =======================
banner("15 · MANUFACTURE NUMBERS — the quantities that pick the process (report §19–§22)")
note("Actuator: toothed parts are micro-MIM ONLY (pressed SMC minimum section 0.8–1.7 mm vs 232 µm "
     "teeth); the 20 µm gap is SET AT ASSEMBLY (force-gauged), not moulded. Cost $0.10–0.25/actuator "
     "at ≈10 M/yr, process-dominated (raw materials $0.0014). Cells: print+plate (prototype) → "
     "electroform (reference) → two-half or thick-wall moulding (volume). Builds A/B/C: report §22.")
head()
inp("walt", "Thickened-wall option for moulding", 0.30, "mm",
    "the §12-proven manufacturability lever — interior stays 3.10")
der("masp", "Mould aspect ratio, as-drawn wall", "={dhc}/{thc}", "-", "depth/wall",
    7.75 / 0.15, "51.7:1 — NOT mouldable single-piece (thin-wall limit ≈15–30:1)")
der("maspt", "Mould aspect ratio, thickened wall", "={dhc}/{walt}", "-", "depth/wall",
    7.75 / 0.30, "25.8:1 — standard thin-wall class; the wall-thickness proof makes this free RF-wise")
der("pitcht", "Pitch at the thickened wall", "={af}+{walt}", "mm", "interior + wall",
    3.1 + 0.30, "array-level cost: broadside grating limit falls 92.2 → 88.2 GHz (Vlad's budget)")
der("skint", "Plating floor at band bottom (10δ at 57.5 GHz)",
    "=10*SQRT({rhop}/(PI()*57.5*1000000000*4*PI()*0.0000001))*1000000", "µm", "10·δ(57.5)",
    10 * math.sqrt(3.0e-8 / (math.pi * 57.5e9 * 4e-7 * math.pi)) * 1e6,
    "≈3.6 µm worst-case plated Cu — why the spec floors at 3–5 µm")

part_banner("PART V — WHO CAN MANUFACTURE + GLOSSARY (contacts live in report Part V)")

# ============================ 16 · glossary + where next =====================
banner("16 · GLOSSARY + WHERE TO GO NEXT")
note("Suppliers with live-verified contacts are in the REPORT, not here: actuator §23 (ten entries), "
     "cells §24 (fourteen entries), RFQ annexes A–E + draft emails §25. Recommended first RFQs: "
     "Thomas Keating + Vitesse (electroformed reference tiles), BMF + Cybershield/SAT (print-and-"
     "plate pair), SWISSto12 (integrated), Hexcel/Plascore slice (week-one bench experiment).")
gloss = [
    ("AF (across-flats)", "hexagon width between opposite flat walls; the INTERIOR AF is the RF dimension"),
    ("Fd", "detent force target = 5·g·Mt — the zero-power holding force the spec demands"),
    ("Mt", "translator mass (0.1577 g) — the moving toothed bar"),
    ("Pm", "permanent-magnet length along the bridge (the trim knob; fixed design 0.243 mm)"),
    ("MMF", "magnetomotive force (ampere-turns) driving flux around the magnetic circuit"),
    ("SMC", "soft magnetic composite — pressed iron powder; CANNOT form 232 µm teeth (why micro-MIM)"),
    ("micro-MIM", "micro metal injection moulding — the process family for the toothed steel parts"),
    ("N42 / N52", "NdFeB magnet grades (Br 1.30 / 1.45 T); N52 = more force, less demag headroom"),
    ("λg (guide wavelength)", "wavelength INSIDE the cell, λ0/√(1−(λ0/λc)²) — sets stroke need λg/2"),
    ("fc (cutoff)", "lowest frequency the cell propagates (53.56 GHz); set ONLY by the interior AF"),
    ("registration", "alignment of the three poles' tooth patterns (as-drawn 374 µm vs exact ⅓-pitch)"),
    ("basin", "a stable detent well; a variant with ONE basin cannot step — the gap-40 kill"),
    ("DUAL drive", "pull the target pole AND cancel the holding pole with reverse current"),
    ("demag gate", "open check: reverse cancel-current vs the magnet's knee — BLOCKS dual drive"),
    ("skin depth (δ)", "how far fields penetrate a conductor (≈0.3 µm here); wall needs ≥≈10δ"),
    ("LPI", "low probability of intercept — the zero-emission-hold property of this aperture"),
    ("VCM", "voice-coil motor (hard-disk actuator) — evidence class for µm-assembled clearances"),
    ("RFQ", "request for quotation — the annex specs Tristan sends to fabricators"),
]
for term, defn in gloss:
    r = _next()
    ws.cell(row=r, column=1, value=term).font = HDR
    c = ws.cell(row=r, column=6, value=defn)
    c.font = SMALL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

for col, w in (("A", 36), ("B", 7), ("C", 14), ("D", 9), ("E", 40), ("F", 62)):
    ws.column_dimensions[col].width = w
for rr in ws.iter_rows(min_col=6, max_col=6):
    for c in rr:
        c.alignment = WRAP
ws.freeze_panes = "A3"

wb.save("out/PHANTM-CALC.xlsx")
print(f"wrote out/PHANTM-CALC.xlsx — {ws.max_row} rows, no defined names")
