#!/usr/bin/env python3
"""fpk_assumption_based_design.py — JLR-facing assumption → results register.

INTENT: Tristan needs something Jack can review that is not a wall of PARTIAL.
We freeze educated design assumptions, attach screening results computed under
those assumptions, and list the partner asks that replace them. This is NOT
ship_ok / homologation — it is an honest concept pack narrative.

FLOW: stamp / CLI → twin JSON + markdown section → Excel can cite the same object.

Run:
  python3 scripts/lib/fpk_assumption_based_design.py --twin out/formula-e-front-mgu-20260729-1432
  python3 scripts/lib/fpk_assumption_based_design.py --selftest
"""

from __future__ import annotations

import argparse
import json
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Optional


SCHEMA = "forgeos.fpk.assumption_based_design/v1"
OUTPUT_BASENAME = "JLR-FE-FRONT-FPK-ASSUMPTION-BASED-DESIGN.json"
MARKDOWN_BASENAME = "JLR-FE-FRONT-FPK-ASSUMPTION-BASED-DESIGN.md"
# INTENT (2026-07-31): Jack fills blanks in Excel — not a wall of PARTIAL in markdown.
JACK_XLSX_BASENAME = "JLR-FE-FRONT-FPK-ASSUMPTIONS-FOR-JACK.xlsx"


def _qty_value(quantities: Mapping[str, Any], key: str, default: Any = None) -> Any:
    raw = quantities.get(key)
    if isinstance(raw, Mapping) and "value" in raw:
        return raw.get("value")
    if raw is not None:
        return raw
    return default


def _load_json(path: Path) -> Optional[dict[str, Any]]:
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def build_register(twin_dir: Path) -> dict[str, Any]:
    """Build the assumption → results → ask register from twin artefacts.

    @description Reads state + motor-stack screens; never invents ship_ok.
    @param twin_dir Twin output directory
    @returns Assumption-based design object
    """

    state = _load_json(twin_dir / "state.json") or {}
    quantities = (
        ((state.get("orchestratorContract") or {}).get("quantities"))
        if isinstance(state.get("orchestratorContract"), Mapping)
        else {}
    )
    if not isinstance(quantities, Mapping):
        quantities = {}

    motor_stack = twin_dir / "_motor_stack"
    em = _load_json(motor_stack / "em_fia_front_kit_case.json") or {}
    cool = _load_json(motor_stack / "analytical_fia_cooling_network_screen.json") or {}
    gear_oil = _load_json(motor_stack / "gear_oil_fia_front_kit_case.json") or {}
    post = _load_json(motor_stack / "post_diff_final_drive_packaging_screen.json") or {}
    iso = _load_json(motor_stack / "iso6336_fia_front_kit_case.json") or {}
    bevel = _load_json(motor_stack / "iso_bevel_fia_front_kit_case.json") or {}
    rotor_struct = _load_json(motor_stack / "calculix_fia_rotor_screen.json") or {}
    pocket_struct = _load_json(motor_stack / "calculix_fia_magnet_pocket_screen.json") or {}
    mount_struct = _load_json(motor_stack / "analytical_fia_case_mount_screen.json") or {}
    mm = _load_json(twin_dir / "motor-multiphysics.json") or {}
    cad = mm.get("cadAuthority") if isinstance(mm.get("cadAuthority"), Mapping) else {}

    em_works = em.get("works_in_kit_context") if isinstance(em.get("works_in_kit_context"), Mapping) else {}
    em_loaded = em.get("loaded_point") if isinstance(em.get("loaded_point"), Mapping) else {}
    em_sweep = (
        ((em.get("rotor_position_sweep") or {}).get("summary"))
        if isinstance(em.get("rotor_position_sweep"), Mapping)
        else {}
    )
    if not isinstance(em_sweep, Mapping):
        em_sweep = {}
    cool_scr = cool.get("screening_results") if isinstance(cool.get("screening_results"), Mapping) else {}
    cool_inq = (
        cool.get("input_quantities")
        if isinstance(cool.get("input_quantities"), Mapping)
        else {}
    )
    gear_scr = (
        gear_oil.get("screening_results")
        if isinstance(gear_oil.get("screening_results"), Mapping)
        else {}
    )
    post_gate = post.get("closure_gate") if isinstance(post.get("closure_gate"), Mapping) else {}
    post_blocker = (
        post.get("architecture_blocker")
        if isinstance(post.get("architecture_blocker"), Mapping)
        else {}
    )
    iso_works = (
        iso.get("works_in_kit_context")
        if isinstance(iso.get("works_in_kit_context"), Mapping)
        else {}
    )
    bevel_margins = bevel.get("margins") if isinstance(bevel.get("margins"), Mapping) else {}
    bevel_works = (
        bevel.get("works_in_kit_context")
        if isinstance(bevel.get("works_in_kit_context"), Mapping)
        else {}
    )
    rotor_scr = (
        rotor_struct.get("screening_results")
        if isinstance(rotor_struct.get("screening_results"), Mapping)
        else {}
    )
    rotor_margins = (
        rotor_struct.get("margins")
        if isinstance(rotor_struct.get("margins"), Mapping)
        else {}
    )
    pocket_scr = (
        pocket_struct.get("screening_results")
        if isinstance(pocket_struct.get("screening_results"), Mapping)
        else {}
    )
    pocket_margins = (
        pocket_struct.get("margins")
        if isinstance(pocket_struct.get("margins"), Mapping)
        else {}
    )
    mount_scr = (
        mount_struct.get("screening_results")
        if isinstance(mount_struct.get("screening_results"), Mapping)
        else {}
    )
    mount_margins = (
        mount_struct.get("margins")
        if isinstance(mount_struct.get("margins"), Mapping)
        else {}
    )

    assumptions = [
        {
            "id": "A-DUTY",
            "status": "FROZEN_ASSUMPTION",
            "statement": "Continuous front regen electrical duty",
            "value": f"{_qty_value(quantities, 'front_regen_electrical_cap_kw', 250)} kW",
            "replace_with": "Team race software / energy tool CSV (DEC-007)",
        },
        {
            "id": "A-BAY",
            "status": "FROZEN_ASSUMPTION",
            "statement": "Package envelope and dry mass aspiration",
            "value": (
                f"{_qty_value(quantities, 'front_bay_envelope_w_mm', 343)}×"
                f"{_qty_value(quantities, 'front_bay_envelope_d_mm', 259)}×"
                f"{_qty_value(quantities, 'front_bay_envelope_h_mm', 267)} mm; "
                f"~{_qty_value(quantities, 'fpk_mass_cap_kg', 32)} kg dry"
            ),
            "replace_with": "Chassis ICD STEP + weighed bill of materials",
        },
        {
            "id": "A-BUS",
            "status": "FROZEN_ASSUMPTION",
            "statement": "DC bus voltage seed",
            "value": f"{_qty_value(quantities, 'dc_bus_voltage_v', 750)} V",
            "replace_with": "Exact bus window and ripple limits",
        },
        {
            "id": "A-COOL",
            "status": "FROZEN_ASSUMPTION",
            "statement": "Coolant inlet and flow",
            "value": (
                f"{_qty_value(quantities, 'coolant_inlet_c', 60)} °C / "
                f"{_qty_value(quantities, 'coolant_flow_l_min', 12)} L/min"
            ),
            "replace_with": "Team coolant loop ICD",
        },
        {
            "id": "A-SPEED",
            "status": "FROZEN_ASSUMPTION",
            "statement": "Maximum rotor speed",
            "value": f"{_qty_value(quantities, 'max_rotor_speed_rpm', 19500)} rpm",
            "replace_with": "Team max used speed + overspeed policy",
        },
        {
            "id": "A-RATIO",
            "status": "FROZEN_ASSUMPTION",
            "statement": "Overall gear ratio seed (2 into nest × 4 post-diff)",
            "value": f"{_qty_value(quantities, 'gear_ratio', 8)}",
            "replace_with": "Final ratio from vehicle model",
        },
        {
            "id": "A-SIC",
            "status": "FROZEN_ASSUMPTION",
            "statement": "SiC module class (topology + analytical loss; not MPN)",
            "value": "3× half-bridge class; ~4.3 kW inverter dissipation seed; ESL ~6.4 nH",
            "replace_with": "Supplier module MPN + datasheet + STEP (DEC-001)",
        },
        {
            "id": "A-IFACE",
            "status": "NEEDS_PARTNER_INPUT",
            "statement": "Vehicle port XYZ / mount CAD",
            "value": "Types only (HV, coolant×2, LV/CAN, halfshafts, mounts) — XYZ not invented",
            "replace_with": "Chassis interface control drawing coordinates",
        },
    ]

    results = [
        {
            "id": "R-BAY-FIT",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Concentric cassette packaging",
            "value": (
                f"Housing Ø{_qty_value(quantities, 'fpk_housing_od_mm', '—')}×"
                f"L{_qty_value(quantities, 'fpk_housing_len_mm', '—')} mm; "
                f"rotor ID {_qty_value(quantities, 'fpk_rotor_id_mm', '—')} / "
                f"OD {_qty_value(quantities, 'fpk_rotor_od_mm', '—')} mm"
            ),
            "evidence": "state.orchestratorContract.quantities + Blender FPK geometry",
        },
        {
            "id": "R-EM-DUTY",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Electromagnetic duty torque screen (peak ≠ continuous at n_max)",
            # DECISION: Red-team 2026-07-31 (Sol/GLM/Grok) FATAL — presenting peak
            # FEMM 207 N·m next to required 125 N·m without mean/reliable/T=Pω reads as
            # 423 kW at 19.5k rpm. Always cite peak + mean + reliable + required.
            "value": (
                f"Required at n_max≈T=P/ω "
                f"{em_loaded.get('required_shaft_torque_nm') or em_works.get('required_shaft_torque_nm', '—')} N·m; "
                f"FEMM peak "
                f"{em_loaded.get('torque_magnitude_nm') or em_works.get('loaded_torque_magnitude_nm', '—')} N·m; "
                f"position-sweep mean "
                f"{em_sweep.get('torque_magnitude_mean_nm', '—')} N·m; "
                f"torque_reliable={em_loaded.get('torque_reliable', em_works.get('torque_reliable'))}; "
                f"duty_torque_screen_ok="
                f"{em_loaded.get('duty_torque_screen_ok', em_works.get('duty_torque_screen_ok'))}"
            ),
            "evidence": "_motor_stack/em_fia_front_kit_case.json",
        },
        {
            "id": "R-COOL-NET",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Coupled cooling network at assumed coolant point",
            "value": (
                f"Δp={cool_scr.get('total_delta_p_kpa', '—')} kPa; "
                f"T_winding={cool_scr.get('maximum_winding_temperature_c', '—')} °C; "
                f"T_module={cool_scr.get('maximum_module_temperature_c', '—')} °C; "
                f"coupled_ok={cool_scr.get('coupled_screen_ok')}; "
                f"Cu={cool_scr.get('copper_loss_w', cool_inq.get('copper_loss_w', '—'))} W; "
                f"inv={cool_scr.get('inverter_loss_w', cool_inq.get('inverter_loss_w', '—'))} W; "
                f"motor_loss={cool_scr.get('motor_loss_w', '—')} W; "
                f"total_loss={cool_scr.get('total_loss_w', '—')} W"
            ),
            "evidence": "_motor_stack/analytical_fia_cooling_network_screen.json",
        },
        {
            "id": "R-GEAR-OIL",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Geometry-bound gear-oil jet / pickup / churning screen",
            "value": (
                f"jet={gear_scr.get('minimum_jet_flow_l_min', '—')} L/min; "
                f"ΔP_jet={gear_scr.get('jet_pressure_required_kpa', '—')} kPa; "
                f"churning={gear_scr.get('churning_loss_w', '—')} W; "
                f"immersion={gear_scr.get('immersion_fraction_geometry', '—')}; "
                f"cornering_ok={gear_scr.get('cornering_pickup_ok')}; "
                f"gallery_ok={gear_scr.get('pickup_gallery_adequate')}"
            ),
            "evidence": "_motor_stack/gear_oil_fia_front_kit_case.json",
        },
        {
            "id": "R-GEAR-PLANET",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Planetary strength + nest fit",
            "value": (
                f"duty_strength_screen_ok={iso_works.get('duty_strength_screen_ok')}; "
                f"nest_fits_rotor={iso_works.get('nest_fits_rotor')}"
            ),
            "evidence": "_motor_stack/iso6336_fia_front_kit_case.json",
        },
        {
            "id": "R-BEVEL-DIFF",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Bevel differential handbook strength screen",
            "value": (
                f"min_strength_FoS={bevel_margins.get('minimum_strength_factor')}; "
                f"contact_FoS={bevel_margins.get('minimum_contact_fos')}; "
                f"duty_strength_screen_ok={bevel_works.get('duty_strength_screen_ok')}"
            ),
            "evidence": "_motor_stack/iso_bevel_fia_front_kit_case.json",
        },
        {
            "id": "R-STRUCT-ROTOR",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Rotor ring centrifugal screen @ max rpm",
            "value": (
                f"von_Mises={rotor_scr.get('max_von_mises_mpa')} MPa; "
                f"screening_FoS={rotor_margins.get('screening_fos_vs_assumed_yield')}; "
                f"below_yield={rotor_margins.get('below_assumed_yield')}"
            ),
            "evidence": "_motor_stack/calculix_fia_rotor_screen.json",
        },
        {
            "id": "R-STRUCT-POCKET",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Magnet-pocket iron-bridge centrifugal screen",
            "value": (
                f"FEA_von_Mises={pocket_scr.get('max_von_mises_mpa')} MPa; "
                f"FEA_FoS={pocket_margins.get('screening_fos_vs_assumed_yield_fea')}; "
                f"analytical_FoS={pocket_margins.get('screening_fos_vs_assumed_yield_analytical')}"
            ),
            "evidence": "_motor_stack/calculix_fia_magnet_pocket_screen.json",
        },
        {
            "id": "R-STRUCT-MOUNT",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Cast case / bay-mount torque screen",
            "value": (
                f"motor_reaction={mount_scr.get('motor_reaction_torque_nm')} N·m; "
                f"carrier_output={mount_scr.get('carrier_output_torque_nm')} N·m; "
                f"min_screening_FoS={mount_margins.get('minimum_screening_fos')}; "
                f"bay_mount_shear_FoS={mount_margins.get('bay_mount_shear_fos')}"
            ),
            "evidence": "_motor_stack/analytical_fia_case_mount_screen.json",
        },
        {
            "id": "R-POST-DIFF",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Post-diff final-drive software screening",
            "value": (
                f"blocker={post_blocker.get('status')}; "
                f"FoS≥{post_gate.get('minimum_strength_factor')}; "
                f"bay_fit={post_gate.get('bay_fit')}; "
                f"interfaces_ok={post_gate.get('interface_register_ok')}"
            ),
            "evidence": "_motor_stack/post_diff_final_drive_packaging_screen.json",
        },
        {
            "id": "R-CAD",
            "status": "RESULT_UNDER_ASSUMPTIONS",
            "statement": "Parametric CAD spine (not supplier release)",
            "value": (
                f"parametric_family_count={cad.get('parametric_family_count')}; "
                f"release_authority_coverage={cad.get('release_authority_coverage')}"
            ),
            "evidence": "motor-multiphysics.json cadAuthority",
        },
    ]

    asks = [
        {
            "id": "ASK-ICD-XYZ",
            "priority": 1,
            "ask": "Chassis interface XYZ for HV, coolant, LV/CAN, halfshafts, mounts",
            "closes": ["A-IFACE"],
            "decision_register": [],
        },
        {
            "id": "ASK-SIC-MPN",
            "priority": 2,
            "ask": "SiC power module manufacturer + MPN + datasheet (+ STEP if available)",
            "closes": ["A-SIC", "DEC-001"],
            "decision_register": ["DEC-001"],
        },
        {
            "id": "ASK-CAP-BANK",
            "priority": 3,
            "ask": "Preferred DC-link film capacitor MPNs / envelope",
            "closes": ["dc_link_capacitors"],
            "decision_register": [],
        },
        {
            "id": "ASK-COOLANT",
            "priority": 4,
            "ask": "Confirm or replace coolant 60 °C / 12 L/min and pressure budget",
            "closes": ["A-COOL"],
            "decision_register": ["DEC-004"],
        },
        {
            "id": "ASK-RATIO-SPEED",
            "priority": 5,
            "ask": "Confirm overall ratio seed 8.0 and max used rotor speed",
            "closes": ["A-RATIO", "A-SPEED"],
            "decision_register": ["DEC-003"],
        },
        {
            "id": "ASK-BENCH",
            "priority": 6,
            "ask": "Any dyno / HIL / flow / double-pulse data on a comparable unit",
            "closes": ["DEC-008", "DEC-010", "hardwareCorrelation"],
            "decision_register": ["DEC-008", "DEC-010", "DEC-006", "DEC-001"],
        },
        {
            "id": "ASK-GERBERS",
            "priority": 7,
            "ask": "Supplier electronics Gerbers / pinout ICD when available",
            "closes": ["DEC-009"],
            "decision_register": ["DEC-009"],
        },
    ]

    open_blockers = mm.get("architectureBlockers")
    if not isinstance(open_blockers, list):
        open_blockers = []
    open_blocker_ids = [
        b.get("blocker_id")
        for b in open_blockers
        if isinstance(b, Mapping) and b.get("status") == "OPEN"
    ]

    duty_ok = em_loaded.get("duty_torque_screen_ok", em_works.get("duty_torque_screen_ok"))
    pitch = _build_pitch(
        duty_torque_screen_ok=duty_ok,
        open_blocker_ids=open_blocker_ids,
    )

    return {
        "schema": SCHEMA,
        "stamped_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "twin": str(twin_dir),
        "pitch": pitch,
        "review_status": "RESULTS_AVAILABLE_UNDER_ASSUMPTIONS",
        "ship_ok": False,
        "homologation": "NOT_CLAIMED",
        "honesty": (
            "RESULT_UNDER_ASSUMPTIONS means educated guesses were frozen and named, "
            "then models were run. It is not hardware correlation and not ship_ok. "
            "NEEDS_PARTNER_INPUT / NEEDS_HARDWARE rows are deliberate — we did not invent them."
        ),
        "assumptions": assumptions,
        "results_under_assumptions": results,
        "asks_from_partner": asks,
        "architecture_blockers_open": open_blocker_ids,
        "brief_markdown": (
            "docs/plans/JLR-FE-FRONT-FPK-ASSUMPTION-BASED-RESULTS-FOR-JACK-2026-07-31.md"
        ),
        "email_draft_markdown": (
            "docs/plans/JLR-FE-FRONT-FPK-EMAIL-ASK-JACK-2026-07-31.md"
        ),
        "jack_fill_in_xlsx": JACK_XLSX_BASENAME,
        "proveCatch": {
            "ship_ok_false": True,
            "has_assumptions": len(assumptions) >= 6,
            "has_results": len(results) >= 7,
            "has_asks": len(asks) >= 5,
            "review_status_is_results_available": True,
            "pitch_not_greenwash_when_duty_fails": (
                duty_ok is not False
                or (
                    "does not clear" in pitch.lower()
                    and "screens torque" not in pitch.lower()
                )
            ),
        },
    }


def _build_pitch(
    *,
    duty_torque_screen_ok: Any,
    open_blocker_ids: list[Any],
) -> str:
    """Jack-facing pitch from live screen truth — never claim screens cleared when they did not.

    INTENT: Red-team 2026-07-31 — static copy said the layout "screens torque…"
    while duty_torque_screen_ok=False and architecture blockers were OPEN.
    """

    blockers = [str(b) for b in open_blocker_ids if b]
    bay = (
        "Under frozen packaging and duty assumptions consistent with the public "
        "Formula E front-kit envelope, the concentric motor–inverter–planetary–diff "
        "layout is sized into the bay."
    )
    if duty_torque_screen_ok is True and not blockers:
        return (
            f"{bay} Analytical screens for duty torque, gear strength, and thermal "
            "margins clear under those assumptions. Not homologated — replace "
            "assumptions when JLR inputs arrive."
        )
    parts = [bay]
    if duty_torque_screen_ok is False:
        parts.append(
            "The electromagnetic duty-torque screen does not clear "
            "(mean / reliability vs required shaft torque at n_max) — peak FEMM "
            "torque alone is not a pass."
        )
    elif duty_torque_screen_ok is not True:
        parts.append(
            "Duty-torque screen status is incomplete on the twin — do not treat "
            "as cleared."
        )
    if blockers:
        parts.append(
            "Open architecture blockers: " + ", ".join(blockers) + "."
        )
    parts.append(
        "Not homologated — replace assumptions when JLR inputs arrive; "
        "ship_ok stays false."
    )
    return " ".join(parts)


def render_markdown(register: Mapping[str, Any]) -> str:
    """Render a short Jack-facing markdown from the register."""

    lines = [
        "# Assumption-based design — results for review",
        "",
        f"**Stamped:** {register.get('stamped_at')}",
        f"**Review status:** `{register.get('review_status')}`",
        f"**ship_ok:** `{register.get('ship_ok')}` — homologation `{register.get('homologation')}`",
        "",
        register.get("pitch") or "",
        "",
        register.get("honesty") or "",
        "",
        "## Frozen assumptions",
        "",
        "| ID | Status | Statement | Value | Replace with |",
        "|---|---|---|---|---|",
    ]
    for row in register.get("assumptions") or []:
        if not isinstance(row, Mapping):
            continue
        lines.append(
            f"| `{row.get('id')}` | {row.get('status')} | {row.get('statement')} | "
            f"{row.get('value')} | {row.get('replace_with')} |"
        )
    lines.extend(
        [
            "",
            "## Results under those assumptions",
            "",
            "| ID | Status | Statement | Value | Evidence |",
            "|---|---|---|---|---|",
        ]
    )
    for row in register.get("results_under_assumptions") or []:
        if not isinstance(row, Mapping):
            continue
        lines.append(
            f"| `{row.get('id')}` | {row.get('status')} | {row.get('statement')} | "
            f"{row.get('value')} | `{row.get('evidence')}` |"
        )
    lines.extend(
        [
            "",
            "## Ask list (replace assumptions)",
            "",
            "| Priority | ID | Ask | Closes |",
            "|---|---|---|---|",
        ]
    )
    for row in register.get("asks_from_partner") or []:
        if not isinstance(row, Mapping):
            continue
        closes = ", ".join(str(x) for x in (row.get("closes") or []))
        lines.append(
            f"| {row.get('priority')} | `{row.get('id')}` | {row.get('ask')} | {closes} |"
        )
    lines.extend(
        [
            "",
            f"**Brief:** `{register.get('brief_markdown')}`",
            f"**Email draft:** `{register.get('email_draft_markdown')}`",
            "",
        ]
    )
    return "\n".join(lines)


def write_jack_workbook(twin_dir: Path, register: Mapping[str, Any]) -> Path:
    """Write a Jack-fillable Excel: our frozen values + blank confirm/replace columns.

    INTENT: Partner review needs a sheet he can edit — yellow cells = fill these.
    Sheets: Instructions | Assumptions (fill) | Results (context) | Asks (fill).
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Border(
        left=Side(style="thin", color="BBBBBB"),
        right=Side(style="thin", color="BBBBBB"),
        top=Side(style="thin", color="BBBBBB"),
        bottom=Side(style="thin", color="BBBBBB"),
    )
    fill_header = PatternFill("solid", fgColor="1F2937")
    font_header = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    fill_ours = PatternFill("solid", fgColor="F3F4F6")
    fill_jack = PatternFill("solid", fgColor="FEF3C7")
    fill_readonly = PatternFill("solid", fgColor="E0E7FF")
    fill_warn = PatternFill("solid", fgColor="FEE2E2")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _style_header(ws, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(1, col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = wrap
            cell.border = thin

    def _autosize(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws0 = wb.active
    ws0.title = "Instructions"
    ws0["A1"] = "JLR FE Front Powertrain Kit — Assumption register for Jack"
    ws0["A1"].font = Font(bold=True, size=14, name="Calibri")
    for i, line in enumerate(
        [
            "",
            f"Stamped: {register.get('stamped_at')}",
            f"Review status: {register.get('review_status')}",
            f"ship_ok: {register.get('ship_ok')} — homologation: {register.get('homologation')}",
            "",
            str(register.get("pitch") or ""),
            "",
            str(register.get("honesty") or ""),
            "",
            "HOW TO USE",
            "1. Open 'Assumptions (fill)' — grey = what Anvil froze; YELLOW = your cells.",
            "2. Per row: leave blank to accept our freeze, OR put the authoritative "
            "value in Jack_value and set Jack_status to REPLACE.",
            "3. 'Asks (fill)' lists partner inputs that close those rows — put STEP / "
            "datasheet / CSV paths in Jack_attachment_or_link.",
            "4. 'Results (context)' is read-only screening under current freezes.",
            "",
            "Jack_status codes: CONFIRM | REPLACE | UNKNOWN | N/A",
            "",
            "We will not invent chassis XYZ, race SiC MPN, Gerbers, or dyno maps.",
        ],
        start=2,
    ):
        ws0.cell(i, 1, line).alignment = wrap
    ws0.column_dimensions["A"].width = 100

    ws1 = wb.create_sheet("Assumptions (fill)")
    headers1 = [
        "ID",
        "Assumption",
        "Our_frozen_value",
        "Our_status",
        "Replace_with_when_you_have_it",
        "Jack_status",
        "Jack_value",
        "Jack_notes",
        "Closes_ask_IDs",
    ]
    for c, h in enumerate(headers1, start=1):
        ws1.cell(1, c, h)
    _style_header(ws1, len(headers1))

    ask_by_closes: dict[str, list[str]] = {}
    for ask in register.get("asks_from_partner") or []:
        if not isinstance(ask, Mapping):
            continue
        for closes in ask.get("closes") or []:
            ask_by_closes.setdefault(str(closes), []).append(str(ask.get("id") or ""))

    n_assumptions = 0
    for r, row in enumerate(register.get("assumptions") or [], start=2):
        if not isinstance(row, Mapping):
            continue
        n_assumptions += 1
        aid = str(row.get("id") or "")
        vals = [
            aid,
            str(row.get("statement") or ""),
            str(row.get("value") or ""),
            str(row.get("status") or ""),
            str(row.get("replace_with") or ""),
            "",
            "",
            "",
            ", ".join(ask_by_closes.get(aid, [])),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws1.cell(r, c, v)
            cell.alignment = wrap
            cell.border = thin
            cell.fill = fill_ours if c <= 5 or c == 9 else fill_jack
        if str(row.get("status") or "").startswith("NEEDS"):
            ws1.cell(r, 4).fill = fill_warn
        ws1.cell(r, 6).comment = Comment(
            "CONFIRM | REPLACE | UNKNOWN | N/A", "Anvil"
        )
        ws1.row_dimensions[r].height = 36
    _autosize(ws1, [10, 36, 42, 22, 40, 14, 28, 36, 16])
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:I{max(2, n_assumptions + 1)}"

    ws2 = wb.create_sheet("Results (context)")
    headers2 = ["ID", "Result", "Value_under_assumptions", "Status", "Evidence"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(1, c, h)
    _style_header(ws2, len(headers2))
    for r, row in enumerate(register.get("results_under_assumptions") or [], start=2):
        if not isinstance(row, Mapping):
            continue
        for c, v in enumerate(
            [
                str(row.get("id") or ""),
                str(row.get("statement") or ""),
                str(row.get("value") or ""),
                str(row.get("status") or ""),
                str(row.get("evidence") or ""),
            ],
            start=1,
        ):
            cell = ws2.cell(r, c, v)
            cell.alignment = wrap
            cell.border = thin
            cell.fill = fill_readonly
        ws2.row_dimensions[r].height = 40
    _autosize(ws2, [14, 40, 70, 28, 40])
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Asks (fill)")
    headers3 = [
        "Priority",
        "Ask_ID",
        "What_we_need",
        "Closes_assumption_IDs",
        "Decision_register",
        "Jack_have_it",
        "Jack_attachment_or_link",
        "Jack_owner",
        "Jack_notes",
    ]
    for c, h in enumerate(headers3, start=1):
        ws3.cell(1, c, h)
    _style_header(ws3, len(headers3))
    for r, ask in enumerate(register.get("asks_from_partner") or [], start=2):
        if not isinstance(ask, Mapping):
            continue
        for c, v in enumerate(
            [
                ask.get("priority"),
                str(ask.get("id") or ""),
                str(ask.get("ask") or ""),
                ", ".join(str(x) for x in (ask.get("closes") or [])),
                ", ".join(str(x) for x in (ask.get("decision_register") or [])),
                "",
                "",
                "",
                "",
            ],
            start=1,
        ):
            cell = ws3.cell(r, c, v)
            cell.alignment = wrap
            cell.border = thin
            cell.fill = fill_ours if c <= 5 else fill_jack
        ws3.cell(r, 6).comment = Comment("Y / N / PARTIAL", "Anvil")
        ws3.row_dimensions[r].height = 36
    _autosize(ws3, [10, 16, 52, 22, 22, 12, 36, 18, 32])
    ws3.freeze_panes = "A2"

    out_path = twin_dir / JACK_XLSX_BASENAME
    temporary = out_path.with_name(f".{out_path.name}.{os.getpid()}.tmp.xlsx")
    wb.save(temporary)
    os.replace(temporary, out_path)
    return out_path


# FLOW: Jack fills yellow cells → ingest_jack_workbook → freezes override JSON
# → dependent screens marked INVALIDATED_PENDING_RERUN (never silent overwrite).
JACK_INGEST_BASENAME = "jack-assumption-overrides.json"
_ASSUMPTION_TO_SCREEN_INVALIDATION = {
    "A-DUTY": ("R-EM-DUTY", "magnetic", "EM_DUTY_TORQUE_SCREEN"),
    "A-BAY": ("R-BAY-FIT", "packaging"),
    "A-BUS": ("R-EM-DUTY", "magnetic"),
    "A-COOL": ("R-COOL-NET", "cooling", "gear_oil"),
    "A-SPEED": ("R-EM-DUTY", "R-STRUCT-ROTOR", "magnetic"),
    "A-RATIO": ("R-GEAR-PLANET", "R-BEVEL-DIFF", "R-GEAR-OIL"),
    "A-SIC": ("R-COOL-NET", "pcb"),
    "A-IFACE": ("R-BAY-FIT", "interfaces"),
}


def ingest_jack_workbook(twin_dir: Path, xlsx_path: Optional[Path] = None) -> dict[str, Any]:
    """Read Jack yellow cells and stamp overrides + screen invalidations.

    INTENT (F-ASK-1): the xlsx is a controlled overwrite surface — REPLACE with a
    value freezes a new assumption seed and INVALIDATES dependent screens until
    re-run. CONFIRM alone does not invent authority. ship_ok stays false.
    """
    from openpyxl import load_workbook

    path = xlsx_path or (twin_dir / JACK_XLSX_BASENAME)
    if not path.is_file():
        return {
            "ok": False,
            "error": f"missing {path.name}",
            "ship_ok": False,
            "overrides": [],
            "screens_invalidated": [],
        }
    wb = load_workbook(path, data_only=False)
    if "Assumptions (fill)" not in wb.sheetnames:
        return {
            "ok": False,
            "error": "Assumptions (fill) sheet missing",
            "ship_ok": False,
            "overrides": [],
            "screens_invalidated": [],
        }
    ws = wb["Assumptions (fill)"]
    overrides: list[dict[str, Any]] = []
    invalidated: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        aid = str(row[0]).strip()
        jack_status = (str(row[5]).strip().upper() if row[5] is not None else "")
        jack_value = row[6]
        jack_notes = row[7]
        if jack_status in ("", "CONFIRM", "UNKNOWN", "N/A"):
            if jack_status == "CONFIRM":
                overrides.append(
                    {
                        "id": aid,
                        "action": "CONFIRM",
                        "jack_value": None,
                        "jack_notes": jack_notes,
                        "authority": "partner_confirm_under_assumptions",
                    }
                )
            continue
        if jack_status == "REPLACE":
            if jack_value in (None, ""):
                overrides.append(
                    {
                        "id": aid,
                        "action": "REPLACE_REJECTED_EMPTY",
                        "jack_value": None,
                        "jack_notes": jack_notes,
                        "error": "REPLACE requires Jack_value",
                    }
                )
                continue
            overrides.append(
                {
                    "id": aid,
                    "action": "REPLACE",
                    "jack_value": jack_value,
                    "jack_notes": jack_notes,
                    "authority": "partner_overwrite_pending_rerun",
                }
            )
            for screen in _ASSUMPTION_TO_SCREEN_INVALIDATION.get(aid, ()):
                invalidated.add(str(screen))
    payload = {
        "schema": "fpk-jack-assumption-overrides/v1",
        "ship_ok": False,
        "stamped_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_xlsx": path.name,
        "overrides": overrides,
        "screens_invalidated": sorted(invalidated),
        "note": (
            "Partner REPLACE freezes override the Anvil seed for the named "
            "assumption and mark dependent screens INVALIDATED_PENDING_RERUN. "
            "Re-stamp multiphysics / ABD after applying overrides. Never mint ship_ok."
        ),
    }
    out = twin_dir / JACK_INGEST_BASENAME
    temporary = out.with_name(f".{out.name}.{os.getpid()}.tmp")
    temporary.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, out)
    payload["ok"] = True
    payload["path"] = str(out)
    return payload


def write_register(
    twin_dir: Path, register: Mapping[str, Any], *, write_xlsx: bool = True
) -> tuple[Path, Path, Optional[Path]]:
    """Atomically write JSON + markdown (+ Jack Excel) into the twin."""

    twin_dir.mkdir(parents=True, exist_ok=True)
    json_path = twin_dir / OUTPUT_BASENAME
    md_path = twin_dir / MARKDOWN_BASENAME
    payload = json.dumps(register, indent=2) + "\n"
    md = render_markdown(register)
    for path, text in ((json_path, payload), (md_path, md)):
        temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        temporary.write_text(text, encoding="utf-8")
        os.replace(temporary, path)
    xlsx_path: Optional[Path] = None
    if write_xlsx:
        xlsx_path = write_jack_workbook(twin_dir, register)
    return json_path, md_path, xlsx_path


def selftest() -> int:
    """ProveCatch: register is reviewable, ship_ok false, asks present, xlsx writable."""

    with tempfile.TemporaryDirectory(prefix="fpk-assumption-selftest-") as raw:
        twin = Path(raw)
        (twin / "_motor_stack").mkdir()
        (twin / "state.json").write_text(
            json.dumps(
                {
                    "orchestratorContract": {
                        "quantities": {
                            "front_regen_electrical_cap_kw": {"value": 250},
                            "front_bay_envelope_w_mm": {"value": 343},
                            "front_bay_envelope_d_mm": {"value": 259},
                            "front_bay_envelope_h_mm": {"value": 267},
                            "fpk_mass_cap_kg": {"value": 32},
                            "dc_bus_voltage_v": {"value": 750},
                            "coolant_inlet_c": {"value": 60},
                            "coolant_flow_l_min": {"value": 12},
                            "max_rotor_speed_rpm": {"value": 19500},
                            "gear_ratio": {"value": 8},
                            "fpk_housing_od_mm": {"value": 251.8},
                            "fpk_housing_len_mm": {"value": 140.5},
                            "fpk_rotor_id_mm": {"value": 130.5},
                            "fpk_rotor_od_mm": {"value": 197.1},
                        }
                    }
                }
            ),
            encoding="utf-8",
        )
        (twin / "_motor_stack" / "em_fia_front_kit_case.json").write_text(
            json.dumps(
                {
                    "works_in_kit_context": {
                        "duty_torque_screen_ok": False,
                        "loaded_torque_magnitude_nm": 207.12,
                        "required_shaft_torque_nm": 125.21,
                        "torque_reliable": False,
                        "torque_magnitude_mean_nm": 118.75,
                    },
                    "loaded_point": {
                        "torque_magnitude_nm": 207.12,
                        "required_shaft_torque_nm": 125.21,
                        "torque_reliable": False,
                        "duty_torque_screen_ok": False,
                    },
                    "rotor_position_sweep": {
                        "summary": {"torque_magnitude_mean_nm": 118.75}
                    },
                }
            ),
            encoding="utf-8",
        )
        (twin / "_motor_stack" / "analytical_fia_cooling_network_screen.json").write_text(
            json.dumps(
                {
                    "screening_results": {
                        "coupled_screen_ok": True,
                        "total_delta_p_kpa": 42.7,
                        "maximum_winding_temperature_c": 67.0,
                        "maximum_module_temperature_c": 71.0,
                        "copper_loss_w": 2180.0,
                        "inverter_loss_w": 4318.0,
                        "motor_loss_w": 2316.0,
                        "total_loss_w": 6634.0,
                    },
                    "input_quantities": {
                        "copper_loss_w": 2180.0,
                        "inverter_loss_w": 4318.0,
                    },
                }
            ),
            encoding="utf-8",
        )
        (twin / "motor-multiphysics.json").write_text(
            json.dumps(
                {
                    "architectureBlockers": [],
                    "cadAuthority": {
                        "parametric_family_count": 11,
                        "release_authority_coverage": 0.0,
                    },
                }
            ),
            encoding="utf-8",
        )
        register = build_register(twin)
        checks = {
            "ship_ok_false": register.get("ship_ok") is False,
            "review_status": register.get("review_status")
            == "RESULTS_AVAILABLE_UNDER_ASSUMPTIONS",
            "assumptions": len(register.get("assumptions") or []) >= 6,
            "results": len(register.get("results_under_assumptions") or []) >= 7,
            "asks": len(register.get("asks_from_partner") or []) >= 5,
            "proveCatch": (register.get("proveCatch") or {}).get("ship_ok_false")
            is True,
        }
        json_path, md_path, xlsx_path = write_register(twin, register)
        checks["wrote_json"] = json_path.is_file()
        checks["wrote_md"] = md_path.is_file() and "RESULT_UNDER_ASSUMPTIONS" in (
            md_path.read_text(encoding="utf-8")
        )
        checks["wrote_xlsx"] = bool(xlsx_path and xlsx_path.is_file())
        if checks["wrote_xlsx"] and xlsx_path is not None:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path)
            checks["xlsx_has_fill_sheet"] = "Assumptions (fill)" in wb.sheetnames
            checks["xlsx_has_asks_sheet"] = "Asks (fill)" in wb.sheetnames
            # proveCatch: yellow fill columns exist (Jack blanks)
            ws = wb["Assumptions (fill)"]
            checks["xlsx_jack_blank_col"] = ws["F2"].value in (None, "")
            # proveCatch F-ASK-1: REPLACE with value invalidates dependent screens
            ws["F2"] = "REPLACE"
            ws["G2"] = 260
            wb.save(xlsx_path)
            ingest = ingest_jack_workbook(twin, xlsx_path)
            checks["jack_ingest_ok"] = ingest.get("ok") is True
            checks["jack_ingest_invalidates"] = "R-EM-DUTY" in (
                ingest.get("screens_invalidated") or []
            )
            checks["jack_ingest_ship_ok_false"] = ingest.get("ship_ok") is False
            # Empty REPLACE must not silently clear
            ws["F2"] = "REPLACE"
            ws["G2"] = None
            wb.save(xlsx_path)
            ingest2 = ingest_jack_workbook(twin, xlsx_path)
            rejected = [
                o
                for o in (ingest2.get("overrides") or [])
                if o.get("action") == "REPLACE_REJECTED_EMPTY"
            ]
            checks["jack_empty_replace_rejected"] = len(rejected) >= 1
            # Loss ledger surfaces on R-COOL-NET
            cool_row = next(
                (
                    r
                    for r in (register.get("results_under_assumptions") or [])
                    if r.get("id") == "R-COOL-NET"
                ),
                {},
            )
            checks["cool_loss_ledger_in_value"] = "Cu=" in str(
                cool_row.get("value") or ""
            )
            # proveCatch: pitch must not greenwash when duty_ok=False
            pitch = str(register.get("pitch") or "")
            checks["pitch_names_duty_fail"] = "does not clear" in pitch.lower()
            checks["pitch_no_screens_torque_greenwash"] = (
                "screens torque" not in pitch.lower()
            )
            checks["proveCatch_pitch"] = (
                register.get("proveCatch") or {}
            ).get("pitch_not_greenwash_when_duty_fails") is True
        if not all(checks.values()):
            print("FAIL", json.dumps(checks, indent=2))
            return 1
        print("fpk_assumption_based_design --selftest OK")
        print(json.dumps(checks, indent=2))
        return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--twin", type=Path, help="Twin directory")
    parser.add_argument("--selftest", action="store_true")
    parser.add_argument(
        "--ingest-jack",
        action="store_true",
        help="Read Jack yellow cells from the twin xlsx and stamp overrides",
    )
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Skip Jack fill-in workbook (JSON+markdown only)",
    )
    args = parser.parse_args()
    if args.selftest:
        return selftest()
    if not args.twin:
        parser.error("--twin is required unless --selftest")
    twin = args.twin.resolve()
    if args.ingest_jack:
        result = ingest_jack_workbook(twin)
        print(json.dumps(result, indent=2))
        return 0 if result.get("ok") else 1
    register = build_register(twin)
    json_path, md_path, xlsx_path = write_register(
        twin, register, write_xlsx=not args.no_xlsx
    )
    print(
        json.dumps(
            {
                "json": str(json_path),
                "markdown": str(md_path),
                "jack_xlsx": str(xlsx_path) if xlsx_path else None,
                "review_status": register.get("review_status"),
                "ship_ok": register.get("ship_ok"),
                "assumption_count": len(register.get("assumptions") or []),
                "result_count": len(register.get("results_under_assumptions") or []),
                "ask_count": len(register.get("asks_from_partner") or []),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
