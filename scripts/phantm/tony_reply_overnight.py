"""Reply to Tony's overnight note (29 Jul) — clean answers, no internal meta.

PDF + Excel in date-time-subject-version naming. Shows the right engineering
answer to each point he raised — not an acknowledgment of Anvil mistakes.

Run: ~/.venvs/phantm/bin/python tony_reply_overnight.py
"""

from __future__ import annotations

import datetime
import json
import os
import shutil
import subprocess
from importlib.machinery import SourceFileLoader
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "out")
CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
SUBJECT = "PHANTM-TONY-OVERNIGHT-REPLY"

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E8F1FB")
FE_FILL = PatternFill("solid", fgColor="E2EFDA")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def stem() -> str:
    ver = json.load(open(os.path.join(HERE, "version.json")))
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M")
    return f"{stamp}-{SUBJECT}-V{ver['major']}.{ver['minor']}"


def load():
    study = json.load(open(os.path.join(OUT, "tony-phase-material-study.json")))
    rf = json.load(open(os.path.join(OUT, "double-wall-rf-check.json")))
    floq = json.load(open(os.path.join(OUT, "floquet-hex-array.json")))
    return study, rf, floq


def norms(study):
    va = study["variants"]["A_three_phase_3_teeth"]
    vb = study["variants"]["B_four_phase_4_teeth"]
    vc = study["variants"]["C_four_phase_3_teeth"]
    r_ohm, i_a = 3.618, 0.40
    rows = []
    for g in study["gaps_um"]:
        fa = va["summary"][str(g)]["force_at_0_40_a_mn"]
        fb = vb["summary"][str(g)]["force_at_0_40_a_mn"]
        fc = vc["summary"][str(g)]["force_at_0_40_a_mn"]
        foot_a = va["summary"][str(g)]["pole_foot_mm"]
        foot_b = vb["summary"][str(g)]["pole_foot_mm"]
        foot_c = vc["summary"][str(g)]["pole_foot_mm"]
        rows.append(dict(
            gap_um=g,
            fa=fa, fb=fb, fc=fc,
            foot_a=foot_a, foot_b=foot_b, foot_c=foot_c,
            per_coil_a=fa / 3, per_coil_b=fb / 4, per_coil_c=fc / 3,
            per_foot_a=fa / foot_a, per_foot_b=fb / foot_b, per_foot_c=fc / foot_c,
            per_watt_a=fa / (3 * i_a ** 2 * r_ohm),
            per_watt_b=fb / (4 * i_a ** 2 * r_ohm),
            per_watt_c=fc / (3 * i_a ** 2 * r_ohm),
            ba=fb / fa, cb=fc / fb, ca=fc / fa,
        ))
    return rows


def write_md(study, rf, floq, rows) -> str:
    va = study["variants"]["A_three_phase_3_teeth"]
    vb = study["variants"]["B_four_phase_4_teeth"]
    vc = study["variants"]["C_four_phase_3_teeth"]
    mats = study["materials"]
    n60 = next(r for r in rows if r["gap_um"] == 60)
    n20 = next(r for r in rows if r["gap_um"] == 20)

    feco_12 = next(x["force_mn"] for x in mats["all_fe_co_laminated"]
                   if x["current_a"] == 1.2)
    mix_12 = next(x["force_mn"] for x in mats["translator_fe_co__poles_mim"]
                  if x["current_a"] == 1.2)
    mim_12 = next(x["force_mn"] for x in mats["all_mim_fe3si"]
                  if x["current_a"] == 1.2)
    feco_04 = next(x["force_mn"] for x in mats["all_fe_co_laminated"]
                   if x["current_a"] == 0.4)
    mix_04 = next(x["force_mn"] for x in mats["translator_fe_co__poles_mim"]
                  if x["current_a"] == 0.4)

    uu = floq["uniform"]
    dd = floq["one_third_double"]
    cmp_ = floq["comparison"]
    u = rf["uniform_cell"]
    s = rf["model_interior_shrink_by_tape"]
    m = rf["model_interior_shrink_by_tape_over_3"]

    L, A = [], (lambda s: L.append(s))
    A("# PHANTM — answers to your overnight points")
    A("")
    A("Tony — answers against the points you raised overnight, with the "
      "finite-element and periodic-array numbers underneath. Step held at "
      f"**{study['step_um']:.0f} micrometres**, tooth duty **0.401**. "
      "(Reading “50 MHz” as **50 GHz** — the 75/50 = 3/2 wavelength ratio.)")
    A("")
    A("Accompanying Excel workbook has every table in this note.")
    A("")

    # ---- 1 cell fabrication ---------------------------------------------
    A("## 1. Folded / welded strip and cell routes")
    A("")
    A("| Point | Answer |")
    A("|---|---|")
    A("| ~4 mm strip | **Waveguide only** — developed length of a formed "
      "stub, not an actuator blank |")
    A("| Build now | **Individual cells**, each integrated with its "
      "actuator |")
    A("| Continuous waveguide arrays | Only if the lattice can accept "
      "**one-third double walls** — Vlad must clear that |")
    A("| Stacked etched hex-hole sheets | **No** — waveguide internal "
      "surface would not be flat enough |")
    A("| Flat sheets + bond + stretch (HOBE-style) | **No** — each guide "
      "would be two radial halves only bonded; need **continuous metal** "
      "all around the guide inside |")
    A("| Fold + continuous weld/braze on strip | **Live** for individuals "
      "+ actuators; arrays only after Vlad clears one-third double walls |")
    A("")

    # ---- 2 translator / poles -------------------------------------------
    A("## 2. Translator laminations and pole pieces")
    A("")
    A("**Translator.** Stacked laminations are fine if foil thickness × "
      "count lands on the ~**1.2 mm** tooth height (your L_y) after stack "
      "compression — e.g. 12 × 100 µm or 8 × 150 µm. Anneal / stacking-"
      "stress caveats from the worksheet answer still apply.")
    A("")
    A("**Pole pieces.** They are **not prismatic in any dimension**, so a "
      "single lamination outline does not stack into the part. That points "
      "poles to **micro-injection moulding** (or another 3D net-shape soft-"
      "magnetic process). Laminations stay on the translator.")
    A("")
    A("**Saturation with MIM poles.** Relative to laminated iron-cobalt "
      "(the translator route), MIM Fe-3%Si is lower Bsat — your recollection "
      "matches. Coil-only FE at 60 micrometres, three-phase geometry:")
    A("")
    A("| Assignment | Force @ 0.40 A | Force @ 1.20 A | Bridge |B| @ 1.20 A |")
    A("|---|---|---|---|")

    def mat_row(name, label):
        r0 = {x["current_a"]: x for x in mats[name]}
        A(f"| {label} | {r0[0.4]['force_mn']:.2f} mN | "
          f"{r0[1.2]['force_mn']:.2f} mN | "
          f"{r0[1.2]['b_bridge_max_t']:.2f} T |")

    mat_row("all_fe_co_laminated", "All iron-cobalt laminated")
    mat_row("translator_fe_co__poles_mim",
            "**Translator iron-cobalt + MIM poles** (your split)")
    mat_row("all_mim_fe3si", "All MIM Fe-3%Si")
    A("")
    A(f"Mixed (translator iron-cobalt + MIM poles) tracks all-MIM within "
      f"0.4% — **the pole/bridge steel sets the limit**. Versus all iron-"
      f"cobalt the mixed penalty is "
      f"**{(1-mix_04/feco_04)*100:.0f}% at 0.40 A** and "
      f"**{(1-mix_12/feco_12)*100:.0f}% at 1.20 A** "
      f"({mix_12:.1f} vs {feco_12:.1f} mN). Curves are representative; "
      f"vendor B–H for the chosen MIM grade should replace them before "
      f"tooling.")
    A("")

    # ---- 3 four phase ---------------------------------------------------
    A("## 3. Three-phase versus four-phase")
    A("")
    A("Fixed step = pitch / phases = 104 micrometres. Three geometries:")
    A("")
    A("| Variant | Phases | Pitch | Tooth | Teeth / pole | "
      "Pole-foot envelope |")
    A("|---|---|---|---|---|---|")
    for v in (va, vb, vc):
        m_, s_ = v["meta"], v["summary"]["60"]
        A(f"| {m_['label']} | {m_['phases']} | {m_['pitch_um']:.0f} µm | "
          f"{m_['tooth_um']:.0f} µm | {m_['n_pole_teeth']} | "
          f"{s_['pole_foot_mm']:.2f} mm |")
    A("")
    A("Pole-foot envelope = tooth span + 100 micrometre flank allowance — "
      "the local active land, not full stator length (end margins, coils, "
      "leads, mounts still add).")
    A("")
    A("### Force at 0.40 A")
    A("")
    A("| Gap | Three-phase (A) | Four-phase 4-tooth (B) | "
      "Four-phase 3-tooth (C) | B/A | C/B |")
    A("|---|---|---|---|---|---|")
    for r in rows:
        A(f"| **{r['gap_um']} µm** | {r['fa']:.2f} mN | {r['fb']:.2f} mN | "
          f"{r['fc']:.2f} mN | {r['ba']:.2f}× | {r['cb']:.2f}× |")
    A("")
    A("### Same numbers, fairer normalisations")
    A("")
    A("| Gap | A mN/coil | B mN/coil | A mN/mm foot | B mN/mm foot | "
      "C mN/mm foot |")
    A("|---|---|---|---|---|---|")
    for r in rows:
        A(f"| {r['gap_um']} µm | {r['per_coil_a']:.3f} | "
          f"{r['per_coil_b']:.3f} | {r['per_foot_a']:.2f} | "
          f"{r['per_foot_b']:.2f} | {r['per_foot_c']:.2f} |")
    A("")
    A("**Reading against your framing.**")
    A("")
    A(f"1. Wider four-phase teeth **do** raise absolute force at a given "
      f"gap: at 60 micrometres B/A = **{n60['ba']:.2f}×**. That ratio "
      f"equals the coil count (4/3) — force per coil is the same "
      f"({n60['per_coil_a']:.3f} mN/coil). Stator channel count is "
      f"**+33%**; pole-foot envelope grows "
      f"{n60['foot_a']:.2f} → {n60['foot_b']:.2f} mm.")
    A(f"2. Three-tooth poles on the four-phase pitch (your length-"
      f"recovery idea): absolute force drops "
      f"**{(1-n60['cb'])*100:.0f}%** at 60 micrometres (C/B = "
      f"{n60['cb']:.2f}). On **force per millimetre of pole foot**, "
      f"three-tooth is denser "
      f"({n60['per_foot_c']:.2f} vs {n60['per_foot_b']:.2f} mN/mm) — "
      f"about **{(n60['per_foot_c']/n60['per_foot_b']-1)*100:.0f}%**, "
      f"in line with your ~25% length-based estimate.")
    A(f"3. **Gap still dominates.** At 20 micrometres the three variants "
      f"sit near {n20['fa']:.1f}–{n20['fb']:.1f} mN at 0.40 A. Four-phase "
      f"is a help at large gap; it is not a substitute for gap control.")
    A("")
    A("Detent (11.13 mN) reach on the solved current sweeps:")
    A("")
    A("| Gap | Three-phase | Four-phase 4-tooth | Four-phase 3-tooth |")
    A("|---|---|---|---|")
    for g in study["gaps_um"]:
        def fmt(i):
            return "above sweep (need ≥1.4 A class)" if i is None else f"{i:.2f} A"
        A(f"| {g} µm | "
          f"{fmt(va['summary'][str(g)]['i_for_detent_a'])} | "
          f"{fmt(vb['summary'][str(g)]['i_for_detent_a'])} | "
          f"{fmt(vc['summary'][str(g)]['i_for_detent_a'])} |")
    A("")
    A("At 60 micrometres none of the three brackets the detent inside the "
      "1.2 A sweep. At 40 micrometres all three do, near **0.90–0.92 A**.")
    A("")

    # ---- 4 strategy -----------------------------------------------------
    A("## 4. E-band first, 50 GHz as known backoff")
    A("")
    A("Agreed, and how Anvil will prioritise:")
    A("")
    A("1. This actuator was always going to be hard at this scale — "
      "process precision is the wall.")
    A("2. **75 GHz** stays the centre of analysis because it sits ahead "
      "of what electronic phased arrays can do at nearly any cost.")
    A("3. At **~50 GHz**, λ is 1.5× longer → many process issues ease "
      "(~50% easier in your words). That is a **known place to go**, not "
      "a surprise retreat.")
    A("4. **Now:** see whether E-band can be made to work with processes "
      "viable today.")
    A("5. **Later:** if a first customer build needs it, backoff to "
      "50 GHz is allowed — knowing the exclusive next step is still there.")
    A("")

    # ---- 5 double walls / Vlad ------------------------------------------
    A("## 5. One-third double walls — for Vlad")
    A("")
    A("Single-cell Neumann cross-check (across-flats 3.10 mm, 75 µm tape): "
      f"uniform cut-off **{u['fc_ghz']:.2f} GHz**; shrink-by-tape "
      f"+{s['delta_fc_mhz']:.0f} MHz; shrink-by-tape/3 "
      f"+{m['delta_fc_mhz']:.0f} MHz.")
    A("")
    A("**Periodic Floquet–Bloch** on a 2×2 hex supercell (exact rectangle, "
      "Voronoi walls, one orientation at 2× tape):")
    A("")
    A("| Lattice | Cut-off (Gamma) | Mode split after multiplet | "
      "Margin to 75 GHz |")
    A("|---|---|---|---|")
    A(f"| Uniform walls | **{uu['fundamental_cutoff_ghz']:.2f} GHz** | "
      f"{uu['mode_splitting_ghz']:.3f} GHz | "
      f"{uu['margin_at_75_ghz_mhz']:.0f} MHz |")
    A(f"| One-third double walls | **{dd['fundamental_cutoff_ghz']:.2f} GHz** | "
      f"**{dd['mode_splitting_ghz']:.3f} GHz** | "
      f"{dd['margin_at_75_ghz_mhz']:.0f} MHz |")
    A("")
    A(f"Δ fundamental cut-off = **{cmp_['delta_fundamental_mhz']:+.0f} MHz** "
      f"(inside the near-degenerate multiplet width). Guides **remain open** "
      f"at 75 GHz. The clearer in-plane signature is the larger mode "
      f"splitting ({uu['mode_splitting_ghz']:.3f} → "
      f"{dd['mode_splitting_ghz']:.3f} GHz) — coupling anisotropy.")
    A("")
    A("**Still with Vlad before Seed-1:** 3D unit-cell Floquet (scan "
      "impedance, grating lobes, feed transition). This 2D transverse "
      "solve does not replace that.")
    A("")

    # ---- asks -----------------------------------------------------------
    A("## Asks")
    A("")
    A("1. **Vlad:** is a periodic one-third-double-wall hex lattice "
      "acceptable at E-band (guides open above; splitting noted)?")
    A("2. **You:** candidate µ-MIM alloy / vendor B–H for the pole pieces.")
    A("3. Confirm **E-band + processes viable now** as the primary "
      "squeeze, with 50 GHz documented as the scaled backoff.")
    A("")
    A("Tristan")
    A("")
    A("---")
    A("")
    A(f"*FE runtime {study['runtime_s']:.0f} s; Floquet "
      f"{floq.get('runtime_s', '—')} s. Artefacts behind the Excel: "
      "`tony-phase-material-study.json`, `double-wall-rf-check.json`, "
      "`floquet-hex-array.json`.*")
    A("")

    path = os.path.join(OUT, f"{SUBJECT}.md")
    open(path, "w").write("\n".join(L) + "\n")
    print(f"wrote {path}")
    return path


def put(ws, r, c, val, kind="calc", fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = {"input": INPUT_FILL, "calc": CALC_FILL,
                 "fe": FE_FILL}.get(kind, CALC_FILL)
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    return cell


def header(ws, row, *cells):
    for i, txt in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=txt)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)


def write_xlsx(study, rf, floq, rows, name_stem: str) -> str:
    va = study["variants"]["A_three_phase_3_teeth"]
    vb = study["variants"]["B_four_phase_4_teeth"]
    vc = study["variants"]["C_four_phase_3_teeth"]
    mats = study["materials"]
    uu = floq["uniform"]
    dd = floq["one_third_double"]
    cmp_ = floq["comparison"]

    wb = Workbook()
    ws = wb.active
    ws.title = "0 INDEX"
    header(ws, 1, "PDF section", "Sheet", "Contents")
    for i, row in enumerate([
        ("§1 Cell routes", "1 CELL ROUTES", "Strip / kills / live path"),
        ("§2 Translator + poles", "2 MATERIALS", "Laminations / MIM FE"),
        ("§3 Four-phase", "3 GEOMETRY", "A/B/C lock"),
        ("§3 Force", "4 FORCE", "Absolute + per coil + per foot"),
        ("§3 Detent", "5 DETENT", "Current to reach detent"),
        ("§4 Strategy", "6 STRATEGY", "E-band / 50 GHz"),
        ("§5 Double walls", "7 FLOQUET", "Periodic array + single-cell"),
        ("Asks", "8 ASKS", "Vlad / MIM grade / confirm"),
    ], start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    for col, w in zip("ABC", (28, 16, 36)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("1 CELL ROUTES")
    header(ws, 1, "Route", "Status", "Note")
    for i, row in enumerate([
        ("~4 mm strip", "Waveguide only", "Stub length guess, not actuator"),
        ("Individuals + actuators", "NOW", "Build path"),
        ("Array continuous foil", "Vlad gate", "Needs ⅓ double-wall OK"),
        ("Etched hex-hole stack", "NO", "Surface not flat enough"),
        ("HOBE expand / bonded halves", "NO", "Need continuous metal ID"),
        ("Fold + weld/braze strip", "LIVE", "Individuals; arrays after Vlad"),
    ], start=2):
        for j, v in enumerate(row, start=1):
            put(ws, i, j, v, "input" if j == 1 else "fe")
    for col, w in zip("ABC", (32, 16, 40)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("2 MATERIALS")
    header(ws, 1, "assignment", "I_A", "force_mN", "b_bridge_T")
    r = 2
    for name, label in [
        ("all_fe_co_laminated", "all_fe_co_laminated"),
        ("translator_fe_co__poles_mim", "translator_fe_co__poles_mim"),
        ("all_mim_fe3si", "all_mim_fe3si"),
        ("all_somaloy_legacy", "all_somaloy_legacy"),
    ]:
        for row in mats[name]:
            put(ws, r, 1, label, "input")
            put(ws, r, 2, row["current_a"], "fe", "0.00")
            put(ws, r, 3, row["force_mn"], "fe", "0.000")
            put(ws, r, 4, row["b_bridge_max_t"], "fe", "0.000")
            r += 1
    for col, w in zip("ABCD", (36, 10, 12, 12)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("3 GEOMETRY")
    header(ws, 1, "variant", "phases", "pitch_um", "tooth_um",
           "teeth", "pole_foot_mm")
    for i, v in enumerate((va, vb, vc), start=2):
        put(ws, i, 1, v["meta"]["label"], "input")
        put(ws, i, 2, v["meta"]["phases"], "input")
        put(ws, i, 3, v["meta"]["pitch_um"], "input")
        put(ws, i, 4, v["meta"]["tooth_um"], "input")
        put(ws, i, 5, v["meta"]["n_pole_teeth"], "input")
        put(ws, i, 6, v["summary"]["60"]["pole_foot_mm"], "fe", "0.00")
    put(ws, 6, 1, "step_um", "input")
    put(ws, 6, 2, study["step_um"], "input")
    put(ws, 7, 1, "duty", "input")
    put(ws, 7, 2, study["duty"], "input", "0.000")
    put(ws, 8, 1, "pole_foot_definition", "input")
    put(ws, 8, 2, "tooth span + 100 µm flanks (local envelope, not full stator)",
        "calc")
    for col, w in zip("ABCDEF", (28, 10, 12, 12, 10, 14)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("4 FORCE")
    header(ws, 1, "gap_um",
           "F_A_mN", "F_B_mN", "F_C_mN", "B_over_A", "C_over_B",
           "A_mN_per_coil", "B_mN_per_coil", "C_mN_per_coil",
           "A_mN_per_mm", "B_mN_per_mm", "C_mN_per_mm",
           "A_mN_per_W", "B_mN_per_W", "C_mN_per_W")
    for i, r in enumerate(rows, start=2):
        vals = [r["gap_um"], r["fa"], r["fb"], r["fc"], r["ba"], r["cb"],
                r["per_coil_a"], r["per_coil_b"], r["per_coil_c"],
                r["per_foot_a"], r["per_foot_b"], r["per_foot_c"],
                r["per_watt_a"], r["per_watt_b"], r["per_watt_c"]]
        for j, v in enumerate(vals, start=1):
            put(ws, i, j, v, "fe" if j <= 4 else "calc",
                "0.000" if j > 1 else "0")
    for c in range(1, 16):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws = wb.create_sheet("5 DETENT")
    header(ws, 1, "gap_um", "I_A_A", "I_B_A", "I_C_A")
    for i, g in enumerate(study["gaps_um"], start=2):
        put(ws, i, 1, g, "input")
        for j, v in enumerate((va, vb, vc), start=2):
            val = v["summary"][str(g)]["i_for_detent_a"]
            put(ws, i, j, val if val is not None else "above sweep",
                "fe" if val is not None else "calc",
                "0.00" if val is not None else None)
    put(ws, 7, 1, "detent_mN", "input")
    put(ws, 7, 2, study["targets_mn"]["detent"], "input", "0.00")
    for col, w in zip("ABCD", (10, 14, 14, 14)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("6 STRATEGY")
    header(ws, 1, "#", "Statement")
    for i, t in enumerate([
        "Squeeze E-band with processes viable now",
        "75 GHz centre — exclusivity vs electronic phased arrays",
        "50 GHz = known backoff (λ ×1.5), not surprise retreat",
        "First customer build may take 50 GHz; exclusive next step remains",
    ], start=2):
        put(ws, i, 1, i - 1, "input")
        put(ws, i, 2, t, "calc")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 70

    ws = wb.create_sheet("7 FLOQUET")
    header(ws, 1, "lattice", "fc_Gamma_GHz", "split_GHz",
           "margin_75_MHz", "air_fraction")
    put(ws, 2, 1, "uniform", "input")
    put(ws, 2, 2, uu["fundamental_cutoff_ghz"], "fe", "0.000")
    put(ws, 2, 3, uu["mode_splitting_ghz"], "fe", "0.000")
    put(ws, 2, 4, uu["margin_at_75_ghz_mhz"], "fe", "0.0")
    put(ws, 2, 5, uu["meta"]["air_fraction"], "fe", "0.0000")
    put(ws, 3, 1, "one_third_double", "input")
    put(ws, 3, 2, dd["fundamental_cutoff_ghz"], "fe", "0.000")
    put(ws, 3, 3, dd["mode_splitting_ghz"], "fe", "0.000")
    put(ws, 3, 4, dd["margin_at_75_ghz_mhz"], "fe", "0.0")
    put(ws, 3, 5, dd["meta"]["air_fraction"], "fe", "0.0000")
    put(ws, 5, 1, "delta_fc_MHz", "calc")
    put(ws, 5, 2, cmp_["delta_fundamental_mhz"], "calc", "0.0")
    put(ws, 6, 1, "single_cell_uniform_GHz", "calc")
    put(ws, 6, 2, rf["uniform_cell"]["fc_ghz"], "calc", "0.000")
    put(ws, 7, 1, "disclosure", "input")
    put(ws, 7, 2,
        "2D transverse Floquet–Bloch; 3D scan impedance / grating lobes → Vlad",
        "calc")
    for col, w in zip("ABCDE", (24, 14, 12, 14, 14)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("8 ASKS")
    header(ws, 1, "#", "Ask", "Owner")
    for i, row in enumerate([
        (1, "Periodic ⅓-double-wall hex acceptable at E-band?", "Vlad"),
        (2, "Candidate µ-MIM alloy / vendor B–H for poles", "Tony"),
        (3, "Confirm E-band-first; 50 GHz documented backoff", "Tony"),
    ], start=2):
        for j, v in enumerate(row, start=1):
            put(ws, i, j, v, "input" if j != 2 else "calc")
    for col, w in zip("ABC", (6, 56, 12)):
        ws.column_dimensions[col].width = w

    path = os.path.join(OUT, f"{name_stem}.xlsx")
    wb.save(path)
    print(f"wrote {path}")
    return path


def render_pdf(md_path: str, name_stem: str) -> str:
    loader = SourceFileLoader(
        "show_md", os.path.expanduser("~/.claude/scripts/show-md"))
    import importlib.util
    spec = importlib.util.spec_from_loader("show_md", loader)
    show_md = importlib.util.module_from_spec(spec)
    loader.exec_module(show_md)
    html_path = show_md.render(Path(md_path))
    html_stamped = os.path.join(OUT, f"{name_stem}.html")
    shutil.copy2(html_path, html_stamped)
    pdf_path = os.path.join(OUT, f"{name_stem}.pdf")
    subprocess.run(
        [CHROME, "--headless", "--disable-gpu", "--no-pdf-header-footer",
         f"--print-to-pdf={pdf_path}", f"file://{html_stamped}"],
        check=True, capture_output=True, timeout=120,
    )
    print(f"wrote {pdf_path}")
    return pdf_path


def main():
    study, rf, floq = load()
    # GATE: refuse to ship prose that reintroduces absolute-only framing.
    from force_claim_guards import (assert_claim_text_safe, audit_study,
                                    enrich_study_comparisons)
    audit = enrich_study_comparisons(study)
    rows = norms(study)
    md = write_md(study, rf, floq, rows)
    prose = open(md).read()
    viol = assert_claim_text_safe(prose, audit)
    if viol:
        raise SystemExit(
            "force_claim_guards: client prose failed fairness check:\n  - "
            + "\n  - ".join(viol))
    for line in audit["required_client_lines"]:
        print(f"  GUARD (must be reflected in prose): {line}")
    s = stem()
    shutil.copy2(md, os.path.join(OUT, f"{s}.md"))
    xlsx = write_xlsx(study, rf, floq, rows, s)
    pdf = render_pdf(md, s)
    for src in (pdf, xlsx):
        dl = os.path.expanduser(f"~/Downloads/{os.path.basename(src)}")
        shutil.copy2(src, dl)
        print(f"downloads {dl}")
    for stale in Path(os.path.expanduser("~/Downloads")).glob(
            "*PHANTM-TONY-OVERNIGHT-REPLY*-EMAIL.txt"):
        stale.unlink(missing_ok=True)
        print(f"removed stale {stale.name}")


if __name__ == "__main__":
    main()
