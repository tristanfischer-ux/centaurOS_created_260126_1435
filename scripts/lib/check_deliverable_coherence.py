#!/usr/bin/env python3
"""Universal deliverable coherence: twin ↔ workbook ↔ design-pack zip.

⭐ UNIVERSAL (2026-08-03). Three defects this campaign shared one shape — a
customer-facing snapshot that drifted from the twin:

  R1  Excel Checks still showed magnet 159.35 after DEC-008 restamped 83.8
  R2  design-pack zip embedded an older xlsx than the DRAFT workbook beside it
  F1  Decision register froze DEC-008 while twin quantities stayed continuous

This guard fails closed when any of those drift. It does NOT rebuild artefacts;
it tells CI/pipeline the rebuild is required.

Checks:
  1. PACK_EQ_WORKBOOK — latest *-design-pack.zip embedded xlsx SHA-256 equals
     latest *DRAFT*engineering-workbook.xlsx (or dossier.xlsx if no DRAFT).
  2. WORKBOOK_EQ_TWIN_MAGNET — Checks tab brief magnet actual equals
     state.orchestratorContract.quantities.mgu_magnet_temp_c (tol 0.15 °C).
  3. FROZEN_DECISIONS_APPLIED — every FROZEN_UNDER_ASSUMPTION with a registered
     handler reports is_applied (via apply_frozen_decisions.plan).
  4. DRAWING_GATES_STATE — if both state.drawingGates and drawing-gates.json
     exist, all_pass must agree.
  5. DECISION_REGISTER_STATE — every id in 10-decision-register.json appears in
     state.decisionRegister (Excel Decision Register tab reads state only).

Usage:
  check_deliverable_coherence.py --twin <dir> [--enforce]
  check_deliverable_coherence.py --selftest
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Optional

_LIB = Path(__file__).resolve().parent
if str(_LIB) not in sys.path:
    sys.path.insert(0, str(_LIB))

from apply_frozen_decisions import (  # noqa: E402
    decision_register_parity,
    plan as plan_frozen,
)

EXIT_INCOHERENT = 49
SCHEMA = "forgeos.fpk.deliverable_coherence/v1"


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _latest(twin: Path, pattern: str) -> Optional[Path]:
    hits = sorted(twin.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    return hits[0] if hits else None


def _find_workbook(twin: Path) -> Optional[Path]:
    # Prefer newest DRAFT named workbook; fall back to dossier.xlsx
    draft = _latest(twin, "*DRAFT*engineering-workbook.xlsx")
    if draft:
        return draft
    dossier = twin / "dossier.xlsx"
    return dossier if dossier.is_file() else None


def _find_pack(twin: Path) -> Optional[Path]:
    return _latest(twin, "*-design-pack.zip")


def _pack_xlsx_bytes(pack: Path) -> Optional[bytes]:
    with zipfile.ZipFile(pack) as z:
        names = [n for n in z.namelist() if n.endswith(".xlsx") and not n.startswith("__")]
        if not names:
            return None
        # Prefer engineering-workbook name
        names.sort(key=lambda n: (0 if "engineering-workbook" in n else 1, n))
        return z.read(names[0])


def _workbook_magnet_actual(xlsx: Path) -> Optional[float]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — not a real xlsx
        return None
    try:
        if "⚠ Checks" not in wb.sheetnames:
            return None
        ws = wb["⚠ Checks"]
        for row in ws.iter_rows(max_row=250, values_only=True):
            if not row:
                continue
            name = str(row[0] or "")
            if name.startswith("Brief target met: magnet_temp"):
                try:
                    return float(row[1])
                except (TypeError, ValueError):
                    return None
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return None


def _twin_magnet(twin: Path) -> Optional[float]:
    st_path = twin / "state.json"
    if not st_path.is_file():
        return None
    st = json.loads(st_path.read_text(encoding="utf-8"))
    q = ((st.get("orchestratorContract") or {}).get("quantities") or {})
    raw = q.get("mgu_magnet_temp_c") if isinstance(q, dict) else None
    try:
        return float(raw.get("value") if isinstance(raw, dict) else raw)
    except (TypeError, ValueError, AttributeError):
        return None


def audit(twin: Path) -> dict[str, Any]:
    twin = twin.resolve()
    findings: list[dict[str, Any]] = []

    wb = _find_workbook(twin)
    pack = _find_pack(twin)

    # ── 1. pack == workbook ────────────────────────────────────────────────
    if wb is None:
        findings.append({
            "kind": "workbook_missing",
            "issue": "no DRAFT engineering-workbook.xlsx or dossier.xlsx in twin",
        })
    if pack is None:
        findings.append({
            "kind": "pack_missing",
            "issue": "no *-design-pack.zip in twin",
        })
    if wb is not None and pack is not None:
        try:
            pack_bytes = _pack_xlsx_bytes(pack)
        except zipfile.BadZipFile as e:
            findings.append({"kind": "pack_corrupt", "issue": str(e)})
            pack_bytes = None
        if pack_bytes is not None:
            wb_sha = _sha256(wb)
            pack_sha = hashlib.sha256(pack_bytes).hexdigest()
            if wb_sha != pack_sha:
                findings.append({
                    "kind": "pack_workbook_mismatch",
                    "issue": (
                        f"design-pack embedded xlsx SHA ≠ workbook SHA — rebuild pack "
                        f"with the same build-excel-export process "
                        f"(workbook={wb.name} pack={pack.name})"
                    ),
                    "workbook_sha": wb_sha[:16],
                    "pack_xlsx_sha": pack_sha[:16],
                    "workbook_bytes": wb.stat().st_size,
                    "pack_xlsx_bytes": len(pack_bytes),
                })

    # ── 2. workbook magnet == twin magnet ──────────────────────────────────
    twin_mag = _twin_magnet(twin)
    if wb is not None and twin_mag is not None:
        wb_mag = _workbook_magnet_actual(wb)
        if wb_mag is None:
            findings.append({
                "kind": "workbook_magnet_row_missing",
                "issue": f"no Brief target met: magnet_temp row in {wb.name}",
            })
        elif abs(wb_mag - twin_mag) > 0.15:
            findings.append({
                "kind": "workbook_twin_magnet_stale",
                "issue": (
                    f"Checks magnet actual {wb_mag} °C ≠ twin mgu_magnet_temp_c "
                    f"{twin_mag} °C — rebuild Excel from twin after restamp"
                ),
                "workbook_magnet_c": wb_mag,
                "twin_magnet_c": twin_mag,
            })

    # ── 3. frozen decisions applied ────────────────────────────────────────
    try:
        frozen_plan = plan_frozen(twin)
    except Exception as e:  # noqa: BLE001
        findings.append({"kind": "frozen_plan_error", "issue": str(e)})
        frozen_plan = []
    for row in frozen_plan:
        if row.get("action") in ("apply", "handler_missing"):
            findings.append({
                "kind": "frozen_decision_not_applied",
                "issue": (
                    f"{row.get('id')} is {row.get('status')} but twin is not restamped "
                    f"(action={row.get('action')}). Run: "
                    f"python3 scripts/lib/apply_frozen_decisions.py --twin <dir>"
                ),
                "decision_id": row.get("id"),
                "action": row.get("action"),
            })
        elif row.get("action") == "no_handler":
            findings.append({
                "kind": "frozen_decision_no_handler",
                "issue": (
                    f"{row.get('id')} is frozen but has no restamp handler registered "
                    f"in apply_frozen_decisions.HANDLERS — add a handler or unfreeze"
                ),
                "decision_id": row.get("id"),
            })

    # ── 4. drawing gates state vs live artefact ────────────────────────────
    gates_path = twin / "drawing-gates.json"
    state_path = twin / "state.json"
    if gates_path.is_file() and state_path.is_file():
        live = json.loads(gates_path.read_text(encoding="utf-8"))
        st = json.loads(state_path.read_text(encoding="utf-8"))
        sg = st.get("drawingGates") or st.get("drawing_gates")
        if isinstance(sg, dict) and "all_pass" in live and "all_pass" in sg:
            if bool(live.get("all_pass")) != bool(sg.get("all_pass")):
                findings.append({
                    "kind": "drawing_gates_state_stale",
                    "issue": (
                        f"drawing-gates.json all_pass={live.get('all_pass')} but "
                        f"state.drawingGates all_pass={sg.get('all_pass')} — "
                        f"re-run drawing_gates and write back to state"
                    ),
                    "live_all_pass": live.get("all_pass"),
                    "state_all_pass": sg.get("all_pass"),
                })

    # ── 5. decision register file ↔ state ──────────────────────────────────
    reg_file = twin / "10-decision-register.json"
    if reg_file.is_file() and state_path.is_file():
        try:
            parity = decision_register_parity(twin)
        except Exception as e:  # noqa: BLE001
            findings.append({"kind": "decision_register_parity_error", "issue": str(e)})
            parity = None
        if parity is not None and parity.get("in_file_not_state"):
            findings.append({
                "kind": "decision_register_state_stale",
                "issue": (
                    f"10-decision-register.json has ids not in state.decisionRegister: "
                    f"{parity['in_file_not_state']} — Excel Decision Register tab will "
                    f"omit them. Run: python3 scripts/lib/apply_frozen_decisions.py "
                    f"--twin <dir> (syncs register even when restamps are already applied)"
                ),
                "missing_ids": parity["in_file_not_state"],
                "file_count": parity.get("file_count"),
                "state_count": parity.get("state_count"),
            })

    return {
        "schema": SCHEMA,
        "twin": str(twin),
        "workbook": str(wb) if wb else None,
        "pack": str(pack) if pack else None,
        "twin_magnet_c": twin_mag,
        "findings": findings,
        "ok": not findings,
        "incoherent_count": len(findings),
    }


def _selftest() -> int:
    fails: list[str] = []

    def ck(name: str, ok: bool, detail: str = "") -> None:
        if not ok:
            fails.append(f"{name}: {detail}")

    # Minimal: empty twin → missing workbook/pack findings
    with tempfile.TemporaryDirectory(prefix="coh-") as raw:
        twin = Path(raw)
        r = audit(twin)
        ck("empty_not_ok", r["ok"] is False, str(r["findings"]))
        kinds = {f["kind"] for f in r["findings"]}
        ck("reports_missing", "workbook_missing" in kinds and "pack_missing" in kinds, str(kinds))

    # Live twin smoke (if present) — should be coherent after V1.281 + DEC-008
    live = Path("/Users/tristanfischer/Developer/CentaurOS-oxccu-efuel/out/formula-e-front-mgu-20260729-1432")
    if live.is_dir() and (live / "state.json").is_file():
        r2 = audit(live)
        # May fail on DEC-009 / drawing state — record but don't hard-require green
        # proveCatch: pack_workbook_mismatch must NOT fire if V1.281 coherent
        kinds2 = {f["kind"] for f in r2["findings"]}
        if (live / "20260803-1845-V1.281-DRAFT-formula-e-front-mgu-engineering-workbook.xlsx").is_file():
            ck("live_no_pack_mismatch", "pack_workbook_mismatch" not in kinds2, str(r2["findings"]))
            ck("live_no_magnet_stale", "workbook_twin_magnet_stale" not in kinds2, str(r2["findings"]))

    # Synthetic mismatch
    with tempfile.TemporaryDirectory(prefix="coh-mis-") as raw:
        twin = Path(raw)
        (twin / "state.json").write_text(json.dumps({
            "ship_ok": False,
            "orchestratorContract": {"quantities": {
                "mgu_magnet_temp_c": {"value": 83.8},
            }},
            "drawingGates": {"all_pass": False, "n_failing": 2},
        }) + "\n")
        (twin / "drawing-gates.json").write_text(json.dumps({
            "all_pass": True, "n_failing": 0, "n_gates": 23,
        }) + "\n")
        # tiny fake xlsx zip members won't openpyxl — skip magnet row via missing sheet
        # Write two different "xlsx" files as raw bytes for sha test
        wb = twin / "20260101-0000-V1.0-DRAFT-test-engineering-workbook.xlsx"
        pk = twin / "20260101-0000-V1.0-test-design-pack.zip"
        wb.write_bytes(b"PK\x03\x04workbook-AAAA")
        # pack with different embedded xlsx
        with zipfile.ZipFile(pk, "w") as z:
            z.writestr("test-design-pack/engineering-workbook.xlsx", b"PK\x03\x04workbook-BBBB")
        r3 = audit(twin)
        kinds3 = {f["kind"] for f in r3["findings"]}
        ck("detects_pack_mismatch", "pack_workbook_mismatch" in kinds3, str(r3["findings"]))
        ck("detects_drawing_stale", "drawing_gates_state_stale" in kinds3, str(r3["findings"]))

    # Decision register file ahead of state
    with tempfile.TemporaryDirectory(prefix="coh-reg-") as raw:
        twin = Path(raw)
        (twin / "state.json").write_text(json.dumps({
            "ship_ok": False,
            "decisionRegister": [
                {"id": "DEC-001", "status": "OPEN", "decision": "old"},
            ],
            "orchestratorContract": {"quantities": {
                "mgu_magnet_temp_c": {"value": 99.4},
            }},
        }) + "\n")
        (twin / "10-decision-register.json").write_text(json.dumps([
            {"id": "DEC-001", "status": "OPEN", "decision": "old"},
            {"id": "DEC-008", "status": "FROZEN_UNDER_ASSUMPTION", "decision": "duty"},
            {"id": "DEC-009", "status": "FROZEN_UNDER_ASSUMPTION", "decision": "24k"},
        ]) + "\n")
        r4 = audit(twin)
        kinds4 = {f["kind"] for f in r4["findings"]}
        ck(
            "detects_register_stale",
            "decision_register_state_stale" in kinds4,
            str(r4["findings"]),
        )

    if fails:
        print("check_deliverable_coherence selftest FAIL:")
        for f in fails:
            print(" ", f)
        return 1
    print("check_deliverable_coherence selftest: OK")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--twin", type=Path)
    ap.add_argument("--enforce", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--json-out", type=Path)
    args = ap.parse_args()
    if args.selftest:
        return _selftest()
    if not args.twin:
        ap.error("--twin required")
    rep = audit(args.twin)
    text = json.dumps(rep, indent=2) + "\n"
    if args.json_out:
        args.json_out.write_text(text, encoding="utf-8")
    print(text)
    print(f"[deliverable_coherence] ok={rep['ok']} findings={rep['incoherent_count']}")
    if args.enforce and not rep["ok"]:
        return EXIT_INCOHERENT
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
