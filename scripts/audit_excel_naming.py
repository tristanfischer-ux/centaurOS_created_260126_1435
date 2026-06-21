#!/usr/bin/env python3
"""
audit_excel_naming.py — verify the canonical-identity work (Tristan 2026-06-21):
  (1) ONE tag per physical part across the drawings (parts-manifest) AND the BoM.
  (2) the "Part names" master is the single name source — every name-bearing tab
      REFERENCES it rather than repeating the string.
  (3) no broken master references; deterministic checks pass.

USAGE
    .venv/bin/python scripts/audit_excel_naming.py out/<run> [<run>/dossier.xlsx]

Per-tab report + an overall PASS/FAIL. READ-ONLY.
"""
from __future__ import annotations
import json
import os
import re
import sys
import unicodedata

import openpyxl


def _norm_name(s) -> str:
    t = unicodedata.normalize("NFKC", str(s or "")).replace("↳", " ")
    for sep in ("·", "•", "—", " - ", " | "):
        if sep in t:
            t = t.split(sep)[0]
    return re.sub(r"\s+", " ", t).strip().lower().strip(" .,:;-")


def _norm_tag(t) -> str:
    return str(t or "").strip().lower()


def main(argv):
    run = argv[0] if argv else "out/ras-5m-62t-v3"
    xlsx = argv[1] if len(argv) > 1 else os.path.join(run, "dossier.xlsx")
    state = json.load(open(os.path.join(run, "state.json")))
    manifest = json.load(open(os.path.join(run, "parts-manifest.json")))["parts"]
    bom = state.get("requirementsBom") or []
    wb = openpyxl.load_workbook(xlsx)

    print(f"AUDIT  run={run}  xlsx={xlsx}")
    print("=" * 78)

    overall = []

    # ── (1) TAG UNIFICATION: manifest equipment_tag == BoM tag, per shared name ──
    bom_tag_by_name = {}
    for r in bom:
        req = str(r.get("requirement", "") or "")
        if str(r.get("tag", "")).find(".") >= 0 or req.strip().startswith("↳"):
            continue
        if "→" in req or re.search(r"\bconnection\b", req, re.I):
            continue
        bom_tag_by_name.setdefault(_norm_name(req), str(r.get("tag", "")))
    man_tag_by_name = {}
    for p in manifest:
        man_tag_by_name.setdefault(_norm_name(p.get("name")), str(p.get("equipment_tag")))
    def _expand_range(tag):
        """A BoM tag may be a quantity-N RANGE ('B-202–203', 'LT-201–205'); expand to the
        set {B-202, B-203} so a manifest single-instance tag (the Nth of the range) counts
        as a MATCH, not a mismatch."""
        s = str(tag or "").strip()
        m = re.match(r"^([A-Za-z]+)-(\d+)[–-](\d+)$", s)
        if m:
            pfx, a, b = m.group(1), int(m.group(2)), int(m.group(3))
            return {f"{pfx}-{n}" for n in range(a, b + 1)}
        return {s}

    shared = sorted(set(bom_tag_by_name) & set(man_tag_by_name))
    mism = [(n, man_tag_by_name[n], bom_tag_by_name[n]) for n in shared
            if man_tag_by_name[n] not in _expand_range(bom_tag_by_name[n])]
    print(f"\n[1] TAG UNIFICATION  (manifest vs BoM, {len(shared)} shared part names)")
    if mism:
        print(f"    FAIL — {len(mism)} name(s) carry a DIFFERENT tag in drawings vs BoM:")
        for n, mt, bt in mism[:20]:
            print(f"      · {n:38s} manifest={mt:10s} BoM={bt}")
    else:
        print(f"    PASS — every shared part has ONE tag across drawings + BoM")
    overall.append(("tag_unification", not mism))

    # ── (2) MASTER NAME SOURCE + per-tab referencing ──
    if "Part names" not in wb.sheetnames:
        print("\n[2] FAIL — no 'Part names' master tab")
        overall.append(("master_present", False))
    else:
        pn = wb["Part names"]
        master_names = set()
        master_rows = {}
        for rr in range(5, pn.max_row + 1):
            nm = pn.cell(rr, 2).value
            if isinstance(nm, str) and nm.strip():
                master_names.add(_norm_name(nm))
                master_rows[rr] = nm
        print(f"\n[2] MASTER 'Part names' — {len(master_rows)} principal names typed once")
        overall.append(("master_present", True))

        # which tabs are expected to reference the master
        NAME_TABS = ["BoM", "Cost", "Spec sheets", "Connection trace", "Line & velocity"]
        print(f"\n[3] PER-TAB REFERENCING (master ref = a cell formula to 'Part names')")
        for t in wb.sheetnames:
            ws = wb[t]
            refs = 0
            literal_hits = 0   # literal cells whose text IS a known master name (missed ref)
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str) and "'Part names'!" in v:
                        refs += 1
                    elif isinstance(v, str) and t not in ("Part names",):
                        if _norm_name(v) in master_names and not v.startswith("="):
                            literal_hits += 1
            flag = ""
            if t in NAME_TABS:
                ok = refs > 0
                flag = "  <== expected refs: " + ("OK" if ok else "MISSING")
                overall.append((f"refs:{t}", ok))
            if refs or literal_hits or t in NAME_TABS:
                print(f"    {t:24s} refs={refs:4d}  unreferenced-name-literals={literal_hits:3d}{flag}")

    # ── (4) BROKEN MASTER REFS ──
    if "Part names" in wb.sheetnames:
        maxr = wb["Part names"].max_row
        bad = 0
        for t in wb.sheetnames:
            for row in wb[t].iter_rows():
                for c in row:
                    if isinstance(c.value, str):
                        for m in re.findall(r"'Part names'!\$[A-Z]\$(\d+)", c.value):
                            if int(m) > maxr or int(m) < 5:
                                bad += 1
        print(f"\n[4] BROKEN MASTER REFS  ->  {bad}")
        overall.append(("no_broken_refs", bad == 0))

    # ── (5) verdict ──
    print("\n" + "=" * 78)
    n_pass = sum(1 for _, ok in overall if ok)
    for k, ok in overall:
        print(f"   {'PASS' if ok else 'FAIL'}  {k}")
    allok = all(ok for _, ok in overall)
    print("=" * 78)
    print(f"OVERALL: {'ALL PASS' if allok else 'FAIL'}  ({n_pass}/{len(overall)})")
    return 0 if allok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
