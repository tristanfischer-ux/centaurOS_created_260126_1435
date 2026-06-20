"""Dump every cell of every tab of an .xlsx as plain text for the cell-by-cell audit
swarm. For each sheet: one line per non-empty cell as `A1\tDISPLAY\t[=FORMULA]`. Loads
TWICE — once data_only=True (cached computed values) and once data_only=False (formulas)
— so the audit sees BOTH what a formula evaluates to AND the formula itself (a formula
cell whose cached value is None = never calculated / a potential #REF). Usage:
  python3 _dump_xlsx_cells.py <file.xlsx> [out_dir]   # writes <out_dir>/tab-<n>-<name>.txt
  python3 _dump_xlsx_cells.py <file.xlsx> --sheet "BoM"   # one sheet to stdout
"""
import os
import re
import sys

import openpyxl


def _slug(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-")[:40]


def dump(path, out_dir=None, only_sheet=None):
    wb_v = openpyxl.load_workbook(path, data_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    manifest = []
    for idx, name in enumerate(wb_v.sheetnames):
        if only_sheet and name != only_sheet:
            continue
        ws_v = wb_v[name]
        ws_f = wb_f[name]
        lines = [f"# SHEET: {name}  ({ws_v.max_row} rows × {ws_v.max_column} cols)"]
        fcells = {}
        for row in ws_f.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    fcells[c.coordinate] = c.value
        ncells = 0
        for row in ws_v.iter_rows():
            for c in row:
                val = c.value
                formula = fcells.get(c.coordinate)
                if val is None and formula is None:
                    continue
                ncells += 1
                disp = "" if val is None else str(val).replace("\t", " ").replace("\n", " ⏎ ")
                if len(disp) > 200:
                    disp = disp[:200] + "…"
                line = f"{c.coordinate}\t{disp}"
                if formula:
                    line += f"\t{formula}"
                elif val is None and formula is None:
                    pass
                lines.append(line)
        body = "\n".join(lines)
        if only_sheet:
            print(body)
            return
        fn = f"tab-{idx:02d}-{_slug(name)}.txt"
        with open(os.path.join(out_dir, fn), "w") as f:
            f.write(body)
        manifest.append((fn, name, ncells))
    if out_dir and not only_sheet:
        with open(os.path.join(out_dir, "_manifest.txt"), "w") as f:
            for fn, name, n in manifest:
                f.write(f"{fn}\t{name}\t{n} cells\n")
        print(f"dumped {len(manifest)} sheets to {out_dir}")
        for fn, name, n in manifest:
            print(f"  {name:28} {n:>5} cells  → {fn}")


if __name__ == "__main__":
    p = sys.argv[1]
    if "--sheet" in sys.argv:
        dump(p, only_sheet=sys.argv[sys.argv.index("--sheet") + 1])
    else:
        od = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(p) or ".", "_cells")
        os.makedirs(od, exist_ok=True)
        dump(p, od)
