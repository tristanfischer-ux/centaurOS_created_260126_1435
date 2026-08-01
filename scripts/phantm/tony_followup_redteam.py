"""Corrected Tony follow-up pack after Sol + Kimi K3 red team + Floquet array.

Reads:
  out/tony-phase-material-study.json
  out/double-wall-rf-check.json
  out/floquet-hex-array.json
  out/council-followup-{sol,kimi}.txt  (for citation only)

Writes stamped markdown + PDF + Excel to out/ and ~/Downloads/
(date-time-subject-version naming). Spells terms out.

Run: ~/.venvs/phantm/bin/python tony_followup_redteam.py
"""

from __future__ import annotations

import datetime
import json
import os
import shutil
import subprocess
from importlib.machinery import SourceFileLoader
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "out")
CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
SUBJECT = "PHANTM-TONY-FOLLOWUP-CALCS-CORRECTED"

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E8F1FB")
WARN_FILL = PatternFill("solid", fgColor="FCE4E4")
FE_FILL = PatternFill("solid", fgColor="E2EFDA")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def stem() -> str:
    ver = json.load(open(os.path.join(HERE, "version.json")))
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M")
    return f"{stamp}-{SUBJECT}-V{ver['major']}.{ver['minor']}"


def load():
    study = json.load(open(os.path.join(OUT, "tony-phase-material-study.json")))
    rf = json.load(open(os.path.join(OUT, "double-wall-rf-check.json")))
    floq_path = os.path.join(OUT, "floquet-hex-array.json")
    floq = json.load(open(floq_path)) if os.path.exists(floq_path) else None
    return study, rf, floq


def force_table(study):
    """Per-pole, per-coil, per-foot, per-watt at 0.40 A."""
    va = study["variants"]["A_three_phase_3_teeth"]
    vb = study["variants"]["B_four_phase_4_teeth"]
    vc = study["variants"]["C_four_phase_3_teeth"]
    r_ohm = 3.618
    i_a = 0.40
    rows = []
    for g in study["gaps_um"]:
        fa = va["summary"][str(g)]["force_at_0_40_a_mn"]
        fb = vb["summary"][str(g)]["force_at_0_40_a_mn"]
        fc = vc["summary"][str(g)]["force_at_0_40_a_mn"]
        foot_a = va["summary"][str(g)]["pole_foot_mm"]
        foot_b = vb["summary"][str(g)]["pole_foot_mm"]
        foot_c = vc["summary"][str(g)]["pole_foot_mm"]
        # per-phase current → copper loss = n_teeth × I²R (one coil per tooth)
        loss_a = 3 * i_a ** 2 * r_ohm
        loss_b = 4 * i_a ** 2 * r_ohm
        loss_c = 3 * i_a ** 2 * r_ohm
        rows.append(dict(
            gap_um=g,
            force_a=fa, force_b=fb, force_c=fc,
            per_coil_a=fa / 3, per_coil_b=fb / 4, per_coil_c=fc / 3,
            per_foot_a=fa / foot_a, per_foot_b=fb / foot_b, per_foot_c=fc / foot_c,
            per_watt_a=fa / loss_a, per_watt_b=fb / loss_b, per_watt_c=fc / loss_c,
            foot_a=foot_a, foot_b=foot_b, foot_c=foot_c,
            ba=fb / fa, cb=fc / fb, ca=fc / fa,
            ba_coil=(fb / 4) / (fa / 3),
            cb_foot=(fc / foot_c) / (fb / foot_b),
        ))
    return rows


def extrapolate_i_for_force(rows, target_mn):
    """Linear extrapolation from last two solved points — optimistic lower bound."""
    i_arr = [r["current_a"] for r in rows]
    f_arr = [r["force_mn"] for r in rows]
    if target_mn <= max(f_arr):
        for a, b, fa, fb in zip(i_arr, i_arr[1:], f_arr, f_arr[1:]):
            if fa <= target_mn <= fb or fb <= target_mn <= fa:
                if fb == fa:
                    return a, "bracket"
                return a + (target_mn - fa) * (b - a) / (fb - fa), "bracket"
    a, b = i_arr[-2], i_arr[-1]
    fa, fb = f_arr[-2], f_arr[-1]
    if fb == fa:
        return None, "flat"
    i_star = b + (target_mn - fb) * (b - a) / (fb - fa)
    return i_star, "extrapolate_linear_optimistic"


def write_md(study, rf, floq, norms) -> str:
    va = study["variants"]["A_three_phase_3_teeth"]
    vb = study["variants"]["B_four_phase_4_teeth"]
    vc = study["variants"]["C_four_phase_3_teeth"]
    mats = study["materials"]
    n60 = next(r for r in norms if r["gap_um"] == 60)
    n20 = next(r for r in norms if r["gap_um"] == 20)

    # detent extrapolation on legacy Somaloy A at 60 µm
    detent = study["targets_mn"]["detent"]
    i_det, how = extrapolate_i_for_force(mats["all_somaloy_legacy"], detent)

    feco_12 = next(x["force_mn"] for x in mats["all_fe_co_laminated"]
                   if x["current_a"] == 1.2)
    mix_12 = next(x["force_mn"] for x in mats["translator_fe_co__poles_mim"]
                  if x["current_a"] == 1.2)
    mim_12 = next(x["force_mn"] for x in mats["all_mim_fe3si"]
                  if x["current_a"] == 1.2)
    mix_04 = next(x["force_mn"] for x in mats["translator_fe_co__poles_mim"]
                  if x["current_a"] == 0.4)
    feco_04 = next(x["force_mn"] for x in mats["all_fe_co_laminated"]
                   if x["current_a"] == 0.4)

    L, A = [], (lambda s: L.append(s))
    A("# PHANTM — follow-up calculations (corrected after red team)")
    A("")
    A("Tony — same finite-element packs as the overnight note, **re-framed "
      "after an independent red team** (OpenAI Sol + Moonshot Kimi K3) and "
      "with a **full 2×2 periodic Floquet array solve** for one-third double "
      "walls. Coil-only reluctance force (permanent-magnet remanence set to "
      "zero). Step held at "
      f"**{study['step_um']:.0f} micrometres** throughout.")
    A("")
    A("## Red-team corrections (read this first)")
    A("")
    A("Both reviewers agreed the arithmetic in the force tables is correct, "
      "but several **client-facing claims were wrongly attributed**. The "
      "corrections below are now the binding reading.")
    A("")
    A("1. **The +33% at 60 micrometres is coil count, not a four-phase "
      "physics bonus.** At 0.40 A and 60 micrometres, force per coil for "
      f"three-phase (A) and four-phase four-tooth (B) is identical to three "
      f"decimals ({n60['per_coil_a']:.3f} mN/coil). B/A = "
      f"**{n60['ba']:.3f}** equals 4/3 within 0.03%. The extra force is the "
      f"extra tooth/coil.")
    A(f"2. **Per unit pole-foot length, four-phase is worse at 60 micrometres.** "
      f"A: {n60['per_foot_a']:.2f} mN/mm; B: {n60['per_foot_b']:.2f} mN/mm "
      f"(about **{100*(n60['per_foot_b']/n60['per_foot_a']-1):.0f}%**). "
      f"If the stator envelope can hold more three-phase poles, three-phase "
      f"wins on force density.")
    A(f"3. **Your ~25% for dropping the fourth tooth is right under the "
      f"length-normalised metric.** Absolute per-pole force C/B = "
      f"{n60['cb']:.3f} (−{(1-n60['cb'])*100:.0f}% at 60 micrometres). "
      f"Force per millimetre of pole foot C/B = **{n60['cb_foot']:.2f}** "
      f"(three-tooth is ~{(n60['cb_foot']-1)*100:.0f}% denser). The earlier "
      f"“~7%, not 25%” line only held for absolute force per pole segment.")
    A("4. **“Pole foot” is the local active tooth envelope (tooth span + "
      "100 micrometre flank allowance), not full stator length.** Full length "
      "still needs end margins, coil bodies, leads, sensors, mounts.")
    A("5. **Coil resistance held at 3.618 ohm is a per-coil model "
      "assumption** (same turns and wire gauge). Wider teeth (167 vs "
      "125 micrometres) lengthen the mean turn slightly — volts at fixed "
      "current are mildly understated for B/C. Total copper loss per pole "
      "at equal per-phase current still scales with tooth count.")
    A("6. **Detent current at 60 micrometres is an optimistic lower bound.** "
      "The sweep tops out at 1.2 A / 8.62 mN (Somaloy-shaped A); detent "
      f"11.13 mN needs linear extrapolation → "
      f"{'~'+format(i_det,'.2f')+' A' if i_det else 'not estimated'} "
      f"({how}). On a saturating curve the true current is **at least** "
      f"that high; a point at ≥1.4 A should be solved before tooling.")
    A("7. **Radio-frequency: single-cell Neumann is not the array answer.** "
      "The Floquet section below is. Three-dimensional scan impedance / "
      "grating lobes still sit with Vlad.")
    A("")

    # ---- geometry -------------------------------------------------------
    A("## 1. Geometry lock")
    A("")
    A("| Variant | Phases | Pitch | Tooth | Teeth on each pole | "
      "Pole-foot envelope | What it isolates |")
    A("|---|---|---|---|---|---|---|")
    A(f"| A three-phase | 3 | {va['meta']['pitch_um']:.0f} µm | "
      f"{va['meta']['tooth_um']:.0f} µm | 3 | "
      f"{va['summary']['60']['pole_foot_mm']:.3f} mm | baseline |")
    A(f"| B four-phase, 4 teeth | 4 | {vb['meta']['pitch_um']:.0f} µm | "
      f"{vb['meta']['tooth_um']:.0f} µm | 4 | "
      f"{vb['summary']['60']['pole_foot_mm']:.3f} mm | "
      f"phase + pitch + tooth count together |")
    A(f"| C four-phase, 3 teeth | 4 | {vc['meta']['pitch_um']:.0f} µm | "
      f"{vc['meta']['tooth_um']:.0f} µm | 3 | "
      f"{vc['summary']['60']['pole_foot_mm']:.3f} mm | "
      f"tooth count at fixed four-phase pitch |")
    A("")
    A("Duty 0.401. Fixed step = pitch / phases = 104 micrometres. Comparing "
      "A to B changes pitch, tooth width, tooth count, and phase count at "
      "once — do not call the whole delta “the four-phase benefit.”")
    A("")

    # ---- force tables ---------------------------------------------------
    A("## 2. Force at 0.40 A — four normalisations")
    A("")
    A("### 2a Absolute force per pole segment (previous table)")
    A("")
    A("| Gap | A | B | C | B/A | C/B | C/A |")
    A("|---|---|---|---|---|---|---|")
    for r in norms:
        A(f"| **{r['gap_um']} µm** | {r['force_a']:.3f} mN | "
          f"{r['force_b']:.3f} mN | {r['force_c']:.3f} mN | "
          f"{r['ba']:.3f} | {r['cb']:.3f} | {r['ca']:.3f} |")
    A("")
    A("### 2b Force per coil (copper-fair at equal turns)")
    A("")
    A("| Gap | A mN/coil | B mN/coil | C mN/coil | B/A per coil |")
    A("|---|---|---|---|---|")
    for r in norms:
        A(f"| {r['gap_um']} µm | {r['per_coil_a']:.3f} | "
          f"{r['per_coil_b']:.3f} | {r['per_coil_c']:.3f} | "
          f"{r['ba_coil']:.3f} |")
    A("")
    A(f"At 60 micrometres B/A per coil = **{n60['ba_coil']:.3f}** (no "
      f"per-copper gain). At 20 micrometres B/A per coil = "
      f"**{n20['ba_coil']:.3f}** (four-phase worse per copper).")
    A("")
    A("### 2c Force per millimetre of pole-foot envelope")
    A("")
    A("| Gap | A mN/mm | B mN/mm | C mN/mm | C/B per mm |")
    A("|---|---|---|---|---|")
    for r in norms:
        A(f"| {r['gap_um']} µm | {r['per_foot_a']:.3f} | "
          f"{r['per_foot_b']:.3f} | {r['per_foot_c']:.3f} | "
          f"{r['cb_foot']:.3f} |")
    A("")
    A("### 2d Force per watt of copper loss (I²R, R = 3.618 ohm per coil)")
    A("")
    A("| Gap | A mN/W | B mN/W | C mN/W |")
    A("|---|---|---|---|")
    for r in norms:
        A(f"| {r['gap_um']} µm | {r['per_watt_a']:.3f} | "
          f"{r['per_watt_b']:.3f} | {r['per_watt_c']:.3f} |")
    A("")
    A(f"**Decomposition at 60 micrometres (fixed tooth count vs pitch).** "
      f"C versus A (same 3 coils, wider pitch/tooth): "
      f"**+{(n60['ca']-1)*100:.1f}%** absolute force — that is the wider-"
      f"tooth geometry, not phase count. B versus C (fourth tooth at same "
      f"pitch): **+{(n60['ba']/n60['ca']-1)*100:.1f}%** absolute — the "
      f"extra coil. Gap still dominates: at 20 micrometres all three "
      f"variants sit near {n20['force_a']:.1f}–{n20['force_b']:.1f} mN.")
    A("")

    # ---- detent ---------------------------------------------------------
    A("## 3. Detent reach — honest about extrapolation")
    A("")
    A(f"Target detent force **{detent:.2f} mN** (coil-only comparison). "
      "Currents reported by the study when the sweep brackets the target:")
    A("")
    A("| Gap | A | B | C | Method |")
    A("|---|---|---|---|---|")
    for g in study["gaps_um"]:
        ia = va["summary"][str(g)]["i_for_detent_a"]
        ib = vb["summary"][str(g)]["i_for_detent_a"]
        ic = vc["summary"][str(g)]["i_for_detent_a"]

        def fmt(i):
            return "not reached on sweep" if i is None else f"{i:.2f} A"

        note = "bracketed on F(I)" if ia is not None else (
            f"linear extrapolate ≥{i_det:.2f} A (optimistic)" if (
                g == 60 and i_det) else "above last solved point")
        A(f"| {g} µm | {fmt(ia)} | {fmt(ib)} | {fmt(ic)} | {note} |")
    A("")
    A("**Corrected claim.** At 60 micrometres the detent is **not reached "
      "on the solved current sweep** (last point 1.2 A). A linear extension "
      "of the last segment suggests about 1.44 A — treat that as a **lower "
      "bound**. At 40 micrometres the bracketed value ~0.92 A is real "
      "(inside the sweep). Gap remains the primary lever.")
    A("")

    # ---- materials ------------------------------------------------------
    A("## 4. Materials (representative curves — not vendor grades)")
    A("")
    A("| Assignment | 0.40 A | 1.00 A | 1.20 A | Bridge |B| @ 1.20 A |")
    A("|---|---|---|---|---|")

    def row(name, label):
        r0 = {x["current_a"]: x for x in mats[name]}
        A(f"| {label} | {r0[0.4]['force_mn']:.2f} mN | "
          f"{r0[1.0]['force_mn']:.2f} mN | {r0[1.2]['force_mn']:.2f} mN | "
          f"{r0[1.2]['b_bridge_max_t']:.2f} T |")

    row("all_fe_co_laminated", "All iron-cobalt laminated")
    row("translator_fe_co__poles_mim",
        "**Translator iron-cobalt + poles MIM Fe-3%Si**")
    row("all_mim_fe3si", "All micro-injection-moulded Fe-3%Si")
    row("all_somaloy_legacy", "Legacy Somaloy-shaped curve")
    A("")
    A(f"Mixed (your split) is within 0.4% of all-MIM at every current shown — "
      f"**the pole/bridge steel sets the limit**, not the translator. Versus "
      f"all iron-cobalt the mixed penalty is "
      f"**{(1-mix_04/feco_04)*100:.0f}% at 0.40 A** and "
      f"**{(1-mix_12/feco_12)*100:.0f}% at 1.20 A** "
      f"({mix_12:.1f} vs {feco_12:.1f} mN). Quote the trend, not a single "
      f"“~13%” as if it were current-independent. Equal-current comparison "
      f"is the right basis for a current-limited drive; vendor B–H must "
      f"replace these curves before tooling.")
    A("")

    # ---- RF + Floquet ---------------------------------------------------
    A("## 5. One-third double walls — single-cell bounds + Floquet array")
    A("")
    u = rf["uniform_cell"]
    s = rf["model_interior_shrink_by_tape"]
    m = rf["model_interior_shrink_by_tape_over_3"]
    h = rf["model_interior_held_exterior_grows"]
    A("### 5a Single-cell Neumann bounds (cross-check only)")
    A("")
    A("| Model | Cut-off | Shift versus uniform |")
    A("|---|---|---|")
    A(f"| Uniform cell (across-flats 3.10 mm) | **{u['fc_ghz']:.2f} GHz** | — |")
    A(f"| Interior shrinks by one tape | {s['fc_ghz']:.2f} GHz | "
      f"**+{s['delta_fc_mhz']:.0f} MHz** |")
    A(f"| Interior shrinks by tape/3 | {m['fc_ghz']:.2f} GHz | "
      f"+{m['delta_fc_mhz']:.0f} MHz |")
    A(f"| Interior held; exterior pitch grows | "
      f"{h['fc_interior_ghz']:.2f} GHz | 0 — pitch "
      f"**+{h['pitch_growth_pct']:.2f}%** on that axis |")
    A("")

    A("### 5b Full periodic Floquet–Bloch array (2×2 supercell)")
    A("")
    if not floq or floq.get("comparison", {}).get("air_pixels_removed", 0) <= 0:
        A("*Floquet artefact missing or mask failed — re-run "
          "`floquet_hex_array.py` before sending.*")
    else:
        uu = floq["uniform"]
        dd = floq["one_third_double"]
        cmp_ = floq["comparison"]
        A(f"Pitch **{floq['geometry_mm']['pitch']} mm**, tape "
          f"**{floq['geometry_mm']['tape']*1000:.0f} micrometres**, "
          f"supercell {floq['geometry_mm']['supercell_cells']} hex centres, "
          f"exact rectangle "
          f"{floq['geometry_mm']['Lx']:.3f} × {floq['geometry_mm']['Ly']:.3f} mm. "
          f"Implied across-flats: uniform "
          f"{floq['geometry_mm']['implied_uniform_across_flats']:.3f} mm; "
          f"double-wall axis "
          f"{floq['geometry_mm']['implied_double_wall_across_flats']:.3f} mm.")
        A("")
        A("| Lattice | Air fraction | Fundamental cut-off (Gamma) | "
          "Margin to 75 GHz | Mode split after multiplet |")
        A("|---|---|---|---|---|")
        A(f"| Uniform walls | {uu['meta']['air_fraction']:.4f} | "
          f"**{uu['fundamental_cutoff_ghz']:.3f} GHz** | "
          f"{uu['margin_at_75_ghz_mhz']:.0f} MHz | "
          f"{uu['mode_splitting_ghz']:.3f} GHz |")
        A(f"| One-third double (class 0) | {dd['meta']['air_fraction']:.4f} | "
          f"**{dd['fundamental_cutoff_ghz']:.3f} GHz** | "
          f"{dd['margin_at_75_ghz_mhz']:.0f} MHz | "
          f"{dd['mode_splitting_ghz']:.3f} GHz |")
        A("")
        A(f"**Δ fundamental cut-off = "
          f"{cmp_['delta_fundamental_mhz']:+.0f} MHz** "
          f"(air pixels removed: {cmp_['air_pixels_removed']}, "
          f"{100*cmp_['air_pixels_removed']/max(uu['meta']['air_pixels'],1):.2f}% "
          f"of uniform air). "
          f"That shift sits inside the uniform near-degenerate multiplet "
          f"width (~{uu['mode_splitting_ghz']*1000:.0f} MHz), so do not "
          f"over-read the sign of Δfc. The clearer array signature is "
          f"**mode splitting after the multiplet: "
          f"{uu['mode_splitting_ghz']:.3f} → {dd['mode_splitting_ghz']:.3f} GHz** "
          f"(coupling anisotropy). "
          f"Lowest sampled Bloch cut-off: uniform "
          f"{cmp_['uniform_min_bloch_cutoff_ghz']:.3f} GHz, double "
          f"{cmp_['double_min_bloch_cutoff_ghz']:.3f} GHz — both remain "
          f"open at 75 GHz with >20 GHz of margin.")
        A("")
        A("**Client-safe claim.** In this 2D periodic transverse model, "
          "one-third double walls do **not** close the guides at 75 GHz; "
          "the fundamental cut-off barely moves versus the multiplet "
          "width, while in-plane coupling anisotropy shows up as larger "
          "mode splitting. **Still required before Seed-1:** Vlad’s 3D "
          "unit-cell Floquet (scan impedance, grating lobes, feed "
          "transition). Finite-array edge cells and the “interior-held / "
          "exterior grows” layout are not representable in a single "
          "periodic cell.")
        A("")
        for line in floq.get("verdict", []):
            A(f"- {line}")
    A("")

    # ---- summary --------------------------------------------------------
    A("## 6. Summary — what to design against")
    A("")
    A("| Question | Corrected answer |")
    A("|---|---|")
    A("| Does four-phase help force? | **Per pole at 60 µm: +33% = one "
      "extra coil.** Per copper / per watt: no gain at 60 µm; worse at "
      "tight gap. Per pole-foot length: three-phase denser. |")
    A("| Cost of dropping the fourth tooth? | **~7% absolute force; "
      "~22–28% on force per millimetre of pole foot** — your 25% matches "
      "the length-normalised reading. |")
    A("| Stator channels / length | Channel count +33%. Pole-foot "
      f"envelope {va['summary']['60']['pole_foot_mm']:.2f} → "
      f"{vb['summary']['60']['pole_foot_mm']:.2f} mm (4-tooth) or "
      f"{vc['summary']['60']['pole_foot_mm']:.2f} mm (3-tooth). Not full "
      f"stator length. |")
    A("| Gap versus phase count | **Gap dominates.** At 20 µm the three "
      "variants nearly meet. |")
    A("| MIM poles | Limit saturated force vs laminated iron-cobalt; "
      f"penalty grows with current (~{(1-mix_04/feco_04)*100:.0f}% → "
      f"~{(1-mix_12/feco_12)*100:.0f}% from 0.4→1.2 A). Translator "
      f"iron-cobalt buys almost nothing once poles are MIM. |")
    delta = (floq or {}).get("comparison", {}).get("delta_fundamental_mhz")
    A("| One-third double walls | Floquet: guides stay open at 75 GHz; "
      f"Δfc "
      f"{'='+format(delta,'+.0f')+' MHz' if delta is not None else 'see §5b'}. "
      f"3D scan still with Vlad. |")
    A("")
    A("---")
    A("")
    A(f"*Finite-element runtime {study['runtime_s']:.0f} s. "
      f"Floquet runtime "
      f"{(floq or {}).get('runtime_s', '—')} s. "
      "Artefacts: `tony-phase-material-study.json`, "
      "`double-wall-rf-check.json`, `floquet-hex-array.json`, "
      "`council-followup-sol.txt`, `council-followup-kimi.txt`. "
      "Material curves are representative; swap for named vendor data "
      "before tooling. Accompanying Excel workbook carries every table.*")
    A("")

    path = os.path.join(OUT, f"{SUBJECT}.md")
    open(path, "w").write("\n".join(L) + "\n")
    print(f"wrote {path}")
    return path


def header(ws, row, *cells):
    for i, txt in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=txt)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def put(ws, r, c, val, kind="calc", fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = {"input": INPUT_FILL, "calc": CALC_FILL,
                 "warn": WARN_FILL, "fe": FE_FILL}.get(kind, CALC_FILL)
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    return cell


def write_xlsx(study, rf, floq, norms, name_stem: str) -> str:
    va = study["variants"]["A_three_phase_3_teeth"]
    vb = study["variants"]["B_four_phase_4_teeth"]
    vc = study["variants"]["C_four_phase_3_teeth"]
    mats = study["materials"]
    detent = study["targets_mn"]["detent"]
    i_det, how = extrapolate_i_for_force(mats["all_somaloy_legacy"], detent)

    wb = Workbook()

    # ---- INDEX ----------------------------------------------------------
    ws = wb.active
    ws.title = "0 INDEX"
    header(ws, 1, "PDF section", "Sheet", "What it holds", "Red-team note")
    idx = [
        ("§0 Red-team corrections", "1 RED TEAM",
         "Binding framing changes", "Sol + Kimi K3"),
        ("§1 Geometry", "2 GEOMETRY", "A/B/C lock", "Foot ≠ full stator"),
        ("§2 Force normalisations", "3 FORCE NORMS",
         "Per pole / coil / foot / watt", "B/A=4/3 at 60 µm"),
        ("§3 Detent currents", "4 DETENT", "Bracket vs extrapolate",
         "1.44 A optimistic"),
        ("§4 Materials", "5 MATERIALS", "Fe-Co / MIM split", "Trend with I"),
        ("§5a Single-cell RF", "6 RF SINGLE CELL", "Neumann bounds",
         "Not the array answer"),
        ("§5b Floquet array", "7 FLOQUET ARRAY", "2×2 periodic",
         "Guides open @ 75 GHz"),
        ("Summary", "8 SUMMARY", "Design choices", "Use norms"),
    ]
    for i, row in enumerate(idx, start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    for col, w in zip("ABCD", (28, 20, 36, 28)):
        ws.column_dimensions[col].width = w

    # ---- RED TEAM -------------------------------------------------------
    ws = wb.create_sheet("1 RED TEAM")
    header(ws, 1, "#", "Correction", "Severity", "Source")
    fixes = [
        (1, "At 60 µm, B/A force = 4/3 = coil count; per-coil force identical",
         "DECISIVE", "Sol + Kimi"),
        (2, "Per pole-foot mm, four-phase B is ~25% below three-phase A at 60 µm",
         "DECISIVE", "Sol + Kimi"),
        (3, "Dropping 4th tooth: ~7% absolute, ~22–28% per foot (Tony ~25% OK)",
         "DECISIVE", "Kimi"),
        (4, "Pole foot = tooth span + 0.1 mm flanks; not full stator length",
         "MATERIAL", "Sol"),
        (5, "R=3.618 Ω is per-coil assumption; wider tooth → longer mean turn",
         "MATERIAL", "Sol + Kimi"),
        (6, "Detent @60 µm needs extrapolation beyond 1.2 A; linear I* optimistic",
         "MATERIAL", "Sol + Kimi"),
        (7, "Single-cell RF ≠ array; Floquet 2D closes in-plane; 3D still Vlad",
         "MATERIAL", "Sol + Kimi"),
    ]
    for i, row in enumerate(fixes, start=2):
        for j, v in enumerate(row, start=1):
            put(ws, i, j, v, kind="warn" if row[2] == "DECISIVE" else "calc")
    for col, w in zip("ABCD", (6, 78, 12, 14)):
        ws.column_dimensions[col].width = w

    # ---- GEOMETRY -------------------------------------------------------
    ws = wb.create_sheet("2 GEOMETRY")
    header(ws, 1, "Variant", "Phases", "Pitch_um", "Tooth_um",
           "Teeth", "Pole_foot_mm", "Note")
    for i, (v, note) in enumerate([
        (va, "baseline three-phase"),
        (vb, "phase+pitch+tooth count"),
        (vc, "tooth count at fixed 4-phase pitch"),
    ], start=2):
        put(ws, i, 1, v["meta"]["label"], "input")
        put(ws, i, 2, v["meta"]["phases"], "input")
        put(ws, i, 3, v["meta"]["pitch_um"], "input")
        put(ws, i, 4, v["meta"]["tooth_um"], "input")
        put(ws, i, 5, v["meta"]["n_pole_teeth"], "input")
        put(ws, i, 6, v["summary"]["60"]["pole_foot_mm"], "fe", "0.000")
        put(ws, i, 7, note)
    put(ws, 6, 1, "step_um", "input")
    put(ws, 6, 2, study["step_um"], "input")
    put(ws, 7, 1, "duty", "input")
    put(ws, 7, 2, study["duty"], "input", "0.000")
    put(ws, 8, 1, "R_coil_ohm (model)", "input")
    put(ws, 8, 2, 3.618, "input", "0.000")
    put(ws, 9, 1, "Foot definition", "warn")
    put(ws, 9, 2,
        "tooth span + 100 µm flank allowance — NOT full stator length", "warn")
    for col, w in zip("ABCDEFG", (28, 10, 12, 12, 10, 14, 36)):
        ws.column_dimensions[col].width = w

    # ---- FORCE NORMS ----------------------------------------------------
    ws = wb.create_sheet("3 FORCE NORMS")
    header(ws, 1,
           "gap_um",
           "F_A_mN", "F_B_mN", "F_C_mN",
           "B/A", "C/B", "C/A",
           "A_mN_per_coil", "B_mN_per_coil", "C_mN_per_coil", "B/A_per_coil",
           "A_mN_per_mm", "B_mN_per_mm", "C_mN_per_mm", "C/B_per_mm",
           "A_mN_per_W", "B_mN_per_W", "C_mN_per_W")
    for i, r in enumerate(norms, start=2):
        vals = [
            r["gap_um"],
            r["force_a"], r["force_b"], r["force_c"],
            r["ba"], r["cb"], r["ca"],
            r["per_coil_a"], r["per_coil_b"], r["per_coil_c"], r["ba_coil"],
            r["per_foot_a"], r["per_foot_b"], r["per_foot_c"], r["cb_foot"],
            r["per_watt_a"], r["per_watt_b"], r["per_watt_c"],
        ]
        for j, v in enumerate(vals, start=1):
            kind = "fe" if j <= 4 else "calc"
            put(ws, i, j, v, kind=kind, fmt="0.000" if j > 1 else "0")
    put(ws, 8, 1, "KEY", "warn")
    put(ws, 8, 2,
        "At 60 µm B/A_per_coil ≈ 1.000 → +33% is coil count. "
        "C/B_per_mm ≈ 1.28 → Tony ~25% length-normalised stands.", "warn")
    for c in range(1, 19):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # ---- DETENT ---------------------------------------------------------
    ws = wb.create_sheet("4 DETENT")
    header(ws, 1, "gap_um", "I_A_A", "I_B_A", "I_C_A", "method")
    for i, g in enumerate(study["gaps_um"], start=2):
        ia = va["summary"][str(g)]["i_for_detent_a"]
        ib = vb["summary"][str(g)]["i_for_detent_a"]
        ic = vc["summary"][str(g)]["i_for_detent_a"]
        put(ws, i, 1, g, "input")
        put(ws, i, 2, ia if ia is not None else "not reached",
            "fe" if ia else "warn")
        put(ws, i, 3, ib if ib is not None else "not reached",
            "fe" if ib else "warn")
        put(ws, i, 4, ic if ic is not None else "not reached",
            "fe" if ic else "warn")
        put(ws, i, 5,
            "bracketed" if ia is not None else
            (f"linear extrapolate {i_det:.3f} A ({how})" if g == 60 else
             "above sweep"),
            "warn" if ia is None else "calc")
    put(ws, 8, 1, "detent_mN", "input")
    put(ws, 8, 2, detent, "input", "0.00")
    put(ws, 9, 1, "NOTE", "warn")
    put(ws, 9, 2,
        "Linear extrapolation underestimates required current on a "
        "saturating curve — treat 60 µm figure as optimistic lower bound.",
        "warn")
    for col, w in zip("ABCDE", (10, 14, 14, 14, 40)):
        ws.column_dimensions[col].width = w

    # ---- MATERIALS ------------------------------------------------------
    ws = wb.create_sheet("5 MATERIALS")
    header(ws, 1, "assignment", "I_A", "force_mN", "b_bridge_T")
    r = 2
    for name, rows in mats.items():
        for row in rows:
            put(ws, r, 1, name, "input")
            put(ws, r, 2, row["current_a"], "fe", "0.00")
            put(ws, r, 3, row["force_mn"], "fe", "0.000")
            put(ws, r, 4, row["b_bridge_max_t"], "fe", "0.000")
            r += 1
    put(ws, r + 1, 1, "penalty_mixed_vs_feco_at_1.2A", "calc")
    mix = next(x["force_mn"] for x in mats["translator_fe_co__poles_mim"]
               if x["current_a"] == 1.2)
    feco = next(x["force_mn"] for x in mats["all_fe_co_laminated"]
                if x["current_a"] == 1.2)
    put(ws, r + 1, 2, 1 - mix / feco, "calc", "0.0%")
    for col, w in zip("ABCD", (36, 10, 12, 12)):
        ws.column_dimensions[col].width = w

    # ---- RF SINGLE ------------------------------------------------------
    ws = wb.create_sheet("6 RF SINGLE CELL")
    header(ws, 1, "model", "fc_GHz", "delta_fc_MHz", "note")
    u = rf["uniform_cell"]
    s = rf["model_interior_shrink_by_tape"]
    m = rf["model_interior_shrink_by_tape_over_3"]
    h = rf["model_interior_held_exterior_grows"]
    rows_rf = [
        ("uniform", u["fc_ghz"], 0, "cross-check only"),
        ("shrink_by_tape", s["fc_ghz"], s["delta_fc_mhz"], "worst single-cell"),
        ("shrink_by_tape/3", m["fc_ghz"], m["delta_fc_mhz"],
         "⅓-double heuristic"),
        ("interior_held", h["fc_interior_ghz"], 0,
         f"pitch +{h['pitch_growth_pct']:.2f}% — needs finite array"),
    ]
    for i, row in enumerate(rows_rf, start=2):
        for j, v in enumerate(row, start=1):
            put(ws, i, j, v, kind="fe" if j < 4 else "warn",
                fmt="0.000" if j in (2, 3) and isinstance(v, float) else None)
    for col, w in zip("ABCD", (22, 12, 14, 40)):
        ws.column_dimensions[col].width = w

    # ---- FLOQUET --------------------------------------------------------
    ws = wb.create_sheet("7 FLOQUET ARRAY")
    header(ws, 1, "lattice", "air_fraction", "fc_Gamma_GHz",
           "margin_75_MHz", "split_GHz", "air_pixels")
    if floq and floq.get("comparison", {}).get("air_pixels_removed", 0) > 0:
        for i, key in enumerate(("uniform", "one_third_double",
                                 "one_third_double_class1"), start=2):
            lat = floq[key]
            put(ws, i, 1, lat["name"], "input")
            put(ws, i, 2, lat["meta"]["air_fraction"], "fe", "0.0000")
            put(ws, i, 3, lat["fundamental_cutoff_ghz"], "fe", "0.000")
            put(ws, i, 4, lat["margin_at_75_ghz_mhz"], "fe", "0.0")
            put(ws, i, 5, lat["mode_splitting_ghz"], "fe", "0.000")
            put(ws, i, 6, lat["meta"]["air_pixels"], "fe")
        put(ws, 6, 1, "delta_fc_MHz", "calc")
        put(ws, 6, 2, floq["comparison"]["delta_fundamental_mhz"], "calc",
            "0.0")
        put(ws, 7, 1, "air_pixels_removed", "calc")
        put(ws, 7, 2, floq["comparison"]["air_pixels_removed"], "calc")
        put(ws, 8, 1, "min_Bloch_uniform_GHz", "calc")
        put(ws, 8, 2, floq["comparison"]["uniform_min_bloch_cutoff_ghz"],
            "calc", "0.000")
        put(ws, 9, 1, "min_Bloch_double_GHz", "calc")
        put(ws, 9, 2, floq["comparison"]["double_min_bloch_cutoff_ghz"],
            "calc", "0.000")
        # band table
        put(ws, 11, 1, "BANDS — Gamma first eight (GHz)", "input")
        put(ws, 12, 1, "uniform")
        for j, f in enumerate(floq["uniform"]["first_eight_gamma_ghz"], start=2):
            put(ws, 12, j, f, "fe", "0.000")
        put(ws, 13, 1, "double_class0")
        for j, f in enumerate(
                floq["one_third_double"]["first_eight_gamma_ghz"], start=2):
            put(ws, 13, j, f, "fe", "0.000")
        put(ws, 15, 1, "DISCLOSURE", "warn")
        put(ws, 15, 2,
            "2D transverse Floquet–Bloch; Neumann TE; not 3D radiating "
            "Floquet. Scan impedance / grating lobes → Vlad.", "warn")
        r = 17
        for line in floq.get("verdict", []):
            put(ws, r, 1, line, "calc")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1
    else:
        put(ws, 2, 1, "Floquet not ready — re-run floquet_hex_array.py", "warn")
    for col, w in zip("ABCDEF", (28, 14, 14, 14, 12, 12)):
        ws.column_dimensions[col].width = w

    # ---- SUMMARY --------------------------------------------------------
    ws = wb.create_sheet("8 SUMMARY")
    header(ws, 1, "Question", "Corrected answer")
    n60 = next(r for r in norms if r["gap_um"] == 60)
    answers = [
        ("Four-phase force help",
         f"Per pole @60 µm +{(n60['ba']-1)*100:.1f}% = coil count; "
         f"per coil {n60['ba_coil']:.3f}; per foot B/A="
         f"{n60['per_foot_b']/n60['per_foot_a']:.3f}"),
        ("Drop 4th tooth",
         f"Absolute −{(1-n60['cb'])*100:.1f}%; per foot C denser by "
         f"{(n60['cb_foot']-1)*100:.1f}% (Tony ~25% length-norm OK)"),
        ("MIM poles",
         "Limit vs Fe-Co; penalty grows with current; ≈ all-MIM"),
        ("Double walls",
         "Floquet: open at 75 GHz; Δfc see sheet 7; 3D still Vlad"),
        ("Primary lever", "Gap control — phase count is secondary"),
    ]
    for i, (q, a) in enumerate(answers, start=2):
        put(ws, i, 1, q, "input")
        put(ws, i, 2, a, "calc")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    path = os.path.join(OUT, f"{name_stem}.xlsx")
    wb.save(path)
    print(f"wrote {path}")
    return path


def render_pdf(md_path: str, name_stem: str) -> str:
    loader = SourceFileLoader(
        "show_md", os.path.expanduser("~/.claude/scripts/show-md"))
    import importlib.util
    spec = importlib.util.spec_from_loader("show_md", loader)
    show_md = importlib.util.module_from_spec(spec)
    loader.exec_module(show_md)
    html_path = show_md.render(Path(md_path))
    html_stamped = os.path.join(OUT, f"{name_stem}.html")
    shutil.copy2(html_path, html_stamped)
    pdf_path = os.path.join(OUT, f"{name_stem}.pdf")
    subprocess.run(
        [CHROME, "--headless", "--disable-gpu", "--no-pdf-header-footer",
         f"--print-to-pdf={pdf_path}", f"file://{html_stamped}"],
        check=True, capture_output=True, timeout=120,
    )
    print(f"wrote {pdf_path}")
    return pdf_path


def main():
    study, rf, floq = load()
    if floq and floq.get("comparison", {}).get("air_pixels_removed", 0) <= 0:
        print("WARNING: floquet artefact looks stale/broken "
              f"(air_pixels_removed="
              f"{floq.get('comparison', {}).get('air_pixels_removed')})")
    norms = force_table(study)
    md = write_md(study, rf, floq, norms)
    s = stem()
    shutil.copy2(md, os.path.join(OUT, f"{s}.md"))
    xlsx = write_xlsx(study, rf, floq, norms, s)
    pdf = render_pdf(md, s)
    for src in (pdf, xlsx):
        dl = os.path.expanduser(f"~/Downloads/{os.path.basename(src)}")
        shutil.copy2(src, dl)
        print(f"downloads {dl}")


if __name__ == "__main__":
    main()
